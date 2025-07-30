import os
from openpyxl import load_workbook
from xsd_parser_service import XSDParser
from excel_export_service import ExcelExporter

# Minimal XSD string with both elements and attributes
test_xsd = '''
<xsd:schema xmlns:xsd="http://www.w3.org/2001/XMLSchema">
  <xsd:complexType name="TestType">
    <xsd:sequence>
      <xsd:element name="Elem1" type="xsd:string"/>
      <xsd:element name="Elem2" type="xsd:string"/>
    </xsd:sequence>
    <xsd:attribute name="Attr1" type="xsd:string"/>
    <xsd:attribute name="Attr2" type="xsd:string"/>
  </xsd:complexType>
</xsd:schema>
'''

# Write XSD to temp file
with open('test_schema.xsd', 'w') as f:
    f.write(test_xsd)

# Parse XSD
parser = XSDParser()
parsed = parser.parse_xsd('test_schema.xsd')
print("Parsed rows:")
for row in parsed['TestType']:
    print(row)  # Print the full row dict for analysis
    print(f"Type: {row.get('Type')}, Levels: {row.get('levels')}")

# Export to Excel
exporter = ExcelExporter()
exporter.export({'TestType': parsed['TestType']}, 'test_output.xlsx')

# Read back Excel and print order
wb = load_workbook('test_output.xlsx')
ws = wb['TestType']
print("\nExcel output order:")
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row[6], row[:2])  # Type, first two levels

# Cleanup
os.remove('test_schema.xsd')
os.remove('test_output.xlsx') 