import openpyxl
from collections import defaultdict


def reorder_attributes_in_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    try:
        for ws in wb.worksheets:
            header_row_1 = [cell.value for cell in ws[1]]
            header_row_2 = [cell.value for cell in ws[2]]
            try:
                level_start = header_row_1.index('Level1_src')
            except ValueError:
                continue  # Not a mapping sheet
            category_col = None
            for cat_col in ['Category_src', 'Category_tgt', 'Category']:
                if cat_col in header_row_1:
                    category_col = header_row_1.index(cat_col)
                    break
            if category_col is None:
                continue
            data_rows = list(ws.iter_rows(min_row=3, values_only=True))
            if not data_rows:
                continue
            # Build a tree: key = tuple(levels), value = (row, children)
            node_map = {}
            root_nodes = []
            for idx, row in enumerate(data_rows):
                levels = row[level_start:level_start+8]
                nonempty_levels = [lvl for lvl in levels if lvl and str(lvl).strip()]
                key = tuple(nonempty_levels)
                node_map[key] = {'row': row, 'children': []}
            # Assign children to parents
            for key in node_map:
                if len(key) == 0:
                    root_nodes.append(key)
                    continue
                parent_key = key[:-1]
                if parent_key in node_map:
                    node_map[parent_key]['children'].append(key)
                else:
                    root_nodes.append(key)  # Orphaned node
            # Recursive function to flatten tree with correct order
            def flatten_tree(node_key):
                node = node_map[node_key]
                result = [node['row']]
                # Reorder children: attributes first, then elements/messages, preserving their order
                attr_keys = []
                elem_keys = []
                for child_key in node['children']:
                    cat_val = str(node_map[child_key]['row'][category_col]).strip().lower() if node_map[child_key]['row'][category_col] else ''
                    if cat_val == 'attribute':
                        attr_keys.append(child_key)
                    else:
                        elem_keys.append(child_key)
                # Keep the original order among attributes and among elements/messages
                for child_key in attr_keys + elem_keys:
                    result.extend(flatten_tree(child_key))
                return result
            # Find all top-level nodes (those with no parent)
            top_level_keys = [key for key in node_map if len(key) == 1 or key not in [child for n in node_map.values() for child in n['children']]]
            # If no top-level nodes found, fallback to all root_nodes
            if not top_level_keys:
                top_level_keys = root_nodes
            # Flatten the tree
            new_data = []
            for key in sorted(top_level_keys, key=lambda k: data_rows.index(node_map[k]['row'])):
                new_data.extend(flatten_tree(key))
            # Clear old data rows
            ws.delete_rows(3, ws.max_row-2)
            for row in new_data:
                ws.append(row)
        wb.save(excel_path)
    finally:
        wb.close()

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 2:
        print("Usage: python reorder_excel_attributes.py <excel_path>")
        sys.exit(1)
    reorder_attributes_in_excel(sys.argv[1]) 