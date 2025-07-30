import json
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re
from pathlib import Path

# --- Heurística de correspondência ---
try:
    from Levenshtein import ratio as levenshtein_ratio
except ImportError:
    def levenshtein_ratio(a, b):
        # Fallback implementation using simple string similarity
        if not a or not b:
            return 0.0
        
        # Convert to lowercase for better matching
        a_lower = a.lower()
        b_lower = b.lower()
        
        # Exact match
        if a_lower == b_lower:
            return 1.0
        
        # Check for common prefixes
        min_len = min(len(a_lower), len(b_lower))
        common_prefix = 0
        for i in range(min_len):
            if a_lower[i] == b_lower[i]:
                common_prefix += 1
            else:
                break
        
        # Check for common suffixes
        common_suffix = 0
        for i in range(1, min_len + 1):
            if a_lower[-i] == b_lower[-i]:
                common_suffix += 1
            else:
                break
        
        # Calculate similarity based on common parts
        total_common = common_prefix + common_suffix
        max_len = max(len(a_lower), len(b_lower))
        
        if max_len == 0:
            return 0.0
        
        # Weight the similarity (prefix is more important than suffix)
        similarity = (common_prefix * 0.7 + common_suffix * 0.3) / max_len
        
        # Add bonus for length similarity
        length_diff = abs(len(a_lower) - len(b_lower))
        length_penalty = length_diff / max_len * 0.2
        
        return max(0.0, similarity - length_penalty)

def normalize_levels(levels):
    # Lower-case, replace 'item' and '[]' with 'ARRAYITEM'
    return [
        'arrayitem' if (isinstance(l, str) and l.lower() in ('item', '[]')) else (l.lower() if isinstance(l, str) else l)
        for l in levels
    ]

def normalized_path_from_levels(levels):
    return '.'.join([
        'arrayitem' if (isinstance(l, str) and l.lower() in ('item', '[]')) else (l.lower() if isinstance(l, str) else l)
        for l in levels
    ])

def map_paths(source_fields, target_fields, threshold=0.7):
    # Build normalized path lists
    source_paths = [normalized_path_from_levels(f['levels']) for f in source_fields]
    target_paths = [normalized_path_from_levels(f['levels']) for f in target_fields]
    mapping = []
    for s, s_field in zip(source_paths, source_fields):
        # Exact match
        for t, t_field in zip(target_paths, target_fields):
            if s == t:
                mapping.append({'source': s, 'target': t, 'similarity': 1.0})
                break
        else:
            # Fuzzy match
            best_match = None
            best_ratio = 0
            for t, t_field in zip(target_paths, target_fields):
                ratio = levenshtein_ratio(s, t)
                if ratio > best_ratio and ratio >= threshold:
                    best_ratio = ratio
                    best_match = t
            if best_match:
                mapping.append({'source': s, 'target': best_match, 'similarity': best_ratio})
            else:
                mapping.append({'source': s, 'target': '', 'similarity': 0.0})
    return mapping

def extract_fields_from_json_schema(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    fields = []
    
    def walk(obj, levels=None, parent_required=None):
        if levels is None:
            levels = []
        if parent_required is None:
            parent_required = []
        
        if isinstance(obj, dict):
            properties = obj.get('properties', {})
            required = obj.get('required', [])
            
            for prop_name, prop_obj in properties.items():
                current_levels = levels + [prop_name]
                current_required = parent_required + required
                
                # Determine if this field is required
                is_required = prop_name in required
                
                # Get field type
                field_type = prop_obj.get('type', 'string')
                if 'array' in field_type:
                    field_type = 'array'
                elif 'object' in field_type:
                    field_type = 'object'
                
                # Add field
                fields.append({
                    'levels': current_levels,
                    'type': field_type,
                    'required': is_required,
                    'description': prop_obj.get('description', '')
                })
                
                # Recursively process nested objects
                if prop_obj.get('type') == 'object' or 'properties' in prop_obj:
                    walk(prop_obj, current_levels, current_required)
                elif prop_obj.get('type') == 'array' and 'items' in prop_obj:
                    # Handle array items
                    items_obj = prop_obj['items']
                    if isinstance(items_obj, dict) and (items_obj.get('type') == 'object' or 'properties' in items_obj):
                        walk(items_obj, current_levels + ['[]'], current_required)
    
    walk(data)
    return fields

def extract_xsd_restrictions(restriction):
    if restriction is None:
        return {}
    restrictions = {}
    for child in restriction:
        if child.tag.endswith('}enumeration'):
            if 'enumerations' not in restrictions:
                restrictions['enumerations'] = []
            restrictions['enumerations'].append(child.get('value'))
        elif child.tag.endswith('}minLength'):
            restrictions['minLength'] = int(child.get('value'))
        elif child.tag.endswith('}maxLength'):
            restrictions['maxLength'] = int(child.get('value'))
        elif child.tag.endswith('}pattern'):
            restrictions['pattern'] = child.get('value')
    return restrictions

def extract_fields_from_xsd(filepath, keep_case=False):
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    fields = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    
    def to_camel_case(s):
        if not s:
            return s
        return s[0].lower() + s[1:]
    
    def get_name(name):
        return name if keep_case else to_camel_case(name)
    
    def get_type_and_details(element):
        typ = element.get('type', '')
        if typ.startswith('xs:'):
            return typ[3:], {}
        elif typ in complex_types:
            return 'complex', {}
        else:
            return 'string', {}
    
    def walk_element(element, levels=None, parent_is_array=False, is_root=False):
        if levels is None:
            levels = []
        
        name = element.get('name', '')
        if not name:
            return
        
        field_name = get_name(name)
        current_levels = levels + [field_name]
        
        if not is_root:
            typ, details = get_type_and_details(element)
            
            # Check if required (not nullable)
            nullable = element.get('nillable', 'false').lower() == 'true'
            is_required = not nullable
            
            fields.append({
                'levels': current_levels,
                'type': typ,
                'required': is_required,
                'description': element.get('annotation/xs:documentation', '')
            })
        
        # Handle complex types
        typ = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        
        if typ in complex_types:
            walk_complex_type(complex_types[typ], current_levels)
        if complex_type is not None:
            walk_complex_type(complex_type, current_levels)
    
    def walk_complex_type(complex_type, levels):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, levels)
    
    for element in root.findall('xs:element', ns):
        walk_element(element, [], False, is_root=True)
    
    return fields

def fields_to_row(fields, max_levels):
    rows = []
    for field in fields:
        row = []
        for i in range(max_levels):
            if i < len(field['levels']):
                row.append(field['levels'][i])
            else:
                row.append('')
        row.extend([field['type'], field['required'], field.get('description', '')])
        rows.append(row)
    return rows

def get_max_levels(fields):
    return max(len(f['levels']) for f in fields) if fields else 0

def build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path):
    def get_req_param(fields, file_name):
        req_fields = [f for f in fields if f['required']]
        req_params = []
        for field in req_fields:
            param = '.'.join(field['levels'])
            req_params.append(f"{file_name}.{param}")
        return req_params
    
    def remove_doubled(levels):
        result = []
        for level in levels:
            if not result or result[-1] != level:
                result.append(level)
        return result
    
    def get_target_hierarchy_row(tgt, tgt_max_levels, prev_tgt_levels):
        if not tgt:
            return [''] * tgt_max_levels + ['', '', '']
        
        tgt_levels = tgt.split('.')
        tgt_levels = remove_doubled(tgt_levels)
        
        row = []
        for i in range(tgt_max_levels):
            if i < len(tgt_levels):
                row.append(tgt_levels[i])
            else:
                row.append('')
        
        # Add type, required, and description columns
        row.extend(['', '', ''])
        return row
    
    src_max_levels = get_max_levels(source_fields)
    tgt_max_levels = get_max_levels(target_fields)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"
    
    # Headers
    headers = []
    for i in range(src_max_levels):
        headers.append(f"Source Level {i+1}")
    for i in range(tgt_max_levels):
        headers.append(f"Target Level {i+1}")
    headers.extend(["Source Type", "Source Required", "Source Description"])
    headers.extend(["Target Type", "Target Required", "Target Description"])
    headers.extend(["Mapping", "Similarity"])
    
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data rows
    prev_tgt_levels = []
    for i, (src_field, map_item) in enumerate(zip(source_fields, mapping)):
        row = []
        
        # Source levels
        for j in range(src_max_levels):
            if j < len(src_field['levels']):
                row.append(src_field['levels'][j])
            else:
                row.append('')
        
        # Target levels (from mapping)
        tgt_levels = map_item['target'].split('.') if map_item['target'] else []
        tgt_levels = remove_doubled(tgt_levels)
        for j in range(tgt_max_levels):
            if j < len(tgt_levels):
                row.append(tgt_levels[j])
            else:
                row.append('')
        
        # Source metadata
        row.extend([src_field['type'], src_field['required'], src_field.get('description', '')])
        
        # Target metadata (find matching target field)
        tgt_field = None
        if map_item['target']:
            for tf in target_fields:
                if '.'.join(tf['levels']) == map_item['target']:
                    tgt_field = tf
                    break
        
        if tgt_field:
            row.extend([tgt_field['type'], tgt_field['required'], tgt_field.get('description', '')])
        else:
            row.extend(['', '', ''])
        
        # Mapping info
        row.extend([map_item['target'], map_item['similarity']])
        
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)

def write_mapping_excel(mapping, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"
    
    # Headers
    headers = ["Source", "Target", "Similarity"]
    ws.append(headers)
    
    # Data
    for item in mapping:
        ws.append([item['source'], item['target'], item['similarity']])
    
    wb.save(output_path)

def run_schema_to_excel_operation(source_path, dest_dir):
    """Convert schema to Excel format"""
    source_path = Path(source_path)
    dest_dir = Path(dest_dir)
    
    if source_path.suffix.lower() == '.json':
        fields = extract_fields_from_json_schema(str(source_path))
    elif source_path.suffix.lower() in ['.xsd', '.xml']:
        fields = extract_fields_from_xsd(str(source_path))
    else:
        raise ValueError(f"Unsupported file format: {source_path.suffix}")
    
    max_levels = get_max_levels(fields)
    rows = fields_to_row(fields, max_levels)
    
    # Create headers
    headers = []
    for i in range(max_levels):
        headers.append(f"Level {i+1}")
    headers.extend(["Type", "Required", "Description"])
    
    # Write Excel file
    output_file = dest_dir / f"{source_path.stem}_schema.xlsx"
    write_excel(rows, str(output_file), headers)

def write_excel(rows, output_file, columns):
    """Write data to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schema"
    
    # Add headers
    ws.append(columns)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    for row in rows:
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)

def run_xsd_to_jsonschema_operation(source_path, dest_dir):
    """Convert XSD to JSON Schema"""
    source_path = Path(source_path)
    dest_dir = Path(dest_dir)
    
    tree = ET.parse(source_path)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    
    def xsd_type_to_json_type(xsd_type):
        type_mapping = {
            'string': 'string',
            'integer': 'integer',
            'int': 'integer',
            'long': 'integer',
            'short': 'integer',
            'byte': 'integer',
            'decimal': 'number',
            'float': 'number',
            'double': 'number',
            'boolean': 'boolean',
            'date': 'string',
            'dateTime': 'string',
            'time': 'string',
            'base64Binary': 'string',
            'hexBinary': 'string',
            'anyURI': 'string',
            'QName': 'string',
            'NOTATION': 'string'
        }
        return type_mapping.get(xsd_type, 'string')
    
    def extract_restrictions(element):
        restrictions = {}
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            if restriction is not None:
                base_type = restriction.get('base', '')
                if base_type.startswith('xs:'):
                    base_type = base_type[3:]
                restrictions['type'] = xsd_type_to_json_type(base_type)
                restrictions.update(extract_xsd_restrictions(restriction))
        return restrictions
    
    def parse_element(element):
        name = element.get('name', '')
        if not name:
            return None
        
        result = {}
        
        # Check for simple type restrictions
        restrictions = extract_restrictions(element)
        if restrictions:
            result.update(restrictions)
        else:
            # Check type attribute
            typ = element.get('type', '')
            if typ.startswith('xs:'):
                xsd_type = typ[3:]
                result['type'] = xsd_type_to_json_type(xsd_type)
            elif typ:
                # Reference to another type
                result['$ref'] = f"#/definitions/{typ}"
            else:
                # Complex type
                complex_type = element.find('xs:complexType', ns)
                if complex_type is not None:
                    result.update(parse_complex_type(complex_type))
                else:
                    result['type'] = 'string'
        
        # Add description if available
        annotation = element.find('xs:annotation', ns)
        if annotation is not None:
            documentation = annotation.find('xs:documentation', ns)
            if documentation is not None and documentation.text:
                result['description'] = documentation.text.strip()
        
        return result
    
    def parse_complex_type(complex_type):
        result = {'type': 'object'}
        properties = {}
        required = []
        
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                name = child.get('name', '')
                if name:
                    child_result = parse_element(child)
                    if child_result:
                        properties[name] = child_result
                        # Check if required (not nullable)
                        if child.get('nillable', 'false').lower() != 'true':
                            required.append(name)
        
        if properties:
            result['properties'] = properties
        if required:
            result['required'] = required
        
        return result
    
    def camelize_json_schema(obj):
        """Convert property names to camelCase"""
        if isinstance(obj, dict):
            result = {}
            for key, value in obj.items():
                if key == 'properties':
                    # Camelize property names
                    camelized_props = {}
                    for prop_name, prop_value in value.items():
                        camel_name = prop_name[0].lower() + prop_name[1:] if prop_name else prop_name
                        camelized_props[camel_name] = camelize_json_schema(prop_value)
                    result[key] = camelized_props
                else:
                    result[key] = camelize_json_schema(value)
            return result
        elif isinstance(obj, list):
            return [camelize_json_schema(item) for item in obj]
        else:
            return obj
    
    # Parse root elements
    schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        "properties": {},
        "required": []
    }
    
    definitions = {}
    
    for element in root.findall('xs:element', ns):
        name = element.get('name', '')
        if name:
            element_schema = parse_element(element)
            if element_schema:
                schema['properties'][name] = element_schema
                # Check if required
                if element.get('nillable', 'false').lower() != 'true':
                    schema['required'].append(name)
    
    # Add complex type definitions
    for complex_type in root.findall('xs:complexType', ns):
        name = complex_type.get('name', '')
        if name:
            definitions[name] = parse_complex_type(complex_type)
    
    if definitions:
        schema['definitions'] = definitions
    
    # Camelize the schema
    schema = camelize_json_schema(schema)
    
    # Write output
    output_file = dest_dir / f"{source_path.stem}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(schema, f, indent=2, ensure_ascii=False)

def run_jsonschema_to_xsd_operation(json_schema_path, dest_dir):
    """Convert JSON Schema to XSD"""
    json_schema_path = Path(json_schema_path)
    dest_dir = Path(dest_dir)
    
    with open(json_schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    
    def to_pascal_case(s):
        if not s:
            return s
        return s[0].upper() + s[1:]
    
    def parse_comment_restrictions(comment):
        """Parse restrictions from comment field"""
        restrictions = {}
        if not comment:
            return restrictions
        
        # Simple pattern matching for common restrictions
        if 'minLength:' in comment:
            match = re.search(r'minLength:\s*(\d+)', comment)
            if match:
                restrictions['minLength'] = int(match.group(1))
        
        if 'maxLength:' in comment:
            match = re.search(r'maxLength:\s*(\d+)', comment)
            if match:
                restrictions['maxLength'] = int(match.group(1))
        
        if 'pattern:' in comment:
            match = re.search(r'pattern:\s*([^\s]+)', comment)
            if match:
                restrictions['pattern'] = match.group(1)
        
        return restrictions
    
    def gen_xsd_element(name, obj, required_fields=None, is_root=False, root_type=None):
        if required_fields is None:
            required_fields = []
        
        element = ET.Element('xs:element')
        element.set('name', name)
        
        if is_root and root_type:
            element.set('type', root_type)
            return element
        
        obj_type = obj.get('type', 'string')
        
        if obj_type == 'array':
            # Handle array type
            element.set('name', name)
            complex_type = ET.SubElement(element, 'xs:complexType')
            sequence = ET.SubElement(complex_type, 'xs:sequence')
            
            items = obj.get('items', {})
            if items.get('type') == 'object':
                # Array of objects
                item_element = ET.SubElement(sequence, 'xs:element')
                item_element.set('name', 'item')
                item_element.set('type', f"{to_pascal_case(name)}Item")
            else:
                # Array of primitives
                item_element = ET.SubElement(sequence, 'xs:element')
                item_element.set('name', 'item')
                item_element.set('type', 'xs:string')  # Default to string
        
        elif obj_type == 'object':
            # Handle object type
            element.set('name', name)
            complex_type = ET.SubElement(element, 'xs:complexType')
            sequence = ET.SubElement(complex_type, 'xs:sequence')
            
            properties = obj.get('properties', {})
            for prop_name, prop_obj in properties.items():
                prop_element = gen_xsd_element(prop_name, prop_obj, required_fields)
                sequence.append(prop_element)
        
        else:
            # Handle primitive types
            element.set('name', name)
            
            # Check if required
            if name in required_fields:
                element.set('nillable', 'false')
            else:
                element.set('nillable', 'true')
            
            # Add type
            type_mapping = {
                'string': 'xs:string',
                'integer': 'xs:integer',
                'number': 'xs:decimal',
                'boolean': 'xs:boolean',
                'date': 'xs:date',
                'dateTime': 'xs:dateTime',
                'time': 'xs:time'
            }
            xsd_type = type_mapping.get(obj_type, 'xs:string')
            element.set('type', xsd_type)
            
            # Add restrictions if any
            restrictions = parse_comment_restrictions(obj.get('description', ''))
            if restrictions:
                simple_type = ET.SubElement(element, 'xs:simpleType')
                restriction = ET.SubElement(simple_type, 'xs:restriction')
                restriction.set('base', xsd_type)
                
                for restriction_type, value in restrictions.items():
                    if restriction_type == 'minLength':
                        min_length = ET.SubElement(restriction, 'xs:minLength')
                        min_length.set('value', str(value))
                    elif restriction_type == 'maxLength':
                        max_length = ET.SubElement(restriction, 'xs:maxLength')
                        max_length.set('value', str(value))
                    elif restriction_type == 'pattern':
                        pattern = ET.SubElement(restriction, 'xs:pattern')
                        pattern.set('value', value)
        
        # Add documentation if available
        if obj.get('description'):
            annotation = ET.SubElement(element, 'xs:annotation')
            documentation = ET.SubElement(annotation, 'xs:documentation')
            documentation.text = obj['description']
        
        return element
    
    def gen_flat_root_element(title, obj, required_fields=None):
        """Generate a flat root element structure"""
        schema = ET.Element('xs:schema')
        schema.set('xmlns:xs', 'http://www.w3.org/2001/XMLSchema')
        
        # Create root element
        root_element = ET.SubElement(schema, 'xs:element')
        root_element.set('name', title)
        
        # Create complex type for root
        root_type = ET.SubElement(schema, 'xs:complexType')
        root_type.set('name', to_pascal_case(title))
        
        sequence = ET.SubElement(root_type, 'xs:sequence')
        
        # Add properties as elements
        properties = obj.get('properties', {})
        for prop_name, prop_obj in properties.items():
            prop_element = gen_xsd_element(prop_name, prop_obj, required_fields)
            sequence.append(prop_element)
        
        # Set root element type
        root_element.set('type', to_pascal_case(title))
        
        return schema
    
    # Generate XSD
    title = json_schema_path.stem
    required_fields = schema.get('required', [])
    
    xsd_schema = gen_flat_root_element(title, schema, required_fields)
    
    # Write output
    output_file = dest_dir / f"{title}.xsd"
    tree = ET.ElementTree(xsd_schema)
    tree.write(output_file, encoding='utf-8', xml_declaration=True) 