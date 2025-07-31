import openpyxl
from openpyxl.styles import Font

class ExcelGenerator:
    def __init__(self, max_level=8):
        self.max_level = max_level

    def excel_sheet_name(self, name):
        return name[:31]

    def export(self, xsd_data_dict, output_file):
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        sheet_count = 0
        for sheet_name, rows in xsd_data_dict.items():
            ws = wb.create_sheet(title=self.excel_sheet_name(sheet_name))
            headers = [f'Level{i+1}' for i in range(self.max_level)] + ['Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Base Type', 'Details', 'Description', 'Category', 'Example']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            # Sort rows: attributes first, then elements
            rows = sorted(rows, key=lambda r: 0 if r.get('Category') == 'attribute' else 1)
            for row in rows:
                ws.append(row['levels'] + [row['Request Parameter'], row['GDPR'], row['Cardinality'], row['Type'], row['Base Type'], row['Details'], row['Description'], row['Category'], row['Example']])
            sheet_count += 1
        if sheet_count > 0:
            wb.remove(default_sheet)
            wb.save(output_file)
            return True
        else:
            return False

    def create_mapping_excel(self, source_fields, target_fields, mapping, output_file):
        """Create Excel file with field mapping"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Field Mapping"
        
        # Headers for mapping
        headers = [
            "Source Field", "Source Type", "Source Path", "Target Field", 
            "Target Type", "Target Path", "Similarity", "Notes"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write mapping data
        row = 2
        for map_item in mapping:
            cells = [
                map_item.get('source', ''),
                map_item.get('source_type', ''),
                map_item.get('source_path', ''),
                map_item.get('target', ''),
                map_item.get('target_type', ''),
                map_item.get('target_path', ''),
                map_item.get('similarity', ''),
                map_item.get('notes', '')
            ]
            
            for col, value in enumerate(cells, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        wb.save(output_file)
        return output_file 