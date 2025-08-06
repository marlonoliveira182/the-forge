import streamlit as st
import os
import tempfile
import json
import xml.etree.ElementTree as ET
from io import BytesIO

import openpyxl
import difflib
from openpyxl.utils import get_column_letter

# Import modern UI libraries

from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode

# Import the microservices
from services.xsd_parser_service import XSDParser
from services.json_schema_parser_service import JSONSchemaParser
from services.excel_export_service import ExcelExporter
from services.wsdl_to_xsd_extractor import merge_xsd_from_wsdl
from services.excel_mapping_service import ExcelMappingService

from services.case_converter_service import pascal_to_camel, camel_to_pascal

from services.converter_service import ConverterService

# Import homepage
from homepage import show_home_page

# Page configuration
st.set_page_config(
    page_title="The Forge - Schema Transformation Tool",
    page_icon="🔨",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com/your-repo/the-forge',
        'Report a bug': 'https://github.com/your-repo/the-forge/issues',
        'About': 'The Forge - Advanced Schema Transformation Tool'
    }
)

# Forge CSS Theme
st.markdown("""
<style>
    /* Forge Color Palette */
    :root {
        --forge-black: #0a0a0a;
        --forge-dark: #1a1a1a;
        --forge-charcoal: #2d2d2d;
        --forge-steel: #3a3a3a;
        --forge-iron: #4a4a4a;
        --forge-coal: #5a5a5a;
        
        --forge-orange: #ff6b35;
        --forge-fire: #ff8c42;
        --forge-flame: #ffa726;
        --forge-ember: #ff7043;
        --forge-glow: #ff5722;
        
        --forge-gold: #ffd700;
        --forge-bronze: #cd7f32;
        --forge-copper: #b87333;
        
        --forge-text: #e0e0e0;
        --forge-text-secondary: #b0b0b0;
        --forge-text-muted: #808080;
        
        --forge-border: #404040;
        --forge-border-light: #505050;
        
        --forge-shadow: 0 4px 8px rgba(0, 0, 0, 0.5);
        --forge-shadow-heavy: 0 8px 16px rgba(0, 0, 0, 0.7);
        --forge-glow-shadow: 0 0 20px rgba(255, 107, 53, 0.3);
        
        --radius-sm: 4px;
        --radius-md: 8px;
        --radius-lg: 12px;
        --radius-xl: 16px;
    }
    
    /* Global Styles */
    .stApp {
        background: linear-gradient(135deg, var(--forge-black) 0%, var(--forge-dark) 100%);
        color: var(--forge-text);
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    /* Main Content Area */
    .main .block-container {
        background: linear-gradient(135deg, var(--forge-dark) 0%, var(--forge-charcoal) 100%);
        color: var(--forge-text);
        padding: 2rem;
        border-radius: var(--radius-lg);
        margin: 1rem;
        box-shadow: var(--forge-shadow);
    }
    
    /* Forge Header */
    .main-header {
        background: linear-gradient(135deg, var(--forge-charcoal) 0%, var(--forge-steel) 100%);
        color: var(--forge-text);
        padding: 3rem 2rem;
        border-radius: var(--radius-xl);
        text-align: center;
        margin-bottom: 2rem;
        border: 2px solid var(--forge-orange);
        box-shadow: var(--forge-shadow-heavy), var(--forge-glow-shadow);
        position: relative;
        overflow: hidden;
    }
    
    .main-header::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><defs><pattern id="forge" width="20" height="20" patternUnits="userSpaceOnUse"><path d="M 20 0 L 0 0 0 20" fill="none" stroke="rgba(255,107,53,0.1)" stroke-width="1"/></pattern></defs><rect width="100" height="100" fill="url(%23forge)"/></svg>');
        opacity: 0.5;
    }
    
    .main-header h1 {
        color: var(--forge-orange);
        margin-bottom: 0.5rem;
        font-size: 2.5rem;
        font-weight: 700;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.8);
        position: relative;
        z-index: 1;
    }
    
    .main-header p {
        color: var(--forge-text);
        font-size: 1.1rem;
        margin-top: 0.5rem;
        position: relative;
        z-index: 1;
    }
    
    /* Ancient Forge Sidebar */
    .css-1d391kg {
        background: linear-gradient(180deg, var(--forge-black) 0%, var(--forge-dark) 100%);
        border-right: 2px solid var(--forge-orange);
        box-shadow: var(--forge-shadow-heavy);
    }
    
    /* Sidebar Header */
    .sidebar-header {
        background: linear-gradient(135deg, var(--forge-charcoal) 0%, var(--forge-steel) 100%);
        color: var(--forge-text);
        padding: 2rem 1.5rem;
        margin: -1rem -1rem 2rem -1rem;
        text-align: center;
        border-bottom: 2px solid var(--forge-orange);
        box-shadow: var(--forge-shadow);
        position: relative;
    }
    
    .sidebar-header h2 {
        margin: 0;
        font-size: 1.4rem;
        font-weight: 700;
        color: var(--forge-orange);
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.8);
    }
    
    /* Navigation Buttons */
    .sidebar .stButton > button {
        background: linear-gradient(135deg, var(--forge-iron) 0%, var(--forge-coal) 100%);
        color: var(--forge-text);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-md);
        padding: 1rem 1.25rem;
        font-weight: 500;
        font-size: 0.95rem;
        transition: all 0.3s ease;
        margin: 0.5rem 0;
        width: 100%;
        text-align: left;
        position: relative;
        box-shadow: var(--forge-shadow);
    }
    
    .sidebar .stButton > button:hover {
        background: linear-gradient(135deg, var(--forge-orange) 0%, var(--forge-fire) 100%);
        color: var(--forge-text);
        transform: translateX(4px);
        box-shadow: var(--forge-shadow), var(--forge-glow-shadow);
        border-color: var(--forge-orange);
    }
    
    /* Active Navigation Button */
    .sidebar .stButton > button[data-active="true"] {
        background: linear-gradient(135deg, var(--forge-orange) 0%, var(--forge-fire) 100%);
        color: var(--forge-text);
        font-weight: 600;
        box-shadow: var(--forge-shadow), var(--forge-glow-shadow);
        border-color: var(--forge-orange);
        transform: translateX(2px);
    }
    
    /* Sidebar Footer */
    .sidebar-footer {
        margin-top: auto;
        padding: 1.5rem 1rem;
        text-align: center;
        color: var(--forge-text-secondary);
        font-size: 0.8rem;
        border-top: 1px solid var(--forge-border);
        background: linear-gradient(135deg, var(--forge-charcoal) 0%, var(--forge-steel) 100%);
        margin: 2rem -1rem -1rem -1rem;
    }
    
    .sidebar-footer .version {
        background: linear-gradient(135deg, var(--forge-orange) 0%, var(--forge-fire) 100%);
        padding: 0.5rem 1rem;
        border-radius: var(--radius-md);
        margin-top: 0.75rem;
        font-family: 'Courier New', monospace;
        font-size: 0.75rem;
        color: var(--forge-text);
        border: 1px solid var(--forge-orange);
        box-shadow: var(--forge-shadow);
    }
    
    /* Section Headers */
    .section-header {
        background: linear-gradient(135deg, var(--forge-charcoal) 0%, var(--forge-steel) 100%);
        color: var(--forge-text);
        padding: 1.5rem 2rem;
        border-radius: var(--radius-lg);
        margin: 2rem 0 1.5rem 0;
        border: 1px solid var(--forge-border);
        box-shadow: var(--forge-shadow);
    }
    
    .section-header h2, .section-header h3 {
        color: var(--forge-orange);
        margin: 0;
        font-weight: 600;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.8);
    }
    
    /* Feature Cards */
    .feature-card {
        background: linear-gradient(135deg, var(--forge-charcoal) 0%, var(--forge-steel) 100%);
        padding: 2rem;
        border-radius: var(--radius-lg);
        border: 1px solid var(--forge-border);
        margin: 1.5rem 0;
        color: var(--forge-text);
        box-shadow: var(--forge-shadow);
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .feature-card:hover {
        transform: translateY(-4px);
        box-shadow: var(--forge-shadow-heavy), var(--forge-glow-shadow);
        border-color: var(--forge-orange);
    }
    
    .feature-card h3 {
        color: var(--forge-orange);
        margin-bottom: 1rem;
        font-weight: 600;
        font-size: 1.3rem;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.8);
    }
    
    .feature-card p {
        color: var(--forge-text-secondary);
        line-height: 1.6;
        margin-bottom: 1rem;
    }
    
    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, var(--forge-orange) 0%, var(--forge-fire) 100%);
        color: var(--forge-text);
        border: none;
        border-radius: var(--radius-md);
        padding: 12px 24px;
        font-weight: 500;
        transition: all 0.2s ease;
        box-shadow: var(--forge-shadow);
    }
    
    .stButton > button:hover {
        background: linear-gradient(135deg, var(--forge-fire) 0%, var(--forge-flame) 100%);
        transform: translateY(-2px);
        box-shadow: var(--forge-shadow), var(--forge-glow-shadow);
    }
    
    /* Upload Area */
    .upload-area {
        border: 2px dashed var(--forge-orange);
        border-radius: var(--radius-lg);
        padding: 3rem 2rem;
        text-align: center;
        background: linear-gradient(135deg, rgba(255, 107, 53, 0.1) 0%, rgba(255, 140, 66, 0.1) 100%);
        margin: 1.5rem 0;
        transition: all 0.3s ease;
    }
    
    .upload-area:hover {
        border-color: var(--forge-fire);
        background: linear-gradient(135deg, rgba(255, 107, 53, 0.2) 0%, rgba(255, 140, 66, 0.2) 100%);
        transform: translateY(-2px);
        box-shadow: var(--forge-shadow), var(--forge-glow-shadow);
    }
    
    /* Message Styles */
    .success-message {
        background: linear-gradient(135deg, var(--forge-gold) 0%, var(--forge-bronze) 100%);
        color: var(--forge-black);
        padding: 1.5rem;
        border-radius: var(--radius-lg);
        margin: 1.5rem 0;
        border: 1px solid var(--forge-gold);
        box-shadow: var(--forge-shadow);
    }
    
    .error-message {
        background: linear-gradient(135deg, var(--forge-ember) 0%, var(--forge-glow) 100%);
        color: var(--forge-text);
        padding: 1.5rem;
        border-radius: var(--radius-lg);
        margin: 1.5rem 0;
        border: 1px solid var(--forge-ember);
        box-shadow: var(--forge-shadow);
    }
    
    .warning-message {
        background: linear-gradient(135deg, var(--forge-copper) 0%, var(--forge-bronze) 100%);
        color: var(--forge-text);
        padding: 1.5rem;
        border-radius: var(--radius-lg);
        margin: 1.5rem 0;
        border: 1px solid var(--forge-copper);
        box-shadow: var(--forge-shadow);
    }
    
    /* Form Elements */
    .stSelectbox, .stSlider, .stCheckbox {
        background: linear-gradient(135deg, var(--forge-iron) 0%, var(--forge-coal) 100%);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-md);
        box-shadow: var(--forge-shadow);
        color: var(--forge-text);
    }
    
    .stFileUploader {
        background: linear-gradient(135deg, var(--forge-iron) 0%, var(--forge-coal) 100%);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-lg);
        box-shadow: var(--forge-shadow);
        padding: 1rem;
    }
    
    /* Code blocks */
    .stCode {
        background: linear-gradient(135deg, var(--forge-black) 0%, var(--forge-dark) 100%);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-md);
        box-shadow: var(--forge-shadow);
        color: var(--forge-text);
    }
    
    /* DataFrames */
    .dataframe {
        background: linear-gradient(135deg, var(--forge-iron) 0%, var(--forge-coal) 100%);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-md);
        box-shadow: var(--forge-shadow);
    }
    
    /* Progress Bars */
    .stProgress > div > div > div > div {
        background: linear-gradient(135deg, var(--forge-orange) 0%, var(--forge-fire) 100%);
        border-radius: var(--radius-sm);
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, var(--forge-iron) 0%, var(--forge-coal) 100%);
        border: 1px solid var(--forge-border);
        border-radius: var(--radius-md);
        padding: 1rem;
        font-weight: 500;
        color: var(--forge-text);
    }
    
    /* Spinner */
    .stSpinner > div {
        border-color: var(--forge-orange);
    }
    

    
    /* Responsive Design */
    @media (max-width: 768px) {
        .main-header {
            padding: 2rem 1rem;
        }
        
        .main-header h1 {
            font-size: 2rem;
        }
        
        .section-header {
            padding: 1rem;
        }
        
        .feature-card {
            padding: 1.5rem;
        }
    }
</style>
""", unsafe_allow_html=True)

# Initialize services with caching
@st.cache_resource
def get_services():
    try:
        services = {
            'xsd_parser': XSDParser(),
            'json_schema_parser': JSONSchemaParser(),
            'excel_exporter': ExcelExporter(),
            'mapping_service': ExcelMappingService(),
            'converter': ConverterService()
        }
        return services
    except Exception as e:
        st.error(f"Error initializing services: {str(e)}")
        # Return empty services dict as fallback
        return {}

def main():
    # Header
    st.markdown('''
        <div class="main-header">
            <h1>🔨 The Forge</h1>
            <p>Schema Transformation & Mapping Tool</p>
        </div>
    ''', unsafe_allow_html=True)
    
    # Get services
    try:
        services = get_services()
        if not services:
            st.error("Failed to initialize services. Please refresh the page.")
            st.stop()
    except Exception as e:
        st.error(f"Error getting services: {str(e)}")
        st.stop()
    
    # Sidebar Header
    st.sidebar.markdown("""
        <div class="sidebar-header">
            <h2>🔨 The Forge</h2>
        </div>
    """, unsafe_allow_html=True)
    
    # Initialize session state if not exists
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "Home"
    
    # Navigation Menu with unique keys
    if st.sidebar.button("🏠 Home", key="nav_home", use_container_width=True):
        st.session_state.current_page = "Home"
    
    if st.sidebar.button("📊 Schema Mapping", key="nav_mapping", use_container_width=True):
        st.session_state.current_page = "Schema Mapping"
    
    if st.sidebar.button("🔧 WSDL to XSD", key="nav_wsdl", use_container_width=True):
        st.session_state.current_page = "WSDL to XSD"
    

    

    
    if st.sidebar.button("🔄 Converter", key="nav_converter", use_container_width=True):
        st.session_state.current_page = "Converter"
    
    if st.sidebar.button("ℹ️ About", key="nav_about", use_container_width=True):
        st.session_state.current_page = "About"
    
    # Sidebar Footer
    st.sidebar.markdown("""
        <div class="sidebar-footer">
            <div>Forged with Streamlit</div>
            <div class="version">v1.0.0</div>
        </div>
    """, unsafe_allow_html=True)
    
    # Display current page
    if st.session_state.current_page == "Home":
        show_home_page()
    elif st.session_state.current_page == "Schema Mapping":
        show_mapping_page(services)
    elif st.session_state.current_page == "WSDL to XSD":
        show_wsdl_to_xsd_page(services)


    elif st.session_state.current_page == "Converter":
        show_converter_page(services)
    elif st.session_state.current_page == "About":
        show_about_page()



def show_mapping_page(services):
    st.markdown('<div class="section-header"><h2>📊 Schema Mapping</h2></div>', unsafe_allow_html=True)
    
    st.markdown("""
    Create field mappings between different schema formats. Upload source and target schema files (XSD, JSON Schema, or JSON Examples) to generate comprehensive mapping documentation.
    
    **New Feature**: JSON Examples are automatically converted to schemas for mapping!
    """)
    
    # Schema type selection
    st.markdown("### 🎯 Schema Type Selection")
    col1, col2 = st.columns(2)
    
    with col1:
        source_schema_type = st.selectbox(
            "Source Schema Type",
            ["XSD", "JSON Schema"],
            help="Select the type of your source schema"
        )
    
    with col2:
        target_schema_type = st.selectbox(
            "Target Schema Type", 
            ["XSD", "JSON Schema"],
            help="Select the type of your target schema"
        )
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📤 Source Schema")
        source_file = st.file_uploader(
            "Upload source schema file",
            type=['xsd', 'xml', 'json'],
            key="source_uploader",
            help="Upload your source schema file (XSD, XML, JSON Schema, or JSON Example)"
        )
        
        if source_file:
            st.markdown(f'<div class="success-message">✅ Uploaded: {source_file.name}</div>', unsafe_allow_html=True)
            # Show file preview
            content = source_file.read()
            source_file.seek(0)  # Reset file pointer
            with st.expander("📄 File Preview"):
                try:
                    decoded_content = content.decode('utf-8')
                    preview = decoded_content[:1000] + "..." if len(decoded_content) > 1000 else decoded_content
                    # Determine language based on file extension
                    if source_file.name.lower().endswith('.json'):
                        st.code(preview, language="json")
                    else:
                        st.code(preview, language="xml")
                except UnicodeDecodeError:
                    st.code(content[:500], language="text")
    
    with col2:
        st.markdown("### 📥 Target Schema")
        target_file = st.file_uploader(
            "Upload target schema file",
            type=['xsd', 'xml', 'json'],
            key="target_uploader",
            help="Upload your target schema file (XSD, XML, JSON Schema, or JSON Example)"
        )
        
        if target_file:
            st.markdown(f'<div class="success-message">✅ Uploaded: {target_file.name}</div>', unsafe_allow_html=True)
            # Show file preview
            content = target_file.read()
            target_file.seek(0)  # Reset file pointer
            with st.expander("📄 File Preview"):
                try:
                    decoded_content = content.decode('utf-8')
                    preview = decoded_content[:1000] + "..." if len(decoded_content) > 1000 else decoded_content
                    # Determine language based on file extension
                    if target_file.name.lower().endswith('.json'):
                        st.code(preview, language="json")
                    else:
                        st.code(preview, language="xml")
                except UnicodeDecodeError:
                    st.code(content[:500], language="text")
    
    # Settings
    st.markdown("### ⚙️ Settings")
    col1, col2, col3 = st.columns(3)
    with col1:
        source_case = st.selectbox("Source Case", ["Original", "PascalCase", "camelCase"], 
                                  help="Convert source field names to specified case")
    with col2:
        target_case = st.selectbox("Target Case", ["Original", "PascalCase", "camelCase"], 
                                  help="Convert target field names to specified case")
    with col3:
        min_match_threshold = st.slider("Minimum Match %", 20, 100, 20, 5,
                                       help="Minimum percentage of fields that must match to generate mapping")
    
    col4, col5 = st.columns(2)
    with col4:
        reorder_attributes = st.checkbox("Reorder Attributes First", value=False,
                                       help="Reorder attributes to appear before elements in each parent structure")
    
    # Generate mapping button
    if st.button("🚀 Generate Mapping", type="primary", use_container_width=True):
        if source_file and target_file:
            with st.spinner("🔄 Generating mapping..."):
                try:
                    # Show conversion status for JSON examples
                    source_is_example = is_json_example(source_file)
                    target_is_example = is_json_example(target_file)
                    
                    if source_is_example or target_is_example:
                        status_text = "Converting JSON examples to schemas..."
                        if source_is_example and target_is_example:
                            status_text = "Converting source and target JSON examples to schemas..."
                        elif source_is_example:
                            status_text = "Converting source JSON example to schema..."
                        elif target_is_example:
                            status_text = "Converting target JSON example to schema..."
                        
                        st.info(f"ℹ️ {status_text}")
                    
                    result = process_mapping(source_file, target_file, services, source_case, target_case, reorder_attributes, min_match_threshold)
                    if result:
                        st.markdown('<div class="success-message">✅ Mapping generated successfully!</div>', unsafe_allow_html=True)
                        st.download_button(
                            label="📥 Download Excel File",
                            data=result,
                            file_name="schema_mapping.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.markdown('<div class="error-message">❌ Failed to generate mapping</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="error-message">❌ Error: {str(e)}</div>', unsafe_allow_html=True)
            
        else:
            st.markdown('<div class="warning-message">⚠️ Please upload both source and target schema files</div>', unsafe_allow_html=True)

def show_wsdl_to_xsd_page(services):
    st.markdown('<div class="section-header"><h2>🔧 WSDL to XSD Extraction</h2></div>', unsafe_allow_html=True)
    
    st.markdown("""
    Extract XSD schemas from WSDL files. Perfect for working with web services and SOAP APIs.
    """)
    
    wsdl_file = st.file_uploader(
        "Upload WSDL file",
        type=['wsdl', 'xml'],
        key="wsdl_uploader",
        help="Upload your WSDL file to extract embedded XSD schemas"
    )
    
    if wsdl_file:
        st.markdown(f'<div class="success-message">✅ Uploaded: {wsdl_file.name}</div>', unsafe_allow_html=True)
        # Show file preview
        content = wsdl_file.read()
        wsdl_file.seek(0)  # Reset file pointer
        with st.expander("📄 WSDL Preview"):
            try:
                decoded_content = content.decode('utf-8')
                preview = decoded_content[:1000] + "..." if len(decoded_content) > 1000 else decoded_content
                st.code(preview, language="xml")
            except UnicodeDecodeError:
                st.code(content[:500], language="text")
    
    if st.button("🔧 Extract XSD", type="primary", use_container_width=True):
        if wsdl_file:
            with st.spinner("🔄 Extracting XSD..."):
                try:
                    result = process_wsdl_to_xsd(wsdl_file, services)
                    if result:
                        st.markdown('<div class="success-message">✅ XSD extracted successfully!</div>', unsafe_allow_html=True)
                        st.download_button(
                            label="📥 Download XSD File",
                            data=result,
                            file_name="extracted_schema.xsd",
                            mime="application/xml",
                            use_container_width=True
                        )
                    else:
                        st.markdown('<div class="error-message">❌ Failed to extract XSD</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="error-message">❌ Error: {str(e)}</div>', unsafe_allow_html=True)
            
        else:
            st.markdown('<div class="warning-message">⚠️ Please upload a WSDL file</div>', unsafe_allow_html=True)



def show_about_page():
    """
    Display the about page with application information and features.
    """
    st.markdown('<div class="section-header"><h2>ℹ️ About The Forge</h2></div>', unsafe_allow_html=True)
    
    st.markdown("""
    **The Forge** is a powerful schema transformation tool that provides comprehensive capabilities for working with schema files.
    
    ### 🎯 Core Features
    
    - **📊 Schema Mapping**: Create field mappings between different schema formats (XSD, JSON Schema)
    - **🔧 WSDL to XSD Extraction**: Extract XSD schemas from WSDL files

    
    ### 📁 Supported Formats
    
    **Input Formats:**
    - XSD (.xsd)
    - XML (.xml)
    - JSON Schema (.json)
    - WSDL (.wsdl, .xml)
    
    **Output Formats:**
    - Excel (.xlsx)
    - XSD (.xsd)
    
    ### 🚀 Key Benefits
    
    - ✅ **No compilation issues** - Pure Python implementation
    - ✅ **Lightweight dependencies** - Minimal external requirements
    - ✅ **Cross-platform compatibility** - Works on Windows, macOS, Linux
    - ✅ **Production ready** - Robust error handling and validation
    - ✅ **Modern web interface** - Clean, responsive Streamlit UI
    
    ### 🔧 Technical Details
    
    This web version is based on The Forge v8 desktop application, adapted for online deployment with Streamlit.
    The application uses microservices architecture for modular functionality and easy maintenance.
    
    ### 📞 Support
    
    For issues, questions, or feature requests, please refer to the project documentation or create an issue in the repository.
    """)
    
    # Version info
    st.markdown("### 📋 Version Information")
    st.code("""
    The Forge Web App v1.0.0
    Built with Streamlit
    Based on The Forge v8 Desktop Application
    """)


def show_converter_page(services):
    """
    Display the converter page with dynamic source/target selection.
    """
    st.markdown('<div class="section-header"><h2>🔄 Converter</h2></div>', unsafe_allow_html=True)
    
    st.markdown("""
    Convert between different schema and example formats. Select your source and target types to see available conversions.
    """)
    
    # Get converter service
    converter_service = services['converter']
    
    # Dynamic source/target selection
    st.markdown("### 🎯 Conversion Selection")
    col1, col2 = st.columns(2)
    
    with col1:
        source_type = st.selectbox(
            "Source Type",
            converter_service.get_source_types(),
            help="Select the type of your source file"
        )
    
    with col2:
        if source_type:
            target_types = converter_service.get_target_types_for_source(source_type)
            target_type = st.selectbox(
                "Target Type",
                target_types,
                help="Select the type you want to convert to"
            )
        else:
            target_type = None
    
    # Get conversion key and info
    if source_type and target_type:
        conversion_key = converter_service.get_conversion_key(source_type, target_type)
        supported_conversions = converter_service.get_supported_conversions()
        conversion_info = supported_conversions.get(conversion_key, {})
        
        if conversion_info:
            st.markdown(f"### 📋 {conversion_info.get('name', 'Conversion')}")
            st.markdown(f"**Description:** {conversion_info.get('description', 'N/A')}")
            st.markdown(f"**Input:** {conversion_info.get('input_type', 'N/A')} → **Output:** {conversion_info.get('output_type', 'N/A')}")
            
            # Show help information
            help_info = converter_service.get_conversion_help(conversion_key)
            with st.expander("ℹ️ Conversion Help"):
                st.markdown(f"**Input Format:** {help_info.get('input_format', 'N/A')}")
                st.markdown(f"**Output Format:** {help_info.get('output_format', 'N/A')}")
                st.markdown(f"**Features:** {help_info.get('features', 'N/A')}")
                st.markdown(f"**Example:** `{help_info.get('example', 'N/A')}`")
        else:
            st.warning(f"⚠️ Conversion from {source_type} to {target_type} is not yet implemented.")
            return
    
        # File upload
        st.markdown("### 📤 Input File")
        
        # Determine file types based on source type
        file_types = []
        if source_type == "json example" or source_type == "json schema":
            file_types = ['json']
        elif source_type == "xsd":
            file_types = ['xsd', 'xml']
        elif source_type == "xml example":
            file_types = ['xml']
        elif source_type == "yaml":
            file_types = ['yaml', 'yml']
        elif source_type == "excel":
            file_types = ['xlsx', 'xls']
        
        uploaded_file = st.file_uploader(
            f"Upload {source_type} file",
            type=file_types,
            key=f"converter_uploader_{conversion_key}",
            help=f"Upload your {source_type} file"
        )
        
        if uploaded_file:
            st.markdown(f'<div class="success-message">✅ Uploaded: {uploaded_file.name}</div>', unsafe_allow_html=True)
            
            # Show file preview
            content = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer
            
            with st.expander("📄 File Preview"):
                try:
                    decoded_content = content.decode('utf-8')
                    preview = decoded_content[:1000] + "..." if len(decoded_content) > 1000 else decoded_content
                    
                    if uploaded_file.name.lower().endswith('.json'):
                        st.code(preview, language="json")
                    elif uploaded_file.name.lower().endswith(('.yaml', '.yml')):
                        st.code(preview, language="yaml")
                    else:
                        st.code(preview, language="xml")
                except UnicodeDecodeError:
                    st.code(content[:500], language="text")
        
        # Conversion options
        st.markdown("### ⚙️ Conversion Options")
        
        if conversion_key == "json_to_schema":
            schema_name = st.text_input(
                "Schema Name",
                value="GeneratedSchema",
                help="Name for the generated JSON schema"
            )
            
            validate_schema = st.checkbox(
                "Validate Generated Schema",
                value=True,
                help="Validate the generated schema using jsonschema library"
            )
            
        elif conversion_key == "xml_to_xsd":
            schema_name = st.text_input(
                "Schema Name",
                value="GeneratedSchema",
                help="Name for the generated XSD schema"
            )
            
            validate_xsd = st.checkbox(
                "Validate Generated XSD",
                value=True,
                help="Validate the generated XSD schema"
            )
            
        elif conversion_key == "xsd_to_xml":
            root_element_name = st.text_input(
                "Root Element Name (Optional)",
                value="",
                help="Specify the root element name (if not specified, will be inferred from XSD)"
            )
            
        elif conversion_key == "json_schema_to_json":
            num_examples = st.number_input(
                "Number of Examples",
                min_value=1,
                max_value=10,
                value=1,
                help="Number of JSON examples to generate"
            )
            
            validate_examples = st.checkbox(
                "Validate Generated Examples",
                value=True,
                help="Validate the generated examples against the schema"
            )
        
        # Process conversion
        if st.button("🔄 Convert", type="primary", use_container_width=True):
            if uploaded_file:
                try:
                    with st.spinner(f"Converting {source_type} to {target_type}..."):
                        # Create temporary file
                        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_file.name.split('.')[-1]}") as temp_file:
                            temp_file.write(uploaded_file.read())
                            temp_file_path = temp_file.name
                        
                        # Prepare conversion parameters
                        conversion_params = {}
                        if conversion_key == "json_to_schema":
                            conversion_params['schema_name'] = schema_name
                        elif conversion_key == "xml_to_xsd":
                            conversion_params['schema_name'] = schema_name
                        elif conversion_key == "xsd_to_xml":
                            if root_element_name:
                                conversion_params['root_element_name'] = root_element_name
                        elif conversion_key == "json_schema_to_json":
                            conversion_params['num_examples'] = num_examples
                        
                        # Handle Excel conversions
                        if conversion_key in ["json_to_excel", "json_schema_to_excel", "xsd_to_excel", "xml_to_excel"]:
                            result = process_excel_conversion(temp_file_path, conversion_key, services)
                        else:
                            # Perform regular conversion
                            result = converter_service.process_file_conversion(
                                temp_file_path, 
                                conversion_key, 
                                **conversion_params
                            )
                        
                        # Validate result if requested
                        if conversion_key == "json_to_schema" and validate_schema:
                            is_valid = converter_service.validate_conversion(conversion_key, None, result)
                            if is_valid:
                                st.markdown('<div class="success-message">✅ Generated schema is valid!</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div class="warning-message">⚠️ Generated schema validation failed</div>', unsafe_allow_html=True)
                        
                        elif conversion_key == "xml_to_xsd" and validate_xsd:
                            is_valid = converter_service.validate_conversion(conversion_key, None, result)
                            if is_valid:
                                st.markdown('<div class="success-message">✅ Generated XSD is valid!</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div class="warning-message">⚠️ Generated XSD validation failed</div>', unsafe_allow_html=True)
                        
                        elif conversion_key == "json_schema_to_json" and validate_examples:
                            is_valid = converter_service.validate_conversion(conversion_key, None, result)
                            if is_valid:
                                st.markdown('<div class="success-message">✅ Generated examples are valid!</div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<div class="warning-message">⚠️ Generated examples validation failed</div>', unsafe_allow_html=True)
                        
                        # Show statistics
                        if conversion_key not in ["json_to_excel", "json_schema_to_excel", "xsd_to_excel", "xml_to_excel"]:
                            stats = converter_service.get_conversion_statistics(conversion_key, result)
                            if stats:
                                st.markdown("#### 📊 Conversion Statistics")
                                col1, col2, col3 = st.columns(3)
                                
                                if conversion_key == "json_to_schema":
                                    with col1:
                                        st.metric("Total Properties", stats.get('total_properties', 0))
                                        st.metric("Required Properties", stats.get('required_properties', 0))
                                    with col2:
                                        st.metric("Object Properties", stats.get('object_properties', 0))
                                        st.metric("Array Properties", stats.get('array_properties', 0))
                                    with col3:
                                        st.metric("Primitive Properties", stats.get('primitive_properties', 0))
                                        st.metric("Max Depth", stats.get('max_depth', 0))
                                
                                elif conversion_key == "xml_to_xsd":
                                    with col1:
                                        st.metric("Total Elements", stats.get('total_elements', 0))
                                        st.metric("Complex Types", stats.get('complex_types', 0))
                                    with col2:
                                        st.metric("Simple Types", stats.get('simple_types', 0))
                                        st.metric("Attributes", stats.get('attributes', 0))
                                    with col3:
                                        st.metric("Max Depth", stats.get('max_depth', 0))
                                
                                elif conversion_key == "xsd_to_xml":
                                    with col1:
                                        st.metric("Total Elements", stats.get('total_elements', 0))
                                        st.metric("Attributes", stats.get('attributes', 0))
                                    with col2:
                                        st.metric("Text Elements", stats.get('text_elements', 0))
                                        st.metric("Max Depth", stats.get('max_depth', 0))
                                
                                elif conversion_key == "json_schema_to_json":
                                    with col1:
                                        st.metric("Total Examples", stats.get('total_examples', 0))
                                        st.metric("Avg Properties", stats.get('avg_object_properties', 0))
                                    with col2:
                                        st.metric("Avg Array Length", stats.get('avg_array_length', 0))
                                        st.metric("Max Depth", stats.get('max_depth', 0))
                        
                        # Display result
                        st.markdown("#### 📄 Conversion Result")
                        
                        if conversion_key in ["json_to_schema", "json_schema_to_json"]:
                            st.json(result)
                        elif conversion_key in ["json_to_excel", "json_schema_to_excel", "xsd_to_excel", "xml_to_excel"]:
                            # For Excel conversions, show success message and download button
                            st.markdown('<div class="success-message">✅ Excel file generated successfully!</div>', unsafe_allow_html=True)
                        else:
                            st.code(result, language="xml")
                        
                        # Download button
                        if conversion_key in ["json_to_schema", "json_schema_to_json"]:
                            result_json = json.dumps(result, indent=2, ensure_ascii=False)
                            file_extension = "json"
                            mime_type = "application/json"
                        elif conversion_key in ["json_to_excel", "json_schema_to_excel", "xsd_to_excel", "xml_to_excel"]:
                            # Excel file download
                            result_json = result
                            file_extension = "xlsx"
                            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        else:
                            result_json = result
                            file_extension = "xsd" if conversion_key == "xml_to_xsd" else "xml"
                            mime_type = "application/xml"
                        
                        st.download_button(
                            label="📥 Download Result",
                            data=result_json,
                            file_name=f"converted.{file_extension}",
                            mime=mime_type,
                            use_container_width=True
                        )
                        
                        # Clean up temporary file
                        os.unlink(temp_file_path)
                        
                except Exception as e:
                    st.markdown(f'<div class="error-message">❌ Error during conversion: {str(e)}</div>', unsafe_allow_html=True)
                    st.error(f"Technical details: {str(e)}")
            else:
                st.markdown('<div class="warning-message">⚠️ Please upload a file to convert</div>', unsafe_allow_html=True)


def process_excel_conversion(file_path: str, conversion_key: str, services: dict) -> bytes:
    """
    Process Excel conversions using the existing ExcelExporter service.
    """
    try:
        excel_exporter = services['excel_exporter']
        
        if conversion_key == "json_to_excel":
            # Convert JSON example to Excel
            with open(file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            # Use ExcelExporter for JSON to Excel conversion
            # First convert JSON to schema, then to Excel
            json_to_schema = services['converter'].json_to_schema
            schema_data = json_to_schema.convert_json_example_to_schema(json_data, "GeneratedSchema")
            # Parse the schema and convert to Excel
            json_schema_parser = services['json_schema_parser']
            # Convert schema to string and use parse_json_schema_string
            schema_string = json.dumps(schema_data)
            parsed_data = json_schema_parser.parse_json_schema_string(schema_string)
            output_buffer = BytesIO()
            excel_exporter.export({'schema': parsed_data}, output_buffer)
            output_buffer.seek(0)
            return output_buffer.getvalue()
        
        elif conversion_key == "json_schema_to_excel":
            # Convert JSON Schema to Excel
            with open(file_path, 'r', encoding='utf-8') as f:
                schema_data = json.load(f)
            # Parse JSON Schema and convert to Excel
            json_schema_parser = services['json_schema_parser']
            # Convert schema to string and use parse_json_schema_string
            schema_string = json.dumps(schema_data)
            parsed_data = json_schema_parser.parse_json_schema_string(schema_string)
            output_buffer = BytesIO()
            excel_exporter.export({'schema': parsed_data}, output_buffer)
            output_buffer.seek(0)
            return output_buffer.getvalue()
        
        elif conversion_key == "xsd_to_excel":
            # Convert XSD to Excel with multiple sheets for multiple messages
            xsd_parser = services['xsd_parser']
            parsed_data = xsd_parser.parse_xsd_file_by_messages(file_path)
            output_buffer = BytesIO()
            excel_exporter.export(parsed_data, output_buffer)
            output_buffer.seek(0)
            return output_buffer.getvalue()
        
        elif conversion_key == "xml_to_excel":
            # Convert XML to Excel (first convert to XSD, then to Excel)
            # This is a simplified approach - in practice, you might want to parse XML directly
            xml_to_xsd = services['converter'].xml_to_xsd
            xsd_content = xml_to_xsd.convert_xml_example_to_xsd(
                open(file_path, 'r', encoding='utf-8').read(), 
                "GeneratedSchema"
            )
            # Save XSD to temp file and convert to Excel
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xsd') as temp_xsd:
                temp_xsd.write(xsd_content.encode('utf-8'))
                temp_xsd_path = temp_xsd.name
            
            try:
                xsd_parser = services['xsd_parser']
                parsed_data = xsd_parser.parse_xsd_file(temp_xsd_path)
                output_buffer = BytesIO()
                excel_exporter.export({'schema': parsed_data}, output_buffer)
                output_buffer.seek(0)
                return output_buffer.getvalue()
            finally:
                os.unlink(temp_xsd_path)
        
        elif conversion_key == "yaml_to_excel":
            # Convert YAML to Excel (first convert to JSON Schema, then to Excel)
            with open(file_path, 'r', encoding='utf-8') as f:
                yaml_content = f.read()
            
            # Convert YAML to JSON Schema
            yaml_to_json_schema = services['converter'].yaml_to_json_schema
            schema_data = yaml_to_json_schema.convert_yaml_to_json_schema(yaml_content, "GeneratedSchema")
            
            # Parse JSON Schema and convert to Excel
            json_schema_parser = services['json_schema_parser']
            schema_string = json.dumps(schema_data)
            parsed_data = json_schema_parser.parse_json_schema_string(schema_string)
            output_buffer = BytesIO()
            excel_exporter.export({'schema': parsed_data}, output_buffer)
            output_buffer.seek(0)
            return output_buffer.getvalue()
        
        else:
            raise ValueError(f"Unsupported Excel conversion: {conversion_key}")
    
    except Exception as e:
        raise Exception(f"Error in Excel conversion {conversion_key}: {str(e)}")




def is_json_example(file):
    """
    Check if a file is a JSON example (not a schema).
    Returns True if it's a JSON example, False otherwise.
    """
    try:
        # Check if it's a JSON file
        if not file.name.lower().endswith('.json'):
            return False
        
        # Read the file content
        content = file.read()
        file.seek(0)  # Reset file pointer
        
        # Try to parse as JSON
        try:
            json_data = json.loads(content.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return False
        
        # Check if it's already a JSON schema (has $schema or properties)
        if isinstance(json_data, dict) and ('$schema' in json_data or 'properties' in json_data):
            return False
        
        # It's a JSON example
        return True
        
    except Exception:
        return False


def is_yaml_file(file):
    """
    Check if a file is a YAML file.
    Returns True if it's a YAML file, False otherwise.
    """
    try:
        # Check if it's a YAML file
        yaml_extensions = ['.yaml', '.yml']
        if not any(file.name.lower().endswith(ext) for ext in yaml_extensions):
            return False
        
        # Read the file content
        content = file.read()
        file.seek(0)  # Reset file pointer
        
        # Try to parse as YAML
        try:
            import yaml
            yaml_data = yaml.safe_load(content.decode('utf-8'))
            return yaml_data is not None
        except (yaml.YAMLError, UnicodeDecodeError):
            return False
        
    except Exception:
        return False


def convert_json_example_if_needed(file_path, services):
    """
    Check if a JSON file is an example (not a schema) and convert it to a schema if needed.
    Returns the original file path if it's already a schema, or a new schema file path if converted.
    """
    try:
        # Check if it's a JSON file
        if not file_path.lower().endswith('.json'):
            return file_path
        
        # Read the file content
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Try to parse as JSON
        try:
            json_data = json.loads(content)
        except json.JSONDecodeError:
            return file_path  # Not valid JSON, return original path
        
        # Check if it's already a JSON schema (has $schema or properties)
        if isinstance(json_data, dict) and ('$schema' in json_data or 'properties' in json_data):
            return file_path  # Already a schema
        
        # It's a JSON example, convert to schema
        converter_service = services.get('converter')
        if converter_service:
            # Generate schema from the JSON example
            schema = converter_service.convert_json_example_to_schema(json_data, "GeneratedSchema")
            
            # Create a temporary schema file
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8') as temp_schema:
                json.dump(schema, temp_schema, indent=2)
                schema_path = temp_schema.name
            
            return schema_path
        else:
            return file_path  # Service not available, return original path
            
    except Exception as e:
        # If any error occurs, return the original file path
        return file_path


def convert_yaml_to_json_schema_if_needed(file_path, services):
    """
    Convert YAML file to JSON Schema.
    Returns the JSON Schema file path.
    """
    try:
        # Check if it's a YAML file
        yaml_extensions = ['.yaml', '.yml']
        if not any(file_path.lower().endswith(ext) for ext in yaml_extensions):
            return file_path
        
        # Read the file content
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Convert YAML to JSON Schema
        converter_service = services.get('converter')
        if converter_service:
            # Generate schema from the YAML content
            schema = converter_service.convert_yaml_to_json_schema(content, "GeneratedSchema")
            
            # Create a temporary schema file
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8') as temp_schema:
                json.dump(schema, temp_schema, indent=2)
                schema_path = temp_schema.name
            
            return schema_path
        else:
            return file_path  # Service not available, return original path
            
    except Exception as e:
        # If any error occurs, return the original file path
        return file_path


def process_mapping(source_file, target_file, services, source_case="Original", target_case="Original", reorder_attributes=False, min_match_threshold=20):
    try:
        # Create temporary files
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{source_file.name.split('.')[-1]}") as source_temp:
            source_temp.write(source_file.read())
            source_temp_path = source_temp.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{target_file.name.split('.')[-1]}") as target_temp:
            target_temp.write(target_file.read())
            target_temp_path = target_temp.name
        
        # --- Enhanced schema parsing logic for XSD, JSON Schema, and JSON Examples ---
        
        # Check if source file is a JSON example or YAML and convert to schema if needed
        source_temp_path = convert_json_example_if_needed(source_temp_path, services)
        source_temp_path = convert_yaml_to_json_schema_if_needed(source_temp_path, services)
        
        # Parse source schema (XSD, JSON Schema, or converted JSON Example/YAML)
        src_rows = parse_schema_file(source_temp_path, services)
        
        # Check if target file is a JSON example or YAML and convert to schema if needed
        target_temp_path = convert_json_example_if_needed(target_temp_path, services)
        target_temp_path = convert_yaml_to_json_schema_if_needed(target_temp_path, services)
        
        # Parse target schema (XSD, JSON Schema, or converted JSON Example)
        tgt_rows = []
        if target_temp_path and os.path.exists(target_temp_path):
            tgt_rows = parse_schema_file(target_temp_path, services)
        
        # Detect if both schemas are JSON schemas
        source_is_json = source_file.name.lower().endswith('.json') or source_temp_path.endswith('.json')
        target_is_json = target_file.name.lower().endswith('.json') or target_temp_path.endswith('.json')
        both_json_schemas = source_is_json and target_is_json
        
        if both_json_schemas:
            # For JSON Schema to JSON Schema mapping, use single sheet approach
            return _process_json_schema_mapping(src_rows, tgt_rows, source_case, target_case, min_match_threshold, source_temp_path, target_temp_path)
        else:
            # For XSD or mixed schema mapping, use multi-sheet approach
            return _process_mixed_schema_mapping(src_rows, tgt_rows, source_case, target_case, reorder_attributes, min_match_threshold, source_temp_path, target_temp_path)
        
    except Exception as e:
        st.error(f"Error in mapping: {str(e)}")
        return None


def _process_json_schema_mapping(src_rows, tgt_rows, source_case, target_case, min_match_threshold, source_temp_path=None, target_temp_path=None):
    """
    Process JSON Schema to JSON Schema mapping using single sheet approach.
    Respects JSON schema logic including restrictions, cardinalities, etc.
    """
    try:
        # Build Excel file with single sheet for JSON schemas
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "JSON Schema Mapping"
        
        # Initialize variables for statistics
        total_source_fields = 0
        matched_fields = 0
        
        # Calculate max levels for both schemas
        max_src_level = max((len(row['levels']) for row in src_rows), default=1)
        max_tgt_level = max((len(row['levels']) for row in tgt_rows), default=1) if tgt_rows else 1
        
        # Define headers with same structure as XSD transformation
        src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
        tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
        headers = src_cols + ['Destination Fields'] + tgt_cols
        ws.append(headers)
        ws.append([''] * len(headers))  # Second header row blank for now
        
        def row_path(row):
            return '.'.join(row['levels'])
        
        tgt_path_dict = {row_path(row): row for row in tgt_rows}
        tgt_paths = list(tgt_path_dict.keys())
        
        # Create a mapping of source paths to target paths based on similarity
        source_to_target_mapping = {}
        
        for src_row in src_rows:
            # Apply case conversion to source levels if needed
            converted_levels = src_row['levels'].copy()
            if source_case == "PascalCase":
                converted_levels = [camel_to_pascal(level) for level in converted_levels]
            elif source_case == "camelCase":
                converted_levels = [pascal_to_camel(level) for level in converted_levels]
            
            src_levels = converted_levels + [''] * (max_src_level - len(converted_levels))
            src_vals = src_levels + [
                src_row.get('Request Parameter',''),
                src_row.get('GDPR',''),
                src_row.get('Cardinality',''),
                src_row.get('Type',''),
                src_row.get('Base Type',''),
                src_row.get('Details',''),
                src_row.get('Description',''),
                src_row.get('Category','element'),
                src_row.get('Example','')
            ]
            
            src_path_str = row_path(src_row)
            
            # Check if we already have a mapping for this source path
            if src_path_str in source_to_target_mapping:
                tgt_row = source_to_target_mapping[src_path_str]
            else:
                # Try to find a match
                tgt_row = tgt_path_dict.get(src_path_str)
                best_match = ''
                
                if not tgt_row and tgt_paths:
                    # Use fuzzy matching with a higher threshold for better accuracy
                    matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.6)
                    if matches:
                        best_match = matches[0]
                        tgt_row = tgt_path_dict[best_match]
                
                # Store the mapping to avoid re-computation
                source_to_target_mapping[src_path_str] = tgt_row
            
            dest_field = ''  # Initialize destination field as empty
            
            # Apply case conversion to target levels if needed
            converted_tgt_levels = []
            if tgt_row:
                converted_tgt_levels = tgt_row['levels'].copy()
                if target_case == "PascalCase":
                    converted_tgt_levels = [camel_to_pascal(level) for level in converted_tgt_levels]
                elif target_case == "camelCase":
                    converted_tgt_levels = [pascal_to_camel(level) for level in converted_tgt_levels]
                # Set destination field to the matched target path
                dest_field = '.'.join([lvl for lvl in converted_tgt_levels if lvl])
                tgt_levels = converted_tgt_levels + [''] * (max_tgt_level - len(converted_tgt_levels))
            else:
                tgt_levels = ['']*max_tgt_level
                # Keep dest_field as empty string for unmatched source fields
            
            tgt_vals = tgt_levels + [
                tgt_row.get('Request Parameter','') if tgt_row else '',
                tgt_row.get('GDPR','') if tgt_row else '',
                tgt_row.get('Cardinality','') if tgt_row else '',
                tgt_row.get('Type','') if tgt_row else '',
                tgt_row.get('Base Type','') if tgt_row else '',
                tgt_row.get('Details','') if tgt_row else '',
                tgt_row.get('Description','') if tgt_row else '',
                tgt_row.get('Category','element') if tgt_row else '',
                tgt_row.get('Example','') if tgt_row else ''
            ]
            
            ws.append(src_vals + [dest_field] + tgt_vals)
        
        # Update statistics
        total_source_fields = len(src_rows)
        matched_fields = sum(1 for src_row in src_rows 
                          if source_to_target_mapping.get(row_path(src_row)) is not None)
        
        # Add summary row at the end
        summary_row = [''] * len(src_vals) + [f'SUMMARY: {matched_fields}/{total_source_fields} fields matched'] + [''] * len(tgt_vals)
        ws.append(summary_row)
        
        # Prune unused columns
        def last_nonempty_level_col(start_col, num_levels):
            for col in range(start_col + num_levels - 1, start_col - 1, -1):
                for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                    if any(cell.value not in (None, '') for cell in row):
                        return col
            return start_col - 1
        
        src_level_start = 1
        last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
        for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
            ws.delete_cols(col)
        
        header_row = [cell.value for cell in ws[1]]
        try:
            tgt_level_start = header_row.index('Level1_tgt') + 1
        except ValueError:
            tgt_level_start = len(header_row) + 1
        
        for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
            col_letter = get_column_letter(col)
            if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                ws.delete_cols(col)
        
        # Calculate overall match percentage
        match_percentage = (matched_fields / total_source_fields * 100) if total_source_fields > 0 else 0
        unmatched_fields = total_source_fields - matched_fields
        
        # Check if we have enough matches to generate a meaningful mapping
        if match_percentage < min_match_threshold:
            # Provide detailed analysis
            st.warning(f"⚠️ **JSON Schemas don't match well enough to generate a mapping**")
            st.markdown(f"""
            **Analysis Results:**
            - **Source fields:** {total_source_fields}
            - **Matched fields:** {matched_fields}
            - **Unmatched fields:** {unmatched_fields}
            - **Match percentage:** {match_percentage:.1f}%
            - **Minimum threshold:** {min_match_threshold}%
            
            **Possible reasons for low match:**
            - Different JSON Schema structures or naming conventions
            - Incompatible data models
            - Different business domains or use cases
            - Missing or extra fields in one of the schemas
            
            **Suggestions:**
            - Check if the JSON schemas are from the same domain/business context
            - Verify that both schemas represent similar data structures
            - Consider using different source/target schemas that are more compatible
            - Review the field names and structure for potential manual mapping
            """)
            return None
        
        # Save to buffer
        output_buffer = BytesIO()
        wb.save(output_buffer)
        
        # Clean up temp files
        if source_temp_path and os.path.exists(source_temp_path):
            os.unlink(source_temp_path)
        if target_temp_path and os.path.exists(target_temp_path):
            os.unlink(target_temp_path)
        
        output_buffer.seek(0)
        
        # Show success message with match statistics
        st.success(f"✅ **JSON Schema mapping generated successfully!** ({match_percentage:.1f}% of fields matched)")
        
        return output_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error in JSON schema mapping: {str(e)}")
        return None


def _process_mixed_schema_mapping(src_rows, tgt_rows, source_case, target_case, reorder_attributes, min_match_threshold, source_temp_path=None, target_temp_path=None):
    """
    Process mixed schema mapping (XSD, JSON Schema, or mixed) using multi-sheet approach.
    This is the original logic for handling XSD and mixed schema types.
    """
    try:
        # Group source rows by message/element name for multi-sheet structure
        src_messages = {}
        current_message = "schema"  # Default message name
        
        for row in src_rows:
            # For JSON Schema, we don't have multiple messages like XSD, so group by first level
            if len(row['levels']) > 0 and row['levels'][0]:
                current_message = row['levels'][0]
            
            if current_message not in src_messages:
                src_messages[current_message] = []
            src_messages[current_message].append(row)
        
        # Build Excel file
        wb = openpyxl.Workbook()
        first = True
        
        # Initialize variables for overall statistics
        total_source_fields = 0
        matched_fields = 0
        
        for msg_name, src_full_rows in src_messages.items():
            if not first:
                ws = wb.create_sheet(title=msg_name[:31])  # Excel sheet name limit
            else:
                ws = wb.active
                ws.title = msg_name[:31]
                first = False
            
            max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
            max_tgt_level = max((len(row['levels']) for row in tgt_rows), default=1) if tgt_rows else 1
            
            src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
            tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
            headers = src_cols + ['Destination Fields'] + tgt_cols
            ws.append(headers)
            ws.append([''] * len(headers))  # Second header row blank for now
            
            def row_path(row):
                return '.'.join(row['levels'])
            
            tgt_path_dict = {row_path(row): row for row in tgt_rows}
            tgt_paths = list(tgt_path_dict.keys())
            
            # Create a mapping of source paths to target paths based on similarity
            # This will help avoid duplicating target fields across multiple source rows
            source_to_target_mapping = {}
            
            for src_row in src_full_rows:
                # Apply case conversion to source levels if needed
                converted_levels = src_row['levels'].copy()
                if source_case == "PascalCase":
                    converted_levels = [camel_to_pascal(level) for level in converted_levels]
                elif source_case == "camelCase":
                    converted_levels = [pascal_to_camel(level) for level in converted_levels]
                
                src_levels = converted_levels + [''] * (max_src_level - len(converted_levels))
                src_vals = src_levels + [
                    src_row.get('Request Parameter',''),
                    src_row.get('GDPR',''),
                    src_row.get('Cardinality',''),
                    src_row.get('Type',''),
                    src_row.get('Base Type',''),
                    src_row.get('Details',''),
                    src_row.get('Description',''),
                    src_row.get('Category','element'),
                    src_row.get('Example','')
                ]
                
                src_path_str = row_path(src_row)
                
                # Check if we already have a mapping for this source path
                if src_path_str in source_to_target_mapping:
                    tgt_row = source_to_target_mapping[src_path_str]
                else:
                    # Try to find a match
                    tgt_row = tgt_path_dict.get(src_path_str)
                    best_match = ''
                    
                    if not tgt_row and tgt_paths:
                        # Use fuzzy matching with a higher threshold for better accuracy
                        matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.6)
                        if matches:
                            best_match = matches[0]
                            tgt_row = tgt_path_dict[best_match]
                    
                    # Store the mapping to avoid re-computation
                    source_to_target_mapping[src_path_str] = tgt_row
                
                dest_field = ''  # Initialize destination field as empty
                
                # Apply case conversion to target levels if needed
                converted_tgt_levels = []
                if tgt_row:
                    converted_tgt_levels = tgt_row['levels'].copy()
                    if target_case == "PascalCase":
                        converted_tgt_levels = [camel_to_pascal(level) for level in converted_tgt_levels]
                    elif target_case == "camelCase":
                        converted_tgt_levels = [pascal_to_camel(level) for level in converted_tgt_levels]
                    # Set destination field to the matched target path
                    dest_field = '.'.join([lvl for lvl in converted_tgt_levels if lvl])
                    tgt_levels = converted_tgt_levels + [''] * (max_tgt_level - len(converted_tgt_levels))
                else:
                    tgt_levels = ['']*max_tgt_level
                    # Keep dest_field as empty string for unmatched source fields
                
                tgt_vals = tgt_levels + [
                    tgt_row.get('Request Parameter','') if tgt_row else '',
                    tgt_row.get('GDPR','') if tgt_row else '',
                    tgt_row.get('Cardinality','') if tgt_row else '',
                    tgt_row.get('Type','') if tgt_row else '',
                    tgt_row.get('Base Type','') if tgt_row else '',
                    tgt_row.get('Details','') if tgt_row else '',
                    tgt_row.get('Description','') if tgt_row else '',
                    tgt_row.get('Category','element') if tgt_row else '',
                    tgt_row.get('Example','') if tgt_row else ''
                ]
                
                ws.append(src_vals + [dest_field] + tgt_vals)
            
            # Update overall statistics
            total_source_fields += len(src_full_rows)
            matched_fields += sum(1 for src_row in src_full_rows 
                              if source_to_target_mapping.get(row_path(src_row)) is not None)
            
            # Add summary row at the end
            summary_row = [''] * len(src_vals) + [f'SUMMARY: {matched_fields}/{total_source_fields} fields matched'] + [''] * len(tgt_vals)
            ws.append(summary_row)
            
            # Prune unused source level columns
            def last_nonempty_level_col(start_col, num_levels):
                for col in range(start_col + max_src_level - 1, start_col - 1, -1):
                    for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                        if any(cell.value not in (None, '') for cell in row):
                            return col
                return start_col - 1
            
            src_level_start = 1
            last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
            for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
                ws.delete_cols(col)
            
            header_row = [cell.value for cell in ws[1]]
            try:
                tgt_level_start = header_row.index('Level1_tgt') + 1
            except ValueError:
                tgt_level_start = len(header_row) + 1
            
            for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
                col_letter = get_column_letter(col)
                if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                    ws.delete_cols(col)
        
        # Calculate overall match percentage
        match_percentage = (matched_fields / total_source_fields * 100) if total_source_fields > 0 else 0
        unmatched_fields = total_source_fields - matched_fields
        
        # Check if we have enough matches to generate a meaningful mapping
        if match_percentage < min_match_threshold:
            # Clean up temp files before returning
            if 'source_temp_path' in locals():
                os.unlink(source_temp_path)
            if 'target_temp_path' in locals():
                os.unlink(target_temp_path)
            
            # Provide detailed analysis
            st.warning(f"⚠️ **Schemas don't match well enough to generate a mapping**")
            st.markdown(f"""
            **Analysis Results:**
            - **Source fields:** {total_source_fields}
            - **Matched fields:** {matched_fields}
            - **Unmatched fields:** {unmatched_fields}
            - **Match percentage:** {match_percentage:.1f}%
            - **Minimum threshold:** {min_match_threshold}%
            
            **Possible reasons for low match:**
            - Different schema structures or naming conventions
            - Incompatible data models
            - Different business domains or use cases
            - Missing or extra fields in one of the schemas
            
            **Suggestions:**
            - Check if the schemas are from the same domain/business context
            - Verify that both schemas represent similar data structures
            - Consider using different source/target schemas that are more compatible
            - Review the field names and structure for potential manual mapping
            """)
            return None
        
        # Save to buffer
        output_buffer = BytesIO()
        wb.save(output_buffer)
        
        # --- Post-processing QA: Excel Output Validator ---
        xsd_path = source_temp_path if 'source_temp_path' in locals() else target_temp_path
        try:
            from services.excel_output_validator import validate_excel_output, _log_messages
            _log_messages.clear()
            # Save to temporary file for validation
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel:
                output_buffer.seek(0)
                temp_excel.write(output_buffer.getvalue())
                temp_excel_path = temp_excel.name
            
            validate_excel_output(xsd_path, temp_excel_path)
            
            # Improved validation display
            if _log_messages:
                # Group messages by type
                success_messages = []
                error_messages = []
                warning_messages = []
                info_messages = []
                
                for line in _log_messages:
                    if line.startswith("[SUCCESS]"):
                        success_messages.append(line)
                    elif line.startswith("[ERROR]"):
                        error_messages.append(line)
                    elif line.startswith("[WARNING]"):
                        warning_messages.append(line)
                    elif line.startswith("[VALIDATE]"):
                        info_messages.append(line)
                    else:
                        info_messages.append(line)
                
                # Display validation summary in a compact format
                with st.expander("🔍 Validation Results", expanded=False):
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        if error_messages:
                            st.error(f"❌ {len(error_messages)} Field Errors")
                        else:
                            st.success("✅ No Field Errors")
                    
                    with col2:
                        if warning_messages:
                            st.warning(f"⚠️ {len(warning_messages)} Field Warnings")
                        else:
                            st.success("✅ No Field Warnings")
                    
                    with col3:
                        if success_messages:
                            st.success(f"✅ {len(success_messages)} Fields Validated")
                    
                    with col4:
                        if info_messages:
                            st.info(f"ℹ️ {len(info_messages)} Validation Steps")
                    
                    # Show detailed messages only if there are issues
                    if error_messages or warning_messages:
                        st.markdown("---")
                        if error_messages:
                            st.markdown("**❌ Field Errors:**")
                            for msg in error_messages[:5]:  # Limit to first 5 errors
                                st.text(f"  • {msg.replace('[ERROR]', '').strip()}")
                            if len(error_messages) > 5:
                                st.text(f"  ... and {len(error_messages) - 5} more field errors")
                        
                        if warning_messages:
                            st.markdown("**⚠️ Field Warnings:**")
                            for msg in warning_messages[:3]:  # Limit to first 3 warnings
                                st.text(f"  • {msg.replace('[WARNING]', '').strip()}")
                            if len(warning_messages) > 3:
                                st.text(f"  ... and {len(warning_messages) - 3} more field warnings")
                    else:
                        st.success("🎉 All validations passed successfully!")
            
            # Clean up temp validation file
            os.unlink(temp_excel_path)
        except Exception as e:
            st.error(f"❌ Error in post-processing validator: {e}")
        
        # --- Attribute reordering if flag is set ---
        if reorder_attributes:
            try:
                from services.reorder_excel_attributes import reorder_attributes_in_excel
                
                # Save to temporary file for reordering
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel:
                    output_buffer.seek(0)
                    temp_excel.write(output_buffer.getvalue())
                    temp_excel_path = temp_excel.name
                
                reorder_attributes_in_excel(temp_excel_path)
                
                # Read back the reordered file
                with open(temp_excel_path, 'rb') as f:
                    output_buffer.seek(0)
                    output_buffer.write(f.read())
                    output_buffer.truncate()
                
                # Clean up temp reordering file
                os.unlink(temp_excel_path)
            except Exception as e:
                st.error(f"Failed to reorder attributes: {str(e)}")
                # Continue without reordering rather than failing the entire process
        
        # Clean up temp files
        if source_temp_path and os.path.exists(source_temp_path):
            os.unlink(source_temp_path)
        if target_temp_path and os.path.exists(target_temp_path):
            os.unlink(target_temp_path)
        
        output_buffer.seek(0)
        
        # Show success message with match statistics
        st.success(f"✅ **Mapping generated successfully!** ({match_percentage:.1f}% of fields matched)")
        
        return output_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error in mixed schema mapping: {str(e)}")
        return None

def process_wsdl_to_xsd(wsdl_file, services):
    try:
        # Read WSDL content
        wsdl_content = wsdl_file.read().decode('utf-8')
        
        # Extract XSD
        xsd_content = merge_xsd_from_wsdl(wsdl_content)
        
        if not xsd_content or xsd_content.startswith("Error"):
            st.error(f"Error extracting XSD: {xsd_content}")
            return None
        
        return xsd_content.encode('utf-8')
        
    except Exception as e:
        st.error(f"Error in WSDL extraction: {str(e)}")
        return None




def parse_schema_file(file_path, services):
    """
    Parse a schema file (XSD or JSON Schema) and return rows in the same format.
    
    Args:
        file_path: Path to the schema file
        services: Dictionary containing parser services
        
    Returns:
        List of dictionaries with the same structure as XSD parser output
    """
    if file_path.endswith('.json'):
        # Handle JSON Schema
        return services['json_schema_parser'].parse_json_schema_file(file_path)
    else:
        # Handle XSD
        return services['xsd_parser'].parse_xsd_file(file_path)

if __name__ == "__main__":
    main()