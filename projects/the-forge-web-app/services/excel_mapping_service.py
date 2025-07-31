import openpyxl
import xml.etree.ElementTree as ET
import difflib

class ExcelMappingService:
    def __init__(self, max_structure_depth=8):
        self.max_structure_depth = max_structure_depth
        self.mapping_data = None
        self.headers = None
        self.sheet_names = []
        self.current_sheet = None
        self.header_rows = 2  # Support two header rows as in the template

    def list_sheets(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def load_mapping_file(self, file_path, sheet_name=None):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet_names = wb.sheetnames
        if sheet_name is None:
            sheet = wb.active
            self.current_sheet = sheet.title
        else:
            sheet = wb[sheet_name]
            self.current_sheet = sheet_name
        # Read two header rows
        header_rows = list(sheet.iter_rows(min_row=1, max_row=self.header_rows, values_only=True))
        self.headers = [f'{header_rows[0][i] or ""}\n{header_rows[1][i] or ""}' for i in range(len(header_rows[0]))]
        self.mapping_data = []
        for row in sheet.iter_rows(min_row=self.header_rows+1, values_only=True):
            row_dict = dict(zip(self.headers, row))
            self.mapping_data.append(row_dict)
        return self.mapping_data

    def get_structures(self, structure_type='source'):
        if not self.mapping_data:
            return []
        key = 'Source Structure' if structure_type == 'source' else 'Target Structure'
        return sorted(set(row.get(key) for row in self.mapping_data if row.get(key)))

    def get_mapping_rows(self, source_sheet=None, target_sheet=None):
        if not self.mapping_data:
            return []
        return self.mapping_data

    @staticmethod
    def flatten_schema(schema, parent_key='', sep='.'):
        # Recursively flatten a dict (JSON Schema or XSD structure) into dot notation paths
        items = []
        if isinstance(schema, dict):
            for k, v in schema.items():
                new_key = f'{parent_key}{sep}{k}' if parent_key else k
                if isinstance(v, dict):
                    items.extend(ExcelMappingService.flatten_schema(v, new_key, sep=sep))
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        items.extend(ExcelMappingService.flatten_schema(item, f'{new_key}[{i}]', sep=sep))
                else:
                    items.append(new_key)
        return items

    def parse_xsd_file(self, file_path):
        """Parse XSD file and return structured data like v8"""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            XSD_NS = '{http://www.w3.org/2001/XMLSchema}'
            
            # Parse simple types
            simple_types = {}
            for st in root.findall(f'.//{XSD_NS}simpleType'):
                name = st.get('name')
                if not name:
                    continue
                restriction = st.find(f'{XSD_NS}restriction')
                base = restriction.get('base') if restriction is not None else None
                restrictions = {}
                if restriction is not None:
                    for cons in restriction:
                        cons_name = cons.tag.replace(XSD_NS, '')
                        val = cons.get('value')
                        if val:
                            restrictions[cons_name] = val
                simple_types[name] = {'base': base, 'restrictions': restrictions}
            
            # Parse complex types
            complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{XSD_NS}complexType') if ct.get('name')}
            
            # Parse elements and create structured rows
            messages = {}
            for elem in root.findall(f'{XSD_NS}element'):
                name = elem.get('name')
                if name:
                    rows = self._parse_element(elem, complex_types, simple_types, 1, category='message')
                    messages[name] = rows
            
            return messages
        except Exception as e:
            print(f"Error parsing XSD file: {e}")
            return {}

    def _parse_element(self, elem, complex_types, simple_types, level, category='element'):
        """Parse an XSD element and return structured rows"""
        rows = []
        name = elem.get('name', '')
        elem_type = elem.get('type', '')
        
        # Create base row
        row = {
            'levels': [name] + [''] * (self.max_structure_depth - 1),
            'Request Parameter': name,
            'GDPR': '',
            'Cardinality': self._get_cardinality(elem),
            'Type': elem_type,
            'Base Type': self._get_base_type(elem_type, simple_types),
            'Details': self._get_details(elem),
            'Description': '',
            'Category': category,
            'Example': ''
        }
        rows.append(row)
        
        # Parse complex type if present
        if elem_type in complex_types:
            complex_type = complex_types[elem_type]
            for child in complex_type:
                if child.tag.endswith('element'):
                    child_rows = self._parse_element(child, complex_types, simple_types, level + 1, 'element')
                    for child_row in child_rows:
                        child_row['levels'] = [name] + child_row['levels'][:self.max_structure_depth - 1]
                    rows.extend(child_rows)
                elif child.tag.endswith('attribute'):
                    attr_name = child.get('name', '')
                    attr_type = child.get('type', '')
                    attr_row = {
                        'levels': [name, attr_name] + [''] * (self.max_structure_depth - 2),
                        'Request Parameter': f"{name}.{attr_name}",
                        'GDPR': '',
                        'Cardinality': self._get_cardinality(child),
                        'Type': attr_type,
                        'Base Type': self._get_base_type(attr_type, simple_types),
                        'Details': self._get_details(child),
                        'Description': '',
                        'Category': 'attribute',
                        'Example': ''
                    }
                    rows.append(attr_row)
        
        return rows

    def _get_cardinality(self, elem):
        """Get cardinality from element"""
        min_occurs = elem.get('minOccurs', '1')
        max_occurs = elem.get('maxOccurs', '1')
        if max_occurs == 'unbounded':
            return '0..*' if min_occurs == '0' else '1..*'
        elif min_occurs == '0' and max_occurs == '1':
            return '0..1'
        elif min_occurs == '1' and max_occurs == '1':
            return '1'
        else:
            return f"{min_occurs}..{max_occurs}"

    def _get_base_type(self, type_name, simple_types):
        """Get base type from simple types"""
        if type_name in simple_types:
            return simple_types[type_name].get('base', type_name)
        return type_name

    def _get_details(self, elem):
        """Get additional details from element"""
        details = []
        if elem.get('nillable') == 'true':
            details.append('nillable')
        if elem.get('abstract') == 'true':
            details.append('abstract')
        return ', '.join(details) if details else ''

    def generate_mapping_from_schemas(self, source_schema, target_schema, source_struct=None, target_struct=None):
        """Generate mapping between source and target schemas like v8"""
        try:
            # Parse source XSD
            source_messages = self.parse_xsd_file(source_schema)
            
            # Parse target XSD
            target_rows = []
            target_messages = self.parse_xsd_file(target_schema)
            for msg_name, rows in target_messages.items():
                target_rows.extend(rows)
            
            # Create mapping structure
            mapping_data = {}
            
            for msg_name, src_full_rows in source_messages.items():
                max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
                max_tgt_level = max((len(row['levels']) for row in target_rows), default=1) if target_rows else 1
                
                # Create headers
                src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
                tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
                headers = src_cols + ['Destination Fields'] + tgt_cols
                
                # Create mapping rows
                mapping_rows = []
                tgt_path_dict = {'.'.join(row['levels']): row for row in target_rows}
                tgt_paths = list(tgt_path_dict.keys())
                
                for src_row in src_full_rows:
                    src_levels = src_row['levels'] + [''] * (max_src_level - len(src_row['levels']))
                    src_vals = src_levels + [
                        src_row.get('Request Parameter', ''),
                        src_row.get('GDPR', ''),
                        src_row.get('Cardinality', ''),
                        src_row.get('Type', ''),
                        src_row.get('Base Type', ''),
                        src_row.get('Details', ''),
                        src_row.get('Description', ''),
                        src_row.get('Category', ''),
                        src_row.get('Example', '')
                    ]
                    
                    # Find best match in target
                    src_path = '.'.join(src_row['levels'])
                    best_match = self._find_best_match(src_path, tgt_paths)
                    tgt_row = tgt_path_dict.get(best_match, {})
                    
                    tgt_levels = tgt_row.get('levels', []) + [''] * (max_tgt_level - len(tgt_row.get('levels', [])))
                    tgt_vals = tgt_levels + [
                        tgt_row.get('Request Parameter', ''),
                        tgt_row.get('GDPR', ''),
                        tgt_row.get('Cardinality', ''),
                        tgt_row.get('Type', ''),
                        tgt_row.get('Base Type', ''),
                        tgt_row.get('Details', ''),
                        tgt_row.get('Description', ''),
                        tgt_row.get('Category', ''),
                        tgt_row.get('Example', '')
                    ]
                    
                    # Create mapping row
                    mapping_row = src_vals + [best_match] + tgt_vals
                    mapping_rows.append(mapping_row)
                
                mapping_data[msg_name] = {
                    'headers': headers,
                    'rows': mapping_rows
                }
            
            return mapping_data
            
        except Exception as e:
            print(f"Error generating mapping: {e}")
            return {}

    def _find_best_match(self, src_path, tgt_paths):
        """Find best matching target path using similarity"""
        if not tgt_paths:
            return ''
        
        # Use difflib to find best match
        matcher = difflib.SequenceMatcher(None, src_path.lower(), '')
        best_match = ''
        best_ratio = 0
        
        for tgt_path in tgt_paths:
            matcher.set_seq2(tgt_path.lower())
            ratio = matcher.ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = tgt_path
        
        return best_match if best_ratio > 0.3 else '' 