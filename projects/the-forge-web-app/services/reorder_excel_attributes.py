import openpyxl
from collections import defaultdict
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Try to get the frontend handler from the main app
try:
    import sys
    if 'app' in sys.modules:
        from app import frontend_handler
        logger.addHandler(frontend_handler)
except ImportError:
    # If app module is not available, just use console logging
    pass

def reorder_attributes_in_excel(excel_path):
    logger.debug(f"Starting reorder_attributes_in_excel with file: {excel_path}")
    try:
        logger.debug("Loading workbook...")
        wb = openpyxl.load_workbook(excel_path)
        if not wb.worksheets:
            logger.error("No worksheets found in Excel file")
            raise Exception("No worksheets found in Excel file")
        
        logger.debug(f"Found {len(wb.worksheets)} worksheets")
        
        for ws in wb.worksheets:
            logger.debug(f"Processing worksheet: {ws.title}")
            header_row_1 = [cell.value for cell in ws[1]]
            header_row_2 = [cell.value for cell in ws[2]]
            logger.debug(f"Header row 1: {header_row_1}")
            
            # Validate that this is a mapping sheet
            try:
                level_start = header_row_1.index('Level1_src')
                logger.debug(f"Found Level1_src at column index: {level_start}")
            except ValueError:
                logger.debug("Level1_src not found - skipping this worksheet")
                continue  # Not a mapping sheet
            
            category_col = None
            for cat_col in ['Category_src', 'Category_tgt', 'Category']:
                if cat_col in header_row_1:
                    category_col = header_row_1.index(cat_col)
                    logger.debug(f"Found category column '{cat_col}' at index: {category_col}")
                    break
            
            if category_col is None:
                logger.debug("No category column found - skipping this worksheet")
                continue  # No category column found
            
            logger.debug("Getting data rows...")
            data_rows = list(ws.iter_rows(min_row=3, values_only=True))
            logger.debug(f"Found {len(data_rows)} data rows")
            
            if not data_rows:
                logger.debug("No data rows found - skipping this worksheet")
                continue  # No data rows found
            
            # Build a tree: key = tuple(levels), value = (row, children)
            logger.debug("Building node map...")
            node_map = {}
            root_nodes = []
            
            for idx, row in enumerate(data_rows):
                logger.debug(f"Processing row {idx + 3}: {row}")
                if len(row) <= level_start:
                    logger.debug(f"Row {idx + 3} too short, skipping")
                    continue  # Skip rows that don't have enough columns
                
                levels = row[level_start:level_start+8]
                nonempty_levels = [lvl for lvl in levels if lvl and str(lvl).strip()]
                key = tuple(nonempty_levels)
                logger.debug(f"Row {idx + 3} key: {key}")
                node_map[key] = {'row': row, 'children': []}
            
            logger.debug(f"Built node map with {len(node_map)} nodes")
            
            # Assign children to parents
            logger.debug("Assigning children to parents...")
            for key in node_map:
                if len(key) == 0:
                    logger.debug(f"Empty key {key} added to root_nodes")
                    root_nodes.append(key)
                    continue
                parent_key = key[:-1]
                if parent_key in node_map:
                    logger.debug(f"Adding {key} as child of {parent_key}")
                    node_map[parent_key]['children'].append(key)
                else:
                    logger.debug(f"Orphaned node {key} added to root_nodes")
                    root_nodes.append(key)  # Orphaned node
            
            # Recursive function to flatten tree with correct order
            def flatten_tree(node_key):
                logger.debug(f"Flattening tree for node: {node_key}")
                node = node_map[node_key]
                result = [node['row']]
                # Reorder children: attributes first, then elements/messages, preserving their order
                attr_keys = []
                elem_keys = []
                
                for child_key in node['children']:
                    child_row = node_map[child_key]['row']
                    logger.debug(f"Processing child {child_key} with row: {child_row}")
                    
                    if len(child_row) <= category_col:
                        logger.debug(f"Child {child_key} row too short for category column")
                        elem_keys.append(child_key)  # Default to element if category column doesn't exist
                        continue
                    
                    cat_val = str(child_row[category_col]).strip().lower() if child_row[category_col] else ''
                    logger.debug(f"Child {child_key} category: '{cat_val}'")
                    
                    if cat_val == 'attribute':
                        logger.debug(f"Adding {child_key} to attributes")
                        attr_keys.append(child_key)
                    else:
                        logger.debug(f"Adding {child_key} to elements")
                        elem_keys.append(child_key)
                
                logger.debug(f"Node {node_key}: {len(attr_keys)} attributes, {len(elem_keys)} elements")
                
                # Keep the original order among attributes and among elements/messages
                for child_key in attr_keys + elem_keys:
                    result.extend(flatten_tree(child_key))
                return result
            
            # Find all top-level nodes (those with no parent)
            logger.debug("Finding top-level nodes...")
            top_level_keys = [key for key in node_map if len(key) == 1 or key not in [child for n in node_map.values() for child in n['children']]]
            logger.debug(f"Found {len(top_level_keys)} top-level keys: {top_level_keys}")
            
            # If no top-level nodes found, fallback to all root_nodes
            if not top_level_keys:
                logger.debug("No top-level keys found, using root_nodes")
                top_level_keys = root_nodes
            
            # Flatten the tree
            logger.debug("Flattening tree...")
            new_data = []
            
            def safe_sort_key(k):
                logger.debug(f"Calculating sort key for: {k}")
                try:
                    index = data_rows.index(node_map[k]['row'])
                    logger.debug(f"Found row {k} at index: {index}")
                    return index
                except ValueError as e:
                    logger.warning(f"Row {k} not found in data_rows, using index 0. Error: {e}")
                    return 0  # Default to beginning if row not found
            
            logger.debug("Sorting top-level keys...")
            sorted_keys = sorted(top_level_keys, key=safe_sort_key)
            logger.debug(f"Sorted keys: {sorted_keys}")
            
            for key in sorted_keys:
                logger.debug(f"Processing sorted key: {key}")
                new_data.extend(flatten_tree(key))
            
            logger.debug(f"Generated {len(new_data)} new data rows")
            
            # Clear old data rows
            logger.debug("Clearing old data rows...")
            ws.delete_rows(3, ws.max_row-2)
            
            # Add new data
            logger.debug("Adding new data rows...")
            for i, row in enumerate(new_data):
                logger.debug(f"Adding row {i+1}: {row}")
                ws.append(row)
        
        logger.debug("Saving workbook...")
        wb.save(excel_path)
        wb.close()
        logger.debug("Attribute reordering completed successfully")
        
    except Exception as e:
        logger.error(f"Exception in reorder_attributes_in_excel: {e}", exc_info=True)
        raise Exception(f"Failed to reorder attributes in Excel: {str(e)}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 2:
        print("Usage: python reorder_excel_attributes.py <excel_path>")
        sys.exit(1)
    reorder_attributes_in_excel(sys.argv[1]) 