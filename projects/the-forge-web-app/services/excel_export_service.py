import openpyxl
from openpyxl.styles import Font

class ExcelExporter:
    def __init__(self, max_level=8):
        self.max_level = max_level

    def excel_sheet_name(self, name):
        return name[:31]

    def export(self, data_dict, output_file):
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        sheet_count = 0
        
        for sheet_name, data in data_dict.items():
            ws = wb.create_sheet(title=self.excel_sheet_name(sheet_name))
            
            if sheet_name == 'mapping':
                # Handle mapping data (new format)
                if 'headers' in data and 'rows' in data:
                    # Add headers
                    ws.append(data['headers'])
                    ws.append([''] * len(data['headers']))  # Second header row blank
                    
                    # Make headers bold
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    
                    # Add data rows
                    for row in data['rows']:
                        ws.append(row)
                else:
                    # Handle old mapping format
                    headers = ['Source Path', 'Target Path']
                    ws.append(headers)
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    for row in data:
                        ws.append([row.get('Source Path', ''), row.get('Target Path', '')])
            else:
                # Handle schema data (existing format)
                headers = [f'Level{i+1}' for i in range(self.max_level)] + ['Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Base Type', 'Details', 'Description', 'Category', 'Example']
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                # Sort rows: attributes first, then elements
                rows = sorted(data, key=lambda r: 0 if r.get('Category') == 'attribute' else 1)
                for row in rows:
                    ws.append(row['levels'] + [row['Request Parameter'], row['GDPR'], row['Cardinality'], row['Type'], row['Base Type'], row['Details'], row['Description'], row['Category'], row['Example']])
            
            sheet_count += 1
        
        if sheet_count > 0:
            wb.remove(default_sheet)
            wb.save(output_file)
            return True
        else:
            return False 