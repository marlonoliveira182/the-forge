import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, PatternFill
import re
from PIL import Image, ImageTk

# --- Funções para extração de caminhos de XSD e JSON Schema ---
def extract_paths_from_json_schema(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    paths = []
    def walk(obj, path=''):
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                new_path = f'{path}.{k}' if path else k
                paths.append(new_path)
                walk(v, new_path)
    walk(data)
    return paths

def extract_paths_from_xsd(filepath, keep_case=False):
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    paths = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    def to_camel_case(s):
        if not s:
            return s
        return s[0].lower() + s[1:]
    def get_name(name):
        return name if keep_case else to_camel_case(name)
    def walk_element(element, path='', is_root=False):
        name = element.get('name', '')
        if not name:
            return
        field_name = get_name(name)
        new_path = field_name if is_root else (f'{path}.{field_name}' if path else field_name)
        if not is_root:
            paths.append(new_path)
        typ = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        if typ in complex_types:
            walk_complex_type(complex_types[typ], new_path)
        if complex_type is not None:
            walk_complex_type(complex_type, new_path)
    def walk_complex_type(complex_type, path):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, path)
    for element in root.findall('xs:element', ns):
        walk_element(element, '', is_root=True)
    return paths

# --- Heurística de correspondência ---
try:
    from Levenshtein import ratio as levenshtein_ratio
except ImportError:
    def levenshtein_ratio(a, b):
        # Fallback simples: proporção de prefixo comum
        min_len = min(len(a), len(b))
        common = 0
        for i in range(min_len):
            if a[i] == b[i]:
                common += 1
            else:
                break
        return common / max(len(a), len(b)) if max(len(a), len(b)) > 0 else 0

def map_paths(source_paths, target_paths, threshold=0.7):
    mapping = []
    for s in source_paths:
        # Correspondência exata
        if s in target_paths:
            mapping.append({'source': s, 'target': s, 'similarity': 1.0})
            continue
        # Similaridade
        best = ('', 0.0)
        for t in target_paths:
            sim = levenshtein_ratio(s, t)
            if sim > best[1]:
                best = (t, sim)
        if best[1] >= threshold:
            mapping.append({'source': s, 'target': best[0], 'similarity': round(best[1], 3)})
        else:
            mapping.append({'source': s, 'target': '', 'similarity': 0.0})
    return mapping

# --- Extração detalhada de campos ---
def extract_fields_from_json_schema(filepath):
    import json
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    fields = []
    def walk(obj, levels=None, level=0):
        if levels is None:
            levels = []
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                lvls = levels + [k]  # Use property name as-is
                typ = v.get('type', '')
                desc = v.get('description', '')
                card = '1' if k in obj.get('required', []) else '0..1'
                details = []
                if 'enum' in v:
                    details.append(f"enum: {v['enum']}")
                if 'pattern' in v:
                    details.append(f"pattern: {v['pattern']}")
                if 'minLength' in v:
                    details.append(f"minLength: {v['minLength']}")
                if 'maxLength' in v:
                    details.append(f"maxLength: {v['maxLength']}")
                if 'minimum' in v:
                    details.append(f"minimum: {v['minimum']}")
                if 'maximum' in v:
                    details.append(f"maximum: {v['maximum']}")
                if 'format' in v:
                    details.append(f"format: {v['format']}")
                details_str = '; '.join(details)
                fields.append({
                    'levels': lvls,
                    'Type': typ,
                    'Description': desc,
                    'Cardinality': card,
                    'Details': details_str
                })
                walk(v, lvls, level+1)
    walk(data)
    return fields

def extract_xsd_restrictions(restriction):
    restr_details = []
    for r in restriction:
        tag = r.tag.split('}')[-1]
        val = r.get('value')
        if tag == 'enumeration' and val is not None:
            restr_details.append(f'enum: {val}')
        elif val is not None:
            restr_details.append(f'{tag}: {val}')
    return '; '.join(restr_details)

def extract_fields_from_xsd(filepath, keep_case=False):
    import xml.etree.ElementTree as ET
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    fields = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
    def to_camel_case(s):
        if not s:
            return s
        return s[0].lower() + s[1:]
    def get_name(name):
        return name if keep_case else to_camel_case(name)
    def get_type_and_details(element):
        typ = element.get('type', '')
        details = ''
        real_type = typ
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        elif typ in simple_types:
            restriction = simple_types[typ].find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        elif element.find('xs:complexType', ns) is not None:
            real_type = 'object'
        return real_type, details
    def walk_element(element, levels=None, omit_root=False):
        if levels is None:
            levels = []
        name = element.get('name', '')
        if not name:
            return
        field_name = get_name(name)
        lvls = levels if omit_root else levels + [field_name]
        desc = ''
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        card = f"{min_occurs}..{max_occurs}" if min_occurs != max_occurs else min_occurs
        typ, details = get_type_and_details(element)
        if not omit_root:
            fields.append({
                'levels': lvls,
                'Type': typ,
                'Description': desc,
                'Cardinality': card,
                'Details': details
            })
        typ_ref = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        if typ_ref in complex_types:
            walk_complex_type(complex_types[typ_ref], lvls)
        if complex_type is not None:
            walk_complex_type(complex_type, lvls)
    def walk_complex_type(complex_type, levels):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, levels)
    for element in root.findall('xs:element', ns):
        walk_element(element, [], omit_root=True)
    return fields

def fields_to_row(fields, max_levels):
    # Converte lista de dicts para lista de listas (linhas), preenchendo níveis
    rows = []
    for f in fields:
        lvls = f['levels']
        row = lvls + [''] * (max_levels - len(lvls))
        row += [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return rows

def get_max_levels(fields):
    return max((len(f['levels']) for f in fields), default=1)

def build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path):
    # Descobrir o máximo de níveis
    src_max_levels = get_max_levels(source_fields)
    tgt_max_levels = get_max_levels(target_fields)
    # v3-style headers
    src_headers = [f'Element Level {i+1}' for i in range(src_max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    tgt_headers = [f'Element Level {i+1}' for i in range(tgt_max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(out_headers)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    def get_req_param(fields, file_name):
        # Se o arquivo for .json, retorna body (json), senão body (xml)
        return 'body (json)' if file_name.lower().endswith('.json') else 'body (xml)'
    src_req_param = get_req_param(source_fields, src_name)
    tgt_req_param = get_req_param(target_fields, tgt_name)
    # Index target fields by full path string (lowercase for comparison)
    tgt_path_map = {'.'.join(f['levels']).lower(): f for f in target_fields}
    prev_src_levels = [''] * src_max_levels
    prev_tgt_levels = [''] * tgt_max_levels
    for i, src in enumerate(source_fields):
        src_path = '.'.join(src['levels'])
        tgt = tgt_path_map.get(src_path.lower(), None)
        # Não repetir parent nos níveis
        src_lvls = src['levels'] + [''] * (src_max_levels - len(src['levels']))
        src_lvls_disp = []
        for idx, val in enumerate(src_lvls):
            src_lvls_disp.append(val if val != prev_src_levels[idx] else '')
        prev_src_levels = [v if v else p for v, p in zip(src_lvls, prev_src_levels)]
        # v3-style row for source
        src_row = (
            src_lvls_disp +
            [src_req_param, '', src.get('Cardinality',''), src.get('Type',''), src.get('Details',''), src.get('Description','')]
        )
        # Preencher o destination field com o caminho do target (mesmo case do target)
        dest_field = '.'.join(tgt['levels']) if tgt else ''
        if tgt:
            tgt_lvls = tgt['levels'] + [''] * (tgt_max_levels - len(tgt['levels']))
            tgt_lvls_disp = []
            for idx, val in enumerate(tgt_lvls):
                tgt_lvls_disp.append(val if val != prev_tgt_levels[idx] else '')
            prev_tgt_levels = [v if v else p for v, p in zip(tgt_lvls, prev_tgt_levels)]
            tgt_row = (
                tgt_lvls_disp +
                [tgt_req_param, '', tgt.get('Cardinality',''), tgt.get('Type',''), tgt.get('Details',''), tgt.get('Description','')]
            )
        else:
            tgt_row = [''] * (tgt_max_levels + 6)
        ws.append(src_row + [dest_field] + tgt_row)
    wb.save(output_path)

# --- Geração do arquivo Excel ---
def write_mapping_excel(mapping, output_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Source Path', 'Target Path', 'Similarity'])
    for row in mapping:
        ws.append([row['source'], row['target'], row['similarity']])
    wb.save(output_path)

def parse_xsd(filepath):
    fields = extract_fields_from_xsd(filepath)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + ['Type', 'Description', 'Cardinality']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels'] + [''] * (max_levels - len(f['levels']))
        lvls_disp = [val if val != prev_levels[idx] else '' for idx, val in enumerate(lvls)]
        prev_levels = [v if v else p for v, p in zip(lvls, prev_levels)]
        row = lvls_disp + [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return columns, rows

def run_mapping_operation(source_path, target_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    target_ext = os.path.splitext(target_path)[1].lower()
    if source_ext == '.json':
        source_paths = extract_paths_from_json_schema(source_path)
        source_fields = extract_fields_from_json_schema(source_path)
    else:
        source_paths = extract_paths_from_xsd(source_path, keep_case=True)
        source_fields = extract_fields_from_xsd(source_path, keep_case=True)
    if target_ext == '.json':
        target_paths = extract_paths_from_json_schema(target_path)
        target_fields = extract_fields_from_json_schema(target_path)
    else:
        target_paths = extract_paths_from_xsd(target_path, keep_case=True)
        target_fields = extract_fields_from_xsd(target_path, keep_case=True)
    mapping = map_paths(source_paths, target_paths)
    src_name = os.path.basename(source_path)
    tgt_name = os.path.basename(target_path)
    output_path = os.path.join(dest_dir, f"mapping_{os.path.splitext(src_name)[0]}_to_{os.path.splitext(tgt_name)[0]}.xlsx")
    build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path)
    print(f"Mapping file saved:\n{output_path}")

def write_excel(rows, output_file, columns):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    # Mesclar células de níveis (apenas as células vazias abaixo do pai)
    num_rows = len(rows)
    level_cols = [i+1 for i, col in enumerate(columns) if col.startswith('Element Level')]
    for col in level_cols:
        row = 2
        while row <= num_rows+1:
            if ws.cell(row=row, column=col).value:
                # Encontrar bloco de células vazias abaixo
                start = row + 1
                end = start
                while end <= num_rows+1 and not ws.cell(row=end, column=col).value:
                    end += 1
                if end-1 > start:
                    ws.merge_cells(start_row=start, start_column=col, end_row=end-1, end_column=col)
                row = end
            else:
                row += 1
    wb.save(output_file)

def run_schema_to_excel_operation(source_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    is_json = source_ext == '.json'
    fields = extract_fields_from_json_schema(source_path) if is_json else extract_fields_from_xsd(source_path)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels']
        # Preencher colunas de níveis, só mostrar valor se diferente do anterior
        level_cells = []
        for idx in range(max_levels):
            val = lvls[idx] if idx < len(lvls) else ''
            disp = val if val != prev_levels[idx] else ''
            level_cells.append(disp)
        prev_levels = [v if v else p for v, p in zip([lvls[i] if i < len(lvls) else '' for i in range(max_levels)], prev_levels)]
        req_param = 'body (json)' if is_json else 'body (xml)'
        row = (
            level_cells +
            [req_param, '', f.get('Cardinality',''), f.get('Type',''), f.get('Details',''), f.get('Description','')]
        )
        rows.append(row)
    src_name = os.path.splitext(os.path.basename(source_path))[0]
    output_path = os.path.join(dest_dir, f"schema_{src_name}.xlsx")
    write_excel(rows, output_path, columns)
    # messagebox.showinfo("Concluído", f"Arquivo Excel salvo em:\n{output_path}")
    print(f"Excel file saved:\n{output_path}")

def to_camel_case(s):
    if not s:
        return s
    return s[0].lower() + s[1:]

def run_xsd_to_jsonschema_operation(source_path, dest_dir):
    import xml.etree.ElementTree as ET
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    tree = ET.parse(source_path)
    root = tree.getroot()
    simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}

    def xsd_type_to_json_type(xsd_type):
        if xsd_type.endswith(':string') or xsd_type == 'string':
            return 'string'
        elif xsd_type.endswith(':int') or xsd_type == 'int' or xsd_type.endswith(':integer') or xsd_type == 'integer':
            return 'integer'
        elif xsd_type.endswith(':boolean') or xsd_type == 'boolean':
            return 'boolean'
        elif xsd_type.endswith(':decimal') or xsd_type == 'decimal' or xsd_type.endswith(':float') or xsd_type == 'float':
            return 'number'
        else:
            return 'string'

    def extract_restrictions(element):
        details = {}
        restriction = element.find('xs:restriction', ns)
        if restriction is not None:
            for r in restriction:
                tag = r.tag.split('}')[-1]
                val = r.get('value')
                if tag == 'enumeration' and val is not None:
                    details.setdefault('enum', []).append(val)
                elif val is not None:
                    details[tag] = val
        return details

    def parse_element(element):
        name = to_camel_case(element.get('name', ''))
        typ = element.get('type', '')
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        # Extract description from annotation/documentation
        desc = ''
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        # Se for complexType referenciado
        if typ in complex_types:
            schema = parse_complex_type(complex_types[typ])
            if desc:
                schema['description'] = desc
            return name, schema
        # Se for complexType inline
        complex_type = element.find('xs:complexType', ns)
        if complex_type is not None:
            schema = parse_complex_type(complex_type)
            if desc:
                schema['description'] = desc
            return name, schema
        # Se for simpleType inline
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            base_type = restriction.get('base') if restriction is not None else 'string'
            prop = {'type': xsd_type_to_json_type(base_type)}
            details = extract_restrictions(simple_type)
            if 'enum' in details:
                prop['enum'] = details['enum']
            if 'pattern' in details:
                prop['pattern'] = details['pattern']
            if 'minLength' in details:
                prop['minLength'] = int(details['minLength'])
            if 'maxLength' in details:
                prop['maxLength'] = int(details['maxLength'])
            if desc:
                prop['description'] = desc
            return name, prop
        # Se for simpleType referenciado
        if typ in simple_types:
            restriction = simple_types[typ].find('xs:restriction', ns)
            base_type = restriction.get('base') if restriction is not None else 'string'
            prop = {'type': xsd_type_to_json_type(base_type)}
            details = extract_restrictions(simple_types[typ])
            if 'enum' in details:
                prop['enum'] = details['enum']
            if 'pattern' in details:
                prop['pattern'] = details['pattern']
            if 'minLength' in details:
                prop['minLength'] = int(details['minLength'])
            if 'maxLength' in details:
                prop['maxLength'] = int(details['maxLength'])
            if desc:
                prop['description'] = desc
            return name, prop
        # Se for tipo simples direto
        prop = {'type': xsd_type_to_json_type(typ)} if typ else {'type': 'string'}
        if desc:
            prop['description'] = desc
        return name, prop

    def parse_complex_type(complex_type):
        props = {}
        required = []
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                child_name, child_schema = parse_element(child)
                # Garantir camelCase recursivo
                child_name = to_camel_case(child_name)
                if child_schema.get('type') == 'object' and 'properties' in child_schema:
                    # Recursivamente garantir camelCase nas propriedades aninhadas
                    child_schema['properties'] = {to_camel_case(k): v for k, v in child_schema['properties'].items()}
                props[child_name] = child_schema
                min_occurs = child.get('minOccurs', '1')
                if min_occurs == '1':
                    required.append(child_name)
        return_schema = {'type': 'object', 'properties': props}
        if required:
            return_schema['required'] = required
        return return_schema

    def camelize_json_schema(obj):
        if isinstance(obj, dict):
            new_obj = {}
            for k, v in obj.items():
                # Always keep JSON Schema meta keys as-is
                if k in ('$schema', 'title', 'type', 'required', 'enum', 'pattern', 'minLength', 'maxLength', 'minimum', 'maximum', 'format', 'description', 'allOf', 'anyOf', 'oneOf', 'not', 'default', 'examples'):
                    new_obj[k] = camelize_json_schema(v)
                elif k == 'properties' and isinstance(v, dict):
                    # Recursively camelCase property keys and their values
                    new_obj[k] = {to_camel_case(pk): camelize_json_schema(pv) for pk, pv in v.items()}
                elif k == 'items' and isinstance(v, dict):
                    new_obj[k] = camelize_json_schema(v)
                else:
                    new_obj[to_camel_case(k)] = camelize_json_schema(v)
            return new_obj
        elif isinstance(obj, list):
            return [camelize_json_schema(i) for i in obj]
        else:
            return obj

    # --- Validation: collect all XSD element names (in camelCase) ---
    def collect_xsd_element_names(element, names=None):
        if names is None:
            names = set()
        name = element.get('name')
        if name:
            names.add(to_camel_case(name))
        # Check for children
        ns_seq = element.find('xs:complexType/xs:sequence', ns)
        if ns_seq is not None:
            for child in ns_seq.findall('xs:element', ns):
                collect_xsd_element_names(child, names)
        # Also check for direct xs:sequence (for complexType elements)
        seq = element.find('xs:sequence', ns)
        if seq is not None:
            for child in seq.findall('xs:element', ns):
                collect_xsd_element_names(child, names)
        return names

    # --- Validation: collect all property names from JSON Schema ---
    def collect_json_schema_property_names(schema, names=None):
        if names is None:
            names = set()
        if isinstance(schema, dict):
            for k, v in schema.items():
                if k == 'properties' and isinstance(v, dict):
                    for pk, pv in v.items():
                        names.add(pk)
                        collect_json_schema_property_names(pv, names)
                else:
                    collect_json_schema_property_names(v, names)
        elif isinstance(schema, list):
            for item in schema:
                collect_json_schema_property_names(item, names)
        return names

    # Encontrar o elemento root
    root_element = None
    for el in root.findall('xs:element', ns):
        root_element = el
        break
    if root_element is None:
        # messagebox.showerror("Erro", "Não foi possível encontrar o elemento root no XSD.")
        print("[ERRO] Não foi possível encontrar o elemento root no XSD.")
        return
    root_name = to_camel_case(root_element.get('name', 'Root'))
    root_schema = parse_element(root_element)[1]
    json_schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": root_name,
        **root_schema
    }
    # Recursively camelCase all property names in the schema
    json_schema = camelize_json_schema(json_schema)

    # --- Validation step ---
    xsd_names = collect_xsd_element_names(root_element)
    # Remove the root element name from the set
    root_camel = to_camel_case(root_element.get('name', ''))
    if root_camel in xsd_names:
        xsd_names.remove(root_camel)
    json_names = collect_json_schema_property_names(json_schema)
    missing = sorted(xsd_names - json_names)
    if missing:
        # messagebox.showerror("Erro de Validação", f"Os seguintes campos do XSD não foram encontrados no JSON Schema gerado (camelCase):\n\n" + "\n".join(missing))
        print("[ERRO] Os seguintes campos do XSD não foram encontrados no JSON Schema gerado (camelCase):\n\n" + "\n".join(missing))
        return

    out_name = root_name + '.schema.json'
    output_path = os.path.join(dest_dir, out_name)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_schema, f, indent=2)
    print('[DEBUG] JSON Schema salvo com sucesso')
    # messagebox.showinfo("Concluído", f"JSON Schema salvo em:\n{output_path}")
    print(f"[INFO] JSON Schema salvo em: {output_path}")

def jsonschema_to_xsd(json_schema_path, dest_dir):
    with open(json_schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    def to_pascal_case(s):
        return s[0].upper() + s[1:] if s else s
    def gen_xsd_element(name, obj):
        t = obj.get('type', 'string')
        # Restrições comuns
        restrictions = []
        if t == 'string':
            if 'maxLength' in obj:
                restrictions.append(f'<xs:maxLength value="{obj["maxLength"]}"/>')
            if 'minLength' in obj:
                restrictions.append(f'<xs:minLength value="{obj["minLength"]}"/>')
            if 'pattern' in obj:
                restrictions.append(f'<xs:pattern value="{obj["pattern"]}"/>')
            if 'enum' in obj:
                for v in obj['enum']:
                    restrictions.append(f'<xs:enumeration value="{v}"/>')
        elif t in ('integer', 'number'):
            if 'minimum' in obj:
                restrictions.append(f'<xs:minInclusive value="{obj["minimum"]}"/>')
            if 'maximum' in obj:
                restrictions.append(f'<xs:maxInclusive value="{obj["maximum"]}"/>')
            if 'enum' in obj:
                for v in obj['enum']:
                    restrictions.append(f'<xs:enumeration value="{v}"/>')
        # Geração do elemento
        if t == 'object':
            props = obj.get('properties', {})
            children = [gen_xsd_element(k, v) for k, v in props.items()]
            return f'<xs:element name="{to_pascal_case(name)}"><xs:complexType><xs:sequence>' + ''.join(children) + '</xs:sequence></xs:complexType></xs:element>'
        elif t == 'array':
            items = obj.get('items', {})
            return f'<xs:element name="{to_pascal_case(name)}" minOccurs="0" maxOccurs="unbounded">' + gen_xsd_element(name, items) + '</xs:element>'
        elif restrictions:
            base_type = 'xs:string' if t == 'string' else ('xs:int' if t == 'integer' else 'xs:decimal')
            restr = ''.join(restrictions)
            return f'<xs:element name="{to_pascal_case(name)}"><xs:simpleType><xs:restriction base="{base_type}">{restr}</xs:restriction></xs:simpleType></xs:element>'
        elif t == 'string':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:string"/>'
        elif t == 'integer':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:int"/>'
        elif t == 'number':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:decimal"/>'
        elif t == 'boolean':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:boolean"/>'
        else:
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:string"/>'
    title = schema.get('title', 'Root')
    xsd = f'<?xml version="1.0" encoding="UTF-8"?>\n<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">' + gen_xsd_element(title, schema) + '</xs:schema>'
    out_name = os.path.splitext(os.path.basename(json_schema_path))[0] + '.fromjson.xsd'
    output_path = os.path.join(dest_dir, out_name)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(xsd)

import tkinter.ttk as ttk
from tkinter import PhotoImage

RECENT_DIRS_FILE = os.path.join(os.path.dirname(__file__), 'recent_dirs.json')
MAX_RECENT_DIRS = 5

def main():
    class ForgeUI(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("The Forge")
            self.geometry("900x600")
            self.minsize(800, 500)
            self.mode = tk.StringVar(value="dark")
            self.logo_paths = {
                'dark': os.path.join(os.path.dirname(__file__), 'assets', 'EDP_Symbol_RGB_Dark_500px.png'),
                'light': os.path.join(os.path.dirname(__file__), 'assets', 'EDP_Symbol_RGB_Light_500px.png')
            }
            self._set_style()
            self._init_menu()
            self._build_header()
            self._build_tabs()

        def _update_logo(self):
            mode = self.mode.get()
            logo_path = self.logo_paths.get(mode, list(self.logo_paths.values())[0])
            img = Image.open(logo_path)
            # Redimensionar mantendo proporção, altura máxima 40px
            max_height = 40
            w, h = img.size
            scale = max_height / h
            new_w = int(w * scale)
            img = img.resize((new_w, max_height), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(img)
            self.logo_label.configure(image=self.logo_img, text="", width=new_w, height=max_height)

        def _get_header_bg(self):
            return "#F7F9FA" if self.mode.get() == "light" else "#212E3E"

        def _set_style(self):
            style = ttk.Style(self)
            if hasattr(self, 'mode') and self.mode.get() == 'light':
                bg = "#F7F9FA"
                fg = "#212E3E"
                accent = "#28FF52"
                tab_bg = "#D4FFDD"
                tab_fg = "#212E3E"
                tab_sel = "#28FF52"
                entry_bg = "#FFFFFF"
                entry_fg = "#212E3E"
                btn_bg = "#28FF52"
                btn_fg = "#212E3E"
                heading_fg = "#263CC8"
            else:
                bg = "#212E3E"
                fg = "#D4FFDD"
                accent = "#28FF52"
                tab_bg = "#424D5B"
                tab_fg = "#D4FFDD"
                tab_sel = "#263CC8"
                entry_bg = "#424D5B"
                entry_fg = "#28FF52"
                btn_bg = "#28FF52"
                btn_fg = "#212E3E"
                heading_fg = "#28FF52"
            self.configure(bg=bg)
            if hasattr(self, 'header'):
                self.header.configure(bg=self._get_header_bg())
            if hasattr(self, 'logo_label'):
                self.logo_label.configure(bg=self._get_header_bg())
            style.theme_use('clam')
            style.configure("EDP.TNotebook", background=bg, borderwidth=0)
            style.configure("EDP.TNotebook.Tab", background=tab_bg, foreground=tab_fg, font=("Segoe UI", 10, "bold"), padding=10)
            style.map("EDP.TNotebook.Tab",
                background=[('selected', tab_sel), ('active', accent)],
                foreground=[('selected', fg), ('active', accent)])
            style.configure("EDP.TFrame", background=bg)
            style.configure("EDP.TLabel", background=bg, foreground=fg, font=("Segoe UI", 10))
            style.configure("EDP.Heading.TLabel", background=bg, foreground=heading_fg, font=("Segoe UI", 13, "bold"))
            style.configure("EDP.TEntry", fieldbackground=entry_bg, foreground=entry_fg, borderwidth=1, relief="flat")
            style.configure("EDP.TButton", background=btn_bg, foreground=btn_fg, font=("Segoe UI", 10, "bold"), borderwidth=0, padding=6)
            style.map("EDP.TButton",
                background=[('active', accent), ('pressed', tab_sel)],
                foreground=[('active', btn_fg), ('pressed', fg)])
            if hasattr(self, 'toggle_btn'):
                if self.mode.get() == 'light':
                    self.toggle_btn.config(text="🌙 Dark Mode")
                else:
                    self.toggle_btn.config(text="☀ Light Mode")
            if hasattr(self, 'logo_label'):
                self._update_logo()

        def _toggle_mode(self):
            self.mode.set('light' if self.mode.get() == 'dark' else 'dark')
            self._set_style()

        def _set_entry_value(self, entry, value, placeholder):
            entry.config(state='normal')
            entry.delete(0, 'end')
            if value:
                entry.insert(0, value)
                entry.config(fg='#212E3E')
            else:
                entry.insert(0, placeholder)
                entry.config(fg='#212E3E')
            entry.config(state='readonly')

        def _add_placeholder(self, entry, placeholder):
            self._set_entry_value(entry, '', placeholder)

        def _add_selectbox_behavior(self, entry, select_callback):
            # Bind click to open file/directory dialog
            entry.bind('<Button-1>', lambda e: select_callback())
            # Change cursor to hand on hover
            entry.config(cursor='hand2')

        def _add_entry_with_icon(self, parent, textvariable, placeholder, select_callback):
            frame = ttk.Frame(parent, style="EDP.TFrame")
            entry = tk.Entry(frame, textvariable=textvariable, width=50, state='readonly')
            entry.pack(side='left', padx=(0,0), expand=True, fill='x')
            self._add_placeholder(entry, placeholder)
            self._add_selectbox_behavior(entry, select_callback)
            # Add folder icon (right side)
            icon_path = os.path.join(os.path.dirname(__file__), 'folder_icon.png')
            if os.path.exists(icon_path):
                icon_img = PhotoImage(file=icon_path)
                icon_label = tk.Label(frame, image=icon_img, borderwidth=0, bg=entry['bg'])
                icon_label.image = icon_img
                icon_label.pack(side='left', padx=(4,2))
                icon_label.bind('<Button-1>', lambda e: select_callback())
                icon_label.config(cursor='hand2')
            return frame, entry

        def _load_recent_dirs(self):
            try:
                with open(RECENT_DIRS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return []

        def _save_recent_dirs(self, dirs):
            try:
                with open(RECENT_DIRS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(dirs[:MAX_RECENT_DIRS], f)
            except Exception:
                pass

        def _add_dir_to_recent(self, dir_path):
            dirs = self._load_recent_dirs()
            if dir_path in dirs:
                dirs.remove(dir_path)
            dirs.insert(0, dir_path)
            self._save_recent_dirs(dirs)

        def _show_recent_dirs_menu(self, entry, var, placeholder):
            menu = tk.Menu(entry, tearoff=0)
            recent_dirs = self._load_recent_dirs()
            for d in recent_dirs[:MAX_RECENT_DIRS]:
                menu.add_command(label=d, command=lambda path=d: self._set_entry_value(entry, path, placeholder) or var.set(path))
            if recent_dirs:
                menu.add_separator()
            menu.add_command(label="Select directory...", command=lambda: self._browse_dir(var, entry, placeholder, from_menu=True))
            # Position menu below entry
            x = entry.winfo_rootx()
            y = entry.winfo_rooty() + entry.winfo_height()
            menu.tk_popup(x, y)

        def _add_entry_with_icon_dir(self, parent, textvariable, placeholder, on_select=None):
            frame = ttk.Frame(parent, style="EDP.TFrame")
            entry = tk.Entry(frame, textvariable=textvariable, width=50, state='readonly')
            entry.pack(side='left', padx=(0,0), expand=True, fill='x')
            self._add_placeholder(entry, placeholder)
            entry.config(cursor='hand2')
            if on_select:
                entry.bind('<Button-1>', lambda e: on_select())
            # Add folder icon (right side)
            icon_path = os.path.join(os.path.dirname(__file__), 'folder_icon.png')
            if os.path.exists(icon_path):
                icon_img = PhotoImage(file=icon_path)
                icon_label = tk.Label(frame, image=icon_img, borderwidth=0, bg=entry['bg'])
                icon_label.image = icon_img
                icon_label.pack(side='left', padx=(4,2))
                if on_select:
                    icon_label.bind('<Button-1>', lambda e: on_select())
                icon_label.config(cursor='hand2')
            return frame, entry

        def _set_recent_menu(self):
            # Add a 'Recent' menu to the main menu bar
            if not hasattr(self, 'recent_menu'):
                self.recent_menu = tk.Menu(self.menu, tearoff=0)
                self.menu.add_cascade(label='Recent', menu=self.recent_menu)
            self.recent_menu.delete(0, 'end')
            recent_dirs = self._load_recent_dirs()
            if not recent_dirs:
                self.recent_menu.add_command(label='No recent directories', state='disabled')
            else:
                for d in recent_dirs:
                    self.recent_menu.add_command(label=d, command=lambda path=d: self._set_recent_dir_to_field(path))

        def _set_recent_dir_to_field(self, dir_path):
            # Fill the currently focused output directory field with the selected recent dir
            if hasattr(self, 'last_output_entry') and self.last_output_entry:
                self._set_entry_value(self.last_output_entry['entry'], dir_path, self.last_output_entry['placeholder'])
                self.last_output_entry['var'].set(dir_path)

        def _browse_file(self, var, entry=None, placeholder=None):
            path = filedialog.askopenfilename(title="Select file", filetypes=[("Schemas", "*.xsd *.json"), ("All", "*.*")])
            if entry and placeholder:
                self._set_entry_value(entry, path, placeholder)
            if path:
                var.set(path)

        def _browse_dir(self, var, entry=None, placeholder=None, from_menu=False):
            path = filedialog.askdirectory(title="Select output directory")
            if entry and placeholder:
                self._set_entry_value(entry, path, placeholder)
            if path:
                var.set(path)
                self._add_dir_to_recent(path)
                self._set_recent_menu()

        def _build_tabs(self):
            self.notebook = ttk.Notebook(self, style="EDP.TNotebook")
            self.notebook.pack(expand=True, fill='both', padx=0, pady=(0,0))
            self.tabs = {}
            # Format conversion
            self.tabs['xsd_to_jsonschema'] = self._build_xsd_to_jsonschema_tab()
            self.notebook.add(self.tabs['xsd_to_jsonschema'], text="XSD to JSON")
            self.tabs['jsonschema_to_xsd'] = self._build_jsonschema_to_xsd_tab()
            self.notebook.add(self.tabs['jsonschema_to_xsd'], text="JSON to XSD")
            # Documentation
            self.tabs['schema_to_excel'] = self._build_schema_to_excel_tab()
            self.notebook.add(self.tabs['schema_to_excel'], text="Schema to Excel")
            # Mapping
            self.tabs['mapping'] = self._build_mapping_tab()
            self.notebook.add(self.tabs['mapping'], text="Mapping")

        # Ajustar padding e espaçamento das abas e campos
        def _build_mapping_tab(self):
            frame = ttk.Frame(self.notebook, style="EDP.TFrame")
            ttk.Label(frame, text="Mapping between two Schemas", style="EDP.Heading.TLabel").pack(pady=(18,4))
            # Source
            src_frame = ttk.Frame(frame, style="EDP.TFrame")
            src_frame.pack(fill='x', padx=8, pady=6)
            src_path = tk.StringVar()
            ttk.Label(src_frame, text="Source Schema:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            src_entry_frame, src_entry = self._add_entry_with_icon(src_frame, src_path, "Select source schema file...", lambda: self._browse_file(src_path, src_entry, "Select source schema file..."))
            src_entry_frame.pack(side='left', fill='x', expand=True)
            # Target
            tgt_frame = ttk.Frame(frame, style="EDP.TFrame")
            tgt_frame.pack(fill='x', padx=8, pady=6)
            tgt_path = tk.StringVar()
            ttk.Label(tgt_frame, text="Target Schema:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            tgt_entry_frame, tgt_entry = self._add_entry_with_icon(tgt_frame, tgt_path, "Select target schema file...", lambda: self._browse_file(tgt_path, tgt_entry, "Select target schema file..."))
            tgt_entry_frame.pack(side='left', fill='x', expand=True)
            # Output dir
            out_frame = ttk.Frame(frame, style="EDP.TFrame")
            out_frame.pack(fill='x', padx=8, pady=6)
            out_dir = tk.StringVar()
            ttk.Label(out_frame, text="Output Directory:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            def open_dir_dialog():
                self.last_output_entry = {'entry': out_entry, 'placeholder': "Select output directory...", 'var': out_dir}
                self._browse_dir(out_dir, out_entry, "Select output directory...")
            out_entry_frame, out_entry = self._add_entry_with_icon_dir(out_frame, out_dir, "Select output directory...", open_dir_dialog)
            out_entry_frame.pack(side='left', fill='x', expand=True)
            ttk.Button(frame, text="Generate Mapping", style="EDP.TButton", command=lambda: self._run_mapping(src_path, tgt_path, out_dir)).pack(pady=18)
            return frame

        def _build_schema_to_excel_tab(self):
            frame = ttk.Frame(self.notebook, style="EDP.TFrame")
            ttk.Label(frame, text="Convert Schema to Excel", style="EDP.Heading.TLabel").pack(pady=(18,4))
            # Source
            src_frame = ttk.Frame(frame, style="EDP.TFrame")
            src_frame.pack(fill='x', padx=8, pady=6)
            src_path = tk.StringVar()
            ttk.Label(src_frame, text="Schema:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            src_entry_frame, src_entry = self._add_entry_with_icon(src_frame, src_path, "Select schema file...", lambda: self._browse_file(src_path, src_entry, "Select schema file..."))
            src_entry_frame.pack(side='left', fill='x', expand=True)
            out_frame = ttk.Frame(frame, style="EDP.TFrame")
            out_frame.pack(fill='x', padx=8, pady=6)
            out_dir = tk.StringVar()
            ttk.Label(out_frame, text="Output Directory:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            def open_dir_dialog():
                self.last_output_entry = {'entry': out_entry, 'placeholder': "Select output directory...", 'var': out_dir}
                self._browse_dir(out_dir, out_entry, "Select output directory...")
            out_entry_frame, out_entry = self._add_entry_with_icon_dir(out_frame, out_dir, "Select output directory...", open_dir_dialog)
            out_entry_frame.pack(side='left', fill='x', expand=True)
            ttk.Button(frame, text="Convert to Excel", style="EDP.TButton", command=lambda: self._run_schema_to_excel(src_path, out_dir)).pack(pady=18)
            return frame

        def _build_xsd_to_jsonschema_tab(self):
            frame = ttk.Frame(self.notebook, style="EDP.TFrame")
            ttk.Label(frame, text="Convert XSD to JSON Schema", style="EDP.Heading.TLabel").pack(pady=(18,4))
            # Source
            src_frame = ttk.Frame(frame, style="EDP.TFrame")
            src_frame.pack(fill='x', padx=8, pady=6)
            src_path = tk.StringVar()
            ttk.Label(src_frame, text="XSD:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            src_entry_frame, src_entry = self._add_entry_with_icon(src_frame, src_path, "Select XSD file...", lambda: self._browse_file(src_path, src_entry, "Select XSD file..."))
            src_entry_frame.pack(side='left', fill='x', expand=True)
            out_frame = ttk.Frame(frame, style="EDP.TFrame")
            out_frame.pack(fill='x', padx=8, pady=6)
            out_dir = tk.StringVar()
            ttk.Label(out_frame, text="Output Directory:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            def open_dir_dialog():
                self.last_output_entry = {'entry': out_entry, 'placeholder': "Select output directory...", 'var': out_dir}
                self._browse_dir(out_dir, out_entry, "Select output directory...")
            out_entry_frame, out_entry = self._add_entry_with_icon_dir(out_frame, out_dir, "Select output directory...", open_dir_dialog)
            out_entry_frame.pack(side='left', fill='x', expand=True)
            ttk.Button(frame, text="Convert to JSON Schema", style="EDP.TButton", command=lambda s=self, sp=src_path, od=out_dir: s._run_xsd_to_jsonschema(sp, od)).pack(pady=18)
            return frame

        def _build_jsonschema_to_xsd_tab(self):
            frame = ttk.Frame(self.notebook, style="EDP.TFrame")
            ttk.Label(frame, text="Convert JSON Schema to XSD", style="EDP.Heading.TLabel").pack(pady=(18,4))
            # Source
            src_frame = ttk.Frame(frame, style="EDP.TFrame")
            src_frame.pack(fill='x', padx=8, pady=6)
            src_path = tk.StringVar()
            ttk.Label(src_frame, text="JSON Schema:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            src_entry_frame, src_entry = self._add_entry_with_icon(src_frame, src_path, "Select JSON Schema file...", lambda: self._browse_file(src_path, src_entry, "Select JSON Schema file..."))
            src_entry_frame.pack(side='left', fill='x', expand=True)
            out_frame = ttk.Frame(frame, style="EDP.TFrame")
            out_frame.pack(fill='x', padx=8, pady=6)
            out_dir = tk.StringVar()
            ttk.Label(out_frame, text="Output Directory:", style="EDP.TLabel").pack(side='left', padx=(0,6))
            def open_dir_dialog():
                self.last_output_entry = {'entry': out_entry, 'placeholder': "Select output directory...", 'var': out_dir}
                self._browse_dir(out_dir, out_entry, "Select output directory...")
            out_entry_frame, out_entry = self._add_entry_with_icon_dir(out_frame, out_dir, "Select output directory...", open_dir_dialog)
            out_entry_frame.pack(side='left', fill='x', expand=True)
            ttk.Button(frame, text="Convert to XSD", style="EDP.TButton", command=lambda: self._run_jsonschema_to_xsd(src_path, out_dir)).pack(pady=18)
            return frame

        def _init_menu(self):
            self.menu = tk.Menu(self)
            self.config(menu=self.menu)
            self.file_menu = tk.Menu(self.menu, tearoff=0)
            self.menu.add_cascade(label='File', menu=self.file_menu)
            self.file_menu.add_command(label='Exit', command=self.destroy)
            self._set_recent_menu()

        def _build_header(self):
            self.header_bg = "#212E3E" if self.mode.get() == 'dark' else "#F7F9FA"
            self.header = tk.Frame(self, bg=self.header_bg, height=64)
            self.header.pack(side='top', fill='x')
            # Logo (from assets)
            assets_dir = os.path.join(os.path.dirname(__file__), 'assets')
            logo_path_light = os.path.join(assets_dir, 'EDP_Symbol_RGB_Light_500px.png')
            logo_path_dark = os.path.join(assets_dir, 'EDP_Symbol_RGB_Dark_500px.png')
            self.logo_img_light = self._load_logo_image(logo_path_light, 'light')
            self.logo_img_dark = self._load_logo_image(logo_path_dark, 'dark')
            self.logo_label = tk.Label(self.header, bg=self.header_bg, borderwidth=0, highlightthickness=0)
            self.logo_label.pack(side='left', padx=(18, 8), pady=8)
            self._update_logo()
            # Spacer
            header_spacer = tk.Frame(self.header, bg=self.header_bg)
            header_spacer.pack(side='left', expand=True, fill='x')
            # Mode toggle
            self.mode_btn = ttk.Button(self.header, text="Dark Mode" if self.mode.get() == 'light' else "Light Mode", style="EDP.TButton", command=self._toggle_mode)
            self.mode_btn.pack(side='right', padx=(0, 18), pady=8)
            # Help button
            self.help_btn = ttk.Button(self.header, text="Help", style="EDP.TButton", command=self._show_help)
            self.help_btn.pack(side='right', padx=(0, 8), pady=8)

        def _load_logo_image(self, path, mode):
            if os.path.exists(path):
                try:
                    img = Image.open(path)
                    max_h = 40
                    if img.height > max_h:
                        scale = max_h / img.height
                        new_w = int(img.width * scale)
                        img = img.resize((new_w, max_h), Image.LANCZOS)
                    return ImageTk.PhotoImage(img)
                except Exception as e:
                    print(f"[LOGO ERROR] Could not load {mode} logo: {e}")
                    return None
            else:
                print(f"[LOGO ERROR] Logo file not found: {path}")
            return None

        def _update_logo(self):
            self.header_bg = "#212E3E" if self.mode.get() == 'dark' else "#F7F9FA"
            self.logo_label.config(bg=self.header_bg)
            logo_img = self.logo_img_dark if self.mode.get() == 'dark' else self.logo_img_light
            if logo_img:
                self.logo_label.config(image=logo_img)
                self.logo_label.image = logo_img
            else:
                self.logo_label.config(image='')
                self.logo_label.image = None

        def _toggle_mode(self):
            if self.mode.get() == 'light':
                self.mode.set('dark')
            else:
                self.mode.set('light')
            self.header_bg = "#212E3E" if self.mode.get() == 'dark' else "#F7F9FA"
            self.header.config(bg=self.header_bg)
            self.logo_label.config(bg=self.header_bg)
            for widget in self.header.winfo_children():
                if isinstance(widget, tk.Frame):
                    widget.config(bg=self.header_bg)
                elif isinstance(widget, tk.Button):
                    widget.config(bg=self.header_bg, fg="#28FF52", activebackground=self.header_bg, activeforeground="#28FF52")
            self._set_style()
            self._update_logo()
            self._refresh_ui_colors()
            self._rebuild_header_colors()
            self.mode_btn.config(text="Dark Mode" if self.mode.get() == 'light' else "Light Mode")

        def _rebuild_header_colors(self):
            header_bg = "#212E3E" if self.mode.get() == 'dark' else "#F7F9FA"
            title_fg = "#28FF52" if self.mode.get() == 'dark' else "#263CC8"
            btn_fg = title_fg
            btn_bg = header_bg
            self.logo_label.config(bg=header_bg)
            # Atualiza todos os filhos do header
            for widget in self.logo_label.master.winfo_children():
                if isinstance(widget, tk.Label) and widget != self.logo_label:
                    widget.config(bg=header_bg, fg=title_fg)
                elif isinstance(widget, tk.Frame):
                    widget.config(bg=header_bg)
                elif isinstance(widget, tk.Button):
                    widget.config(bg=btn_bg, fg=btn_fg, activebackground=btn_bg, activeforeground=btn_fg)

        def _refresh_ui_colors(self):
            # Redraw/recolor widgets as needed (minimal for ttk, but can be extended)
            self.update_idletasks()

        def _show_help(self):
            help_data = [
                ("XSD to JSON Schema", "Converts an XSD (XML Schema) file into a JSON Schema file. Useful for integrating XML-based systems with JSON-based APIs.", [
                    "Select the XSD file.",
                    "Choose the output directory.",
                    "Click 'Convert to JSON Schema'."
                ]),
                ("JSON Schema to XSD", "Converts a JSON Schema file into an XSD (XML Schema) file. Useful for generating XML validation schemas from JSON-based definitions.", [
                    "Select the JSON Schema file.",
                    "Choose the output directory.",
                    "Click 'Convert to XSD'."
                ]),
                ("Schema to Excel", "Generates an Excel spreadsheet that documents the structure of a schema (XSD or JSON Schema). Useful for analysis, documentation, or sharing with non-technical stakeholders.", [
                    "Select the schema file (XSD or JSON Schema).",
                    "Choose the output directory.",
                    "Click 'Convert to Excel'."
                ]),
                ("Mapping", "Creates a mapping (correspondence) between two schemas (XSD or JSON Schema). Useful for data migration, integration, or transformation projects.", [
                    "Select the source and target schemas (XSD or JSON Schema).",
                    "Choose the output directory.",
                    "Click 'Generate Mapping'."
                ])
            ]
            win = tk.Toplevel(self)
            win.title("Help - How to use")
            win.resizable(False, False)
            win.geometry("650x600")
            bg = "#212E3E" if self.mode.get()=="dark" else "#F7F9FA"
            fg = "#28FF52" if self.mode.get()=="dark" else "#263CC8"
            title_fg = fg
            section_title_fg = fg
            desc_fg = "#F7F9FA" if self.mode.get()=="dark" else "#212E3E"
            step_fg = fg
            win.configure(bg=bg)
            # Título
            title = tk.Label(win, text="How to use each operation", font=("Segoe UI", 16, "bold"), fg=title_fg, bg=bg)
            title.pack(pady=(18, 8))
            # Frame de rolagem
            scroll_frame = tk.Frame(win, bg=bg)
            scroll_frame.pack(expand=True, fill='both', padx=18, pady=(0,0))
            canvas = tk.Canvas(scroll_frame, bg=bg, highlightthickness=0, borderwidth=0)
            scrollbar = tk.Scrollbar(scroll_frame, orient='vertical', command=canvas.yview)
            scrollable = tk.Frame(canvas, bg=bg)
            scrollable_id = canvas.create_window((0,0), window=scrollable, anchor='nw')
            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            canvas.configure(yscrollcommand=scrollbar.set)
            def on_configure(event):
                canvas.configure(scrollregion=canvas.bbox('all'))
            scrollable.bind('<Configure>', on_configure)
            # Expande/colapsa
            expanded = {}
            def toggle(idx):
                expanded[idx] = not expanded.get(idx, False)
                for i, (frame, content) in enumerate(op_frames):
                    if i == idx:
                        if expanded[idx]:
                            content.pack(fill='x', pady=(0,12))
                            op_btns[i].config(text='▼ ' + help_data[i][0])
                        else:
                            content.pack_forget()
                            op_btns[i].config(text='► ' + help_data[i][0])
            op_frames = []
            op_btns = []
            for idx, (op_title, op_desc, op_steps) in enumerate(help_data):
                op_frame = tk.Frame(scrollable, bg=bg)
                op_frame.pack(fill='x', pady=(0,0))
                btn = tk.Button(op_frame, text='► ' + op_title, font=("Segoe UI", 13, "bold"), fg=section_title_fg, bg=bg, bd=0, activebackground=bg, activeforeground=section_title_fg, cursor='hand2', anchor='w', command=lambda idx=idx: toggle(idx))
                btn.pack(fill='x', pady=(8,0), padx=0)
                op_btns.append(btn)
                content = tk.Frame(op_frame, bg=bg)
                desc = tk.Label(content, text=op_desc, font=("Segoe UI", 11), fg=desc_fg, bg=bg, anchor='w', justify='left', wraplength=600)
                desc.pack(fill='x', padx=(18,0), pady=(2,4), anchor='w')
                steps_title = tk.Label(content, text='Steps:', font=("Segoe UI", 11, "bold"), fg=step_fg, bg=bg, anchor='w', wraplength=600)
                steps_title.pack(fill='x', padx=(30,0), anchor='w')
                for i, step in enumerate(op_steps, 1):
                    step_lbl = tk.Label(content, text=f"  {i}. {step}", font=("Segoe UI", 11), fg=desc_fg, bg=bg, anchor='w', justify='left', wraplength=600)
                    step_lbl.pack(fill='x', padx=(42,0), anchor='w')
                op_frames.append((op_frame, content))
                expanded[idx] = False
            # Botão Close centralizado
            btn_frame = tk.Frame(win, bg=bg)
            btn_frame.pack(fill='x', pady=(0,18))
            close_btn = ttk.Button(btn_frame, text="Close", style="EDP.TButton", command=win.destroy)
            close_btn.pack(pady=0)
            win.transient(self)
            win.grab_set()

        def _run_mapping(self, src_path_var, tgt_path_var, out_dir_var):
            src_path = src_path_var.get()
            tgt_path = tgt_path_var.get()
            out_dir = out_dir_var.get()
            if not src_path or not tgt_path or not out_dir:
                messagebox.showerror("Error", "Please select all three files (source schema, target schema, and output directory).")
                return
            try:
                run_mapping_operation(src_path, tgt_path, out_dir)
                messagebox.showinfo("Success", f"Mapping file generated in:\n{out_dir}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate mapping:\n{e}")

        def _run_schema_to_excel(self, src_path_var, out_dir_var):
            src_path = src_path_var.get()
            out_dir = out_dir_var.get()
            if not src_path or not out_dir:
                messagebox.showerror("Error", "Please select both the schema file and output directory.")
                return
            try:
                run_schema_to_excel_operation(src_path, out_dir)
                messagebox.showinfo("Success", f"Excel file generated in:\n{out_dir}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to convert schema to Excel:\n{e}")

        def _run_xsd_to_jsonschema(self, src_path_var, out_dir_var):
            src_path = src_path_var.get()
            out_dir = out_dir_var.get()
            if not src_path or not out_dir:
                messagebox.showerror("Error", "Please select both the XSD file and output directory.")
        return
            try:
                run_xsd_to_jsonschema_operation(src_path, out_dir)
                messagebox.showinfo("Success", f"JSON Schema file generated in:\n{out_dir}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to convert XSD to JSON Schema:\n{e}")

        def _run_jsonschema_to_xsd(self, src_path_var, out_dir_var):
            src_path = src_path_var.get()
            out_dir = out_dir_var.get()
            if not src_path or not out_dir:
                messagebox.showerror("Error", "Please select both the JSON Schema file and output directory.")
                return
            try:
                jsonschema_to_xsd(src_path, out_dir)
                messagebox.showinfo("Success", f"XSD file generated in:\n{out_dir}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to convert JSON Schema to XSD:\n{e}")

    app = ForgeUI()
    app.mainloop()

if __name__ == "__main__":
    main() 