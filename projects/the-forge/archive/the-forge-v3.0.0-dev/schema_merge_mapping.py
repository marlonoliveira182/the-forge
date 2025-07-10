import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, PatternFill

# --- Funções para extração de caminhos de XSD e JSON Schema ---
def extract_paths_from_json_schema(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    paths = []
    def walk(obj, path=''):
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                new_path = f'{path}.{k}' if path else k
                paths.append(new_path)
                walk(v, new_path)
    walk(data)
    return paths

def extract_paths_from_xsd(filepath):
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    paths = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    def walk_element(element, path=''):
        name = element.get('name', '')
        if not name:
            return
        new_path = f'{path}.{name}' if path else name
        paths.append(new_path)
        typ = element.get('type', '')
        # Inline complexType
        complex_type = element.find('xs:complexType', ns)
        if typ in complex_types:
            walk_complex_type(complex_types[typ], new_path)
        if complex_type is not None:
            walk_complex_type(complex_type, new_path)
    def walk_complex_type(complex_type, path):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, path)
    for element in root.findall('xs:element', ns):
        walk_element(element, '')
    return paths

# --- Heurística de correspondência ---
try:
    from Levenshtein import ratio as levenshtein_ratio
except ImportError:
    def levenshtein_ratio(a, b):
        # Fallback simples: proporção de prefixo comum
        min_len = min(len(a), len(b))
        common = 0
        for i in range(min_len):
            if a[i] == b[i]:
                common += 1
            else:
                break
        return common / max(len(a), len(b)) if max(len(a), len(b)) > 0 else 0

def map_paths(source_paths, target_paths, threshold=0.7):
    mapping = []
    for s in source_paths:
        # Correspondência exata
        if s in target_paths:
            mapping.append({'source': s, 'target': s, 'similarity': 1.0})
            continue
        # Similaridade
        best = ('', 0.0)
        for t in target_paths:
            sim = levenshtein_ratio(s, t)
            if sim > best[1]:
                best = (t, sim)
        if best[1] >= threshold:
            mapping.append({'source': s, 'target': best[0], 'similarity': round(best[1], 3)})
        else:
            mapping.append({'source': s, 'target': '', 'similarity': 0.0})
    return mapping

# --- Extração detalhada de campos ---
def extract_fields_from_json_schema(filepath):
    import json
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    fields = []
    def walk(obj, levels=None, level=0):
        if levels is None:
            levels = []
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                lvls = levels + [k]
                typ = v.get('type', '')
                desc = v.get('description', '')
                card = '1' if k in obj.get('required', []) else '0..1'
                details = []
                if 'enum' in v:
                    details.append(f"enum: {v['enum']}")
                if 'pattern' in v:
                    details.append(f"pattern: {v['pattern']}")
                if 'minLength' in v:
                    details.append(f"minLength: {v['minLength']}")
                if 'maxLength' in v:
                    details.append(f"maxLength: {v['maxLength']}")
                if 'minimum' in v:
                    details.append(f"minimum: {v['minimum']}")
                if 'maximum' in v:
                    details.append(f"maximum: {v['maximum']}")
                if 'format' in v:
                    details.append(f"format: {v['format']}")
                details_str = '; '.join(details)
                fields.append({
                    'levels': lvls,
                    'Type': typ,
                    'Description': desc,
                    'Cardinality': card,
                    'Details': details_str
                })
                walk(v, lvls, level+1)
    walk(data)
    return fields

def extract_xsd_restrictions(restriction):
    restr_details = []
    for r in restriction:
        tag = r.tag.split('}')[-1]
        val = r.get('value')
        if tag == 'enumeration' and val is not None:
            restr_details.append(f'enum: {val}')
        elif val is not None:
            restr_details.append(f'{tag}: {val}')
    return '; '.join(restr_details)

def extract_fields_from_xsd(filepath):
    import xml.etree.ElementTree as ET
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    fields = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
    def get_type_and_details(element):
        typ = element.get('type', '')
        details = ''
        real_type = typ
        # Inline simpleType
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        # Referenciado simpleType
        elif typ in simple_types:
            restriction = simple_types[typ].find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        # Inline complexType
        elif element.find('xs:complexType', ns) is not None:
            real_type = 'object'
        return real_type, details
    def walk_element(element, levels=None):
        if levels is None:
            levels = []
        name = element.get('name', '')
        if not name:
            return
        lvls = levels + [name]
        desc = ''
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        card = f"{min_occurs}..{max_occurs}" if min_occurs != max_occurs else min_occurs
        typ, details = get_type_and_details(element)
        fields.append({
            'levels': lvls,
            'Type': typ,
            'Description': desc,
            'Cardinality': card,
            'Details': details
        })
        typ_ref = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        if typ_ref in complex_types:
            walk_complex_type(complex_types[typ_ref], lvls)
        if complex_type is not None:
            walk_complex_type(complex_type, lvls)
    def walk_complex_type(complex_type, levels):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, levels)
    for element in root.findall('xs:element', ns):
        walk_element(element, [])
    return fields

def fields_to_row(fields, max_levels):
    # Converte lista de dicts para lista de listas (linhas), preenchendo níveis
    rows = []
    for f in fields:
        lvls = f['levels']
        row = lvls + [''] * (max_levels - len(lvls))
        row += [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return rows

def get_max_levels(fields):
    return max((len(f['levels']) for f in fields), default=1)

def build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path):
    # Descobrir o máximo de níveis
    src_max_levels = get_max_levels(source_fields)
    tgt_max_levels = get_max_levels(target_fields)
    src_headers = [f'Element Level {i+1}' for i in range(src_max_levels)] + ['Type', 'Description', 'Cardinality']
    tgt_headers = [f'Element Level {i+1}' for i in range(tgt_max_levels)] + ['Type', 'Description', 'Cardinality']
    out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(out_headers)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    # Index target fields by path
    tgt_path_map = {'.'.join(f['levels']): f for f in target_fields}
    prev_src_levels = [''] * src_max_levels
    prev_tgt_levels = [''] * tgt_max_levels
    for i, src in enumerate(source_fields):
        src_path = '.'.join(src['levels'])
        # Mapping: encontrar target correspondente
        m = mapping[i]
        tgt_path = m['target']
        tgt = tgt_path_map.get(tgt_path, None)
        # Não repetir parent nos níveis
        src_lvls = src['levels'] + [''] * (src_max_levels - len(src['levels']))
        src_lvls_disp = []
        for idx, val in enumerate(src_lvls):
            src_lvls_disp.append(val if val != prev_src_levels[idx] else '')
        prev_src_levels = [v if v else p for v, p in zip(src_lvls, prev_src_levels)]
        src_row = src_lvls_disp + [src.get('Type',''), src.get('Description',''), src.get('Cardinality','')]
        dest_field = tgt_path if tgt else ''
        if tgt:
            tgt_lvls = tgt['levels'] + [''] * (tgt_max_levels - len(tgt['levels']))
            tgt_lvls_disp = []
            for idx, val in enumerate(tgt_lvls):
                tgt_lvls_disp.append(val if val != prev_tgt_levels[idx] else '')
            prev_tgt_levels = [v if v else p for v, p in zip(tgt_lvls, prev_tgt_levels)]
            tgt_row = tgt_lvls_disp + [tgt.get('Type',''), tgt.get('Description',''), tgt.get('Cardinality','')]
        else:
            tgt_row = [''] * (tgt_max_levels + 3)
        ws.append(src_row + [dest_field] + tgt_row)
    wb.save(output_path)

# --- Geração do arquivo Excel ---
def write_mapping_excel(mapping, output_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Source Path', 'Target Path', 'Similarity'])
    for row in mapping:
        ws.append([row['source'], row['target'], row['similarity']])
    wb.save(output_path)

def parse_xsd(filepath):
    fields = extract_fields_from_xsd(filepath)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + ['Type', 'Description', 'Cardinality']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels'] + [''] * (max_levels - len(f['levels']))
        lvls_disp = [val if val != prev_levels[idx] else '' for idx, val in enumerate(lvls)]
        prev_levels = [v if v else p for v, p in zip(lvls, prev_levels)]
        row = lvls_disp + [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return columns, rows

def run_mapping_operation(source_path, target_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    target_ext = os.path.splitext(target_path)[1].lower()
    if source_ext == '.json':
        source_paths = extract_paths_from_json_schema(source_path)
        source_fields = extract_fields_from_json_schema(source_path)
    else:
        source_paths = extract_paths_from_xsd(source_path)
        source_fields = extract_fields_from_xsd(source_path)
    if target_ext == '.json':
        target_paths = extract_paths_from_json_schema(target_path)
        target_fields = extract_fields_from_json_schema(target_path)
    else:
        target_paths = extract_paths_from_xsd(target_path)
        target_fields = extract_fields_from_xsd(target_path)
    mapping = map_paths(source_paths, target_paths)
    src_name = os.path.splitext(os.path.basename(source_path))[0]
    tgt_name = os.path.splitext(os.path.basename(target_path))[0]
    output_path = os.path.join(dest_dir, f"mapping_{src_name}_to_{tgt_name}.xlsx")
    build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path)
    messagebox.showinfo("Concluído", f"Mapping file salvo em:\n{output_path}")

def write_excel(rows, output_file, columns):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    # Mesclar células de níveis (apenas as células vazias abaixo do pai)
    num_rows = len(rows)
    level_cols = [i+1 for i, col in enumerate(columns) if col.startswith('Element Level')]
    for col in level_cols:
        row = 2
        while row <= num_rows+1:
            if ws.cell(row=row, column=col).value:
                # Encontrar bloco de células vazias abaixo
                start = row + 1
                end = start
                while end <= num_rows+1 and not ws.cell(row=end, column=col).value:
                    end += 1
                if end-1 > start:
                    ws.merge_cells(start_row=start, start_column=col, end_row=end-1, end_column=col)
                row = end
            else:
                row += 1
    wb.save(output_file)

def run_schema_to_excel_operation(source_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    is_json = source_ext == '.json'
    fields = extract_fields_from_json_schema(source_path) if is_json else extract_fields_from_xsd(source_path)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels']
        # Preencher colunas de níveis, só mostrar valor se diferente do anterior
        level_cells = []
        for idx in range(max_levels):
            val = lvls[idx] if idx < len(lvls) else ''
            disp = val if val != prev_levels[idx] else ''
            level_cells.append(disp)
        prev_levels = [v if v else p for v, p in zip([lvls[i] if i < len(lvls) else '' for i in range(max_levels)], prev_levels)]
        req_param = 'body (json)' if is_json else 'body (xml)'
        row = (
            level_cells +
            [req_param, '', f.get('Cardinality',''), f.get('Type',''), f.get('Details',''), f.get('Description','')]
        )
        rows.append(row)
    src_name = os.path.splitext(os.path.basename(source_path))[0]
    output_path = os.path.join(dest_dir, f"schema_{src_name}.xlsx")
    write_excel(rows, output_path, columns)
    messagebox.showinfo("Concluído", f"Arquivo Excel salvo em:\n{output_path}")

# --- Interface gráfica ---
def main():
    class UnifiedTool(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("The Forge - Unified Tool")
            self.geometry("650x350")
            self.resizable(False, False)
            self.source_path = tk.StringVar()
            self.target_path = tk.StringVar()
            self.dest_dir = tk.StringVar()
            self.op_map = {
                "Gerar Mapping entre Schemas (XSD/JSON)": "mapping",
                "Converter Schema para Excel": "schema_to_excel"
            }
            self.operation = tk.StringVar(value="Selecione a operação...")
            op_combo = ttk.Combobox(self, textvariable=self.operation, state="readonly", width=40)
            op_combo['values'] = ("Selecione a operação...",) + tuple(self.op_map.keys())
            op_combo.pack(pady=(0,10))
            op_combo.bind("<<ComboboxSelected>>", self.update_fields)
            # Campos dinâmicos
            self.frame_fields = tk.Frame(self)
            self.frame_fields.pack(fill='x', padx=10)
            self.build_fields()
            tk.Button(self, text="Executar", command=self.on_continue).pack(pady=15)
            self.result = None
        def build_fields(self):
            for widget in self.frame_fields.winfo_children():
                widget.destroy()
            op = self.op_map.get(self.operation.get(), None)
            if op == "mapping":
                # Source
                tk.Label(self.frame_fields, text="Source schema (XSD ou JSON Schema):").pack(pady=(5,0))
                frame1 = tk.Frame(self.frame_fields)
                frame1.pack(fill='x')
                tk.Entry(frame1, textvariable=self.source_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame1, text="Procurar", command=self.browse_source).pack(side='left', padx=5)
                # Target
                tk.Label(self.frame_fields, text="Target schema (XSD ou JSON Schema):").pack(pady=(5,0))
                frame2 = tk.Frame(self.frame_fields)
                frame2.pack(fill='x')
                tk.Entry(frame2, textvariable=self.target_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame2, text="Procurar", command=self.browse_target).pack(side='left', padx=5)
            elif op == "schema_to_excel":
                # Apenas source
                tk.Label(self.frame_fields, text="Schema de entrada (XSD ou JSON Schema):").pack(pady=(5,0))
                frame1 = tk.Frame(self.frame_fields)
                frame1.pack(fill='x')
                tk.Entry(frame1, textvariable=self.source_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame1, text="Procurar", command=self.browse_source).pack(side='left', padx=5)
            # Destino
            if op:
                tk.Label(self.frame_fields, text="Diretório de saída:").pack(pady=(5,0))
                frame3 = tk.Frame(self.frame_fields)
                frame3.pack(fill='x')
                tk.Entry(frame3, textvariable=self.dest_dir, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame3, text="Procurar", command=self.browse_dest).pack(side='left', padx=5)
        def update_fields(self, event=None):
            self.build_fields()
        def browse_source(self):
            path = filedialog.askopenfilename(title="Selecione o schema SOURCE", filetypes=[("Schemas", "*.xsd *.json"), ("Todos", "*.*")])
            if path:
                self.source_path.set(path)
        def browse_target(self):
            path = filedialog.askopenfilename(title="Selecione o schema TARGET", filetypes=[("Schemas", "*.xsd *.json"), ("Todos", "*.*")])
            if path:
                self.target_path.set(path)
        def browse_dest(self):
            path = filedialog.askdirectory(title="Selecione o diretório de saída")
            if path:
                self.dest_dir.set(path)
        def on_continue(self):
            op = self.op_map.get(self.operation.get(), None)
            if not op:
                messagebox.showerror("Erro", "Selecione a operação desejada.")
                return
            if not self.source_path.get() or not self.dest_dir.get() or (op=="mapping" and not self.target_path.get()):
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios.")
                return
            self.result = (op, self.source_path.get(), self.target_path.get() if op=="mapping" else None, self.dest_dir.get())
            self.destroy()

    selector = UnifiedTool()
    selector.mainloop()
    if not selector.result:
        print("Operação cancelada.")
        return
    op, source_path, target_path, dest_dir = selector.result
    if op == "mapping":
        run_mapping_operation(source_path, target_path, dest_dir)
    elif op == "schema_to_excel":
        run_schema_to_excel_operation(source_path, dest_dir)

if __name__ == "__main__":
    main() 