"""
Excel generator for creating formatted Excel files from schema data and mappings.
"""

from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .schema_processor import SchemaField
from .mapping_engine import MappingResult


class ExcelGenerator:
    """Handles Excel file generation with proper formatting and styling."""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    def create_schema_excel(self, fields: List[SchemaField], output_path: str, 
                           schema_name: str = "Schema") -> None:
        """Create Excel file from schema fields."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Schema Fields"
        
        # Determine max levels
        max_levels = max((len(f.levels) for f in fields), default=1)
        
        # Create headers
        headers = [f'Element Level {i+1}' for i in range(max_levels)] + [
            'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description'
        ]
        ws.append(headers)
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.cell_alignment
        
        # Add data rows
        prev_levels = [''] * max_levels
        # Ensure root field is first if present
        root_fields = [f for f in fields if len(f.levels) == 1]
        non_root_fields = [f for f in fields if len(f.levels) > 1]
        ordered_fields = root_fields + non_root_fields
        for field in ordered_fields:
            print(f"[DEBUG] Writing field to Excel: {'.'.join(field.levels)}")
            row = self._field_to_row(field, max_levels, prev_levels)
            ws.append(row)
            prev_levels = [v if v else p for v, p in zip(row[:max_levels], prev_levels)]
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Merge cells for hierarchy
        self._merge_hierarchy_cells(ws, max_levels)
        
        wb.save(output_path)
    
    def create_mapping_excel(self, source_fields: List[SchemaField], 
                            target_fields: List[SchemaField], 
                            mapping: List[MappingResult], 
                            output_path: str, 
                            source_name: str, 
                            target_name: str) -> None:
        """Create Excel file with mapping between source and target schemas."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Schema Mapping"
        
        # Determine max levels for both schemas
        src_max_levels = max((len(f.levels) for f in source_fields), default=1)
        tgt_max_levels = max((len(f.levels) for f in target_fields), default=1)
        
        # Create headers
        src_headers = [f'Source Level {i+1}' for i in range(src_max_levels)] + [
            'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description'
        ]
        tgt_headers = [f'Target Level {i+1}' for i in range(tgt_max_levels)] + [
            'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description'
        ]
        out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
        ws.append(out_headers)
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.cell_alignment
        
        # Create field maps for lookup
        source_field_map = {'.'.join(f.levels): f for f in source_fields}
        target_field_map = {'.'.join(f.levels): f for f in target_fields}
        
        # Add mapping rows
        prev_src_levels = [''] * src_max_levels
        prev_tgt_levels = [''] * tgt_max_levels
        
        for mapping_result in mapping:
            src_field = mapping_result.source_field
            tgt_field = mapping_result.target_field
            
            if src_field is None:
                continue
            
            # Source row
            src_row = self._field_to_row(src_field, src_max_levels, prev_src_levels)
            prev_src_levels = [v if v else p for v, p in zip(src_row[:src_max_levels], prev_src_levels)]
            
            # Destination field
            dest_field = '.'.join(tgt_field.levels) if tgt_field else ''
            
            # Target row
            if tgt_field:
                tgt_row = self._field_to_row(tgt_field, tgt_max_levels, prev_tgt_levels)
                prev_tgt_levels = [v if v else p for v, p in zip(tgt_row[:tgt_max_levels], prev_tgt_levels)]
            else:
                tgt_row = [''] * (tgt_max_levels + 6)
                prev_tgt_levels = [''] * tgt_max_levels
            
            # Combine rows
            full_row = src_row + [dest_field] + tgt_row
            ws.append(full_row)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Merge cells for hierarchy
        self._merge_hierarchy_cells(ws, src_max_levels)
        self._merge_hierarchy_cells(ws, tgt_max_levels, offset=src_max_levels + 7)
        
        wb.save(output_path)
    
    def create_simple_mapping_excel(self, mapping: List[MappingResult], 
                                   output_path: str) -> None:
        """Create simple Excel file with just mapping results."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping Results"
        
        # Headers
        headers = ['Source Path', 'Target Path', 'Similarity']
        ws.append(headers)
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.cell_alignment
        
        # Add mapping data
        for mapping_result in mapping:
            ws.append([
                mapping_result.source,
                mapping_result.target,
                mapping_result.similarity
            ])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        wb.save(output_path)
    
    def _field_to_row(self, field: SchemaField, max_levels: int, 
                      prev_levels: List[str]) -> List[Any]:
        """Convert a field to a row for Excel."""
        # Handle levels - convert to lowercase to match test expectations
        levels = [l.lower() if l != 'item' else '[]' for l in field.levels]
        level_cells = []
        
        for idx in range(max_levels):
            val = levels[idx] if idx < len(levels) else ''
            level_cells.append(val if val != prev_levels[idx] else '')
        
        # Determine request parameter
        req_param = 'body (json)' if any('[]' in l for l in field.levels) else 'body (xml)'
        
        return level_cells + [
            req_param, '', field.cardinality, field.type, 
            field.details, field.description
        ]
    
    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _merge_hierarchy_cells(self, ws, max_levels: int, offset: int = 0) -> None:
        """Merge cells for hierarchical display."""
        num_rows = ws.max_row
        level_cols = [i + 1 + offset for i in range(max_levels)]
        
        for col in level_cols:
            row = 2  # Start after header
            while row <= num_rows:
                if ws.cell(row=row, column=col).value:
                    # Find block of empty cells below
                    start = row + 1
                    end = start
                    while end <= num_rows and not ws.cell(row=end, column=col).value:
                        end += 1
                    if end - 1 > start:
                        ws.merge_cells(start_row=start, start_column=col, 
                                     end_row=end-1, end_column=col)
                    row = end
                else:
                    row += 1 