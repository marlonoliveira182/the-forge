"""
Excel Utilities Module

Provides Excel-specific utility functions:
- Excel file operations
- Cell formatting utilities
- Data extraction helpers
- Excel validation
"""

import os
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelUtils:
    """
    Utility class for Excel-specific operations.
    
    Provides methods for:
    - Excel file operations and validation
    - Cell formatting and styling
    - Data extraction and manipulation
    - Excel-specific validation
    """
    
    def __init__(self):
        self.supported_formats = {'.xlsx', '.xls'}
        self.default_styles = {
            'header_font': Font(bold=True, color="FFFFFF", size=12),
            'header_fill': PatternFill(
                start_color="212E3E",  # Marine Blue
                end_color="212E3E",
                fill_type="solid"
            ),
            'data_font': Font(size=10),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def load_excel_file(self, file_path: str) -> Optional[Workbook]:
        """
        Load an Excel file safely.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Workbook object or None if error
        """
        try:
            if not os.path.exists(file_path):
                return None
            
            if not self.is_excel_file(file_path):
                return None
            
            return load_workbook(file_path, data_only=True)
        except Exception:
            return None
    
    def is_excel_file(self, file_path: str) -> bool:
        """
        Check if a file is a valid Excel file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if file is a valid Excel file
        """
        if not file_path:
            return False
        
        extension = os.path.splitext(file_path)[1].lower()
        return extension in self.supported_formats
    
    def get_worksheet_data(self, worksheet: Optional[Worksheet]) -> Tuple[List[str], List[List[str]]]:
        """
        Extract data from a worksheet.
        
        Args:
            worksheet: The worksheet to extract data from
            
        Returns:
            Tuple of (headers, rows)
        """
        headers = []
        rows = []
        
        if not worksheet:
            return headers, rows
        
        # Get headers from first row
        for cell in worksheet[1]:
            headers.append(str(cell.value) if cell.value is not None else "")
        
        # Get data rows
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_data)
        
        return headers, rows
    
    def create_worksheet_data(
        self,
        headers: List[str],
        rows: List[List[str]]
    ) -> Tuple[List[str], List[List[str]]]:
        """
        Create standardized worksheet data.
        
        Args:
            headers: Column headers
            rows: Data rows
            
        Returns:
            Tuple of (normalized_headers, normalized_rows)
        """
        # Normalize headers
        normalized_headers = []
        for header in headers:
            if header:
                normalized_headers.append(str(header).strip())
            else:
                normalized_headers.append("")
        
        # Normalize rows
        normalized_rows = []
        for row in rows:
            normalized_row = []
            for cell in row:
                if cell is not None:
                    normalized_row.append(str(cell).strip())
                else:
                    normalized_row.append("")
            normalized_rows.append(normalized_row)
        
        return normalized_headers, normalized_rows
    
    def apply_cell_formatting(
        self,
        worksheet: Optional[Worksheet],
        cell: Any,
        style_type: str = "data"
    ) -> None:
        """
        Apply formatting to a cell.
        
        Args:
            worksheet: The worksheet containing the cell
            cell: The cell to format
            style_type: Type of formatting to apply
        """
        if not worksheet or not cell:
            return
            
        if style_type == "header":
            cell.font = self.default_styles['header_font']
            cell.fill = self.default_styles['header_fill']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.font = self.default_styles['data_font']
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        cell.border = self.default_styles['border']
    
    def auto_size_columns(self, worksheet: Worksheet) -> None:
        """
        Auto-size columns based on content.
        
        Args:
            worksheet: The worksheet to auto-size
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = None
            
            # Get column letter
            if column[0].column is not None:
                column_letter = get_column_letter(column[0].column)
                
                # Find maximum length in column
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set column width (with limits)
                if column_letter:
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def find_column_index(self, headers: List[str], column_name: str) -> int:
        """
        Find the index of a column by name.
        
        Args:
            headers: List of column headers
            column_name: Name of the column to find
            
        Returns:
            Column index or -1 if not found
        """
        for i, header in enumerate(headers):
            if header.lower() == column_name.lower():
                return i
        return -1
    
    def extract_column_data(
        self,
        headers: List[str],
        rows: List[List[str]],
        column_name: str
    ) -> List[str]:
        """
        Extract data from a specific column.
        
        Args:
            headers: Column headers
            rows: Data rows
            column_name: Name of the column to extract
            
        Returns:
            List of values from the specified column
        """
        column_index = self.find_column_index(headers, column_name)
        if column_index == -1:
            return []
        
        column_data = []
        for row in rows:
            if column_index < len(row):
                column_data.append(row[column_index])
            else:
                column_data.append("")
        
        return column_data
    
    def validate_excel_structure(
        self,
        headers: List[str],
        rows: List[List[str]]
    ) -> Tuple[bool, List[str]]:
        """
        Validate Excel data structure.
        
        Args:
            headers: Column headers
            rows: Data rows
            
        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []
        
        # Check for empty headers
        if not headers:
            errors.append("No headers found")
            return False, errors
        
        # Check for empty header cells
        for i, header in enumerate(headers):
            if not header.strip():
                errors.append(f"Empty header at column {i + 1}")
        
        # Check for consistent row lengths
        expected_length = len(headers)
        for i, row in enumerate(rows, 2):  # Start from row 2 (after headers)
            if len(row) != expected_length:
                errors.append(f"Row {i} has {len(row)} columns, expected {expected_length}")
        
        return len(errors) == 0, errors
    
    def create_summary_sheet(
        self,
        workbook: Workbook,
        summary_data: Dict[str, Any],
        sheet_name: str = "Summary"
    ) -> Worksheet:
        """
        Create a summary sheet with statistics.
        
        Args:
            workbook: The workbook to add sheet to
            summary_data: Dictionary with summary information
            sheet_name: Name for the summary sheet
            
        Returns:
            The created worksheet
        """
        # Remove existing sheet if it exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Write headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            self.apply_cell_formatting(worksheet, cell, "header")
        
        # Write data
        for row_idx, (key, value) in enumerate(summary_data.items(), 2):
            worksheet.cell(row=row_idx, column=1, value=key)
            worksheet.cell(row=row_idx, column=2, value=str(value))
        
        # Auto-size columns
        self.auto_size_columns(worksheet)
        
        return worksheet
    
    def merge_excel_files(
        self,
        source_path: str,
        target_path: Optional[str],
        output_path: str
    ) -> bool:
        """
        Merge Excel files with path-based mapping.
        
        Args:
            source_path: Path to source Excel file
            target_path: Path to target Excel file (optional)
            output_path: Path for output merged file
            
        Returns:
            True if merge was successful
        """
        try:
            # Load source file
            source_wb = self.load_excel_file(source_path)
            if not source_wb:
                return False
            
            source_ws = source_wb.active
            source_headers, source_rows = self.get_worksheet_data(source_ws)
            
            # Create output workbook
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.title = "Merged"
            
            if target_path and os.path.exists(target_path):
                # Load target file
                target_wb = self.load_excel_file(target_path)
                if target_wb and target_wb.active:
                    target_ws = target_wb.active
                    target_headers, target_rows = self.get_worksheet_data(target_ws)
                    
                    # Create merged headers
                    merged_headers = source_headers + ['Destination Field (Target Path)'] + target_headers
                    
                    # Write headers
                    for col, header in enumerate(merged_headers, 1):
                        cell = output_ws.cell(row=1, column=col, value=header)
                        self.apply_cell_formatting(output_ws, cell, "header")
                    
                    # Write merged data
                    for row_idx, source_row in enumerate(source_rows, 2):
                        # Add source row data
                        for col_idx, value in enumerate(source_row, 1):
                            output_ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Add destination field (empty for now)
                        dest_col = len(source_headers) + 1
                        output_ws.cell(row=row_idx, column=dest_col, value="")
                        
                        # Add target row data if available
                        if row_idx - 2 < len(target_rows):
                            target_row = target_rows[row_idx - 2]
                            for col_idx, value in enumerate(target_row, dest_col + 1):
                                output_ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    # No target file, just copy source
                    for col, header in enumerate(source_headers, 1):
                        cell = output_ws.cell(row=1, column=col, value=header)
                        self.apply_cell_formatting(output_ws, cell, "header")
                    
                    for row_idx, row in enumerate(source_rows, 2):
                        for col_idx, value in enumerate(row, 1):
                            output_ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                # No target file, just copy source
                for col, header in enumerate(source_headers, 1):
                    cell = output_ws.cell(row=1, column=col, value=header)
                    self.apply_cell_formatting(output_ws, cell, "header")
                
                for row_idx, row in enumerate(source_rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        output_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size columns
            self.auto_size_columns(output_ws)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            output_wb.save(output_path)
            return True
            
        except Exception:
            return False
    
    def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get comprehensive information about an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with Excel file information
        """
        info = {
            'path': file_path,
            'exists': False,
            'is_excel': False,
            'sheets': [],
            'total_rows': 0,
            'total_columns': 0,
            'file_size': None
        }
        
        if not os.path.exists(file_path):
            return info
        
        info['exists'] = True
        info['is_excel'] = self.is_excel_file(file_path)
        
        if info['is_excel']:
            try:
                info['file_size'] = os.path.getsize(file_path)
                workbook = self.load_excel_file(file_path)
                
                if workbook:
                    info['sheets'] = workbook.sheetnames
                    
                    # Get data from first sheet
                    if workbook.sheetnames:
                        worksheet = workbook[workbook.sheetnames[0]]
                        headers, rows = self.get_worksheet_data(worksheet)
                        info['total_rows'] = len(rows)
                        info['total_columns'] = len(headers)
                
            except Exception:
                pass
        
        return info 