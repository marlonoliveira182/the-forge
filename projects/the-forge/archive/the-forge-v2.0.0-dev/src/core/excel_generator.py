"""
Excel Generator Module

Handles creation and formatting of Excel files with:
- Professional styling and colors
- Header formatting
- Data validation
- Auto-sizing columns
- Multiple sheet support
"""

import os
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelGenerator:
    """
    Handles Excel file creation with professional formatting and styling.
    
    Provides methods for creating well-formatted Excel files with:
    - Header styling with EDP color scheme
    - Auto-sized columns
    - Data validation
    - Multiple sheet support
    """
    
    def __init__(self):
        # EDP Color Scheme (from memories)
        self.colors = {
            'marine_blue': '212E3E',
            'electric_green': '28FF52',
            'slate_grey': '7C9599',
            'seaweed_green': '225E66',
            'cobalt_blue': '263CC8',
            'violet_purple': '6D32FF',
            'ice_blue': '0CD3F8'
        }
        
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(
            start_color=self.colors['marine_blue'],
            end_color=self.colors['marine_blue'],
            fill_type="solid"
        )
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_workbook(self) -> Workbook:
        """Create a new Excel workbook."""
        return Workbook()
    
    def create_sheet(self, workbook: Workbook, sheet_name: str = "Schema") -> Worksheet:
        """
        Create a new worksheet with proper setup.
        
        Args:
            workbook: The Excel workbook
            sheet_name: Name for the worksheet
            
        Returns:
            The created worksheet
        """
        if sheet_name in workbook.sheetnames:
            # Remove existing sheet with same name
            workbook.remove(workbook[sheet_name])
        
        worksheet = workbook.create_sheet(title=sheet_name)
        return worksheet
    
    def write_headers(self, worksheet: Worksheet, headers: List[str]) -> None:
        """
        Write headers to worksheet with professional formatting.
        
        Args:
            worksheet: The worksheet to write to
            headers: List of header strings
        """
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def write_data(self, worksheet: Worksheet, data: List[List[str]], start_row: int = 2) -> None:
        """
        Write data rows to worksheet with formatting.
        
        Args:
            worksheet: The worksheet to write to
            data: List of data rows
            start_row: Starting row number (default: 2, after headers)
        """
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.data_font
                cell.border = self.border
                
                # Center align headers and specific columns
                if col_idx <= len(row_data):  # Ensure we don't exceed data length
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def auto_size_columns(self, worksheet: Worksheet) -> None:
        """
        Auto-size columns based on content.
        
        Args:
            worksheet: The worksheet to auto-size
        """
        for column in worksheet.columns:
            max_length = 0
            if column[0].column is not None:
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum and maximum width
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def apply_conditional_formatting(self, worksheet: Worksheet, headers: List[str]) -> None:
        """
        Apply conditional formatting based on column types.
        
        Args:
            worksheet: The worksheet to format
            headers: List of headers to determine column types
        """
        # Find type and cardinality columns for special formatting
        type_col = None
        cardinality_col = None
        
        for i, header in enumerate(headers, 1):
            if header.lower() == 'type':
                type_col = i
            elif header.lower() == 'cardinality':
                cardinality_col = i
        
        # Apply special formatting to type column
        if type_col:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=type_col)
                if cell.value:
                    cell_value = str(cell.value).lower()
                    # Color code different types
                    if cell_value in ['string', 'text']:
                        cell.fill = PatternFill(
                            start_color='E8F5E8',  # Light green
                            end_color='E8F5E8',
                            fill_type="solid"
                        )
                    elif cell_value in ['integer', 'number', 'int']:
                        cell.fill = PatternFill(
                            start_color='E8F0F8',  # Light blue
                            end_color='E8F0F8',
                            fill_type="solid"
                        )
                    elif cell_value == 'object':
                        cell.fill = PatternFill(
                            start_color='FFF2CC',  # Light yellow
                            end_color='FFF2CC',
                            fill_type="solid"
                        )
    
    def create_schema_excel(
        self,
        headers: List[str],
        data: List[List[str]],
        output_path: str,
        sheet_name: str = "Schema"
    ) -> None:
        """
        Create a complete Excel file with schema data.
        
        Args:
            headers: Column headers
            data: Data rows
            output_path: Output file path
            sheet_name: Name for the worksheet
        """
        # Create workbook and worksheet
        workbook = self.create_workbook()
        worksheet = self.create_sheet(workbook, sheet_name)
        
        # Write headers and data
        self.write_headers(worksheet, headers)
        self.write_data(worksheet, data)
        
        # Apply formatting
        self.auto_size_columns(worksheet)
        self.apply_conditional_formatting(worksheet, headers)
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save workbook
        workbook.save(output_path)
    
    def create_merged_excel(
        self,
        source_headers: List[str],
        source_data: List[List[str]],
        target_headers: List[str],
        target_data: List[List[str]],
        mapping_data: List[List[str]],
        output_path: str,
        sheet_name: str = "Merged"
    ) -> None:
        """
        Create a merged Excel file with source, target, and mapping data.
        
        Args:
            source_headers: Source schema headers
            source_data: Source schema data
            target_headers: Target schema headers
            target_data: Target schema data
            mapping_data: Mapping data between schemas
            output_path: Output file path
            sheet_name: Name for the worksheet
        """
        # Create workbook and worksheet
        workbook = self.create_workbook()
        worksheet = self.create_sheet(workbook, sheet_name)
        
        # Create merged headers
        merged_headers = source_headers + ['Destination Field (Target Path)'] + target_headers
        
        # Create merged data
        merged_data = []
        for i, source_row in enumerate(source_data):
            target_row = target_data[i] if i < len(target_data) else [''] * len(target_headers)
            mapping_row = mapping_data[i] if i < len(mapping_data) else ['']
            
            merged_row = source_row + mapping_row + target_row
            merged_data.append(merged_row)
        
        # Write headers and data
        self.write_headers(worksheet, merged_headers)
        self.write_data(worksheet, merged_data)
        
        # Apply formatting
        self.auto_size_columns(worksheet)
        self.apply_conditional_formatting(worksheet, merged_headers)
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save workbook
        workbook.save(output_path)
    
    def add_validation_sheet(
        self,
        workbook: Workbook,
        validation_issues: List[str],
        sheet_name: str = "Validation"
    ) -> None:
        """
        Add a validation sheet with issues found.
        
        Args:
            workbook: The workbook to add sheet to
            validation_issues: List of validation issues
            sheet_name: Name for the validation sheet
        """
        worksheet = self.create_sheet(workbook, sheet_name)
        
        # Write validation headers
        headers = ['Row', 'Issue']
        self.write_headers(worksheet, headers)
        
        # Write validation data
        validation_data = []
        for issue in validation_issues:
            # Extract row number if present
            row_num = "N/A"
            if "Row" in issue:
                try:
                    row_num = issue.split("Row")[1].split(":")[0].strip()
                except:
                    pass
            
            validation_data.append([row_num, issue])
        
        self.write_data(worksheet, validation_data)
        self.auto_size_columns(worksheet)
    
    def create_summary_sheet(
        self,
        workbook: Workbook,
        summary_data: Dict[str, Any],
        sheet_name: str = "Summary"
    ) -> None:
        """
        Create a summary sheet with processing statistics.
        
        Args:
            workbook: The workbook to add sheet to
            summary_data: Dictionary with summary information
            sheet_name: Name for the summary sheet
        """
        worksheet = self.create_sheet(workbook, sheet_name)
        
        # Write summary headers
        headers = ['Metric', 'Value']
        self.write_headers(worksheet, headers)
        
        # Write summary data
        summary_rows = []
        for key, value in summary_data.items():
            summary_rows.append([key, str(value)])
        
        self.write_data(worksheet, summary_rows)
        self.auto_size_columns(worksheet)
    
    def apply_custom_styling(self, worksheet: Worksheet, style_config: Dict[str, Any]) -> None:
        """
        Apply custom styling based on configuration.
        
        Args:
            worksheet: The worksheet to style
            style_config: Configuration dictionary for styling
        """
        # Apply custom colors if specified
        if 'header_color' in style_config:
            header_fill = PatternFill(
                start_color=style_config['header_color'],
                end_color=style_config['header_color'],
                fill_type="solid"
            )
            for cell in worksheet[1]:
                cell.fill = header_fill
        
        # Apply custom fonts if specified
        if 'header_font_size' in style_config:
            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=style_config['header_font_size']
            )
            for cell in worksheet[1]:
                cell.font = header_font
        
        # Apply custom borders if specified
        if 'border_style' in style_config:
            border_style = style_config['border_style']
            border = Border(
                left=Side(style=border_style),
                right=Side(style=border_style),
                top=Side(style=border_style),
                bottom=Side(style=border_style)
            )
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border 