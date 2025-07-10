import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

# === CONFIGURATION ===
# Remover diretórios fixos
# SOURCE_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\source"
# TARGET_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\target"
# OUTPUT_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\mapping-file"

# === UTILITY FUNCTIONS ===
def build_path_from_levels(row, level_indices, type_index=None, all_rows=None, row_idx=None):
    # Para cada coluna de nível, busca para cima se estiver vazia
    path_parts = []
    for idx in level_indices:
        val = row[idx]
        if (val is None or str(val).strip() == '') and all_rows is not None and row_idx is not None:
            # Busca para cima
            for up_idx in range(row_idx-1, -1, -1):
                up_val = all_rows[up_idx][idx]
                if up_val is not None and str(up_val).strip() != '':
                    val = up_val
                    break
        if val is not None and str(val).strip() != '':
            path_parts.append(str(val).strip())
    return '.'.join(path_parts)

def normalize_path(path):
    # Lowercase, trim, skip empty, and collapse multiple dots
    parts = [p.strip().lower() for p in path.split('.') if p.strip()]
    return '.'.join(parts)

def get_level_indices(headers):
    return [i for i, h in enumerate(headers) if h.strip().startswith('Element Level')]

def get_destination_field(norm_src_path, tgt_path_display, tgt_row, tgt_rows=None, tgt_level_indices=None, tgt_row_idx=None):
    # Se houver tgt_row, construa o path buscando valores não vazios acima para cada coluna 'Element Level X'
    if tgt_row and tgt_rows is not None and tgt_level_indices is not None and tgt_row_idx is not None:
        path_parts = []
        for col_idx in tgt_level_indices:
            val = tgt_row[col_idx]
            if val is None or str(val).strip() == '':
                # Busca para cima
                for up_idx in range(tgt_row_idx-1, -1, -1):
                    up_val = tgt_rows[up_idx][col_idx]
                    if up_val is not None and str(up_val).strip() != '':
                        val = up_val
                        break
            if val is not None and str(val).strip() != '':
                path_parts.append(str(val).strip())
        return '.'.join(path_parts)
    else:
        return ''

def merge_xlsx(source_path, target_path, output_path):
    src_wb = load_workbook(source_path)
    src_ws = src_wb.active
    src_headers = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]
    src_level_indices = get_level_indices(src_headers)
    src_type_index = src_headers.index('Type') if 'Type' in src_headers else None

    # Build source path map (normalized)
    src_rows = list(src_ws.iter_rows(min_row=2, values_only=True))
    src_path_map = {}
    for idx, row in enumerate(src_rows):
        path = build_path_from_levels(row, src_level_indices, src_type_index, all_rows=src_rows, row_idx=idx)
        norm_path = normalize_path(path)
        src_path_map[norm_path] = row

    # Prepare target
    tgt_rows = []
    tgt_headers = []
    tgt_level_indices = []
    tgt_type_index = None
    tgt_path_to_row = {}
    tgt_path_display = {}
    if target_path and os.path.exists(target_path):
        tgt_wb = load_workbook(target_path)
        tgt_ws = tgt_wb.active
        tgt_headers = [cell.value for cell in next(tgt_ws.iter_rows(min_row=1, max_row=1))]
        tgt_level_indices = get_level_indices(tgt_headers)
        tgt_type_index = tgt_headers.index('Type') if 'Type' in tgt_headers else None
        tgt_rows = list(tgt_ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(tgt_rows):
            tgt_path = build_path_from_levels(row, tgt_level_indices, tgt_type_index, all_rows=tgt_rows, row_idx=idx)
            norm_tgt_path = normalize_path(tgt_path)
            tgt_path_to_row[norm_tgt_path] = row
            tgt_path_display[norm_tgt_path] = tgt_path

    # Prepare output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Merged'
    out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
    out_ws.append(out_headers)

    # Apply header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in out_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # For each source row, find the matching target row by normalized path
    for norm_src_path, src_row in src_path_map.items():
        tgt_row = tgt_path_to_row.get(norm_src_path, None)
        dest_field = get_destination_field(norm_src_path, tgt_path_display, tgt_row, tgt_rows, tgt_level_indices, tgt_rows.index(tgt_row) if tgt_row else None)
        out_row = list(src_row) + [dest_field] + (list(tgt_row) if tgt_row else [''] * len(tgt_headers))
        out_ws.append(out_row)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    out_wb.save(output_path)
    print(f"Merged file saved to: {output_path}")

def get_base_filename(filename):
    # Remove sufixos comuns e extensão, e coloca em minúsculas
    name = os.path.splitext(filename)[0]
    name = re.sub(r'[_-](source|target)$', '', name, flags=re.IGNORECASE)
    return name.lower()

def main():
    class MergerSelector(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Selecionar arquivos e destino")
            self.geometry("550x240")
            self.resizable(False, False)
            self.source_path = tk.StringVar()
            self.target_path = tk.StringVar()
            self.dest_dir = tk.StringVar()
            # Fonte
            tk.Label(self, text="Arquivo FONTE (.xlsx):").pack(pady=(10,0))
            frame1 = tk.Frame(self)
            frame1.pack(fill='x', padx=10)
            tk.Entry(frame1, textvariable=self.source_path, width=55, state='readonly').pack(side='left', expand=True, fill='x')
            tk.Button(frame1, text="Procurar", command=self.browse_source).pack(side='left', padx=5)
            # Alvo
            tk.Label(self, text="Arquivo ALVO (.xlsx) [opcional]:").pack(pady=(10,0))
            frame2 = tk.Frame(self)
            frame2.pack(fill='x', padx=10)
            tk.Entry(frame2, textvariable=self.target_path, width=55, state='readonly').pack(side='left', expand=True, fill='x')
            tk.Button(frame2, text="Procurar", command=self.browse_target).pack(side='left', padx=5)
            # Destino
            tk.Label(self, text="Diretório de destino:").pack(pady=(10,0))
            frame3 = tk.Frame(self)
            frame3.pack(fill='x', padx=10)
            tk.Entry(frame3, textvariable=self.dest_dir, width=55, state='readonly').pack(side='left', expand=True, fill='x')
            tk.Button(frame3, text="Procurar", command=self.browse_dest).pack(side='left', padx=5)
            tk.Button(self, text="Continuar", command=self.on_continue).pack(pady=15)
            self.result = None
        def browse_source(self):
            path = filedialog.askopenfilename(title="Selecione o arquivo FONTE", filetypes=[("Excel files", "*.xlsx")])
            if path:
                self.source_path.set(path)
        def browse_target(self):
            path = filedialog.askopenfilename(title="Selecione o arquivo ALVO (opcional)", filetypes=[("Excel files", "*.xlsx")])
            if path:
                self.target_path.set(path)
        def browse_dest(self):
            path = filedialog.askdirectory(title="Selecione o diretório de destino")
            if path:
                self.dest_dir.set(path)
        def on_continue(self):
            if not self.source_path.get() or not self.dest_dir.get():
                messagebox.showerror("Erro", "Selecione o arquivo fonte e o diretório de destino.")
                return
            self.result = (self.source_path.get(), self.target_path.get() if self.target_path.get() else None, self.dest_dir.get())
            self.destroy()

    selector = MergerSelector()
    selector.mainloop()
    if not selector.result:
        print("Operação cancelada.")
        return
    source_path, target_path, dest_dir = selector.result
    base_name = os.path.splitext(os.path.basename(source_path))[0]
    output_path = os.path.join(dest_dir, base_name + "_merged.xlsx")
    merge_xlsx(source_path, target_path, output_path)
    messagebox.showinfo("Concluído", f"Arquivo mesclado salvo em:\n{output_path}")

if __name__ == "__main__":
    main() 