"""
Excel utilities for Excel-specific operations and formatting.
"""

from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelUtils:
    """Utility class for Excel operations."""
    
    # Default styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    CELL_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def create_workbook() -> Workbook:
        """Create a new workbook with default settings."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        return wb
    
    @staticmethod
    def add_worksheet(wb: Workbook, title: str) -> Worksheet:
        """Add a new worksheet to workbook."""
        ws = wb.create_sheet(title=title)
        return ws
    
    @staticmethod
    def apply_header_style(ws: Worksheet, row: int = 1) -> None:
        """Apply header styling to a row."""
        for cell in ws[row]:
            cell.font = ExcelUtils.HEADER_FONT
            cell.fill = ExcelUtils.HEADER_FILL
            cell.alignment = ExcelUtils.CELL_ALIGNMENT
            cell.border = ExcelUtils.BORDER
    
    @staticmethod
    def apply_cell_style(ws: Worksheet, row: int, column: int, 
                        font: Optional[Font] = None, 
                        fill: Optional[PatternFill] = None,
                        alignment: Optional[Alignment] = None,
                        border: Optional[Border] = None) -> None:
        """Apply styling to a specific cell."""
        cell = ws.cell(row=row, column=column)
        
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
    
    @staticmethod
    def auto_adjust_columns(ws: Worksheet, min_width: int = 8, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def merge_hierarchy_cells(ws: Worksheet, max_levels: int, offset: int = 0) -> None:
        """Merge cells for hierarchical display."""
        num_rows = ws.max_row
        level_cols = [i + 1 + offset for i in range(max_levels)]
        
        for col in level_cols:
            row = 2  # Start after header
            while row <= num_rows:
                if ws.cell(row=row, column=col).value:
                    # Find block of empty cells below
                    start = row + 1
                    end = start
                    while end <= num_rows and not ws.cell(row=end, column=col).value:
                        end += 1
                    if end - 1 > start:
                        ws.merge_cells(start_row=start, start_column=col, 
                                     end_row=end-1, end_column=col)
                    row = end
                else:
                    row += 1
    
    @staticmethod
    def add_conditional_formatting(ws: Worksheet, column: int, 
                                 condition: str, color: str) -> None:
        """Add conditional formatting to a column."""
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Apply conditional formatting to the entire column
        ws.conditional_formatting.add(
            f'{get_column_letter(column)}2:{get_column_letter(column)}{ws.max_row}',
            CellIsRule(operator=condition, fill=fill)
        )
    
    @staticmethod
    def create_summary_sheet(wb: Workbook, data: Dict[str, Any]) -> Worksheet:
        """Create a summary sheet with statistics."""
        ws = ExcelUtils.add_worksheet(wb, "Summary")
        
        # Add title
        ws['A1'] = "Schema Processing Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # Add data
        row = 3
        for key, value in data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
        
        # Apply styling
        ExcelUtils.apply_header_style(ws, 1)
        ExcelUtils.auto_adjust_columns(ws)
        
        return ws
    
    @staticmethod
    def add_data_validation(ws: Worksheet, column: int, 
                           validation_type: str, formula1: str, 
                           formula2: Optional[str] = None) -> None:
        """Add data validation to a column."""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        dv = DataValidation(type=validation_type, formula1=formula1, formula2=formula2)
        dv.add(f'{get_column_letter(column)}2:{get_column_letter(column)}1000')
        ws.add_data_validation(dv)
    
    @staticmethod
    def create_pivot_table(ws: Worksheet, data_range: str, 
                          row_fields: List[str], value_fields: List[str]) -> None:
        """Create a pivot table (simplified version)."""
        # This is a placeholder - actual pivot table creation would require
        # more complex logic and might need additional libraries
        pass
    
    @staticmethod
    def export_to_csv(ws: Worksheet, output_path: str) -> None:
        """Export worksheet to CSV format."""
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
    
    @staticmethod
    def read_excel_data(filepath: str, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """Read data from Excel file."""
        wb = load_workbook(filepath, read_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        wb.close()
        return data
    
    @staticmethod
    def validate_excel_structure(filepath: str) -> Dict[str, Any]:
        """Validate Excel file structure and return analysis."""
        try:
            wb = load_workbook(filepath, read_only=True)
            
            analysis = {
                'sheet_count': len(wb.sheetnames),
                'sheets': [],
                'total_rows': 0,
                'total_columns': 0,
                'is_valid': True,
                'errors': []
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_info = {
                    'name': sheet_name,
                    'rows': ws.max_row,
                    'columns': ws.max_column,
                    'has_data': ws.max_row > 0 and ws.max_column > 0
                }
                analysis['sheets'].append(sheet_info)
                analysis['total_rows'] = max(analysis['total_rows'], ws.max_row)
                analysis['total_columns'] = max(analysis['total_columns'], ws.max_column)
            
            wb.close()
            return analysis
            
        except Exception as e:
            return {
                'is_valid': False,
                'errors': [str(e)]
            }
    
    @staticmethod
    def create_chart(ws: Worksheet, chart_type: str, 
                    data_range: str, title: str) -> None:
        """Create a chart in the worksheet."""
        from openpyxl.chart import BarChart, LineChart, PieChart
        
        chart_map = {
            'bar': BarChart,
            'line': LineChart,
            'pie': PieChart
        }
        
        if chart_type not in chart_map:
            raise ValueError(f"Unsupported chart type: {chart_type}")
        
        chart_class = chart_map[chart_type]
        chart = chart_class()
        chart.title = title
        chart.add_data(ws, data_range)
        ws.add_chart(chart, "E1")
    
    @staticmethod
    def apply_number_format(ws: Worksheet, column: int, 
                          number_format: str) -> None:
        """Apply number formatting to a column."""
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=column)
            cell.number_format = number_format
    
    @staticmethod
    def freeze_panes(ws: Worksheet, row: int = 1, column: int = 0) -> None:
        """Freeze panes in worksheet."""
        ws.freeze_panes = ws.cell(row=row, column=column)
    
    @staticmethod
    def add_filter(ws: Worksheet, row: int = 1) -> None:
        """Add filter to worksheet."""
        ws.auto_filter.ref = ws.dimensions
    
    @staticmethod
    def create_hyperlink(ws: Worksheet, row: int, column: int, 
                        url: str, text: str) -> None:
        """Create a hyperlink in a cell."""
        from openpyxl.styles import Font
        
        cell = ws.cell(row=row, column=column)
        cell.hyperlink = url
        cell.value = text
        cell.font = Font(color="0000FF", underline="single")
    
    @staticmethod
    def protect_worksheet(ws: Worksheet, password: Optional[str] = None) -> None:
        """Protect worksheet with optional password."""
        ws.protection.sheet = True
        if password:
            ws.protection.password = password
    
    @staticmethod
    def get_cell_value(ws: Worksheet, row: int, column: int) -> Any:
        """Get cell value with error handling."""
        try:
            return ws.cell(row=row, column=column).value
        except:
            return None
    
    @staticmethod
    def set_cell_value(ws: Worksheet, row: int, column: int, value: Any) -> None:
        """Set cell value with type handling."""
        cell = ws.cell(row=row, column=column)
        
        if isinstance(value, (int, float)):
            cell.value = value
        elif isinstance(value, str):
            cell.value = value
        elif value is None:
            cell.value = ""
        else:
            cell.value = str(value) 