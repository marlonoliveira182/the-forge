import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), 'microservices'))
from microservices.xsd_parser_service import XSDParser
from microservices.json_to_excel_service import JSONToExcelService
from microservices.xsd_json_converter_service import XSDJSONConverterService
from microservices.case_converter_service import pascal_to_camel, camel_to_pascal
import json
import openpyxl
import logging

# Set up logging
log_path = os.path.join(os.path.dirname(__file__), 'schema_to_excel_gui.log')
logging.basicConfig(
    filename=log_path,
    filemode='a',
    format='%(asctime)s %(levelname)s: %(message)s',
    level=logging.INFO
)

class SchemaToExcelGUI:
    def __init__(self, root):
        logging.info('Starting SchemaToExcelGUI')
        self.root = root
        self.root.title('Schema to Mapping Tool')
        self.schema_source_path = tk.StringVar()
        self.schema_target_path = tk.StringVar()
        self.mapping_output_path = tk.StringVar()
        self.schema_source_type = tk.StringVar(value='xsd')
        self.schema_target_type = tk.StringVar(value='xsd')
        self.source_case = tk.StringVar(value='none')
        self.target_case = tk.StringVar(value='none')
        self.json_to_excel_service = JSONToExcelService()
        self.xsd_json_converter = XSDJSONConverterService()
        self.create_widgets()

    def create_widgets(self):
        frame_schema_map = tk.Frame(self.root, padx=10, pady=10)
        frame_schema_map.pack(fill=tk.BOTH, expand=True)
        tk.Label(frame_schema_map, text='Source Schema:').grid(row=0, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_source_path, width=40).grid(row=0, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=lambda: self.browse_schema(self.schema_source_path)).grid(row=0, column=2)
        tk.Label(frame_schema_map, text='Type:').grid(row=0, column=3, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_source_type, width=8, state='readonly').grid(row=0, column=4)
        tk.Label(frame_schema_map, text='Source Case:').grid(row=0, column=5, sticky='w')
        ttk.Combobox(frame_schema_map, textvariable=self.source_case, values=['none', 'pascal', 'camel'], state='readonly', width=8).grid(row=0, column=6)
        tk.Label(frame_schema_map, text='Target Schema:').grid(row=1, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_target_path, width=40).grid(row=1, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=lambda: self.browse_schema(self.schema_target_path)).grid(row=1, column=2)
        tk.Label(frame_schema_map, text='Type:').grid(row=1, column=3, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_target_type, width=8, state='readonly').grid(row=1, column=4)
        tk.Label(frame_schema_map, text='Target Case:').grid(row=1, column=5, sticky='w')
        ttk.Combobox(frame_schema_map, textvariable=self.target_case, values=['none', 'pascal', 'camel'], state='readonly', width=8).grid(row=1, column=6)
        tk.Label(frame_schema_map, text='Output Mapping Excel:').grid(row=2, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.mapping_output_path, width=40).grid(row=2, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=self.browse_mapping_output).grid(row=2, column=2)
        tk.Button(frame_schema_map, text='Generate Mapping', command=self.generate_schema_mapping).grid(row=3, column=1, pady=10)
        self.schema_map_info = tk.Text(frame_schema_map, height=10, width=80)
        self.schema_map_info.grid(row=4, column=0, columnspan=7, pady=10)

    def browse_schema(self, var):
        try:
            file_path = filedialog.askopenfilename(filetypes=[('Schema Files', '*.xsd *.json')])
            if file_path:
                var.set(file_path)
                ext = os.path.splitext(file_path)[1].lower()
                logging.info(f'User selected schema file: {file_path}')
                # Auto-detect type based on file extension
                if var == self.schema_source_path:
                    if ext == '.json':
                        self.schema_source_type.set('json')
                    elif ext == '.xsd':
                        self.schema_source_type.set('xsd')
                elif var == self.schema_target_path:
                    if ext == '.json':
                        self.schema_target_type.set('json')
                    elif ext == '.xsd':
                        self.schema_target_type.set('xsd')
        except Exception as e:
            logging.error(f'Error in browse_schema: {e}')
            raise
    def browse_mapping_output(self):
        try:
            file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
            if file_path:
                self.mapping_output_path.set(file_path)
                logging.info(f'User selected output mapping file: {file_path}')
        except Exception as e:
            logging.error(f'Error in browse_mapping_output: {e}')
            raise
    def convert_case(self, levels, mode):
        if mode == 'none':
            return levels
        elif mode == 'pascal':
            return [camel_to_pascal(lvl) if lvl else '' for lvl in levels]
        elif mode == 'camel':
            return [pascal_to_camel(lvl) if lvl else '' for lvl in levels]
        return levels
    def get_xsd_rows(self, path):
        parser = XSDParser()
        return parser.parse_xsd_file(path)
    def get_json_rows(self, path):
        # Improved JSON Schema logic: recursively walk and extract all fields/structures
        def walk_json_schema(schema, parent_levels=None, max_depth=20):
            if parent_levels is None:
                parent_levels = []
            rows = []
            if 'type' in schema and schema['type'] == 'object' and 'properties' in schema:
                for prop, prop_schema in schema['properties'].items():
                    levels = parent_levels + [prop]
                    row = {
                        'levels': levels,
                        'Request Parameter': '',
                        'GDPR': '',
                        'Cardinality': '',
                        'Type': prop_schema.get('type', ''),
                        'Details': '',
                        'Description': prop_schema.get('description', '')
                    }
                    rows.append(row)
                    # Recurse into nested objects
                    if prop_schema.get('type') == 'object':
                        rows.extend(walk_json_schema(prop_schema, levels, max_depth))
                    # Handle arrays of objects
                    elif prop_schema.get('type') == 'array' and 'items' in prop_schema:
                        item_schema = prop_schema['items']
                        array_levels = levels + ['[]']
                        row_array = {
                            'levels': array_levels,
                            'Request Parameter': '',
                            'GDPR': '',
                            'Cardinality': '',
                            'Type': item_schema.get('type', ''),
                            'Details': '',
                            'Description': item_schema.get('description', '')
                        }
                        rows.append(row_array)
                        if item_schema.get('type') == 'object':
                            rows.extend(walk_json_schema(item_schema, array_levels, max_depth))
            return rows
        schema = self.json_to_excel_service.load_json_schema(path)
        return walk_json_schema(schema)
    def generate_schema_mapping(self):
        import difflib
        src_type = self.schema_source_type.get()
        tgt_type = self.schema_target_type.get()
        src_path = self.schema_source_path.get()
        tgt_path = self.schema_target_path.get()
        out_path = self.mapping_output_path.get()
        src_case = self.source_case.get()
        tgt_case = self.target_case.get()
        try:
            logging.info(f'Generating mapping: src={src_path} ({src_type}), tgt={tgt_path} ({tgt_type}), out={out_path}')
            # Step 1: Get source rows
            if src_type == 'xsd':
                src_full_rows = self.get_xsd_rows(src_path)
            else:
                src_full_rows = self.get_json_rows(src_path)
            # Apply case conversion to source levels
            for row in src_full_rows:
                row['levels'] = self.convert_case(row['levels'], src_case)

            # Step 2: Get target rows
            if tgt_type == 'xsd':
                tgt_full_rows = self.get_xsd_rows(tgt_path)
            else:
                tgt_full_rows = self.get_json_rows(tgt_path)
            # Apply case conversion to target levels
            for row in tgt_full_rows:
                row['levels'] = self.convert_case(row['levels'], tgt_case)

            # Step 3: Determine max level for source and target
            max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
            max_tgt_level = max((len(row['levels']) for row in tgt_full_rows), default=1)
            src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Details_src', 'Description_src']
            tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Details_tgt', 'Description_tgt']
            headers = src_cols + ['Destination Fields'] + tgt_cols

            def row_path(row):
                return '.'.join(row['levels'])
            tgt_path_dict = {row_path(row): row for row in tgt_full_rows}
            tgt_paths = list(tgt_path_dict.keys())
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
            ws.append([''] * len(headers))  # Second header row blank for now
            for src_row in src_full_rows:
                src_levels = src_row['levels'] + [''] * (max_src_level - len(src_row['levels']))
                src_vals = src_levels + [src_row.get('Request Parameter',''), src_row.get('GDPR',''), src_row.get('Cardinality',''), src_row.get('Type',''), src_row.get('Details',''), src_row.get('Description','')]
                # Fast dictionary lookup for target row
                src_path_str = row_path(src_row)
                tgt_row = tgt_path_dict.get(src_path_str)
                best_match = ''
                if not tgt_row and tgt_paths:
                    # Fallback to difflib for fuzzy match only if no exact match
                    import difflib
                    matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.0)
                    if matches:
                        best_match = matches[0]
                        tgt_row = tgt_path_dict[best_match]
                tgt_levels = tgt_row['levels'] + [''] * (max_tgt_level - len(tgt_row['levels'])) if tgt_row else ['']*max_tgt_level
                tgt_vals = tgt_levels + [tgt_row.get('Request Parameter','') if tgt_row else '', tgt_row.get('GDPR','') if tgt_row else '', tgt_row.get('Cardinality','') if tgt_row else '', tgt_row.get('Type','') if tgt_row else '', tgt_row.get('Details','') if tgt_row else '', tgt_row.get('Description','') if tgt_row else '']
                dest_field = '.'.join([lvl for lvl in tgt_row['levels'] if lvl]) if tgt_row else ''
                ws.append(src_vals + [dest_field] + tgt_vals)

            # Remove trailing LevelN columns that are empty for all rows (source and target independently)
            from openpyxl.utils import get_column_letter
            def last_nonempty_level_col(start_col, num_levels):
                for col in range(start_col + num_levels - 1, start_col - 1, -1):
                    for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                        if any(cell.value not in (None, '') for cell in row):
                            return col
                return start_col - 1
            # Source Level columns
            src_level_start = 1
            last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
            for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
                ws.delete_cols(col)
            # Dynamically find the starting index of the target Level columns after source columns are deleted
            header_row = [cell.value for cell in ws[1]]
            try:
                tgt_level_start = header_row.index('Level1_tgt') + 1  # 1-based index for openpyxl
            except ValueError:
                tgt_level_start = len(header_row) + 1  # fallback: no target levels found
            for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
                col_letter = get_column_letter(col)
                if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                    ws.delete_cols(col)

            wb.save(out_path)
            logging.info(f'Mapping file generated: {out_path}, Rows: {len(src_full_rows)}')
            self.schema_map_info.delete('1.0', tk.END)
            self.schema_map_info.insert(tk.END, f'Mapping file generated: {out_path}\nRows: {len(src_full_rows)}')
        except Exception as e:
            logging.error(f'Failed to generate mapping: {e}', exc_info=True)
            messagebox.showerror('Error', f'Failed to generate mapping: {e}')

if __name__ == '__main__':
    root = tk.Tk()
    app = SchemaToExcelGUI(root)
    root.mainloop()
