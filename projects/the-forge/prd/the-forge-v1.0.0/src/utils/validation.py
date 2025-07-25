"""
Validation utilities for schema validation and data integrity checks.
"""

import json
import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path


class ValidationUtils:
    """Utility class for validation operations."""
    
    @staticmethod
    def validate_json_schema(filepath: str) -> Tuple[bool, str]:
        """Validate JSON Schema file structure."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Check required JSON Schema fields
            if not isinstance(data, dict):
                return False, "JSON Schema must be a JSON object"
            
            if '$schema' not in data:
                return False, "JSON Schema must have $schema field"
            
            if 'type' not in data and 'properties' not in data:
                return False, "JSON Schema must have either 'type' or 'properties' field"
            
            # Validate properties if present
            if 'properties' in data:
                if not isinstance(data['properties'], dict):
                    return False, "Properties must be an object"
                
                for prop_name, prop_schema in data['properties'].items():
                    if not isinstance(prop_schema, dict):
                        return False, f"Property '{prop_name}' must be an object"
                    
                    if 'type' not in prop_schema:
                        return False, f"Property '{prop_name}' must have a 'type' field"
            
            return True, "JSON Schema is valid"
            
        except json.JSONDecodeError as e:
            return False, f"Invalid JSON: {str(e)}"
        except Exception as e:
            return False, f"Validation error: {str(e)}"
    
    @staticmethod
    def validate_xsd_schema(filepath: str) -> Tuple[bool, str]:
        """Validate XSD schema file structure."""
        try:
            tree = ET.parse(filepath)
            root = tree.getroot()
            
            # Check namespace
            if 'xs' not in root.nsmap and 'http://www.w3.org/2001/XMLSchema' not in root.nsmap.values():
                return False, "XSD must use XML Schema namespace"
            
            # Check for root element
            root_elements = root.findall('xs:element', {'xs': 'http://www.w3.org/2001/XMLSchema'})
            if not root_elements:
                return False, "XSD must have at least one root element"
            
            # Check for complex types
            complex_types = root.findall('xs:complexType', {'xs': 'http://www.w3.org/2001/XMLSchema'})
            simple_types = root.findall('xs:simpleType', {'xs': 'http://www.w3.org/2001/XMLSchema'})
            
            if not complex_types and not simple_types:
                return False, "XSD must have at least one type definition"
            
            return True, "XSD schema is valid"
            
        except ET.ParseError as e:
            return False, f"Invalid XML: {str(e)}"
        except Exception as e:
            return False, f"Validation error: {str(e)}"
    
    @staticmethod
    def validate_schema_fields(fields: List[Dict]) -> Tuple[bool, str]:
        """Validate schema fields structure."""
        if not isinstance(fields, list):
            return False, "Fields must be a list"
        
        required_fields = ['levels', 'type', 'description', 'cardinality', 'details']
        
        for i, field in enumerate(fields):
            if not isinstance(field, dict):
                return False, f"Field {i} must be a dictionary"
            
            for required_field in required_fields:
                if required_field not in field:
                    return False, f"Field {i} missing required field: {required_field}"
            
            if not isinstance(field['levels'], list):
                return False, f"Field {i} levels must be a list"
            
            if not isinstance(field['type'], str):
                return False, f"Field {i} type must be a string"
            
            if not isinstance(field['description'], str):
                return False, f"Field {i} description must be a string"
            
            if not isinstance(field['cardinality'], str):
                return False, f"Field {i} cardinality must be a string"
            
            if not isinstance(field['details'], str):
                return False, f"Field {i} details must be a string"
        
        return True, "Schema fields are valid"
    
    @staticmethod
    def validate_mapping_results(mapping: List[Dict]) -> Tuple[bool, str]:
        """Validate mapping results structure."""
        if not isinstance(mapping, list):
            return False, "Mapping must be a list"
        
        required_fields = ['source', 'target', 'similarity']
        
        for i, mapping_item in enumerate(mapping):
            if not isinstance(mapping_item, dict):
                return False, f"Mapping item {i} must be a dictionary"
            
            for required_field in required_fields:
                if required_field not in mapping_item:
                    return False, f"Mapping item {i} missing required field: {required_field}"
            
            if not isinstance(mapping_item['source'], str):
                return False, f"Mapping item {i} source must be a string"
            
            if not isinstance(mapping_item['target'], str):
                return False, f"Mapping item {i} target must be a string"
            
            if not isinstance(mapping_item['similarity'], (int, float)):
                return False, f"Mapping item {i} similarity must be a number"
            
            if not 0 <= mapping_item['similarity'] <= 1:
                return False, f"Mapping item {i} similarity must be between 0 and 1"
        
        return True, "Mapping results are valid"
    
    @staticmethod
    def validate_excel_file(filepath: str) -> Tuple[bool, str]:
        """Validate Excel file structure."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            
            if not wb.sheetnames:
                return False, "Excel file must have at least one worksheet"
            
            ws = wb.active
            if ws.max_row < 1:
                return False, "Excel file must have at least one row (header)"
            
            if ws.max_column < 1:
                return False, "Excel file must have at least one column"
            
            wb.close()
            return True, "Excel file is valid"
            
        except Exception as e:
            return False, f"Excel validation error: {str(e)}"
    
    @staticmethod
    def compare_schema_structures(schema1_fields: List[Dict], 
                                schema2_fields: List[Dict]) -> Dict[str, Any]:
        """Compare two schema structures and return analysis."""
        schema1_paths = {'.'.join(f['levels']) for f in schema1_fields}
        schema2_paths = {'.'.join(f['levels']) for f in schema2_fields}
        
        common_paths = schema1_paths & schema2_paths
        schema1_only = schema1_paths - schema2_paths
        schema2_only = schema2_paths - schema1_paths
        
        return {
            'common_paths': len(common_paths),
            'schema1_only': len(schema1_only),
            'schema2_only': len(schema2_only),
            'total_schema1': len(schema1_paths),
            'total_schema2': len(schema2_paths),
            'similarity_percentage': round(len(common_paths) / max(len(schema1_paths), len(schema2_paths)) * 100, 2) if max(len(schema1_paths), len(schema2_paths)) > 0 else 0,
            'common_paths_list': list(common_paths),
            'schema1_only_list': list(schema1_only),
            'schema2_only_list': list(schema2_only)
        }
    
    @staticmethod
    def validate_conversion_roundtrip(original_path: str, converted_path: str, 
                                     conversion_type: str) -> Tuple[bool, str]:
        """Validate that a conversion roundtrip produces consistent results."""
        try:
            if conversion_type == 'xsd_to_json':
                # Convert back to XSD and compare
                from ..core.converter import SchemaConverter
                converter = SchemaConverter()
                
                # Convert JSON back to XSD
                temp_xsd_path = converted_path.replace('.json', '_roundtrip.xsd')
                converter.json_schema_to_xsd(converted_path, temp_xsd_path)
                
                # Compare original and roundtrip XSD
                original_tree = ET.parse(original_path)
                roundtrip_tree = ET.parse(temp_xsd_path)
                
                # Simple comparison of root elements
                original_root = original_tree.getroot()
                roundtrip_root = roundtrip_tree.getroot()
                
                original_elements = original_root.findall('xs:element', {'xs': 'http://www.w3.org/2001/XMLSchema'})
                roundtrip_elements = roundtrip_root.findall('xs:element', {'xs': 'http://www.w3.org/2001/XMLSchema'})
                
                if len(original_elements) != len(roundtrip_elements):
                    return False, f"Roundtrip conversion failed: different number of root elements ({len(original_elements)} vs {len(roundtrip_elements)})"
                
                # Clean up temp file
                Path(temp_xsd_path).unlink(missing_ok=True)
                
                return True, "Roundtrip conversion successful"
            
            elif conversion_type == 'json_to_xsd':
                # Convert back to JSON and compare
                from ..core.converter import SchemaConverter
                converter = SchemaConverter()
                
                # Convert XSD back to JSON
                temp_json_path = converted_path.replace('.xsd', '_roundtrip.json')
                converter.xsd_to_json_schema(converted_path, temp_json_path)
                
                # Compare original and roundtrip JSON
                with open(original_path, 'r') as f:
                    original_json = json.load(f)
                with open(temp_json_path, 'r') as f:
                    roundtrip_json = json.load(f)
                
                # Simple comparison of structure
                if original_json.get('type') != roundtrip_json.get('type'):
                    return False, "Roundtrip conversion failed: different root types"
                
                # Clean up temp file
                Path(temp_json_path).unlink(missing_ok=True)
                
                return True, "Roundtrip conversion successful"
            
            else:
                return False, f"Unknown conversion type: {conversion_type}"
                
        except Exception as e:
            return False, f"Roundtrip validation error: {str(e)}"
    
    @staticmethod
    def validate_mapping_coverage(mapping: List[Dict], 
                                 source_fields: List[Dict], 
                                 target_fields: List[Dict]) -> Dict[str, Any]:
        """Validate mapping coverage and quality."""
        source_paths = {'.'.join(f['levels']) for f in source_fields}
        target_paths = {'.'.join(f['levels']) for f in target_fields}
        
        mapped_source = {m['source'] for m in mapping if m['similarity'] >= 0.7}
        mapped_target = {m['target'] for m in mapping if m['similarity'] >= 0.7}
        
        unmapped_source = source_paths - mapped_source
        unmapped_target = target_paths - mapped_target
        
        exact_matches = sum(1 for m in mapping if m['similarity'] == 1.0)
        good_matches = sum(1 for m in mapping if m['similarity'] >= 0.8)
        poor_matches = sum(1 for m in mapping if m['similarity'] < 0.5)
        
        avg_similarity = sum(m['similarity'] for m in mapping) / len(mapping) if mapping else 0
        
        return {
            'total_mappings': len(mapping),
            'exact_matches': exact_matches,
            'good_matches': good_matches,
            'poor_matches': poor_matches,
            'average_similarity': round(avg_similarity, 3),
            'source_coverage': round(len(mapped_source) / len(source_paths) * 100, 1) if source_paths else 0,
            'target_coverage': round(len(mapped_target) / len(target_paths) * 100, 1) if target_paths else 0,
            'unmapped_source_count': len(unmapped_source),
            'unmapped_target_count': len(unmapped_target),
            'unmapped_source_paths': list(unmapped_source),
            'unmapped_target_paths': list(unmapped_target)
        } 