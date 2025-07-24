"""
Excel Generator - Creates formatted Excel files for schema documentation and mapping
Enhanced for Request/Response structure with detailed metadata
"""

from typing import Dict, List, Any, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .schema_field import SchemaField
from .mapping_engine import FieldMapping
from openpyxl.cell.cell import MergedCell

class ExcelGenerator:
    """Generates formatted Excel files for schema documentation and mapping"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Color coding for different mapping statuses
        self.exact_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        self.good_match_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
        self.weak_match_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Red
        self.unmapped_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")    # Gray
    
    def create_mapping_excel_request_response(self, source_fields: List[SchemaField], target_fields: List[SchemaField], 
                                           mappings: List[FieldMapping], transformations: Dict[str, Dict[str, Any]], 
                                           output_path: str, sender_name: str = "Sender", receiver_name: str = "Receiver") -> bool:
        """Create Excel file with a single flow (Request) for confirmed mappings, with dynamic headers"""
        try:
            wb = Workbook()
            # Create only the Request sheet
            request_ws = wb.active
            if request_ws:
                request_ws.title = "Request"
            self._create_request_response_sheet(request_ws, source_fields, mappings, transformations, "Request", sender_name, receiver_name)
            # Optionally, keep summary, transformations, unmapped sheets
            self._create_summary_sheet(wb, source_fields, target_fields, mappings)
            if transformations:
                self._create_transformations_sheet(wb, mappings, transformations)
            self._create_unmapped_sheet(wb, source_fields, target_fields, mappings)
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error creating Request/Response Excel: {e}")
            return False

    def _create_request_response_sheet(self, ws, fields: List[SchemaField], mappings: List[FieldMapping], 
                                     transformations: Dict[str, Dict[str, Any]], sheet_type: str, sender_name: str, receiver_name: str):
        """Create Request or Response sheet with detailed field information and two header rows"""
        # Define columns and group headers
        group_headers = [
            sender_name, sender_name, sender_name, sender_name, sender_name, sender_name,
            "Mappings", "Mappings", "Mappings", receiver_name, receiver_name
        ]
        columns = [
            "Element", "Request Parame", "GDPR", "Cardinal", "Type", "Details",
            "Description", "Transformation Mapping", "Destination Fields", "Element", "Description"
        ]
        # First header row (group headers)
        for col, group in enumerate(group_headers, 1):
            cell = ws.cell(row=1, column=col, value=group)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Second header row (column names)
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Merge cells for group headers
        col_ranges = [(1,6), (7,9), (10,11)]
        for start, end in col_ranges:
            if start != end:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        # Fill data rows (starting from row 3)
        row = 3
        for field in fields:
            # Exemplo de preenchimento, adaptar conforme necessário
            row_data = [
                field.name, "", "", field.cardinality, field.type, field.description,
                "", "", "", "", ""
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1
        self._auto_adjust_columns(ws)
    
    def _get_transformation_logic(self, field_path: str, transformations: Dict[str, Dict[str, Any]]) -> str:
        """Get transformation logic for a field"""
        if field_path not in transformations:
            return ""
        
        rules = transformations[field_path]
        logic_parts = []
        
        if rules.get("type_conversion") and rules["type_conversion"] != "No conversion":
            logic_parts.append(f"Type: {rules['type_conversion']}")
        
        if rules.get("format_pattern"):
            logic_parts.append(f"Format: {rules['format_pattern']}")
        
        if rules.get("default_value"):
            logic_parts.append(f"Default: {rules['default_value']}")
        
        if rules.get("apply_validation"):
            logic_parts.append("Validation: Yes")
        
        if rules.get("custom_code"):
            logic_parts.append(f"Custom: {rules['custom_code'][:50]}...")
        
        return "; ".join(logic_parts)
    
    def _format_constraints(self, constraints: Dict[str, Any]) -> str:
        """Format constraints dictionary as string"""
        if not constraints:
            return ""
        
        constraint_parts = []
        for key, value in constraints.items():
            if isinstance(value, list):
                constraint_parts.append(f"{key}: {', '.join(map(str, value))}")
            else:
                constraint_parts.append(f"{key}: {value}")
        
        return "; ".join(constraint_parts)
    
    def _create_summary_sheet(self, wb: Workbook, source_fields: List[SchemaField], 
                            target_fields: List[SchemaField], mappings: List[FieldMapping]) -> None:
        """Create summary sheet with mapping statistics"""
        ws = wb.create_sheet("Mapping Summary")
        
        # Summary statistics
        total_source = len(source_fields)
        total_target = len(target_fields)
        exact_matches = len([m for m in mappings if m.is_exact_match])
        good_matches = len([m for m in mappings if m.is_good_match])
        weak_matches = len([m for m in mappings if m.is_weak_match])
        unmapped = len([m for m in mappings if m.is_unmapped])
        mapped_fields = len([m for m in mappings if not m.is_unmapped])
        
        # Headers
        headers = ["Metric", "Value", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        data = [
            ["Total Source Fields", total_source, "100%"],
            ["Total Target Fields", total_target, "100%"],
            ["Mapped Fields", mapped_fields, f"{mapped_fields/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Exact Matches", exact_matches, f"{exact_matches/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Good Matches", good_matches, f"{good_matches/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Weak Matches", weak_matches, f"{weak_matches/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Unmapped Fields", unmapped, f"{unmapped/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Average Similarity", f"{sum(m.similarity for m in mappings)/len(mappings):.3f}" if mappings else "0.000", ""],
            ["Source Coverage", f"{mapped_fields}/{total_source}", f"{mapped_fields/total_source*100:.1f}%" if total_source > 0 else "0%"],
            ["Target Coverage", f"{len(set(m.target_field.path for m in mappings if not m.is_unmapped))}/{total_target}", f"{len(set(m.target_field.path for m in mappings if not m.is_unmapped))/total_target*100:.1f}%" if total_target > 0 else "0%"]
        ]
        
        for row, (metric, value, percentage) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            ws.cell(row=row, column=3, value=percentage).border = self.border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_transformations_sheet(self, wb: Workbook, mappings: List[FieldMapping], 
                                    transformations: Dict[str, Dict[str, Any]]) -> None:
        """Create transformations sheet with detailed transformation rules"""
        ws = wb.create_sheet("Transformations")
        
        # Headers
        headers = [
            "Source Field", "Target Field", "Type Conversion", "Format Pattern", 
            "Default Value", "Apply Validation", "Custom Code", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        row = 2
        for source_path, rules in transformations.items():
            # Find the mapping for this source field
            mapping = next((m for m in mappings if m.source_field.path == source_path), None)
            
            if mapping:
                ws.cell(row=row, column=1, value=source_path).border = self.border
                ws.cell(row=row, column=2, value=mapping.target_field.path).border = self.border
                ws.cell(row=row, column=3, value=rules.get("type_conversion", "")).border = self.border
                ws.cell(row=row, column=4, value=rules.get("format_pattern", "")).border = self.border
                ws.cell(row=row, column=5, value=rules.get("default_value", "")).border = self.border
                ws.cell(row=row, column=6, value="Yes" if rules.get("apply_validation", False) else "No").border = self.border
                ws.cell(row=row, column=7, value=rules.get("custom_code", "")).border = self.border
                ws.cell(row=row, column=8, value=mapping.mapping_notes).border = self.border
            
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_unmapped_sheet(self, wb: Workbook, source_fields: List[SchemaField], 
                              target_fields: List[SchemaField], mappings: List[FieldMapping]) -> None:
        """Create sheet showing unmapped fields"""
        ws = wb.create_sheet("Unmapped Fields")
        
        # Headers
        headers = ["Schema Type", "Field Path", "Field Name", "Data Type", "Cardinality", "Required", "Description", "Constraints"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Unmapped source fields
        row = 2
        for mapping in mappings:
            if mapping.is_unmapped:
                field = mapping.source_field
                ws.cell(row=row, column=1, value="Source").border = self.border
                ws.cell(row=row, column=2, value=field.path).border = self.border
                ws.cell(row=row, column=3, value=field.name).border = self.border
                ws.cell(row=row, column=4, value=field.type).border = self.border
                ws.cell(row=row, column=5, value=field.cardinality).border = self.border
                ws.cell(row=row, column=6, value="Yes" if field.required else "No").border = self.border
                ws.cell(row=row, column=7, value=field.description).border = self.border
                ws.cell(row=row, column=8, value=self._format_constraints(field.constraints)).border = self.border
                row += 1
        
        # Unmapped target fields
        mapped_target_paths = {m.target_field.path for m in mappings if not m.is_unmapped}
        for field in target_fields:
            if field.path not in mapped_target_paths:
                ws.cell(row=row, column=1, value="Target").border = self.border
                ws.cell(row=row, column=2, value=field.path).border = self.border
                ws.cell(row=row, column=3, value=field.name).border = self.border
                ws.cell(row=row, column=4, value=field.type).border = self.border
                ws.cell(row=row, column=5, value=field.cardinality).border = self.border
                ws.cell(row=row, column=6, value="Yes" if field.required else "No").border = self.border
                ws.cell(row=row, column=7, value=field.description).border = self.border
                ws.cell(row=row, column=8, value=self._format_constraints(field.constraints)).border = self.border
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths, skipping merged cells"""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Legacy methods for backward compatibility
    def create_schema_excel(self, fields: List[SchemaField], output_path: str, schema_name: str = "Schema") -> bool:
        """Create Excel file with schema field documentation (legacy method)"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            if ws:
                ws.title = "Schema Fields"
            
            # Define columns
            columns = [
                "Element Level 1", "Element Level 2", "Element Level 3", 
                "Element Level 4", "Element Level 5", "Element Level 6",
                "Type", "Cardinality", "Description", "Details", "JSON Path"
            ]
            
            # Add headers
            for col, header in enumerate(columns, 1):
                if ws:
                    cell = ws.cell(row=1, column=col, value=header)
                    if cell:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.border = self.border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Process fields
            row = 2
            for field in fields:
                # Split path into levels
                path_parts = field.path.split('.')
                levels = path_parts + [''] * (6 - len(path_parts))
                
                # Create row data
                row_data = levels + [
                    field.type,
                    field.cardinality,
                    field.description,
                    field.path
                ]
                
                # Add row to worksheet
                for col, value in enumerate(row_data, 1):
                    if ws:
                        cell = ws.cell(row=row, column=col, value=value)
                        if cell:
                            cell.border = self.border
                            cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                row += 1
            
            # Auto-adjust column widths
            if ws:
                self._auto_adjust_columns(ws)
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error creating schema Excel: {e}")
            return False
    
    def create_mapping_excel(self, source_fields: List[SchemaField], target_fields: List[SchemaField], 
                           mappings: List[FieldMapping], output_path: str) -> bool:
        """Create Excel file with field mappings between schemas (legacy method)"""
        # Use the new Request/Response structure
        return self.create_mapping_excel_request_response(source_fields, target_fields, mappings, {}, output_path) 

    def create_field_level_mapping_excel(self, mappings: List[FieldMapping], output_path: str) -> bool:
        """Export all fields (complex and leaf) from both source and target, aligning mapped fields, and including unmapped fields with hierarchical indentation. Cardinality is always taken from the field data, never derived."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Field Mapping"

            def split_path(path):
                return path.split(".") if path else []

            # Collect all unique source and target fields (complex and leaf)
            all_source_fields = []
            all_target_fields = []
            seen_src = set()
            seen_tgt = set()
            for m in mappings:
                f = m.source_field
                t = m.target_field
                if f and f.path not in seen_src:
                    all_source_fields.append(f)
                    seen_src.add(f.path)
                if t and t.path and t.path not in seen_tgt:
                    all_target_fields.append(t)
                    seen_tgt.add(t.path)
            # Add unmapped target fields
            for m in mappings:
                t = m.target_field
                if t and t.path and t.path not in seen_tgt:
                    all_target_fields.append(t)
                    seen_tgt.add(t.path)
            # Add any target fields not in mappings (fully unmapped)
            mapping_target_paths = {m.target_field.path for m in mappings if m.target_field and m.target_field.path}
            for t in all_target_fields:
                if t.path not in mapping_target_paths:
                    all_target_fields.append(t)

            # Determine max depth for source and target paths
            max_src_depth = max((len(split_path(f.path)) for f in all_source_fields), default=1)
            max_tgt_depth = max((len(split_path(f.path)) for f in all_target_fields), default=1)

            src_levels = [f"Source Level{i+1}" for i in range(max_src_depth)]
            tgt_levels = [f"Target Level{i+1}" for i in range(max_tgt_depth)]
            src_cols = src_levels + ["Request Parameter", "GDPR", "Cardinality", "Type", "Description"]
            map_cols = ["Transformation Mapping", "Destination Fields"]
            tgt_cols = tgt_levels + ["Request Parameter", "GDPR", "Cardinality", "Type", "Description"]
            headers = src_cols + map_cols + tgt_cols
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Build mappings for quick lookup
            mapping_by_src = {m.source_field.path: m for m in mappings if m.source_field}
            mapping_by_tgt = {m.target_field.path: m for m in mappings if m.target_field and m.target_field.path}

            # Prepare rows: mapped, unmapped source, unmapped target
            rows = []
            used_targets = set()
            # 1. Mapped and unmapped source fields
            for src_field in all_source_fields:
                mapping = mapping_by_src.get(src_field.path)
                if mapping and mapping.target_field and mapping.target_field.path:
                    tgt_field = mapping.target_field
                    used_targets.add(tgt_field.path)
                else:
                    tgt_field = None
                rows.append((src_field, mapping, tgt_field))
            # 2. Unmapped target fields
            for tgt_field in all_target_fields:
                if tgt_field.path not in used_targets:
                    rows.append((None, None, tgt_field))

            # For hierarchical indentation: track last values for each level (source and target)
            last_src_levels = [None] * max_src_depth
            last_tgt_levels = [None] * max_tgt_depth

            for row_idx, (src_field, mapping, tgt_field) in enumerate(rows, 2):
                # Source columns
                if src_field:
                    src_path_parts = split_path(src_field.path)
                    src_row = []
                    for i in range(max_src_depth):
                        val = src_path_parts[i] if i < len(src_path_parts) else ""
                        if val != last_src_levels[i]:
                            src_row.append(val)
                            last_src_levels[i] = val
                            for j in range(i+1, max_src_depth):
                                last_src_levels[j] = None
                        else:
                            src_row.append("")
                    src_row += [
                        getattr(src_field, 'request_parameter', ""),
                        getattr(src_field, 'gdpr', ""),
                        src_field.cardinality,
                        src_field.type,
                        src_field.description
                    ]
                else:
                    src_row = [""] * (max_src_depth + 5)
                # Mapping columns
                if mapping and mapping.target_field and mapping.target_field.path:
                    transformation = mapping.mapping_notes or ""
                    if hasattr(self, '_get_transformation_logic'):
                        logic = self._get_transformation_logic(mapping.source_field.path, {})
                        if logic:
                            transformation = logic
                    dest_fields = mapping.target_field.path
                else:
                    transformation = ""
                    dest_fields = tgt_field.path if tgt_field else ""
                map_row = [transformation, dest_fields]
                # Target columns
                if tgt_field:
                    tgt_path_parts = split_path(tgt_field.path)
                    tgt_row = []
                    for i in range(max_tgt_depth):
                        val = tgt_path_parts[i] if i < len(tgt_path_parts) else ""
                        if val != last_tgt_levels[i]:
                            tgt_row.append(val)
                            last_tgt_levels[i] = val
                            for j in range(i+1, max_tgt_depth):
                                last_tgt_levels[j] = None
                        else:
                            tgt_row.append("")
                    tgt_row += [
                        getattr(tgt_field, 'request_parameter', ""),
                        getattr(tgt_field, 'gdpr', ""),
                        tgt_field.cardinality,
                        tgt_field.type,
                        tgt_field.description
                    ]
                else:
                    tgt_row = [""] * (max_tgt_depth + 5)
                row_data = src_row + map_row + tgt_row
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    if cell is not None:
                        cell.border = self.border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
            self._auto_adjust_columns(ws)
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error creating Field Mapping Excel: {e}")
            return False 