import openpyxl
from openpyxl.styles import Font
import os
import json
import sys

class ExcelExporter:
    def __init__(self, max_level=8):
        self.max_level = max_level

    def export(self, xsd_data_dict, output_file):
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        sheet_count = 0
        for sheet_name, rows in xsd_data_dict.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            headers = [f'Level{i+1}' for i in range(self.max_level)] + ['Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append(row['levels'] + [row['Request Parameter'], row['GDPR'], row['Cardinality'], row['Type'], row['Details'], row['Description']])
            sheet_count += 1
        if sheet_count > 0:
            wb.remove(default_sheet)
            wb.save(output_file)
            return True
        else:
            return False

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Export parsed XSD data (JSON) to Excel.')
    parser.add_argument('input', help='Input JSON file (from XSDParser)')
    parser.add_argument('output', help='Output Excel file')
    args = parser.parse_args()

    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)
    exporter = ExcelExporter()
    success = exporter.export(data, args.output)
    if success:
        print(f'Excel file exported to {args.output}')
    else:
        print('No sheets were created. No Excel file exported.') 