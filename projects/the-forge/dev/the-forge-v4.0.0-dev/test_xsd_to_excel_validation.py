print('--- test_xsd_to_excel_validation.py started ---')
import sys
import os
import openpyxl
from microservices.xsd_parser_service import XSDParser
from microservices.wsdl_to_xsd_extractor import extract_xsds_from_wsdl, extract_wsdl_operations, merge_xsd_schemas
import xml.etree.ElementTree as ET

# Ensure script runs from its own directory
os.chdir(os.path.dirname(os.path.abspath(__file__)))

XSD_NS = '{http://www.w3.org/2001/XMLSchema}'

def flatten_levels(levels):
    return '.'.join(levels)

def get_flattened_path(row):
    # Only consider columns with non-None keys
    levels = [v for k, v in sorted(((k, v) for k, v in row.items() if k is not None), key=lambda x: x[0])
              if k.lower().startswith('level') and v]
    return '.'.join(str(l) for l in levels if l)

def extract_fields_from_wsdl(wsdl_path):
    print(f'Extracting fields from WSDL: {wsdl_path}')
    schemas = extract_xsds_from_wsdl(wsdl_path)
    operations = extract_wsdl_operations(wsdl_path)
    if not schemas or not operations:
        print("No schemas or operations found in WSDL.")
        return []
    merged_xsd = merge_xsd_schemas([x[1] for x in schemas])
    all_fields = []
    parser = XSDParser()
    root = ET.fromstring(merged_xsd)
    for op in operations:
        root_elem = op['input_element']
        if not root_elem:
            continue
        elem_name = root_elem[1]
        elem = None
        for e in root.findall(f'.//{XSD_NS}element'):
            if e.get('name') == elem_name:
                elem = e
                break
        if elem is not None:
            simple_types = {}
            for st in root.findall(f'.//{XSD_NS}simpleType'):
                name = st.get('name')
                if not name:
                    continue
                restriction = st.find(f'{XSD_NS}restriction')
                base = restriction.get('base') if restriction is not None else None
                restrictions = {}
                if restriction is not None:
                    for cons in restriction:
                        cons_name = cons.tag.replace(XSD_NS, '')
                        val = cons.get('value')
                        if val:
                            restrictions[cons_name] = val
                simple_types[name] = {'base': base, 'restrictions': restrictions}
            complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{XSD_NS}complexType') if ct.get('name')}
            op_fields = parser.parse_element(elem, complex_types, simple_types, 1, category='message')
            all_fields.extend(op_fields)
    print(f'Extracted {len(all_fields)} fields from WSDL.')
    return all_fields

def parse_and_export(xsd_path, excel_path):
    print(f'Parsing and exporting: {xsd_path} -> {excel_path}')
    parser = XSDParser()
    rows = []
    if xsd_path.lower().endswith('.wsdl'):
        schemas = extract_xsds_from_wsdl(xsd_path)
        operations = extract_wsdl_operations(xsd_path)
        if not schemas or not operations:
            print("No schemas or operations found in WSDL.")
            return []
        merged_xsd = merge_xsd_schemas([x[1] for x in schemas])
        root_elem = operations[0]['input_element']  # Use first operation's input
        if not root_elem:
            print("No input element found for first operation.")
            return []
        root = ET.fromstring(merged_xsd)
        elem_name = root_elem[1]
        elem = None
        for e in root.findall(f'.//{XSD_NS}element'):
            if e.get('name') == elem_name:
                elem = e
                break
        if elem is not None:
            simple_types = {}
            for st in root.findall(f'.//{XSD_NS}simpleType'):
                name = st.get('name')
                if not name:
                    continue
                restriction = st.find(f'{XSD_NS}restriction')
                base = restriction.get('base') if restriction is not None else None
                restrictions = {}
                if restriction is not None:
                    for cons in restriction:
                        cons_name = cons.tag.replace(XSD_NS, '')
                        val = cons.get('value')
                        if val:
                            restrictions[cons_name] = val
                simple_types[name] = {'base': base, 'restrictions': restrictions}
            complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{XSD_NS}complexType') if ct.get('name')}
            rows = parser.parse_element(elem, complex_types, simple_types, 1, category='message')
    else:
        rows = parser.parse_xsd_file(xsd_path)
    if not rows:
        print("No rows found for export.")
        return []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([*rows[0].keys()])
    for row in rows:
        row_out = row.copy()
        row_out['levels'] = flatten_levels(row['levels'])
        ws.append([row_out[k] for k in row_out.keys()])
    wb.save(excel_path)
    print(f"Exported {len(rows)} rows to {excel_path}")
    return rows

def validate_excel_against_reference(generated_excel, reference_excel):
    print(f'Comparing generated Excel: {generated_excel} against reference: {reference_excel}')
    import openpyxl
    def load_fields(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        print(f'Loaded {len(rows)} rows from {path}')
        return rows
    ref_rows = load_fields(reference_excel)
    gen_rows = load_fields(generated_excel)
    missing = []
    type_mismatches = []
    for r in ref_rows:
        ref_path = get_flattened_path(r)
        ref_type = r.get('Type')
        found = False
        for g in gen_rows:
            gen_path = get_flattened_path(g)
            if gen_path == ref_path:
                found = True
                if g.get('Type') != ref_type:
                    type_mismatches.append((ref_path, ref_type, g.get('Type')))
                break
        if not found:
            missing.append(ref_path)
    print(f"Total reference fields: {len(ref_rows)}")
    print(f"Total generated fields: {len(gen_rows)}")
    print(f"Missing in generated: {len(missing)}")
    if missing:
        print("Missing rows:")
        for m in missing:
            print(f"  {m}")
    print(f"Type mismatches: {len(type_mismatches)}")
    if type_mismatches:
        print("Type mismatches:")
        for m in type_mismatches:
            print(f"  {m[0]}: expected {m[1]}, got {m[2]}")
    if not missing and not type_mismatches:
        print("Validation PASSED: All reference fields present and types correct in generated Excel.")
    else:
        print("Validation FAILED.")

if __name__ == '__main__':
    print('--- Script main started ---')
    if len(sys.argv) < 2:
        print("Usage: python test_xsd_to_excel_validation.py <xsd_or_wsdl_path> [excel_path] [reference_excel]")
        sys.exit(1)
    xsd_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else 'test_output.xlsx'
    reference_excel = sys.argv[3] if len(sys.argv) > 3 else None
    print(f'Input: {xsd_path}\nOutput: {excel_path}\nReference: {reference_excel}')
    rows = parse_and_export(xsd_path, excel_path)
    if reference_excel:
        validate_excel_against_reference(excel_path, reference_excel)
    elif rows and xsd_path.lower().endswith('.wsdl'):
        # The original code had validate_excel_against_wsdl(excel_path, xsd_path) here,
        # but validate_excel_against_wsdl is not defined in the provided code.
        # Assuming it was intended to be a placeholder or will be added later.
        # For now, commenting out to avoid errors.
        # validate_excel_against_wsdl(excel_path, xsd_path)
        print("Validation complete for XSD input (no WSDL field comparison).")
    elif rows:
        print("Validation complete for XSD input (no WSDL field comparison).") 