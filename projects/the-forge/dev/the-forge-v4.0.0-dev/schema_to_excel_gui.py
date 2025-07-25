import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), 'microservices'))
from microservices.xsd_parser_service import XSDParser
from microservices.json_to_excel_service import JSONToExcelService
from microservices.xsd_json_converter_service import XSDJSONConverterService
from microservices.case_converter_service import pascal_to_camel, camel_to_pascal
import json
import openpyxl
import logging
import xml.etree.ElementTree as ET

# Set up logging
log_path = os.path.join(os.path.dirname(__file__), 'schema_to_excel_gui.log')
logging.basicConfig(
    filename=log_path,
    filemode='a',
    format='%(asctime)s %(levelname)s: %(message)s',
    level=logging.INFO
)

XSD_NS = '{http://www.w3.org/2001/XMLSchema}'

class SchemaToExcelGUI:
    def __init__(self, root):
        logging.info('Starting SchemaToExcelGUI')
        self.root = root
        self.root.title('Schema to Mapping Tool')
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.schema_source_path = tk.StringVar()
        self.schema_target_path = tk.StringVar()
        self.mapping_output_path = tk.StringVar()
        self.schema_source_type = tk.StringVar(value='xsd')
        self.schema_target_type = tk.StringVar(value='xsd')
        self.source_case = tk.StringVar(value='none')
        self.target_case = tk.StringVar(value='none')
        self.json_to_excel_service = JSONToExcelService()
        self.xsd_json_converter = XSDJSONConverterService()
        self.create_widgets()
        # self.create_wsdl_to_xsd_tab()  # Remove WSDL tab

    def create_widgets(self):
        frame_schema_map = tk.Frame(self.notebook, padx=10, pady=10)
        self.notebook.add(frame_schema_map, text='Schema Mapping')
        tk.Label(frame_schema_map, text='Source Schema:').grid(row=0, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_source_path, width=40).grid(row=0, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=lambda: self.browse_schema(self.schema_source_path)).grid(row=0, column=2)
        tk.Label(frame_schema_map, text='Type:').grid(row=0, column=3, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_source_type, width=8, state='readonly').grid(row=0, column=4)
        tk.Label(frame_schema_map, text='Source Case:').grid(row=0, column=5, sticky='w')
        ttk.Combobox(frame_schema_map, textvariable=self.source_case, values=['none', 'pascal', 'camel'], state='readonly', width=8).grid(row=0, column=6)
        tk.Label(frame_schema_map, text='Target Schema:').grid(row=1, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_target_path, width=40).grid(row=1, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=lambda: self.browse_schema(self.schema_target_path)).grid(row=1, column=2)
        tk.Label(frame_schema_map, text='Type:').grid(row=1, column=3, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.schema_target_type, width=8, state='readonly').grid(row=1, column=4)
        tk.Label(frame_schema_map, text='Target Case:').grid(row=1, column=5, sticky='w')
        ttk.Combobox(frame_schema_map, textvariable=self.target_case, values=['none', 'pascal', 'camel'], state='readonly', width=8).grid(row=1, column=6)
        tk.Label(frame_schema_map, text='Output Mapping Excel:').grid(row=2, column=0, sticky='w')
        tk.Entry(frame_schema_map, textvariable=self.mapping_output_path, width=40).grid(row=2, column=1, padx=5)
        tk.Button(frame_schema_map, text='Browse...', command=self.browse_mapping_output).grid(row=2, column=2)
        tk.Button(frame_schema_map, text='Generate Mapping', command=self.generate_schema_mapping).grid(row=3, column=1, pady=10)
        self.schema_map_info = tk.Text(frame_schema_map, height=10, width=80)
        self.schema_map_info.grid(row=4, column=0, columnspan=7, pady=10)

    def browse_schema(self, var):
        file_path = filedialog.askopenfilename(filetypes=[('Schema Files', '*.xsd *.json')])
        if file_path:
            var.set(file_path)
            ext = os.path.splitext(file_path)[1].lower()
            if var == self.schema_source_path:
                if ext == '.json':
                    self.schema_source_type.set('json')
                elif ext == '.xsd':
                    self.schema_source_type.set('xsd')
            elif var == self.schema_target_path:
                if ext == '.json':
                    self.schema_target_type.set('json')
                elif ext == '.xsd':
                    self.schema_target_type.set('xsd')

    def browse_mapping_output(self):
        try:
            file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
            if file_path:
                self.mapping_output_path.set(file_path)
                logging.info(f'User selected output mapping file: {file_path}')
        except Exception as e:
            logging.error(f'Error in browse_mapping_output: {e}')
            raise
    def convert_case(self, levels, mode):
        if mode == 'none':
            return levels
        elif mode == 'pascal':
            return [camel_to_pascal(lvl) if lvl else '' for lvl in levels]
        elif mode == 'camel':
            return [pascal_to_camel(lvl) if lvl else '' for lvl in levels]
        return levels
    def load_wsdl_schemas(self, wsdl_path, is_source):
        try:
            schemas = extract_xsds_from_wsdl(wsdl_path)
            # Extract all messages
            tree = ET.parse(wsdl_path)
            root = tree.getroot()
            ns = {'wsdl': 'http://schemas.xmlsoap.org/wsdl/'}
            messages = []
            for msg in root.findall('wsdl:message', ns):
                msg_name = msg.get('name')
                for part in msg.findall('wsdl:part', ns):
                    elem = part.get('element')
                    if elem:
                        messages.append((msg_name, elem))
            if not schemas:
                messagebox.showerror('Error', 'No XSD schemas found in WSDL.')
                return
            if not messages:
                messagebox.showerror('Error', 'No messages found in WSDL.')
                return
            msg_options = [f"{name} ({elem})" for name, elem in messages]
            if is_source:
                self.wsdl_source_schemas = schemas
                self.wsdl_source_messages = messages
                self.selected_wsdl_source_message.set(msg_options[0])
            else:
                self.wsdl_target_schemas = schemas
                self.wsdl_target_messages = messages
                self.selected_wsdl_target_message.set(msg_options[0])
            self.show_wsdl_message_selector(is_source, msg_options)
        except Exception as e:
            messagebox.showerror('Error', f'Failed to extract XSDs/messages from WSDL: {e}')

    def show_wsdl_message_selector(self, is_source, msg_options):
        selector = tk.Toplevel(self.root)
        selector.title('Select WSDL Message')
        tk.Label(selector, text='Select message to use for mapping:').pack(pady=5)
        var = self.selected_wsdl_source_message if is_source else self.selected_wsdl_target_message
        combo = ttk.Combobox(selector, textvariable=var, values=msg_options, state='readonly', width=50)
        combo.pack(padx=10, pady=10)
        def on_ok():
            # After message selection, merge schemas and set root element
            msg_idx = combo.current()
            schemas = self.wsdl_source_schemas if is_source else self.wsdl_target_schemas
            messages = self.wsdl_source_messages if is_source else self.wsdl_target_messages
            msg_name, elem_qname = messages[msg_idx]
            merged_xsd = merge_xsd_schemas([x[1] for x in schemas])
            # Parse the QName to get the local name
            if ':' in elem_qname:
                _, elem_local = elem_qname.split(':', 1)
            else:
                elem_local = elem_qname
            if is_source:
                self.wsdl_source_merged_xsd = merged_xsd
                self.wsdl_source_root_element = (None, elem_local)
            else:
                self.wsdl_target_merged_xsd = merged_xsd
                self.wsdl_target_root_element = (None, elem_local)
            selector.destroy()
        tk.Button(selector, text='OK', command=on_ok).pack(pady=5)

    def get_xsd_rows(self, path, wsdl_schemas=None, wsdl_selected=None, merged_xsd=None, root_element=None):
        parser = XSDParser()
        if merged_xsd and root_element:
            root = ET.fromstring(merged_xsd)
            elem_name = root_element[1] if root_element else None
            elem = None
            for e in root.findall(f'.//{XSD_NS}element'):
                if e.get('name') == elem_name:
                    elem = e
                    break
            if elem is not None:
                simple_types = {}
                for st in root.findall(f'.//{XSD_NS}simpleType'):
                    name = st.get('name')
                    if not name:
                        continue
                    restriction = st.find(f'{XSD_NS}restriction')
                    base = restriction.get('base') if restriction is not None else None
                    restrictions = {}
                    if restriction is not None:
                        for cons in restriction:
                            cons_name = cons.tag.replace(XSD_NS, '')
                            val = cons.get('value')
                            if val:
                                restrictions[cons_name] = val
                    simple_types[name] = {'base': base, 'restrictions': restrictions}
                complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{XSD_NS}complexType') if ct.get('name')}
                return parser.parse_element(elem, complex_types, simple_types, 1, category='message')
            else:
                return []
        if wsdl_schemas and wsdl_selected:
            idx = int(wsdl_selected.split(':')[0]) - 1
            xsd_str = wsdl_schemas[idx][1]
            return parser.parse_xsd_string(xsd_str)
        return parser.parse_xsd_file(path)
    def get_json_rows(self, path):
        # Improved JSON Schema logic: recursively walk and extract all fields/structures
        def walk_json_schema(schema, parent_levels=None, max_depth=20):
            if parent_levels is None:
                parent_levels = []
            rows = []
            if 'type' in schema and schema['type'] == 'object' and 'properties' in schema:
                for prop, prop_schema in schema['properties'].items():
                    levels = parent_levels + [prop]
                    row = {
                        'levels': levels,
                        'Request Parameter': '',
                        'GDPR': '',
                        'Cardinality': '',
                        'Type': prop_schema.get('type', ''),
                        'Details': '',
                        'Description': prop_schema.get('description', ''),
                        'Category': 'element'
                    }
                    rows.append(row)
                    # Recurse into nested objects
                    if prop_schema.get('type') == 'object':
                        rows.extend(walk_json_schema(prop_schema, levels, max_depth))
                    # Handle arrays of objects
                    elif prop_schema.get('type') == 'array' and 'items' in prop_schema:
                        item_schema = prop_schema['items']
                        array_levels = levels + ['[]']
                        row_array = {
                            'levels': array_levels,
                            'Request Parameter': '',
                            'GDPR': '',
                            'Cardinality': '',
                            'Type': item_schema.get('type', ''),
                            'Details': '',
                            'Description': item_schema.get('description', ''),
                            'Category': 'element'
                        }
                        rows.append(row_array)
                        if item_schema.get('type') == 'object':
                            rows.extend(walk_json_schema(item_schema, array_levels, max_depth))
            return rows
        schema = self.json_to_excel_service.load_json_schema(path)
        return walk_json_schema(schema)
    def generate_schema_mapping(self):
        import difflib
        src_type = self.schema_source_type.get()
        tgt_type = self.schema_target_type.get()
        src_path = self.schema_source_path.get()
        tgt_path = self.schema_target_path.get()
        out_path = self.mapping_output_path.get()
        src_case = self.source_case.get()
        tgt_case = self.target_case.get()
        try:
            logging.info(f'Generating mapping: src={src_path} ({src_type}), tgt={tgt_path} ({tgt_type}), out={out_path}')
            # Step 1: Get source rows (multi-message support, preserve order)
            parser = XSDParser()
            src_messages = {}
            if src_type == 'xsd':
                tree = ET.parse(src_path)
                root = tree.getroot()
                simple_types = {}
                for st in root.findall(f'.//{XSD_NS}simpleType'):
                    name = st.get('name')
                    if not name:
                        continue
                    restriction = st.find(XSD_NS+'restriction')
                    base = restriction.get('base') if restriction is not None else None
                    restrictions = {}
                    if restriction is not None:
                        for cons in restriction:
                            cons_name = cons.tag.replace(XSD_NS, '')
                            val = cons.get('value')
                            if val:
                                restrictions[cons_name] = val
                    simple_types[name] = {'base': base, 'restrictions': restrictions}
                complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{XSD_NS}complexType') if ct.get('name')}
                for elem in root.findall(f'{XSD_NS}element'):
                    name = elem.get('name')
                    if name:
                        rows = parser.parse_element(elem, complex_types, simple_types, 1, category='message')
                        for row in rows:
                            row['levels'] = self.convert_case(row['levels'], src_case)
                        src_messages[name] = rows  # rows are in XSD order
            else:
                src_full_rows = self.get_xsd_rows(src_path) if src_type == 'xsd' else self.get_json_rows(src_path)
                src_messages['Sheet1'] = src_full_rows
            # Step 2: Get target rows (single sheet for now)
            tgt_full_rows = self.get_xsd_rows(tgt_path) if tgt_type == 'xsd' else self.get_json_rows(tgt_path)
            for row in tgt_full_rows:
                row['levels'] = self.convert_case(row['levels'], tgt_case)
            max_tgt_level = max((len(row['levels']) for row in tgt_full_rows), default=1)
            tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
            # Step 3: Write to Excel, one sheet per message, preserving order
            wb = openpyxl.Workbook()
            first = True
            for msg_name, src_full_rows in src_messages.items():
                if not first:
                    ws = wb.create_sheet(title=msg_name[:31])
                else:
                    ws = wb.active
                    ws.title = msg_name[:31]
                    first = False
                max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
                src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
                headers = src_cols + ['Destination Fields'] + tgt_cols
                def row_path(row):
                    return '.'.join(row['levels'])
                tgt_path_dict = {row_path(row): row for row in tgt_full_rows}
                tgt_paths = list(tgt_path_dict.keys())
                ws.append(headers)
                ws.append([''] * len(headers))  # Second header row blank for now
                for src_row in src_full_rows:  # order preserved
                    src_levels = src_row['levels'] + [''] * (max_src_level - len(src_row['levels']))
                    src_vals = src_levels + [
                        src_row.get('Request Parameter',''),
                        src_row.get('GDPR',''),
                        src_row.get('Cardinality',''),
                        src_row.get('Type',''),
                        src_row.get('Base Type',''),
                        src_row.get('Details',''),
                        src_row.get('Description',''),
                        src_row.get('Category','element'),
                        src_row.get('Example','')
                    ]
                    src_path_str = '.'.join(src_row['levels'])
                    tgt_row = tgt_path_dict.get(src_path_str)
                    best_match = ''
                    if not tgt_row and tgt_paths:
                        import difflib
                        matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.0)
                        if matches:
                            best_match = matches[0]
                            tgt_row = tgt_path_dict[best_match]
                    tgt_levels = tgt_row['levels'] + [''] * (max_tgt_level - len(tgt_row['levels'])) if tgt_row else ['']*max_tgt_level
                    tgt_vals = tgt_levels + [
                        tgt_row.get('Request Parameter','') if tgt_row else '',
                        tgt_row.get('GDPR','') if tgt_row else '',
                        tgt_row.get('Cardinality','') if tgt_row else '',
                        tgt_row.get('Type','') if tgt_row else '',
                        tgt_row.get('Base Type','') if tgt_row else '',
                        tgt_row.get('Details','') if tgt_row else '',
                        tgt_row.get('Description','') if tgt_row else '',
                        tgt_row.get('Category','element') if tgt_row else '',
                        tgt_row.get('Example','') if tgt_row else ''
                    ]
                    dest_field = '.'.join([lvl for lvl in tgt_row['levels'] if lvl]) if tgt_row else ''
                    ws.append(src_vals + [dest_field] + tgt_vals)
                from openpyxl.utils import get_column_letter
                def last_nonempty_level_col(start_col, num_levels):
                    for col in range(start_col + num_levels - 1, start_col - 1, -1):
                        for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                            if any(cell.value not in (None, '') for cell in row):
                                return col
                    return start_col - 1
                src_level_start = 1
                last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
                for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
                    ws.delete_cols(col)
                header_row = [cell.value for cell in ws[1]]
                try:
                    tgt_level_start = header_row.index('Level1_tgt') + 1
                except ValueError:
                    tgt_level_start = len(header_row) + 1
                for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
                    col_letter = get_column_letter(col)
                    if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                        ws.delete_cols(col)
            wb.save(out_path)
            logging.info(f'Mapping file generated: {out_path}, Sheets: {list(src_messages.keys())}')
            self.schema_map_info.delete('1.0', tk.END)
            self.schema_map_info.insert(tk.END, f'Mapping file generated: {out_path}\nSheets: {list(src_messages.keys())}')
        except Exception as e:
            logging.error(f'Failed to generate mapping: {e}', exc_info=True)
            messagebox.showerror('Error', f'Failed to generate mapping: {e}')

if __name__ == '__main__':
    root = tk.Tk()
    app = SchemaToExcelGUI(root)
    root.mainloop()
