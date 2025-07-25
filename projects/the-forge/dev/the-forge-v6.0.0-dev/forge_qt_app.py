import sys
import os
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QFileDialog, QTextEdit, QFrame, QToolButton, QCheckBox, QMessageBox, QStackedWidget, QListWidget, QListWidgetItem,
    QSplitter, QStatusBar, QSizePolicy
)
from PySide6.QtGui import QIcon, QFont, QFontDatabase
from PySide6.QtCore import Qt, QSize, Signal
from microservices.wsdl_to_xsd_extractor import merge_xsd_from_wsdl
import traceback
from microservices.excel_export_service import ExcelExporter
from microservices.xsd_parser_service import XSDParser
import difflib
from openpyxl.utils import get_column_letter
from microservices.excel_output_validator import validate_excel_output, log_to_ui as validator_log_to_ui
import threading

ASSETS_DIR = os.path.join(os.path.dirname(__file__), 'assets')
ICONS_DIR = os.path.join(ASSETS_DIR, 'Icons_SVG', '_M', 'Actions')
FONT_DIR = os.path.join(ASSETS_DIR, 'Mulish', 'static')
LOGO_PATH = os.path.join(ASSETS_DIR, 'anvil.ico')

EDP_COLORS = {
    'marine_blue': '#212E3E',
    'marine_blue_light': '#2A3647',
    'electric_green': '#28FF52',
    'cobalt_blue': '#263CC8',
    'slate_grey': '#BECACC',
    'white': '#FFFFFF',
    'black': '#212E3E',
    'grey': '#A3B5B8',
    'ice_blue': '#0CD3F8',
}

SECTIONS = [
    ("Mapping", 'actions_edit_M.svg'),
    ("WSDL to XSD", 'actions_file_download_M.svg'),
    ("About", 'actions_more_info_M.svg'),
]

def get_icon(icon_name):
    path = os.path.join(ICONS_DIR, icon_name)
    if not os.path.exists(path):
        path = os.path.join(ICONS_DIR, 'actions_edit_M.svg')
    return QIcon(path)

class ForgeMainWindow(QMainWindow):
    log_signal = Signal(str, str)  # message, level
    def __init__(self):
        super().__init__()
        self.setWindowTitle("The Forge - EDP Modern UI")
        self.setMinimumSize(900, 600)
        self.setWindowIcon(QIcon(LOGO_PATH) if os.path.exists(LOGO_PATH) else QIcon())
        self._load_fonts()
        self._init_ui()
        self.source_path_full = ''
        self.target_path_full = ''
        self.output_path_full = ''
        self.log_signal.connect(self._log)

    def _load_fonts(self):
        for font_file in os.listdir(FONT_DIR):
            if font_file.endswith('.ttf'):
                QFontDatabase.addApplicationFont(os.path.join(FONT_DIR, font_file))
        self.setFont(QFont('Mulish', 11))

    def _init_ui(self):
        splitter = QSplitter()
        splitter.setOrientation(Qt.Horizontal)
        self.setCentralWidget(splitter)

        # Sidebar
        sidebar = QWidget()
        sidebar.setObjectName("sidebarPanel")
        sidebar.setFixedWidth(200)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)
        logo_row = QHBoxLayout()
        logo = QLabel()
        if os.path.exists(LOGO_PATH):
            logo.setPixmap(QIcon(LOGO_PATH).pixmap(36, 36))
        logo_row.addWidget(logo)
        app_name = QLabel("The Forge")
        app_name.setFont(QFont('Mulish', 16, QFont.Bold))
        app_name.setStyleSheet(f"color: {EDP_COLORS['electric_green']}; margin-left: 8px;")
        logo_row.addWidget(app_name)
        logo_row.addStretch()
        logo_widget = QWidget()
        logo_widget.setLayout(logo_row)
        logo_widget.setFixedHeight(56)
        sidebar_layout.addWidget(logo_widget)
        self.nav_list = QListWidget()
        self.nav_list.setObjectName("sidebarNav")
        self.nav_list.setIconSize(QSize(24, 24))
        for section, icon in SECTIONS:
            self.nav_list.addItem(QListWidgetItem(get_icon(icon), section))
        self.nav_list.setCurrentRow(0)
        self.nav_list.currentRowChanged.connect(self._on_nav_changed)
        self.nav_list.setStyleSheet(f"font-size: 15px; border: none;")
        sidebar_layout.addWidget(self.nav_list)
        sidebar_layout.addStretch()
        splitter.addWidget(sidebar)

        # Main content area (stacked)
        self.stack = QStackedWidget()
        self.stack.setObjectName("mainPanel")
        splitter.addWidget(self.stack)
        splitter.setStretchFactor(1, 1)

        # Top bar
        topbar = QFrame()
        topbar.setFrameShape(QFrame.StyledPanel)
        topbar.setStyleSheet(f"background: {EDP_COLORS['marine_blue_light']};")
        topbar_layout = QHBoxLayout(topbar)
        topbar_layout.setContentsMargins(24, 8, 24, 8)
        self.breadcrumb = QLabel(SECTIONS[0][0])
        self.breadcrumb.setFont(QFont('Mulish', 12, QFont.Bold))
        self.breadcrumb.setStyleSheet(f"color: {EDP_COLORS['electric_green']};")
        topbar_layout.addWidget(self.breadcrumb)
        topbar_layout.addStretch()
        self.theme_toggle = QCheckBox("Dark Mode")
        self.theme_toggle.setChecked(True)
        self.theme_toggle.stateChanged.connect(self._toggle_theme)
        self.theme_toggle.setStyleSheet(f"font-size: 14px; color: {EDP_COLORS['white']}; margin-right: 12px;")
        topbar_layout.addWidget(self.theme_toggle)
        help_btn = QToolButton()
        help_btn.setIcon(get_icon('actions_more_info_M.svg'))
        help_btn.setToolTip("About")
        help_btn.clicked.connect(lambda: self._on_nav_changed(2))
        topbar_layout.addWidget(help_btn)
        topbar.setMaximumHeight(48)
        self.addToolBar(Qt.TopToolBarArea, self._wrap_toolbar(topbar))

        # Section widgets (order must match SECTIONS)
        self.mapping_widget = self._build_mapping_section()
        self.wsdl2xsd_widget = self._build_wsdl2xsd_section()
        self.about_widget = self._build_about_section()
        self.stack.addWidget(self.mapping_widget)
        self.stack.addWidget(self.wsdl2xsd_widget)
        self.stack.addWidget(self.about_widget)

        # Status bar
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._apply_theme()

    def _toggle_theme(self):
        self._apply_theme()

    def _on_nav_changed(self, idx):
        self.stack.setCurrentIndex(idx)
        self.breadcrumb.setText(SECTIONS[idx][0])
        self.status.showMessage(f"{SECTIONS[idx][0]} selected", 2000)

    def _apply_theme(self):
        if self.theme_toggle.isChecked():
            self.setStyleSheet(f"""
                QMainWindow {{ background: {EDP_COLORS['marine_blue']}; }}
                QLabel, QLineEdit, QTextEdit {{ color: {EDP_COLORS['white']}; font-family: 'Mulish'; font-size: 15px; }}
                #sidebarPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
                #mainPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
                QListWidget#sidebarNav {{ background: transparent; font-size: 15px; border: none; }}
                QListWidget#sidebarNav::item:selected {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-left: 4px solid {EDP_COLORS['cobalt_blue']}; }}
                QPushButton {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-radius: 8px; font-family: 'Mulish'; font-weight: bold; font-size: 15px; min-height: 40px; padding: 0 20px; border: none; }}
                QPushButton:disabled {{ background: {EDP_COLORS['grey']}; color: {EDP_COLORS['marine_blue_light']}; }}
                QLineEdit, QTextEdit {{ background: {EDP_COLORS['marine_blue']}; border: 1.5px solid {EDP_COLORS['marine_blue_light']}; border-radius: 8px; font-family: 'Mulish'; font-size: 14px; padding: 8px; }}
                QToolButton {{ background: transparent; border: none; }}
                QStatusBar {{ background: {EDP_COLORS['marine_blue_light']}; color: {EDP_COLORS['electric_green']}; font-size: 13px; border: none; }}
            """)
        else:
            self.setStyleSheet(f"""
                QMainWindow {{ background: #F0F4F8; }}
                QLabel {{ color: {EDP_COLORS['marine_blue']}; font-family: 'Mulish'; font-size: 15px; }}
                #sidebarPanel {{ background: #F7F9FA; border-radius: 16px; border: 1.5px solid #E0E4EA; }}
                #mainPanel {{ background: #FFFFFF; border-radius: 18px; border: 2px solid #E0E4EA; }}
                QListWidget#sidebarNav {{ background: transparent; font-size: 15px; border: none; }}
                QListWidget#sidebarNav::item:selected {{ background: #E6F9F0; color: {EDP_COLORS['marine_blue']}; border-left: 4px solid {EDP_COLORS['electric_green']}; }}
                QPushButton {{ background: #FFFFFF; color: {EDP_COLORS['cobalt_blue']}; border-radius: 8px; font-family: 'Mulish'; font-weight: 600; font-size: 15px; min-height: 40px; padding: 0 20px; border: 1.5px solid #E0E4EA; }}
                QPushButton:disabled {{ background: #F0F4F8; color: #B0B8C1; border: 1.5px solid #E0E4EA; }}
                QPushButton:focus {{ border: 2px solid {EDP_COLORS['electric_green']}; }}
                QPushButton#primary {{ background: {EDP_COLORS['cobalt_blue']}; color: #FFFFFF; border: none; }}
                QLineEdit, QTextEdit {{ background: #FFFFFF; border: 1.5px solid #E0E4EA; border-radius: 8px; font-family: 'Mulish'; font-size: 14px; padding: 8px; }}
                QToolButton {{ background: transparent; border: none; }}
                QStatusBar {{ background: #F7F9FA; color: {EDP_COLORS['cobalt_blue']}; font-size: 13px; border-top: 1px solid #E0E4EA; }}
            """)

    def _build_mapping_section(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(48, 48, 48, 48)
        layout.setSpacing(20)
        # File pickers row
        file_row = QHBoxLayout()
        file_row.setSpacing(16)
        self.source_path = QLineEdit()
        self.source_path.setPlaceholderText("Select source schema (.xsd)")
        self.source_path.setReadOnly(True)
        file_row.addWidget(self._file_picker(self.source_path, self._pick_source, "Source", get_icon('actions_file_upload_M.svg')))
        self.target_path = QLineEdit()
        self.target_path.setPlaceholderText("Select target schema (.xsd)")
        self.target_path.setReadOnly(True)
        file_row.addWidget(self._file_picker(self.target_path, self._pick_target, "Target", get_icon('actions_file_upload_M.svg')))
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("Select output mapping file (.xlsx)")
        self.output_path.setReadOnly(True)
        file_row.addWidget(self._file_picker(self.output_path, self._pick_output, "Output", get_icon('actions_file_download_M.svg'), save=True))
        layout.addLayout(file_row)
        # Generate button
        self.generate_btn = QPushButton("Generate Mapping")
        self.generate_btn.setFont(QFont('Mulish', 13, QFont.Bold))
        self.generate_btn.clicked.connect(self._on_generate)
        layout.addWidget(self.generate_btn)
        # Status label
        self.status_label = QLabel("")
        self.status_label.setFont(QFont('Mulish', 11, QFont.Medium))
        layout.addWidget(self.status_label)
        # Log output
        self.log_text = QTextEdit()
        self.log_text.setFont(QFont('Mulish', 10))
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(100)
        layout.addWidget(self.log_text)
        layout.addStretch()
        return w

    def _file_picker(self, line_edit, pick_cmd, label, icon, save=False):
        w = QWidget()
        l = QHBoxLayout(w)
        l.setContentsMargins(0, 0, 0, 0)
        icon_label = QLabel()
        icon_label.setPixmap(icon.pixmap(20, 20))
        l.addWidget(icon_label)
        lbl = QLabel(label)
        lbl.setFont(QFont('Mulish', 11, QFont.Bold))
        l.addWidget(lbl)
        l.addWidget(line_edit)
        browse_btn = QPushButton("Browse")
        browse_btn.setFont(QFont('Mulish', 10))
        browse_btn.clicked.connect(pick_cmd)
        l.addWidget(browse_btn)
        return w

    def _pick_source(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Source Schema", "", "XSD Files (*.xsd)")
        if path:
            self.source_path_full = path
            self.source_path.setText(self._truncate_path(path))

    def _pick_target(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Target Schema", "", "XSD Files (*.xsd)")
        if path:
            self.target_path_full = path
            self.target_path.setText(self._truncate_path(path))

    def _pick_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Select Output Mapping File", "", "Excel Files (*.xlsx)")
        if path:
            self.output_path_full = path
            self.output_path.setText(self._truncate_path(path))

    def _on_generate(self):
        self._log("[INFO] Generate Mapping clicked.", level="info")
        self.status_label.setText("Generating mapping...")
        self.status_label.setStyleSheet(f"color: {EDP_COLORS['electric_green']};")
        self.generate_btn.setEnabled(False)
        QApplication.instance().processEvents()
        try:
            src = self.source_path_full
            tgt = self.target_path_full
            out = self.output_path_full
            self._log(f"[INFO] Source: {src}", level="info")
            self._log(f"[INFO] Target: {tgt}", level="info")
            self._log(f"[INFO] Output: {out}", level="info")

            # Ensure output directory exists
            output_dir = os.path.dirname(out)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            # --- v4 logic: parse source and target, multi-message, row matching, column structure ---
            from microservices.xsd_parser_service import XSDParser
            import xml.etree.ElementTree as ET
            import openpyxl
            XSD_NS = '{http://www.w3.org/2001/XMLSchema}'
            parser = XSDParser()
            # Parse source XSD
            tree = ET.parse(src)
            root = tree.getroot()
            simple_types = {}
            for st in root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}simpleType'):
                name = st.get('name')
                if not name:
                    continue
                restriction = st.find('{http://www.w3.org/2001/XMLSchema}restriction')
                base = restriction.get('base') if restriction is not None else None
                restrictions = {}
                if restriction is not None:
                    for cons in restriction:
                        cons_name = cons.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                        val = cons.get('value')
                        if val:
                            restrictions[cons_name] = val
                simple_types[name] = {'base': base, 'restrictions': restrictions}
            complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}complexType') if ct.get('name')}
            src_messages = {}
            for elem in root.findall('{http://www.w3.org/2001/XMLSchema}element'):
                name = elem.get('name')
                if name:
                    rows = parser.parse_element(elem, complex_types, simple_types, 1, category='message')
                    src_messages[name] = rows
            # Parse target XSD (single sheet)
            tgt_rows = []
            if tgt and os.path.exists(tgt):
                tgt_tree = ET.parse(tgt)
                tgt_root = tgt_tree.getroot()
                tgt_simple_types = {}
                for st in tgt_root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}simpleType'):
                    name = st.get('name')
                    if not name:
                        continue
                    restriction = st.find('{http://www.w3.org/2001/XMLSchema}restriction')
                    base = restriction.get('base') if restriction is not None else None
                    restrictions = {}
                    if restriction is not None:
                        for cons in restriction:
                            cons_name = cons.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                            val = cons.get('value')
                            if val:
                                restrictions[cons_name] = val
                    tgt_simple_types[name] = {'base': base, 'restrictions': restrictions}
                tgt_complex_types = {ct.get('name'): ct for ct in tgt_root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}complexType') if ct.get('name')}
                for elem in tgt_root.findall('{http://www.w3.org/2001/XMLSchema}element'):
                    rows = parser.parse_element(elem, tgt_complex_types, tgt_simple_types, 1, category='message')
                    tgt_rows.extend(rows)
            # Build Excel file
            wb = openpyxl.Workbook()
            first = True
            for msg_name, src_full_rows in src_messages.items():
                if not first:
                    ws = wb.create_sheet(title=msg_name[:31])
                else:
                    ws = wb.active
                    ws.title = msg_name[:31]
                    first = False
                max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
                max_tgt_level = max((len(row['levels']) for row in tgt_rows), default=1) if tgt_rows else 1
                src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
                tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
                headers = src_cols + ['Destination Fields'] + tgt_cols
                ws.append(headers)
                ws.append([''] * len(headers))  # Second header row blank for now
                def row_path(row):
                    return '.'.join(row['levels'])
                tgt_path_dict = {row_path(row): row for row in tgt_rows}
                tgt_paths = list(tgt_path_dict.keys())
                for src_row in src_full_rows:
                    src_levels = src_row['levels'] + [''] * (max_src_level - len(src_row['levels']))
                    src_vals = src_levels + [
                        src_row.get('Request Parameter',''),
                        src_row.get('GDPR',''),
                        src_row.get('Cardinality',''),
                        src_row.get('Type',''),
                        src_row.get('Base Type',''),
                        src_row.get('Details',''),
                        src_row.get('Description',''),
                        src_row.get('Category','element'),
                        src_row.get('Example','')
                    ]
                    src_path_str = row_path(src_row)
                    tgt_row = tgt_path_dict.get(src_path_str)
                    best_match = ''
                    if not tgt_row and tgt_paths:
                        matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.0)
                        if matches:
                            best_match = matches[0]
                            tgt_row = tgt_path_dict[best_match]
                    tgt_levels = tgt_row['levels'] + [''] * (max_tgt_level - len(tgt_row['levels'])) if tgt_row else ['']*max_tgt_level
                    tgt_vals = tgt_levels + [
                        tgt_row.get('Request Parameter','') if tgt_row else '',
                        tgt_row.get('GDPR','') if tgt_row else '',
                        tgt_row.get('Cardinality','') if tgt_row else '',
                        tgt_row.get('Type','') if tgt_row else '',
                        tgt_row.get('Base Type','') if tgt_row else '',
                        tgt_row.get('Details','') if tgt_row else '',
                        tgt_row.get('Description','') if tgt_row else '',
                        tgt_row.get('Category','element') if tgt_row else '',
                        tgt_row.get('Example','') if tgt_row else ''
                    ]
                    dest_field = '.'.join([lvl for lvl in tgt_row['levels'] if lvl]) if tgt_row else ''
                    ws.append(src_vals + [dest_field] + tgt_vals)
                # Prune unused source level columns
                def last_nonempty_level_col(start_col, num_levels):
                    for col in range(start_col + num_levels - 1, start_col - 1, -1):
                        for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                            if any(cell.value not in (None, '') for cell in row):
                                return col
                    return start_col - 1
                src_level_start = 1
                last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
                for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
                    ws.delete_cols(col)
                header_row = [cell.value for cell in ws[1]]
                try:
                    tgt_level_start = header_row.index('Level1_tgt') + 1
                except ValueError:
                    tgt_level_start = len(header_row) + 1
                for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
                    col_letter = get_column_letter(col)
                    if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                        ws.delete_cols(col)
            wb.save(out)
            self._log(f"[SUCCESS] Output file created: {out}", level="info")
            # --- Post-processing QA: Excel Output Validator ---
            def run_validator():
                # Patch validator's log_to_ui to use the UI log via signal
                def ui_logger(msg):
                    if msg.startswith("✅") or msg.startswith("🔎"):
                        level = "info"
                    elif msg.startswith("❌"):
                        level = "error"
                    else:
                        level = "warning"
                    self.log_signal.emit(msg, level)
                validator_log_to_ui.__globals__['log_to_ui'] = ui_logger
                # Use source XSD for both source and target if only one is provided
                xsd_path = src if src else tgt
                try:
                    validate_excel_output(xsd_path, out)
                except Exception as e:
                    self.log_signal.emit(f"❌ Error in post-processing validator: {e}", "error")
            threading.Thread(target=run_validator, daemon=True).start()
            self._finish_generate()
        except Exception as e:
            import traceback
            error_msg = f"Mapping Error: {e}\n{traceback.format_exc()}"
            self._log(error_msg, level="error")
            self.status_label.setText("Mapping failed!")
            self.status_label.setStyleSheet("color: #FF3333;")
        self.generate_btn.setEnabled(True)

    def _finish_generate(self):
        self.status_label.setText("Mapping generated successfully!")
        self.status_label.setStyleSheet(f"color: {EDP_COLORS['electric_green']};")
        self._log("[INFO] Mapping generated.", level="info")

    def _log(self, msg, level="info"):
        color = {"info": EDP_COLORS['marine_blue_light'], "warning": EDP_COLORS['ice_blue'], "error": "#FF3333"}.get(level, EDP_COLORS['marine_blue_light'])
        self.log_text.setTextColor(Qt.black if level == "info" else Qt.red)
        self.log_text.append(msg)
        self.log_text.setTextColor(Qt.black)
        self.log_text.ensureCursorVisible()

    def _truncate_path(self, path, maxlen=60):
        if len(path) <= maxlen:
            return path
        return "..." + path[-(maxlen-3):]

    def _build_about_section(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(48, 48, 48, 48)
        layout.setSpacing(20)
        title = QLabel("About The Forge")
        title.setFont(QFont('Mulish', 16, QFont.Bold))
        layout.addWidget(title)
        subtitle = QLabel("EDP Schema Mapping Tool\nVersion 5.0.0\nBuilt with PySide6 and EDP Design System.")
        subtitle.setFont(QFont('Mulish', 11))
        layout.addWidget(subtitle)
        layout.addStretch()
        return w

    def _build_wsdl2xsd_section(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(48, 48, 48, 48)
        layout.setSpacing(20)
        title = QLabel("WSDL to XSD")
        title.setFont(QFont('Mulish', 16, QFont.Bold))
        layout.addWidget(title)
        subtitle = QLabel("Convert WSDL to XSD for EDP Schema Mapping.")
        subtitle.setFont(QFont('Mulish', 11))
        layout.addWidget(subtitle)
        # WSDL input area
        self.wsdl_input = QTextEdit()
        self.wsdl_input.setPlaceholderText("Paste WSDL content here...")
        self.wsdl_input.setMinimumHeight(120)
        layout.addWidget(self.wsdl_input)
        # File upload
        file_row = QHBoxLayout()
        self.wsdl_file_path = QLineEdit()
        self.wsdl_file_path.setPlaceholderText("Or select a .wsdl file...")
        self.wsdl_file_path.setReadOnly(True)
        file_row.addWidget(self.wsdl_file_path)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self._pick_wsdl_file)
        file_row.addWidget(browse_btn)
        layout.addLayout(file_row)
        # Extract button
        extract_btn = QPushButton("Extract XSD")
        extract_btn.setObjectName("primary")
        extract_btn.clicked.connect(self._on_extract_xsd)
        layout.addWidget(extract_btn)
        # Output area
        self.xsd_output = QTextEdit()
        self.xsd_output.setPlaceholderText("Merged XSD will appear here...")
        self.xsd_output.setMinimumHeight(180)
        layout.addWidget(self.xsd_output)
        # Log output area for this tab
        self.wsdl2xsd_log_text = QTextEdit()
        self.wsdl2xsd_log_text.setReadOnly(True)
        self.wsdl2xsd_log_text.setMinimumHeight(80)
        self.wsdl2xsd_log_text.setPlaceholderText("Log output for WSDL to XSD extraction...")
        layout.addWidget(self.wsdl2xsd_log_text)
        # Download button
        download_btn = QPushButton("Download XSD")
        download_btn.clicked.connect(self._on_download_xsd)
        layout.addWidget(download_btn)
        layout.addStretch()
        return w

    def _pick_wsdl_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select WSDL File", "", "WSDL Files (*.wsdl)")
        if path:
            self.wsdl_file_path.setText(path)
            with open(path, 'r', encoding='utf-8') as f:
                self.wsdl_input.setPlainText(f.read())

    def _on_extract_xsd(self):
        wsdl_content = self.wsdl_input.toPlainText()
        if not wsdl_content.strip():
            QMessageBox.warning(self, "No WSDL", "Please paste WSDL content or select a file.")
            return
        try:
            merged_xsd = merge_xsd_from_wsdl(wsdl_content)
            self.xsd_output.setPlainText(merged_xsd)
        except Exception as e:
            import traceback
            error_msg = f"Extraction Error: {e}\n{traceback.format_exc()}"
            self.wsdl2xsd_log_text.append(error_msg)
            QMessageBox.critical(self, "Extraction Error", str(e))

    def _on_download_xsd(self):
        xsd_content = self.xsd_output.toPlainText()
        if not xsd_content.strip():
            QMessageBox.warning(self, "No XSD", "No XSD to download. Please extract first.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save XSD File", "merged.xsd", "XSD Files (*.xsd)")
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(xsd_content)

    def _wrap_toolbar(self, widget):
        from PySide6.QtWidgets import QToolBar
        tb = QToolBar()
        tb.setMovable(False)
        tb.setFloatable(False)
        tb.addWidget(widget)
        return tb

def main():
    app = QApplication(sys.argv)
    window = ForgeMainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main() 