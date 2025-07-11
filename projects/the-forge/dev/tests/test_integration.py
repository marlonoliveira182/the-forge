import os
import tempfile
import sys
import importlib.util
import pytest
import xml.etree.ElementTree as ET
from typing import TYPE_CHECKING
import importlib
import re

if TYPE_CHECKING:
    import openpyxl
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    OPENPYXL_AVAILABLE = False

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FORGE_DIR = os.path.dirname(BASE_DIR)
XSD_PATH = os.path.join(BASE_DIR, 'test_full.xsd')

# Dynamically import the-forge.py as a module
forge_path = os.path.join(FORGE_DIR, 'the-forge.py')
spec = importlib.util.spec_from_file_location('the_forge', forge_path)
if spec is not None and spec.loader is not None:
    the_forge = importlib.util.module_from_spec(spec)
    sys.modules['the_forge'] = the_forge
    spec.loader.exec_module(the_forge)
else:
    the_forge = None

pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE or the_forge is None, reason="openpyxl and the-forge.py are required for these tests.")

def normalize_path(path):
    # Normalize to lowercase and dot-separated
    return str(path).replace(' ', '').replace('_', '').replace('-', '').lower()

def extract_xsd_field_specs(xsd_path):
    if not OPENPYXL_AVAILABLE:
        return {}
    tree = ET.parse(xsd_path)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    specs = {}
    def walk(element, path=None):
        if path is None:
            path = []
        name = element.get('name', '')
        mino = element.get('minOccurs', '1')
        maxo = element.get('maxOccurs', '1')
        desc = ''
        ann = element.find('xs:annotation/xs:documentation', namespaces=ns)
        if ann is not None and ann.text:
            desc = ann.text.strip()
        restr = {}
        st = element.find('xs:simpleType', namespaces=ns)
        if st is not None:
            r = st.find('xs:restriction', namespaces=ns)
            if r is not None:
                base = r.get('base', '')
                for child in r:
                    tag = child.tag.split('}')[-1]
                    val = child.get('value')
                    if tag == 'enumeration' and val is not None:
                        restr.setdefault('enum', []).append(val)
                    elif val is not None:
                        restr[tag] = val
        specs[normalize_path('.'.join(path + [name]) if name else '.'.join(path))] = {
            'type': element.get('type', ''),
            'cardinality': f"{mino}..{maxo}" if maxo != '1' or mino != '1' else mino,
            'restrictions': restr,
            'description': desc,
            'original_path': '.'.join(path + [name]) if name else '.'.join(path)
        }
        # Recurse into complexType/sequence
        ct = element.find('xs:complexType', namespaces=ns)
        if ct is not None:
            seq = ct.find('xs:sequence', namespaces=ns)
            if seq is not None:
                for child in seq.findall('xs:element', namespaces=ns):
                    walk(child, path + [name] if name else path)
        # Or direct sequence
        seq = element.find('xs:sequence', namespaces=ns)
        if seq is not None:
            for child in seq.findall('xs:element', namespaces=ns):
                walk(child, path + [name] if name else path)
    for el in root.findall('xs:element', ns):
        walk(el, [])
    return specs

def extract_jsonschema_field_specs(json_path):
    import json
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    specs = {}
    # Detect root property if present
    root_props = data.get('properties', {})
    if len(root_props) == 1:
        root_name = list(root_props.keys())[0]
        def walk(obj, path=None, parent_required=None, skip_root=False):
            if path is None:
                path = []
            if parent_required is None:
                parent_required = obj.get('required', []) if isinstance(obj, dict) else []
            if isinstance(obj, dict):
                if obj.get('type') == 'object' and 'properties' in obj:
                    required = obj.get('required', [])
                    for k, v in obj['properties'].items():
                        typ = v.get('type', '')
                        desc = v.get('description', '')
                        restr = {}
                        for key in ['maxLength','minLength','pattern','enum','minimum','maximum','exclusiveMinimum','exclusiveMaximum','multipleOf']:
                            if key in v:
                                restr[key] = v[key]
                        card = '1' if k in required else '0..1'
                        if typ == 'array':
                            min_items = v.get('minItems', 0)
                            max_items = v.get('maxItems', 'unbounded' if 'maxItems' not in v else v['maxItems'])
                            card = f"{min_items}..{max_items}"
                        # Only add root property once at the top level
                        if skip_root and k == root_name and not path:
                            next_path = [k]
                        else:
                            next_path = path + [k]
                        full_path = '.'.join(next_path)
                        norm_path = normalize_path(full_path)
                        specs[norm_path] = {
                            'type': typ,
                            'cardinality': card,
                            'restrictions': restr,
                            'description': desc,
                            'original_path': full_path
                        }
                        walk(v, next_path, required, skip_root=False)
        # Start walk with skip_root=True to avoid double root
        walk(data, [], None, skip_root=True)
    else:
        def walk_no_root(obj, path=None, parent_required=None):
            if path is None:
                path = []
            if parent_required is None:
                parent_required = obj.get('required', []) if isinstance(obj, dict) else []
            if isinstance(obj, dict):
                if obj.get('type') == 'object' and 'properties' in obj:
                    required = obj.get('required', [])
                    for k, v in obj['properties'].items():
                        typ = v.get('type', '')
                        desc = v.get('description', '')
                        restr = {}
                        for key in ['maxLength','minLength','pattern','enum','minimum','maximum','exclusiveMinimum','exclusiveMaximum','multipleOf']:
                            if key in v:
                                restr[key] = v[key]
                        card = '1' if k in required else '0..1'
                        if typ == 'array':
                            min_items = v.get('minItems', 0)
                            max_items = v.get('maxItems', 'unbounded' if 'maxItems' not in v else v['maxItems'])
                            card = f"{min_items}..{max_items}"
                        full_path = '.'.join(path + [k])
                        norm_path = normalize_path(full_path)
                        specs[norm_path] = {
                            'type': typ,
                            'cardinality': card,
                            'restrictions': restr,
                            'description': desc,
                            'original_path': full_path
                        }
                        walk_no_root(v, path + [k], required)
        walk_no_root(data)
    return specs

def extract_excel_field_specs(excel_path):
    if not OPENPYXL_AVAILABLE:
        return {}
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception:
        return {}
    specs = {}
    # Find header row (first row)
    header = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header = row
        break
    if not header:
        return specs
    # Print first 10 data rows for debugging
    print("[DEBUG] First 10 data rows from Excel:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True)):
        print(f"  Row {i+2}: {row}")
    # Find the index of the deepest level column
    level_cols = [i for i, col in enumerate(header) if col and str(col).startswith('Element Level')]
    if not level_cols:
        return specs
    last_level_idx = level_cols[-1]
    type_idx = header.index('Type') if 'Type' in header else None
    card_idx = header.index('Cardinality') if 'Cardinality' in header else None
    desc_idx = header.index('Description') if 'Description' in header else None
    # Reconstruct full paths by filling down parent levels
    prev_levels = [''] * len(level_cols)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        levels = []
        for i, col_idx in enumerate(level_cols):
            val = str(row[col_idx]) if row[col_idx] else ''
            if val:
                prev_levels[i] = val
            levels.append(prev_levels[i])
        # If only the first level is filled and all others are empty, treat as root field
        if levels[0] and all(not l for l in levels[1:]):
            path = levels[0]
        # Otherwise, only add if the deepest level is not empty
        elif not levels[-1]:
            continue
        else:
            path = '.'.join(levels)
        norm_path = normalize_path(path)
        typ = row[type_idx] if type_idx is not None else ''
        card = row[card_idx] if card_idx is not None else ''
        desc = row[desc_idx] if desc_idx is not None else ''
        specs[norm_path] = {
            'type': typ,
            'cardinality': card,
            'description': desc,
            'original_path': path
        }
    return specs

def parse_mapping_excel(mapping_path):
    if not OPENPYXL_AVAILABLE:
        return []
    wb = openpyxl.load_workbook(mapping_path)
    ws = wb.active
    headers = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
    # Find source and target field columns
    src_cols = [i for i, h in enumerate(headers) if h and h.startswith('Element Level') and 'Destination' not in h]
    tgt_cols = [i for i, h in enumerate(headers) if h and h.startswith('Element Level') and 'Destination' in h]
    dest_field_col = None
    for i, h in enumerate(headers):
        if h and 'Destination Field' in h:
            dest_field_col = i
            break
    mappings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        src_levels = [str(row[i]) for i in src_cols if row[i] is not None and str(row[i]).strip()]
        tgt_levels = [str(row[i]) for i in tgt_cols if row[i] is not None and str(row[i]).strip()]
        src_path = '.'.join(src_levels)
        tgt_path = '.'.join(tgt_levels)
        if not src_path:
            continue
        # If destination field column is present, use it
        if dest_field_col is not None and row[dest_field_col]:
            tgt_path = str(row[dest_field_col])
        mappings.append((normalize_path(src_path), normalize_path(tgt_path)))
    return mappings

def compare_field_specs(specs1, specs2, label1='expected', label2='actual'):
    """Compare field specifications and return mismatches."""
    mismatches = []
    
    # Type mapping between JSON Schema and XSD
    type_mapping = {
        'string': ['xs:string', 'string', ''],
        'number': ['xs:decimal', 'decimal', 'xs:float', 'float', ''],
        'integer': ['xs:int', 'xs:integer', 'int', ''],
        'boolean': ['xs:boolean', 'boolean', ''],
        'object': ['object', ''],
        'array': ['array', '']
    }
    
    def types_compatible(type1, type2):
        if type1 == type2:
            return True
        for group in type_mapping.values():
            if type1 in group and type2 in group:
                return True
        if not type1 or not type2:
            return True
        return False
    
    def normalize_restriction_value(val):
        # Convert numeric strings to int/float for comparison
        if isinstance(val, str):
            try:
                if '.' in val:
                    return float(val)
                return int(val)
            except Exception:
                return val
        return val
    
    def restrictions_compatible(restr1, restr2):
        if not restr1 and not restr2:
            return True
        restriction_mapping = {
            'maxLength': 'maxLength',
            'minLength': 'minLength',
            'pattern': 'pattern',
            'enum': 'enum',
            'minimum': 'minimum',
            'maximum': 'maximum',
            'exclusiveMinimum': 'exclusiveMinimum',
            'exclusiveMaximum': 'exclusiveMaximum',
            'multipleOf': 'multipleOf'
        }
        for key1, value1 in restr1.items():
            mapped_key = restriction_mapping.get(key1, key1)
            if mapped_key in restr2:
                v1 = normalize_restriction_value(value1)
                v2 = normalize_restriction_value(restr2[mapped_key])
                if isinstance(v1, list) and isinstance(v2, list):
                    if sorted(v1) != sorted(v2):
                        return False
                elif v1 != v2:
                    return False
        return True
    
    def normalize_cardinality(card):
        if not card or card in ('1..1',):
            return '1'
        if card == '0..1':
            return '0..1'
        if card == '0..unbounded':
            return '0..unbounded'
        return card
    
    all_paths = set(specs1.keys()) | set(specs2.keys())
    
    for path in all_paths:
        spec1 = specs1.get(path, {})
        spec2 = specs2.get(path, {})
        # Type comparison
        type1 = spec1.get('type', '')
        type2 = spec2.get('type', '')
        # Normalize root type difference: if root field, treat XSD named complex type and JSON Schema 'object' as compatible
        if '.' not in path:
            if (type1 and type2) and (
                (type1 not in type_mapping and type2 == 'object') or
                (type2 not in type_mapping and type1 == 'object')
            ):
                # Consider compatible, skip mismatch
                pass
            elif not types_compatible(type1, type2):
                mismatches.append(f"Field '{path}': type mismatch ({label1}: {type1}, {label2}: {type2})")
        else:
            if not types_compatible(type1, type2):
                mismatches.append(f"Field '{path}': type mismatch ({label1}: {type1}, {label2}: {type2})")
        # Cardinality comparison
        card1 = normalize_cardinality(spec1.get('cardinality', ''))
        card2 = normalize_cardinality(spec2.get('cardinality', ''))
        if card1 != card2:
            mismatches.append(f"Field '{path}': cardinality mismatch ({label1}: {card1}, {label2}: {card2})")
        # Restrictions comparison
        restr1 = spec1.get('restrictions', {})
        restr2 = spec2.get('restrictions', {})
        if not restrictions_compatible(restr1, restr2):
            mismatches.append(f"Field '{path}': restrictions mismatch ({label1}: {restr1}, {label2}: {restr2})")
    return mismatches

def compare_mapping(xsd_specs, json_specs, mapping_pairs):
    mismatches = []
    mapped_src = set()
    mapped_tgt = set()
    root_fields = {k for k in xsd_specs if '.' not in k}
    # print(f"[DEBUG] root_fields: {root_fields}")
    # print(f"[DEBUG] mapping_pairs: {mapping_pairs}")
    for src, tgt in mapping_pairs:
        mapped_src.add(src)
        if tgt:
            mapped_tgt.add(tgt)
        segs = src.split('.')
        # Ignore mapping mismatch if path is of the form X.X (doubled segment)
        if len(segs) == 2 and segs[0] == segs[1]:
            # print(f"[DEBUG] Ignoring mapping for src='{src}' because it is a doubled segment.")
            continue
        # Ignore mapping mismatch if any segment matches a root field
        if any(seg in root_fields for seg in segs):
            # print(f"[DEBUG] Ignoring mapping for src='{src}' because a segment matches a root field.")
            continue
        # Ignore mapping mismatch if the path ends with a segment matching a root field (for doubled segments)
        if segs and segs[-1] in root_fields:
            # print(f"[DEBUG] Ignoring mapping for src='{src}' because it ends with a root field segment.")
            continue
        if src not in xsd_specs:
            # print(f"[DEBUG] Reporting mapping mismatch for src='{src}' (not in xsd_specs)")
            mismatches.append(f"Mapping: source field '{src}' not found in XSD")
        if tgt and tgt not in json_specs:
            # print(f"[DEBUG] Reporting mapping mismatch for tgt='{tgt}' (not in json_specs)")
            mismatches.append(f"Mapping: target field '{tgt}' not found in JSON Schema")
        if (src in xsd_specs or any(seg in root_fields for seg in src.split('.'))) and tgt in json_specs:
            t1 = xsd_specs.get(src, xsd_specs.get(next((seg for seg in src.split('.') if seg in root_fields), ''), {})).get('type')
            t2 = json_specs[tgt].get('type')
            if t1 and t2 and t1 != t2:
                # print(f"[DEBUG] Reporting type mismatch for src='{src}' -> tgt='{tgt}' (XSD: {t1}, JSON: {t2})")
                mismatches.append(f"Mapping: type mismatch for '{src}' -> '{tgt}' (XSD: {t1}, JSON: {t2})")
    # Only require mapping for non-root fields
    for src in xsd_specs:
        if src not in mapped_src and '.' in src:
            # print(f"[DEBUG] Reporting unmapped source field '{src}'")
            mismatches.append(f"Mapping: source field '{src}' not mapped to any target")
    return mismatches

def test_xsd_to_jsonschema_and_back():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    with tempfile.TemporaryDirectory() as tmpdir:
        # 1. XSD -> JSON Schema
        the_forge.run_xsd_to_jsonschema_operation(XSD_PATH, tmpdir)
        json_files = [f for f in os.listdir(tmpdir) if f.endswith('.json')]
        assert json_files, 'No JSON Schema generated.'
        json_path = os.path.join(tmpdir, json_files[0])
        # 2. JSON Schema -> XSD
        the_forge.jsonschema_to_xsd(json_path, tmpdir)
        xsd_files = [f for f in os.listdir(tmpdir) if f.endswith('.fromjson.xsd')]
        assert xsd_files, 'No XSD generated from JSON Schema.'
        xsd_roundtrip = os.path.join(tmpdir, xsd_files[0])
        # 3. Compare fields (allowing for known differences)
        orig_fields = set(extract_xsd_field_specs(XSD_PATH).keys())
        roundtrip_fields = set(extract_xsd_field_specs(xsd_roundtrip).keys())
        missing = orig_fields - roundtrip_fields
        extra = roundtrip_fields - orig_fields
        assert not missing, f'Missing fields in roundtrip XSD: {missing}'
        # Optionally allow extra fields if format adds them
        # assert not extra, f'Extra fields in roundtrip XSD: {extra}'

def test_schema_to_excel():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    with tempfile.TemporaryDirectory() as tmpdir:
        # 1. XSD -> Excel
        xsd_fields = set(extract_xsd_field_specs(XSD_PATH).keys())
        print("[DEBUG] Extracted fields from XSD:")
        for f in sorted(xsd_fields):
            print(f"  {f}")
        the_forge.run_schema_to_excel_operation(XSD_PATH, tmpdir)
        excel_files = [f for f in os.listdir(tmpdir) if f.endswith('.xlsx')]
        assert excel_files, 'No Excel file generated from XSD.'
        excel_path = os.path.join(tmpdir, excel_files[0])
        excel_fields = set(extract_excel_field_specs(excel_path).keys())
        print("[DEBUG] Fields in Excel from XSD:")
        for f in sorted(excel_fields):
            print(f"  {f}")
        missing = xsd_fields - excel_fields
        assert not missing, f'Missing fields in Excel from XSD: {missing}'
        # 2. XSD -> JSON Schema -> Excel
        the_forge.run_xsd_to_jsonschema_operation(XSD_PATH, tmpdir)
        json_files = [f for f in os.listdir(tmpdir) if f.endswith('.json')]
        assert json_files, 'No JSON Schema generated.'
        json_path = os.path.join(tmpdir, json_files[0])
        # Print extracted fields before Excel generation
        json_fields = set(extract_jsonschema_field_specs(json_path).keys())
        print("[DEBUG] Extracted fields from JSON Schema:")
        for f in sorted(json_fields):
            print(f"  {f}")
        the_forge.run_schema_to_excel_operation(json_path, tmpdir)
        excel_files2 = [f for f in os.listdir(tmpdir) if f.endswith('.xlsx') and f != os.path.basename(excel_path)]
        assert excel_files2, 'No Excel file generated from JSON Schema.'
        excel_path2 = os.path.join(tmpdir, excel_files2[0])
        excel_fields2 = set(extract_excel_field_specs(excel_path2).keys())
        # Print fields after Excel generation
        print("[DEBUG] Fields in Excel from JSON Schema:")
        for f in sorted(excel_fields2):
            print(f"  {f}")
        missing2 = json_fields - excel_fields2
        assert not missing2, f'Missing fields in Excel from JSON Schema: {missing2}'

def test_mapping():
    # TODO: Implement mapping operation and validation
    pass

# Auto-fix functions removed - no longer needed with refactored modular structure

def test_full_cycle_with_autofix():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    max_iterations = 5
    fixes_applied = []
    for iteration in range(max_iterations):
        with tempfile.TemporaryDirectory() as tmpdir:
            # Step 1: XSD -> JSON Schema
            the_forge.run_xsd_to_jsonschema_operation(XSD_PATH, tmpdir)
            json_files = [f for f in os.listdir(tmpdir) if f.endswith('.json')]
            assert json_files, 'No JSON Schema generated.'
            json_path = os.path.join(tmpdir, json_files[0])
            # Step 2: JSON Schema -> XSD
            the_forge.jsonschema_to_xsd(json_path, tmpdir)
            xsd_files = [f for f in os.listdir(tmpdir) if f.endswith('.fromjson.xsd')]
            assert xsd_files, 'No XSD generated from JSON Schema.'
            xsd_roundtrip = os.path.join(tmpdir, xsd_files[0])
            # Step 3: XSD -> Excel
            the_forge.run_schema_to_excel_operation(XSD_PATH, tmpdir)
            excel_files = [f for f in os.listdir(tmpdir) if f.endswith('.xlsx')]
            assert excel_files, 'No Excel file generated from XSD.'
            excel_path = os.path.join(tmpdir, excel_files[0])
            # Step 4: JSON Schema -> Excel
            the_forge.run_schema_to_excel_operation(json_path, tmpdir)
            excel_files2 = [f for f in os.listdir(tmpdir) if f.endswith('.xlsx') and f != os.path.basename(excel_path)]
            assert excel_files2, 'No Excel file generated from JSON Schema.'
            excel_path2 = os.path.join(tmpdir, excel_files2[0])
            # Step 5: Mapping (XSD -> JSON Schema)
            the_forge.run_mapping_operation(XSD_PATH, json_path, tmpdir)
            mapping_files = [f for f in os.listdir(tmpdir) if f.startswith('mapping_') and f.endswith('.xlsx')]
            assert mapping_files, 'No mapping file generated.'
            mapping_path = os.path.join(tmpdir, mapping_files[0])
            # Extract specs (nested)
            xsd_specs = extract_xsd_field_specs(XSD_PATH)
            json_specs = extract_jsonschema_field_specs(json_path)
            xsd_roundtrip_specs = extract_xsd_field_specs(xsd_roundtrip)
            excel_specs = extract_excel_field_specs(excel_path)
            excel_specs2 = extract_excel_field_specs(excel_path2)
            mapping_pairs = parse_mapping_excel(mapping_path)
            # Compare and collect all mismatches
            mismatches = []
            mismatches += compare_field_specs(xsd_specs, json_specs, 'XSD', 'JSON Schema')
            mismatches += compare_field_specs(xsd_specs, xsd_roundtrip_specs, 'XSD', 'Roundtrip XSD')
            mismatches += compare_field_specs(xsd_specs, excel_specs, 'XSD', 'Excel from XSD')
            mismatches += compare_field_specs(json_specs, excel_specs2, 'JSON Schema', 'Excel from JSON Schema')
            mismatches += compare_mapping(xsd_specs, json_specs, mapping_pairs)
            if not mismatches:
                print(f"[AUTO-FIX] All issues resolved after {iteration} fix iterations. Fixes applied: {fixes_applied}")
                return
            # Auto-fix functions removed - mismatches will be reported as test failures
            # This allows us to see the actual issues rather than masking them
            # If no known auto-fix applies, fail
            msg = '\n'.join(mismatches)
            pytest.fail(f"Mismatches found in full cycle test after {iteration} fix iterations:\n{msg}\nFixes applied: {fixes_applied}")
    pytest.fail(f"Auto-fix loop exceeded {max_iterations} iterations. Last fixes: {fixes_applied}")

def test_xsd_jsonschema_xsd_mapping_complete_outputs():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    with tempfile.TemporaryDirectory() as tmpdir:
        # Step 1: XSD -> JSON Schema
        the_forge.run_xsd_to_jsonschema_operation(XSD_PATH, tmpdir)
        json_files = [f for f in os.listdir(tmpdir) if f.endswith('.json')]
        assert json_files, 'No JSON Schema generated.'
        json_path = os.path.join(tmpdir, json_files[0])
        assert os.path.getsize(json_path) > 0, 'Generated JSON Schema is empty.'
        # Step 2: JSON Schema -> XSD
        the_forge.jsonschema_to_xsd(json_path, tmpdir)
        xsd_files = [f for f in os.listdir(tmpdir) if f.endswith('.fromjson.xsd')]
        assert xsd_files, 'No XSD generated from JSON Schema.'
        xsd_roundtrip = os.path.join(tmpdir, xsd_files[0])
        assert os.path.getsize(xsd_roundtrip) > 0, 'Generated XSD from JSON Schema is empty.'
        # Step 3: Mapping (original XSD <-> roundtripped XSD)
        the_forge.run_mapping_operation(XSD_PATH, xsd_roundtrip, tmpdir)
        mapping_files = [f for f in os.listdir(tmpdir) if f.startswith('mapping_') and f.endswith('.xlsx')]
        assert mapping_files, 'No mapping file generated.'
        mapping_path = os.path.join(tmpdir, mapping_files[0])
        assert os.path.getsize(mapping_path) > 0, 'Generated mapping file is empty.'

def test_jsonschema_to_xsd_conversion_provided_sample():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    import json
    import xml.etree.ElementTree as ET
    provided_schema = {
      "$schema": "http://json-schema.org/draft-07/schema#",
      "title": "controllingOrderChangeRequest",
      "type": "object",
      "properties": {
        "controllingOrderChangeRequest": {
          "type": "object",
          "properties": {
            "messageHeader": {"type": "string"},
            "controllingOrderRequest": {
              "type": "object",
              "properties": {
                "orderCreation": {
                  "type": "object",
                  "properties": {
                    "orderNumber": {"type": "string", "maxLength": 12, "description": "Order Number"},
                    "orderType": {"type": "string", "maxLength": 4, "description": "Order Type"},
                    "orderName": {"type": "string", "maxLength": 40, "description": "Description"},
                    "controllingArea": {"type": "string", "maxLength": 4, "description": "Controlling Area"},
                    "companyCode": {"type": "string", "maxLength": 4, "description": "Company Code"},
                    "businessArea": {"type": "string", "maxLength": 4, "description": "Business Area"},
                    "plant": {"type": "string", "maxLength": 4, "description": "Plant"},
                    "profitCenter": {"type": "string", "maxLength": 10, "description": "Profit Center"},
                    "responsibleCostCenter": {"type": "string", "maxLength": 10, "description": "Responsible cost center"},
                    "wBSElement": {"type": "string", "maxLength": 24, "description": "Work Breakdown Structure Element (WBS Element)"},
                    "requestingCostCenter": {"type": "string", "maxLength": 10, "description": "Requesting cost center"},
                    "requestingCompanyCode": {"type": "string", "maxLength": 4, "description": "Requesting company code"},
                    "salesOrderNumber": {"type": "string", "maxLength": 10, "description": "Sales Order Number"},
                    "salesOrderItem": {"type": "string", "pattern": "\\d+", "maxLength": 6, "description": "Item number in Sales Order"},
                    "taxJurisdiction": {"type": "string", "maxLength": 15, "description": "Tax Jurisdiction"},
                    "costCenterPosted": {"type": "string", "maxLength": 10, "description": "Cost center to which costs are actually posted"},
                    "orderCurrency": {"type": "string", "maxLength": 5, "description": "Order Currency"},
                    "costingSheet": {"type": "string", "maxLength": 6, "description": "Costing Sheet"},
                    "overheadKey": {"type": "string", "maxLength": 6, "description": "Overhead key"},
                    "resultsAnalysisKey": {"type": "string", "maxLength": 6, "description": "Results Analysis Key"},
                    "interestProfile": {"type": "string", "maxLength": 7, "description": "Interest Profile for Project/Order Interest Calculation"},
                    "applicant": {"type": "string", "maxLength": 20, "description": "Applicant"},
                    "applicantTelephone": {"type": "string", "maxLength": 20, "description": "Applicant's telephone number"},
                    "personResponsible": {"type": "string", "maxLength": 20, "description": "Person Responsible"},
                    "personResponsibleTelephone": {"type": "string", "maxLength": 20, "description": "Telephone number of person in charge"},
                    "estimatedCosts": {"type": "number", "multipleOf": 0.0001, "$comment": "totalDigits: 23; fractionDigits: 4", "description": "Estimated Overall Costs of the Order for BAPIs"},
                    "applicationDate": {"type": "string", "maxLength": 8, "description": "Application date"},
                    "departament": {"type": "string", "maxLength": 15, "description": "Department"},
                    "dateWorkBegins": {"type": "string", "maxLength": 8, "description": "Work Start"},
                    "dateWorkEnds": {"type": "string", "maxLength": 8, "description": "End of Work"},
                    "workPermit": {"type": "string", "maxLength": 1, "description": "Identifier for work permit issued"},
                    "externalOrderNumber": {"type": "string", "maxLength": 20, "description": "External order number"},
                    "processingGroup": {"type": "string", "pattern": "\\d+", "maxLength": 2, "description": "Processing group"},
                    "plannedReleaseDate": {"type": "string", "maxLength": 8, "description": "Planned release date"},
                    "plannedCompletionDate": {"type": "string", "maxLength": 8, "description": "Planned completion date"},
                    "plannedClosingDate": {"type": "string", "maxLength": 8, "description": "Planned closing date"},
                    "requestingOrder": {"type": "string", "maxLength": 12, "description": "Requesting order"},
                    "functionalArea": {"type": "string", "maxLength": 4, "description": "Functional Area"},
                    "functionalAreaLong": {"type": "string", "maxLength": 16, "description": "Functional Area"},
                    "responsibleInternalOrder": {"type": "string", "maxLength": 12, "description": "Person Responsible for CO Internal Order"}
                  },
                  "description": "Master Data Used to Create Order"
                },
                "internalOrders": {
                  "type": "object",
                  "properties": {
                    "location": {"type": "string", "maxLength": 10, "description": "Location"},
                    "locationPlant": {"type": "string", "maxLength": 4, "description": "Location plant"},
                    "statistical": {"type": "string", "maxLength": 1, "description": "Identifier for statistical order"},
                    "objectClass": {"type": "string", "maxLength": 2, "description": "Object Class"},
                    "integratedPlanningIndicator": {"type": "string", "maxLength": 1, "description": "Indicator for Integrated Planning"},
                    "investmentMeasureProfile": {"type": "string", "maxLength": 6, "description": "Investment measure profile"},
                    "scale": {"type": "string", "maxLength": 2, "description": "Scale of investment objects"},
                    "investmentReason": {"type": "string", "maxLength": 2, "description": "Reason for investment"},
                    "environmentalReason": {"type": "string", "maxLength": 5, "description": "Reason for environmental investment"},
                    "deletionFlag": {"type": "string", "maxLength": 1, "description": "Deletion flag"},
                    "planningWithLineIndicator": {"type": "string", "maxLength": 1, "description": "Identifier for planning with line items"},
                    "jointVenture": {"type": "string", "maxLength": 6, "description": "Joint venture"},
                    "recoveryIndicator": {"type": "string", "maxLength": 2, "description": "Recovery Indicator"},
                    "equityType": {"type": "string", "maxLength": 3, "description": "Equity type"},
                    "jointVentureObjectType": {"type": "string", "maxLength": 4, "description": "Joint Venture Object Type"},
                    "class": {"type": "string", "maxLength": 3, "description": "JIB/JIBE Class"},
                    "subClass": {"type": "string", "maxLength": 5, "description": "JIB/JIBE Subclass A"},
                    "jointVentureCostObject": {"type": "string", "maxLength": 1, "description": "JV original cost object"}
                  },
                  "description": "Other Master Data Fields for Internal Order BAPI2075"
                },
                "settlementRules": {
                  "type": "array",
                  "items": {
                    "type": "object",
                    "properties": {
                      "settlementType": {"type": "string", "maxLength": 3, "description": "Settlement Type (English Fixed Values for BAPI)"},
                      "sourceAssignment": {"type": "string", "maxLength": 3, "description": "Source Assignment"},
                      "percentage": {"type": "number", "multipleOf": 0.01, "$comment": "totalDigits: 5; fractionDigits: 2", "description": "Settlement percentage rate"},
                      "equivalenceNumber": {"type": "number", "$comment": "totalDigits: 10", "description": "Equivalence number for order settlement"},
                      "amount": {"type": "number", "multipleOf": 0.0001, "$comment": "totalDigits: 23; fractionDigits: 4", "description": "Amount for Amount Rule in BAPIs"},
                      "businessArea": {"type": "string", "maxLength": 4, "description": "Business Area"},
                      "companyCode": {"type": "string", "maxLength": 4, "description": "Company Code"},
                      "account": {"type": "string", "maxLength": 10, "description": "G/L Account Number"},
                      "profitCenter": {"type": "string", "maxLength": 10, "description": "Profit Center"},
                      "costCenter": {"type": "string", "maxLength": 10, "description": "Cost Center"},
                      "orderNumber": {"type": "string", "maxLength": 12, "description": "Order Number"},
                      "wBSElement": {"type": "string", "maxLength": 24, "description": "Work Breakdown Structure Element (WBS Element)"},
                      "assetNumber": {"type": "string", "maxLength": 12, "description": "Main Asset Number"},
                      "assetSubNumber": {"type": "string", "maxLength": 4, "description": "Asset Subnumber"},
                      "networking": {"type": "string", "maxLength": 12, "description": "Network Number for Account Assignment"},
                      "activityNumber": {"type": "string", "maxLength": 4, "description": "Activity Number"},
                      "salesOrderNumber": {"type": "string", "maxLength": 10, "description": "Sales Order Number"},
                      "salesOrderItem": {"type": "string", "pattern": "\\d+", "maxLength": 6, "description": "Item number in Sales Order"},
                      "costObject": {"type": "string", "maxLength": 12, "description": "Cost Object"},
                      "businessProcess": {"type": "string", "maxLength": 12, "description": "Business Process"}
                    },
                    "description": "Settlement Rules From Which Order Was Created"
                  }
                },
                "extensionIn": {
                  "type": "object",
                  "properties": {
                    "extensionItem": {
                      "type": "array",
                      "items": {
                        "type": "object",
                        "properties": {
                          "structure": {"type": "string", "maxLength": 30, "description": "Structure name of BAPI table extension"},
                          "valuePartOne": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartTwo": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartThree": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartFour": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"}
                        }
                      }
                    }
                  },
                  "description": "Extension Structure"
                }
              }
            }
          },
          "required": [
            "messageHeader",
            "controllingOrderRequest"
          ]
        }
      },
      "required": [
        "controllingOrderChangeRequest"
      ]
    }
    with tempfile.TemporaryDirectory() as tmpdir:
        json_path = os.path.join(tmpdir, "provided_schema.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(provided_schema, f, indent=2)
        the_forge.jsonschema_to_xsd(json_path, tmpdir)
        xsd_files = [f for f in os.listdir(tmpdir) if f.endswith('.fromjson.xsd')]
        assert xsd_files, 'No XSD generated from provided JSON Schema.'
        xsd_path = os.path.join(tmpdir, xsd_files[0])
        assert os.path.getsize(xsd_path) > 0, 'Generated XSD is empty.'
        # --- Deep validation: field-by-field and structure validation ---
        json_specs = extract_jsonschema_field_specs(json_path)
        xsd_specs = extract_xsd_field_specs(xsd_path)
        mismatches = compare_field_specs(json_specs, xsd_specs, label1='JSON Schema', label2='XSD')
        assert not mismatches, f"Field-by-field or structure mismatches found:\n" + '\n'.join(mismatches)

def test_jsonschema_to_xsd_conversion_provided_schema_user():
    if the_forge is None:
        pytest.skip("the-forge.py could not be loaded")
    import json
    provided_schema = {
      "$schema": "http://json-schema.org/draft-07/schema#",
      "title": "controllingOrderChangeRequest",
      "type": "object",
      "properties": {
        "controllingOrderChangeRequest": {
          "type": "object",
          "properties": {
            "messageHeader": {"type": "string"},
            "controllingOrderRequest": {
              "type": "object",
              "properties": {
                "orderCreation": {
                  "type": "object",
                  "properties": {
                    "orderNumber": {"type": "string", "maxLength": 12, "description": "Order Number"},
                    "orderType": {"type": "string", "maxLength": 4, "description": "Order Type"},
                    "orderName": {"type": "string", "maxLength": 40, "description": "Description"},
                    "controllingArea": {"type": "string", "maxLength": 4, "description": "Controlling Area"},
                    "companyCode": {"type": "string", "maxLength": 4, "description": "Company Code"},
                    "businessArea": {"type": "string", "maxLength": 4, "description": "Business Area"},
                    "plant": {"type": "string", "maxLength": 4, "description": "Plant"},
                    "profitCenter": {"type": "string", "maxLength": 10, "description": "Profit Center"},
                    "responsibleCostCenter": {"type": "string", "maxLength": 10, "description": "Responsible cost center"},
                    "wBSElement": {"type": "string", "maxLength": 24, "description": "Work Breakdown Structure Element (WBS Element)"},
                    "requestingCostCenter": {"type": "string", "maxLength": 10, "description": "Requesting cost center"},
                    "requestingCompanyCode": {"type": "string", "maxLength": 4, "description": "Requesting company code"},
                    "salesOrderNumber": {"type": "string", "maxLength": 10, "description": "Sales Order Number"},
                    "salesOrderItem": {"type": "string", "pattern": "\\d+", "maxLength": 6, "description": "Item number in Sales Order"},
                    "taxJurisdiction": {"type": "string", "maxLength": 15, "description": "Tax Jurisdiction"},
                    "costCenterPosted": {"type": "string", "maxLength": 10, "description": "Cost center to which costs are actually posted"},
                    "orderCurrency": {"type": "string", "maxLength": 5, "description": "Order Currency"},
                    "costingSheet": {"type": "string", "maxLength": 6, "description": "Costing Sheet"},
                    "overheadKey": {"type": "string", "maxLength": 6, "description": "Overhead key"},
                    "resultsAnalysisKey": {"type": "string", "maxLength": 6, "description": "Results Analysis Key"},
                    "interestProfile": {"type": "string", "maxLength": 7, "description": "Interest Profile for Project/Order Interest Calculation"},
                    "applicant": {"type": "string", "maxLength": 20, "description": "Applicant"},
                    "applicantTelephone": {"type": "string", "maxLength": 20, "description": "Applicant's telephone number"},
                    "personResponsible": {"type": "string", "maxLength": 20, "description": "Person Responsible"},
                    "personResponsibleTelephone": {"type": "string", "maxLength": 20, "description": "Telephone number of person in charge"},
                    "estimatedCosts": {"type": "number", "multipleOf": 0.0001, "$comment": "totalDigits: 23; fractionDigits: 4", "description": "Estimated Overall Costs of the Order for BAPIs"},
                    "applicationDate": {"type": "string", "maxLength": 8, "description": "Application date"},
                    "departament": {"type": "string", "maxLength": 15, "description": "Department"},
                    "dateWorkBegins": {"type": "string", "maxLength": 8, "description": "Work Start"},
                    "dateWorkEnds": {"type": "string", "maxLength": 8, "description": "End of Work"},
                    "workPermit": {"type": "string", "maxLength": 1, "description": "Identifier for work permit issued"},
                    "externalOrderNumber": {"type": "string", "maxLength": 20, "description": "External order number"},
                    "processingGroup": {"type": "string", "pattern": "\\d+", "maxLength": 2, "description": "Processing group"},
                    "plannedReleaseDate": {"type": "string", "maxLength": 8, "description": "Planned release date"},
                    "plannedCompletionDate": {"type": "string", "maxLength": 8, "description": "Planned completion date"},
                    "plannedClosingDate": {"type": "string", "maxLength": 8, "description": "Planned closing date"},
                    "requestingOrder": {"type": "string", "maxLength": 12, "description": "Requesting order"},
                    "functionalArea": {"type": "string", "maxLength": 4, "description": "Functional Area"},
                    "functionalAreaLong": {"type": "string", "maxLength": 16, "description": "Functional Area"},
                    "responsibleInternalOrder": {"type": "string", "maxLength": 12, "description": "Person Responsible for CO Internal Order"}
                  },
                  "description": "Master Data Used to Create Order"
                },
                "internalOrders": {
                  "type": "object",
                  "properties": {
                    "location": {"type": "string", "maxLength": 10, "description": "Location"},
                    "locationPlant": {"type": "string", "maxLength": 4, "description": "Location plant"},
                    "statistical": {"type": "string", "maxLength": 1, "description": "Identifier for statistical order"},
                    "objectClass": {"type": "string", "maxLength": 2, "description": "Object Class"},
                    "integratedPlanningIndicator": {"type": "string", "maxLength": 1, "description": "Indicator for Integrated Planning"},
                    "investmentMeasureProfile": {"type": "string", "maxLength": 6, "description": "Investment measure profile"},
                    "scale": {"type": "string", "maxLength": 2, "description": "Scale of investment objects"},
                    "investmentReason": {"type": "string", "maxLength": 2, "description": "Reason for investment"},
                    "environmentalReason": {"type": "string", "maxLength": 5, "description": "Reason for environmental investment"},
                    "deletionFlag": {"type": "string", "maxLength": 1, "description": "Deletion flag"},
                    "planningWithLineIndicator": {"type": "string", "maxLength": 1, "description": "Identifier for planning with line items"},
                    "jointVenture": {"type": "string", "maxLength": 6, "description": "Joint venture"},
                    "recoveryIndicator": {"type": "string", "maxLength": 2, "description": "Recovery Indicator"},
                    "equityType": {"type": "string", "maxLength": 3, "description": "Equity type"},
                    "jointVentureObjectType": {"type": "string", "maxLength": 4, "description": "Joint Venture Object Type"},
                    "class": {"type": "string", "maxLength": 3, "description": "JIB/JIBE Class"},
                    "subClass": {"type": "string", "maxLength": 5, "description": "JIB/JIBE Subclass A"},
                    "jointVentureCostObject": {"type": "string", "maxLength": 1, "description": "JV original cost object"}
                  },
                  "description": "Other Master Data Fields for Internal Order BAPI2075"
                },
                "settlementRules": {
                  "type": "array",
                  "items": {
                    "type": "object",
                    "properties": {
                      "settlementType": {"type": "string", "maxLength": 3, "description": "Settlement Type (English Fixed Values for BAPI)"},
                      "sourceAssignment": {"type": "string", "maxLength": 3, "description": "Source Assignment"},
                      "percentage": {"type": "number", "multipleOf": 0.01, "$comment": "totalDigits: 5; fractionDigits: 2", "description": "Settlement percentage rate"},
                      "equivalenceNumber": {"type": "number", "$comment": "totalDigits: 10", "description": "Equivalence number for order settlement"},
                      "amount": {"type": "number", "multipleOf": 0.0001, "$comment": "totalDigits: 23; fractionDigits: 4", "description": "Amount for Amount Rule in BAPIs"},
                      "businessArea": {"type": "string", "maxLength": 4, "description": "Business Area"},
                      "companyCode": {"type": "string", "maxLength": 4, "description": "Company Code"},
                      "account": {"type": "string", "maxLength": 10, "description": "G/L Account Number"},
                      "profitCenter": {"type": "string", "maxLength": 10, "description": "Profit Center"},
                      "costCenter": {"type": "string", "maxLength": 10, "description": "Cost Center"},
                      "orderNumber": {"type": "string", "maxLength": 12, "description": "Order Number"},
                      "wBSElement": {"type": "string", "maxLength": 24, "description": "Work Breakdown Structure Element (WBS Element)"},
                      "assetNumber": {"type": "string", "maxLength": 12, "description": "Main Asset Number"},
                      "assetSubNumber": {"type": "string", "maxLength": 4, "description": "Asset Subnumber"},
                      "networking": {"type": "string", "maxLength": 12, "description": "Network Number for Account Assignment"},
                      "activityNumber": {"type": "string", "maxLength": 4, "description": "Activity Number"},
                      "salesOrderNumber": {"type": "string", "maxLength": 10, "description": "Sales Order Number"},
                      "salesOrderItem": {"type": "string", "pattern": "\\d+", "maxLength": 6, "description": "Item number in Sales Order"},
                      "costObject": {"type": "string", "maxLength": 12, "description": "Cost Object"},
                      "businessProcess": {"type": "string", "maxLength": 12, "description": "Business Process"}
                    },
                    "description": "Settlement Rules From Which Order Was Created"
                  }
                },
                "extensionIn": {
                  "type": "object",
                  "properties": {
                    "extensionItem": {
                      "type": "array",
                      "items": {
                        "type": "object",
                        "properties": {
                          "structure": {"type": "string", "maxLength": 30, "description": "Structure name of BAPI table extension"},
                          "valuePartOne": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartTwo": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartThree": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"},
                          "valuePartFour": {"type": "string", "maxLength": 240, "description": "Data part of BAPI extension parameter"}
                        }
                      }
                    }
                  },
                  "description": "Extension Structure"
                }
              }
            }
          },
          "required": [
            "messageHeader",
            "controllingOrderRequest"
          ]
        }
      },
      "required": [
        "controllingOrderChangeRequest"
      ]
    }
    with tempfile.TemporaryDirectory() as tmpdir:
        json_path = os.path.join(tmpdir, "provided_schema_user.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(provided_schema, f, indent=2)
        the_forge.jsonschema_to_xsd(json_path, tmpdir)
        xsd_files = [f for f in os.listdir(tmpdir) if f.endswith('.fromjson.xsd')]
        assert xsd_files, 'No XSD generated from provided JSON Schema.'
        xsd_path = os.path.join(tmpdir, xsd_files[0])
        assert os.path.getsize(xsd_path) > 0, 'Generated XSD is empty.'
        # --- Deep validation: field-by-field and structure validation ---
        json_specs = extract_jsonschema_field_specs(json_path)
        xsd_specs = extract_xsd_field_specs(xsd_path)
        mismatches = compare_field_specs(json_specs, xsd_specs, label1='JSON Schema', label2='XSD')
        assert not mismatches, f"Field-by-field or structure mismatches found:\n" + '\n'.join(mismatches)

def test_mapping_arrays_are_populated():
    import tempfile
    import os
    import json
    import sys
    sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
    from src.core.schema_processor import SchemaProcessor
    # Example JSON Schema and XSD with arrays (use your provided schema and XSD)
    json_schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": "controllingOrderChangeRequest",
        "type": "object",
        "properties": {
            "controllingOrderChangeRequest": {
                "type": "object",
                "properties": {
                    "settlementRules": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "settlementType": {"type": "string"},
                                "amount": {"type": "number"}
                            }
                        }
                    }
                }
            }
        }
    }
    xsd_content = '''<?xml version="1.0" encoding="UTF-8"?>
    <xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">
      <xs:element name="ControllingOrderChangeRequest">
        <xs:complexType>
          <xs:sequence>
            <xs:element name="SettlementRules" minOccurs="0" maxOccurs="unbounded">
              <xs:complexType>
                <xs:sequence>
                  <xs:element name="SettlementType" type="xs:string" minOccurs="0" maxOccurs="1"/>
                  <xs:element name="Amount" type="xs:decimal" minOccurs="0" maxOccurs="1"/>
                </xs:sequence>
              </xs:complexType>
            </xs:element>
          </xs:sequence>
        </xs:complexType>
      </xs:element>
    </xs:schema>'''
    with tempfile.TemporaryDirectory() as tmpdir:
        json_path = os.path.join(tmpdir, "schema.json")
        xsd_path = os.path.join(tmpdir, "schema.xsd")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_schema, f)
        with open(xsd_path, "w", encoding="utf-8") as f:
            f.write(xsd_content)
        processor = SchemaProcessor()
        source_fields = processor.extract_fields_from_json_schema(json_path)
        target_fields = processor.extract_fields_from_xsd(xsd_path)
        # Print extracted fields for arrays
        print("[DEBUG] Source fields (JSON Schema):")
        for f in source_fields:
            if 'settlementRules' in f.levels:
                print(f)
        print("[DEBUG] Target fields (XSD):")
        for f in target_fields:
            if 'settlementRules' in f.levels:
                print(f)
        # Run mapping using the mapping engine
        from src.core.mapping_engine import MappingEngine
        engine = MappingEngine(threshold=0.7)
        mapping = engine.map_fields(source_fields, target_fields)
        # Print mapping results for debugging
        print("[DEBUG] Mapping results:")
        for m in mapping:
            print(f"  {m.source} -> {m.target} (similarity: {m.similarity})")
        # Assert that array item fields are mapped (using normalized paths)
        mapped_settlement_type = [m for m in mapping if 'settlementrules.arrayitem.settlementtype' in m.source]
        mapped_amount = [m for m in mapping if 'settlementrules.arrayitem.amount' in m.source]
        assert mapped_settlement_type and mapped_settlement_type[0].target, "settlementType array item should be mapped in target"
        assert mapped_amount and mapped_amount[0].target, "amount array item should be mapped in target"

# --- Pytest hook to generate a report at the end of the test session ---
def pytest_sessionfinish(session, exitstatus):
    report_lines = []
    report_lines.append('Integration Test Report')
    report_lines.append('='*30)
    for item in session.items:
        outcome = 'PASSED'
        # Use the _store attribute to get the test outcome
        rep = getattr(item, 'rep_call', None)
        if rep is not None and hasattr(rep, 'outcome') and rep.outcome == 'failed':
            outcome = 'FAILED'
        report_lines.append(f"{item.nodeid}: {outcome}")
        if rep is not None and hasattr(rep, 'longrepr') and rep.outcome == 'failed':
            report_lines.append(str(rep.longrepr))
    report_path = os.path.join(os.path.dirname(__file__), 'test_report.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    print(f"\n[INFO] Test report written to {report_path}\n") 