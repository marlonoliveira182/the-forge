import sys
import os
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QFileDialog, QTextEdit, QFrame, QToolButton, QCheckBox, QMessageBox, QSplitter, QStatusBar, QSizePolicy, QTabWidget, QSpacerItem, QComboBox
)
from PySide6.QtGui import QIcon, QFont, QFontDatabase, QPixmap
from PySide6.QtCore import Qt, Signal, QTimer, QSize
from openpyxl.utils import get_column_letter

ASSETS_DIR = os.path.join(os.path.dirname(__file__), 'assets')
ICONS_DIR = os.path.join(ASSETS_DIR, 'Icons_SVG', '_M', 'Actions')
FONT_DIR = os.path.join(ASSETS_DIR, 'Mulish', 'static')
LOGO_PATH = os.path.join(ASSETS_DIR, 'anvil.ico')

EDP_COLORS = {
    'marine_blue': '#212E3E',
    'marine_blue_light': '#2A3647',
    'electric_green': '#28FF52',
    'cobalt_blue': '#263CC8',
    'slate_grey': '#BECACC',
    'white': '#FFFFFF',
    'black': '#212E3E',
    'grey': '#A3B5B8',
    'ice_blue': '#0CD3F8',
}

SECTIONS = [
    ("Mapping", 'actions_edit_M.svg'),
    ("WSDL to XSD", 'actions_file_download_M.svg'),
]

def get_icon(icon_name):
    path = os.path.join(ICONS_DIR, icon_name)
    if not os.path.exists(path):
        path = os.path.join(ICONS_DIR, 'actions_edit_M.svg')
    return QIcon(path)

class DraggableLineEdit(QLineEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)
        self._on_file_dropped = None
    def set_on_file_dropped(self, callback):
        self._on_file_dropped = callback
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)
    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                path = urls[0].toLocalFile()
                self.setText(path)
                if self._on_file_dropped:
                    self._on_file_dropped(path)
            event.acceptProposedAction()
        else:
            super().dropEvent(event)

class DropZone(QFrame):
    def __init__(self, label_text, file_ext, on_file_selected, parent=None):
        super().__init__(parent)
        self.icon_label = QLabel()  # Ensure this is created first
        self.icon_label.setAlignment(Qt.AlignCenter)
        self.icon_label.setStyleSheet('border: none; background: none;')
        self.setFrameShape(QFrame.StyledPanel)
        self.setAcceptDrops(True)
        self.file_filter = file_ext
        self.on_file_selected = on_file_selected
        self.file_path = None
        # Use a modern upload icon from assets (fallback to folder if not found)
        icon_path = os.path.join(os.path.dirname(__file__), 'assets', 'upload.png')
        if not os.path.exists(icon_path):
            icon_path = os.path.join(os.path.dirname(__file__), 'assets', 'folder.png')
        # Add to the end of _init_ui or after main window setup
        self.setStyleSheet(self.styleSheet() + """
            QPushButton {
                border-radius: 5px;
                background: #263CC8;
                color: #fff;
                font-family: 'Mulish';
                font-size: 13px;
                padding: 6px 14px;
                border: none;
            }
            QPushButton#primary {
                background: #28FF52;
                color: #212E3E;
            }
            QPushButton:hover {
                background: #4A90E2;
                color: #fff;
            }
            QPushButton:focus {
                outline: 2px solid #4A90E2;
            }
            QFrame:focus {
                outline: 2px solid #4A90E2;
            }
            QTextEdit[logarea="true"] {
                background: #f7f7fa;
                border-radius: 6px;
                border: 1.2px solid #e0e4ea;
                color: #263CC8;
                font-family: 'Mulish';
                font-size: 11px;
            }
        """)
        # In DropZone, prefer SVG if available
        icon_path_svg = os.path.join(os.path.dirname(__file__), 'assets', 'upload.svg')
        if os.path.exists(icon_path_svg):
            pixmap = QPixmap(icon_path_svg).scaled(28, 28, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.icon_label.setPixmap(pixmap)
        elif os.path.exists(icon_path):
            pixmap = QPixmap(icon_path).scaled(28, 28, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.icon_label.setPixmap(pixmap)
        else:
            self.icon_label.clear()
        # Main label: modern, readable, EDP colors
        self.label_main = QLabel()
        self.label_main.setAlignment(Qt.AlignCenter)
        self.label_main.setText(
            '<span style="color:#263CC8;font-size:13px;">Drop files here or '
            '<b style="color:#212E3E;">Browse files</b></span>'
        )
        self.label_main.setTextFormat(Qt.RichText)
        self.label_main.setFont(QFont('Mulish', 13))
        self.label_main.setStyleSheet('border: none; background: none;')
        # Layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 16, 8, 16)
        layout.setSpacing(8)
        if self.icon_label.pixmap() is not None and not self.icon_label.pixmap().isNull():
            layout.addStretch(1)
            layout.addWidget(self.icon_label, alignment=Qt.AlignHCenter)
            layout.addWidget(self.label_main, alignment=Qt.AlignHCenter)
            layout.addStretch(1)
        else:
            layout.addStretch(1)
            layout.addWidget(self.label_main, alignment=Qt.AlignHCenter)
            layout.addStretch(1)
        self.setMinimumHeight(40)
        self.setMinimumWidth(180)
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip("Drag and drop an XSD file here, or click to select from disk.")
        self._set_default_style()
    def _set_default_style(self):
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #28FF52;
                border-radius: 8px;
                background: #fff;
            }
        """)
        self.label_main.setStyleSheet('border: none; background: none; color: #263CC8;')
    def _set_hover_style(self):
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #263CC8;
                border-radius: 8px;
                background: #f8faff;
            }
        """)
        self.label_main.setStyleSheet('border: none; background: none; color: #263CC8;')
    def _set_active_style(self):
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #212E3E;
                border-radius: 8px;
                background: #f7f7fa;
            }
        """)
        self.label_main.setStyleSheet('border: none; background: none; color: #212E3E;')
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().lower().endswith(self.file_filter):
                event.acceptProposedAction()
                self._set_hover_style()
            else:
                event.ignore()
        else:
            event.ignore()
    def dragLeaveEvent(self, event):
        self._set_default_style()
    def dropEvent(self, event):
        self._set_active_style()
        if event.mimeData().hasUrls():
            path = event.mimeData().urls()[0].toLocalFile()
            if path.lower().endswith(self.file_filter):
                self.set_file(path)
        QTimer.singleShot(400, self._set_default_style)
    def mousePressEvent(self, event):
        path, _ = QFileDialog.getOpenFileName(self, "Browse files", os.path.expanduser("~"), f"XSD Files (*{self.file_filter})")
        if path:
            self.set_file(path)
    def set_file(self, path):
        self.file_path = path
        name = os.path.basename(path)
        # Use high-contrast color for file name
        self.label_main.setText(
            f'<span style="color:#212E3E;font-size:13px;">{name}</span>'
        )
        self.label_main.setTextFormat(Qt.RichText)
        self.label_main.setStyleSheet('border: none; background: none; color: #212E3E;')
        self._set_active_style()
        if self.on_file_selected:
            self.on_file_selected(path)
    def clear(self):
        self.file_path = None
        self.label_main.setText(
            '<span style="color:#263CC8;font-size:13px;">Drop files here or '
            '<b style="color:#212E3E;">Browse files</b></span>'
        )
        self.label_main.setTextFormat(Qt.RichText)
        self.label_main.setStyleSheet('border: none; background: none; color: #263CC8;')
        self._set_default_style()

class ForgeMainWindow(QMainWindow):
    log_signal = Signal(str, str)  # message, level
    def __init__(self):
        super().__init__()
        self.setWindowTitle("The Forge")
        self.setMinimumSize(900, 600)
        self.setWindowIcon(QIcon(LOGO_PATH) if os.path.exists(LOGO_PATH) else QIcon())
        self._load_fonts()
        self._init_ui()
        self.source_path_full = ''
        self.target_path_full = ''
        self.output_path_full = ''
        self.log_signal.connect(self._log)
        self.working_folder = ''
        self.field_case = "PascalCase" # Initialize field_case
        # Set initial theme
        # self._toggle_theme(force_dark=True) # Removed theme feature

    def _load_fonts(self):
        # Only load the fonts actually used (Regular and Bold)
        for font_file in os.listdir(FONT_DIR):
            if font_file in ("Mulish-Regular.ttf", "Mulish-Bold.ttf"):
                QFontDatabase.addApplicationFont(os.path.join(FONT_DIR, font_file))
        self.setFont(QFont('Mulish', 11))

    def _init_ui(self):
        splitter = QSplitter()
        splitter.setOrientation(Qt.Horizontal)
        self.setCentralWidget(splitter)

        # --- Topbar with Centered Tabs ---
        # Remove the entire topbar and its layout, including tabs and theme toggle
        # Instead, add the tabs directly to the main layout or splitter
        # Remove all code related to topbar, theme_checkbox, theme toggling, and topbar_layout
        # Add tabs directly to splitter
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setMovable(False)
        self.tabs.setDocumentMode(True)
        self.tabs.setStyleSheet('''
            QTabBar::tab {
                background: #212E3E;
                color: #fff;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 4px 18px;
                margin-right: 2px;
                font-family: "Mulish";
                font-size: 13px;
                min-width: 80px;
                min-height: 24px;
                border: none;
            }
            QTabBar::tab:selected {
                background: #fff;
                color: #263CC8;
                font-weight: bold;
                border-bottom: 2px solid #28FF52;
            }
            QTabBar::tab:!selected {
                background: #212E3E;
                color: #fff;
                font-weight: normal;
                border-bottom: 2px solid #212E3E;
            }
            QTabWidget::pane {
                border: none;
                background: #212E3E;
                top: 0px;
            }
            QTabBar {
                background: #212E3E;
                border: none;
                margin-top: 8px;
            }
        ''')
        # Section widgets (order must match SECTIONS)
        self.mapping_widget = self._build_mapping_section()
        self.wsdl2xsd_widget = self._build_wsdl2xsd_section()
        self.tabs.addTab(self.mapping_widget, "Mapping")
        self.tabs.addTab(self.wsdl2xsd_widget, "WSDL to XSD")
        self.tabs.currentChanged.connect(self._on_tab_changed)
        splitter.addWidget(self.tabs)
        splitter.setStretchFactor(0, 1)
        # Status bar
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        # Set the default (dark mode) colors for the whole app
        self.setStyleSheet(f"""
            QMainWindow {{ background: {EDP_COLORS['marine_blue']}; }}
            QLabel, QLineEdit, QTextEdit {{ color: {EDP_COLORS['white']}; font-family: 'Mulish'; font-size: 15px; }}
            #sidebarPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
            #mainPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
            QListWidget#sidebarNav {{ background: transparent; font-size: 15px; border: none; }}
            QListWidget#sidebarNav::item:selected {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-left: 4px solid {EDP_COLORS['cobalt_blue']}; }}
            QPushButton {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-radius: 8px; font-family: 'Mulish'; font-weight: bold; font-size: 15px; min-height: 40px; padding: 0 20px; border: none; }}
            QPushButton:disabled {{ background: {EDP_COLORS['grey']}; color: {EDP_COLORS['marine_blue_light']}; }}
            QLineEdit, QTextEdit {{ background: {EDP_COLORS['marine_blue']}; border: 1.5px solid {EDP_COLORS['marine_blue_light']}; border-radius: 8px; font-family: 'Mulish'; font-size: 14px; padding: 8px; }}
            QToolButton {{ background: transparent; border: none; }}
            QStatusBar {{ background: {EDP_COLORS['marine_blue_light']}; color: {EDP_COLORS['electric_green']}; font-size: 13px; border: none; }}
        """)

    # Add the handler for the checkbox
    # def _on_theme_checkbox_toggled(self, checked):
    #     self._toggle_theme(force_dark=checked) # Removed theme feature

    # def _toggle_theme(self, force_dark=None):
    #     # force_dark: True for dark, False for light
    #     dark = force_dark if force_dark is not None else (self.theme_checkbox.isChecked())
    #     if dark:
    #         self.setStyleSheet(f"""
    #             QMainWindow {{ background: {EDP_COLORS['marine_blue']}; }}
    #             QLabel, QLineEdit, QTextEdit {{ color: {EDP_COLORS['white']}; font-family: 'Mulish'; font-size: 15px; }}
    #             #sidebarPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
    #             #mainPanel {{ background: {EDP_COLORS['marine_blue_light']}; border-radius: 12px; }}
    #             QListWidget#sidebarNav {{ background: transparent; font-size: 15px; border: none; }}
    #             QListWidget#sidebarNav::item:selected {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-left: 4px solid {EDP_COLORS['cobalt_blue']}; }}
    #             QPushButton {{ background: {EDP_COLORS['electric_green']}; color: {EDP_COLORS['marine_blue']}; border-radius: 8px; font-family: 'Mulish'; font-weight: bold; font-size: 15px; min-height: 40px; padding: 0 20px; border: none; }}
    #             QPushButton:disabled {{ background: {EDP_COLORS['grey']}; color: {EDP_COLORS['marine_blue_light']}; }}
    #             QLineEdit, QTextEdit {{ background: {EDP_COLORS['marine_blue']}; border: 1.5px solid {EDP_COLORS['marine_blue_light']}; border-radius: 8px; font-family: 'Mulish'; font-size: 14px; padding: 8px; }}
    #             QToolButton {{ background: transparent; border: none; }}
    #             QStatusBar {{ background: {EDP_COLORS['marine_blue_light']}; color: {EDP_COLORS['electric_green']}; font-size: 13px; border: none; }}
    #         """)
    #     else:
    #         self.setStyleSheet(f"""
    #             QMainWindow {{ background: #F0F4F8; }}
    #             QLabel {{ color: {EDP_COLORS['marine_blue']}; font-family: 'Mulish'; font-size: 15px; }}
    #             #sidebarPanel {{ background: #F7F9FA; border-radius: 16px; border: 1.5px solid #E0E4EA; }}
    #             #mainPanel {{ background: #FFFFFF; border-radius: 18px; border: 2px solid #E0E4EA; }}
    #             QListWidget#sidebarNav {{ background: transparent; font-size: 15px; border: none; }}
    #             QListWidget#sidebarNav::item:selected {{ background: #E6F9F0; color: {EDP_COLORS['marine_blue']}; border-left: 4px solid {EDP_COLORS['electric_green']}; }}
    #             QPushButton {{ background: #FFFFFF; color: {EDP_COLORS['cobalt_blue']}; border-radius: 8px; font-family: 'Mulish'; font-weight: 600; font-size: 15px; min-height: 40px; padding: 0 20px; border: 1.5px solid #E0E4EA; }}
    #             QPushButton:disabled {{ background: #F0F4F8; color: #B0B8C1; border: 1.5px solid #E0E4EA; }}
    #             QPushButton:focus {{ border: 2px solid {EDP_COLORS['electric_green']}; }}
    #             QPushButton#primary {{ background: {EDP_COLORS['cobalt_blue']}; color: #FFFFFF; border: none; }}
    #             QLineEdit, QTextEdit {{ background: #FFFFFF; border: 1.5px solid #E0E4EA; border-radius: 8px; font-family: 'Mulish'; font-size: 14px; padding: 8px; }}
    #             QToolButton {{ background: transparent; border: none; }}
    #             QStatusBar {{ background: #F7F9FA; color: {EDP_COLORS['cobalt_blue']}; font-size: 13px; border-top: 1px solid #E0E4EA; }}
    #         """)

    def _on_tab_changed(self, idx):
        self.status.showMessage(f"{self.tabs.tabText(idx)} selected", 2000)

    def _build_mapping_section(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(16, 16, 16, 16)  # Reduced margins
        layout.setSpacing(10)  # Reduced spacing
        # --- Dropzones ---
        drop_row = QHBoxLayout()
        drop_row.setSpacing(8)
        self.src_drop = DropZone("Drop files here or Browse files", ".xsd", self._on_src_selected)
        self.tgt_drop = DropZone("Drop files here or Browse files", ".xsd", self._on_tgt_selected)
        self.src_drop.setMinimumHeight(60)
        self.tgt_drop.setMinimumHeight(60)
        drop_row.addWidget(self.src_drop)
        drop_row.addWidget(self.tgt_drop)
        layout.addLayout(drop_row)
        # Output folder selection
        output_row = QHBoxLayout()
        output_row.setSpacing(6)
        self.output_folder_line = QLineEdit()
        self.output_folder_line.setReadOnly(True)
        self.output_folder_line.setPlaceholderText("Select output folder...")
        self.output_folder_line.setMinimumHeight(24)
        output_folder_btn = QPushButton("Select Output Folder")
        output_folder_btn.setMinimumHeight(24)
        output_folder_btn.clicked.connect(self._select_output_folder)
        output_row.addWidget(self.output_folder_line)
        output_row.addWidget(output_folder_btn)
        layout.addLayout(output_row)
        # Output path (auto-generated)
        self.output_path = QLineEdit()
        self.output_path.setReadOnly(True)
        self.output_path.setPlaceholderText("Output mapping file will be auto-named here.")
        self.output_path.setMinimumHeight(24)
        layout.addWidget(QLabel("Output Mapping File:"))
        layout.addWidget(self.output_path)
        # Generate button
        self.generate_btn = QPushButton("Generate Mapping")
        self.generate_btn.setFont(QFont('Mulish', 11, QFont.Bold))
        self.generate_btn.setMinimumHeight(28)
        self.generate_btn.clicked.connect(self._on_generate)
        self.generate_btn.setEnabled(False)
        layout.addWidget(self.generate_btn)
        self.status_label = QLabel("")
        self.status_label.setFont(QFont('Mulish', 10, QFont.Medium))
        layout.addWidget(self.status_label)
        self.log_text = QTextEdit()
        self.log_text.setFont(QFont('Mulish', 9))
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(60)
        self.log_text.setProperty('logarea', True)
        layout.addWidget(self.log_text)
        # layout.addStretch()  # Remove excessive stretch
        return w

    def _select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", os.path.expanduser("~"))
        if folder:
            self.output_folder_line.setText(folder)
            self.output_folder = folder
            self._update_output_path()

    def _on_src_selected(self, path):
        self.source_path_full = path
        self._update_output_path()

    def _on_tgt_selected(self, path):
        self.target_path_full = path
        self._update_output_path()

    def _update_output_path(self):
        src = getattr(self, 'source_path_full', None)
        tgt = getattr(self, 'target_path_full', None)
        output_folder = getattr(self, 'output_folder', None)
        # Allow source and target to be the same
        if src and tgt and output_folder:
            src_name = os.path.splitext(os.path.basename(src))[0]
            tgt_name = os.path.splitext(os.path.basename(tgt))[0]
            out_name = f"mapping_{src_name}_to_{tgt_name}.xlsx"
            out_path = os.path.join(output_folder, out_name)
            self.output_path.setText(out_path)
            self.output_path_full = out_path
            self.generate_btn.setEnabled(True)
        else:
            self.output_path.clear()
            self.generate_btn.setEnabled(False)

    def _file_picker(self, line_edit, pick_cmd, label, icon, save=False):
        w = QWidget()
        l = QHBoxLayout(w)
        l.setContentsMargins(0, 0, 0, 0)
        icon_label = QLabel()
        icon_label.setPixmap(icon.pixmap(20, 20))
        l.addWidget(icon_label)
        lbl = QLabel(label)
        lbl.setFont(QFont('Mulish', 11, QFont.Bold))
        l.addWidget(lbl)
        l.addWidget(line_edit)
        browse_btn = QPushButton("Browse")
        browse_btn.setFont(QFont('Mulish', 10))
        browse_btn.clicked.connect(pick_cmd)
        l.addWidget(browse_btn)
        return w

    def _pick_source(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Source Schema", "", "XSD Files (*.xsd)")
        if path:
            self.source_path_full = path
            self.source_path.setText(self._truncate_path(path))

    def _pick_target(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Target Schema", "", "XSD Files (*.xsd)")
        if path:
            self.target_path_full = path
            self.target_path.setText(self._truncate_path(path))

    def _pick_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Select Output Mapping File", "", "Excel Files (*.xlsx)")
        if path:
            self.output_path_full = path
            self.output_path.setText(self._truncate_path(path))

    def _on_generate(self):
        self._log("[INFO] Generate Mapping clicked.", level="info")
        self.status_label.setText("Generating mapping...")
        self.status_label.setStyleSheet(f"color: {EDP_COLORS['electric_green']};")
        self.generate_btn.setEnabled(False)
        QApplication.instance().processEvents()
        try:
            src = self.source_path_full
            tgt = self.target_path_full
            out = self.output_path_full
            self._log(f"[INFO] Source: {src}", level="info")
            self._log(f"[INFO] Target: {tgt}", level="info")
            self._log(f"[INFO] Output: {out}", level="info")

            # Ensure output directory exists
            output_dir = os.path.dirname(out)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            # --- v4 logic: parse source and target, multi-message, row matching, column structure ---
            from microservices.xsd_parser_service import XSDParser
            import xml.etree.ElementTree as ET
            import openpyxl
            import difflib
            XSD_NS = '{http://www.w3.org/2001/XMLSchema}'
            parser = XSDParser()
            # Parse source XSD
            tree = ET.parse(src)
            root = tree.getroot()
            simple_types = {}
            for st in root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}simpleType'):
                name = st.get('name')
                if not name:
                    continue
                restriction = st.find('{http://www.w3.org/2001/XMLSchema}restriction')
                base = restriction.get('base') if restriction is not None else None
                restrictions = {}
                if restriction is not None:
                    for cons in restriction:
                        cons_name = cons.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                        val = cons.get('value')
                        if val:
                            restrictions[cons_name] = val
                simple_types[name] = {'base': base, 'restrictions': restrictions}
            complex_types = {ct.get('name'): ct for ct in root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}complexType') if ct.get('name')}
            src_messages = {}
            for elem in root.findall('{http://www.w3.org/2001/XMLSchema}element'):
                name = elem.get('name')
                if name:
                    rows = parser.parse_element(elem, complex_types, simple_types, 1, category='message')
                    src_messages[name] = rows
            # Parse target XSD (single sheet)
            tgt_rows = []
            if tgt and os.path.exists(tgt):
                tgt_tree = ET.parse(tgt)
                tgt_root = tgt_tree.getroot()
                tgt_simple_types = {}
                for st in tgt_root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}simpleType'):
                    name = st.get('name')
                    if not name:
                        continue
                    restriction = st.find('{http://www.w3.org/2001/XMLSchema}restriction')
                    base = restriction.get('base') if restriction is not None else None
                    restrictions = {}
                    if restriction is not None:
                        for cons in restriction:
                            cons_name = cons.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                            val = cons.get('value')
                            if val:
                                restrictions[cons_name] = val
                    tgt_simple_types[name] = {'base': base, 'restrictions': restrictions}
                tgt_complex_types = {ct.get('name'): ct for ct in tgt_root.findall(f'.//{{http://www.w3.org/2001/XMLSchema}}complexType') if ct.get('name')}
                for elem in tgt_root.findall('{http://www.w3.org/2001/XMLSchema}element'):
                    rows = parser.parse_element(elem, tgt_complex_types, tgt_simple_types, 1, category='message')
                    tgt_rows.extend(rows)

            # --- Remove case conversion logic ---
            # (No conversion of row['levels'] for source or target)

            # Build Excel file
            wb = openpyxl.Workbook()
            first = True
            for msg_name, src_full_rows in src_messages.items():
                if not first:
                    ws = wb.create_sheet(title=excel_sheet_name(msg_name))
                else:
                    ws = wb.active
                    ws.title = excel_sheet_name(msg_name)
                    first = False
                max_src_level = max((len(row['levels']) for row in src_full_rows), default=1)
                max_tgt_level = max((len(row['levels']) for row in tgt_rows), default=1) if tgt_rows else 1
                src_cols = [f'Level{i+1}_src' for i in range(max_src_level)] + ['Request Parameter_src', 'GDPR_src', 'Cardinality_src', 'Type_src', 'Base Type_src', 'Details_src', 'Description_src', 'Category_src', 'Example_src']
                tgt_cols = [f'Level{i+1}_tgt' for i in range(max_tgt_level)] + ['Request Parameter_tgt', 'GDPR_tgt', 'Cardinality_tgt', 'Type_tgt', 'Base Type_tgt', 'Details_tgt', 'Description_tgt', 'Category_tgt', 'Example_tgt']
                headers = src_cols + ['Destination Fields'] + tgt_cols
                ws.append(headers)
                ws.append([''] * len(headers))  # Second header row blank for now
                def row_path(row):
                    return '.'.join(row['levels'])
                tgt_path_dict = {row_path(row): row for row in tgt_rows}
                tgt_paths = list(tgt_path_dict.keys())
                for src_row in src_full_rows:
                    src_levels = src_row['levels'] + [''] * (max_src_level - len(src_row['levels']))
                    src_vals = src_levels + [
                        src_row.get('Request Parameter',''),
                        src_row.get('GDPR',''),
                        src_row.get('Cardinality',''),
                        src_row.get('Type',''),
                        src_row.get('Base Type',''),
                        src_row.get('Details',''),
                        src_row.get('Description',''),
                        src_row.get('Category','element'),
                        src_row.get('Example','')
                    ]
                    src_path_str = row_path(src_row)
                    tgt_row = tgt_path_dict.get(src_path_str)
                    best_match = ''
                    if not tgt_row and tgt_paths:
                        matches = difflib.get_close_matches(src_path_str, tgt_paths, n=1, cutoff=0.0)
                        if matches:
                            best_match = matches[0]
                            tgt_row = tgt_path_dict[best_match]
                    tgt_levels = tgt_row['levels'] + [''] * (max_tgt_level - len(tgt_row['levels'])) if tgt_row else ['']*max_tgt_level
                    tgt_vals = tgt_levels + [
                        tgt_row.get('Request Parameter','') if tgt_row else '',
                        tgt_row.get('GDPR','') if tgt_row else '',
                        tgt_row.get('Cardinality','') if tgt_row else '',
                        tgt_row.get('Type','') if tgt_row else '',
                        tgt_row.get('Base Type','') if tgt_row else '',
                        tgt_row.get('Details','') if tgt_row else '',
                        tgt_row.get('Description','') if tgt_row else '',
                        tgt_row.get('Category','element') if tgt_row else '',
                        tgt_row.get('Example','') if tgt_row else ''
                    ]
                    dest_field = '.'.join([lvl for lvl in tgt_row['levels'] if lvl]) if tgt_row else ''
                    ws.append(src_vals + [dest_field] + tgt_vals)
                # Prune unused source level columns
                def last_nonempty_level_col(start_col, num_levels):
                    for col in range(start_col + num_levels - 1, start_col - 1, -1):
                        for row in ws.iter_rows(min_row=3, min_col=col, max_col=col):
                            if any(cell.value not in (None, '') for cell in row):
                                return col
                    return start_col - 1
                src_level_start = 1
                last_src_col = last_nonempty_level_col(src_level_start, max_src_level)
                for col in range(src_level_start + max_src_level - 1, last_src_col, -1):
                    ws.delete_cols(col)
                header_row = [cell.value for cell in ws[1]]
                try:
                    tgt_level_start = header_row.index('Level1_tgt') + 1
                except ValueError:
                    tgt_level_start = len(header_row) + 1
                for col in range(tgt_level_start + max_tgt_level - 1, tgt_level_start - 1, -1):
                    col_letter = get_column_letter(col)
                    if col > tgt_level_start and all((ws.cell(row=row, column=col).value in (None, '')) for row in range(3, ws.max_row + 1)):
                        ws.delete_cols(col)
            wb.save(out)
            self._log(f"[SUCCESS] Output file created: {out}", level="info")
            # --- Post-processing QA: Excel Output Validator ---
            import threading
            def run_validator():
                from microservices.excel_output_validator import validate_excel_output, log_to_ui as validator_log_to_ui
                # Patch validator's log_to_ui to use the UI log via signal
                def ui_logger(msg):
                    if msg.startswith("✅") or msg.startswith("🔎"):
                        level = "info"
                    elif msg.startswith("❌"):
                        level = "error"
                    else:
                        level = "warning"
                    self.log_signal.emit(msg, level)
                validator_log_to_ui.__globals__['log_to_ui'] = ui_logger
                # Use source XSD for both source and target if only one is provided
                xsd_path = src if src else tgt
                try:
                    validate_excel_output(xsd_path, out)
                except Exception as e:
                    self.log_signal.emit(f"❌ Error in post-processing validator: {e}", "error")
            threading.Thread(target=run_validator, daemon=True).start()
            self._finish_generate()
        except Exception as e:
            import traceback
            error_msg = f"Mapping Error: {e}\n{traceback.format_exc()}"
            self._log(error_msg, level="error")
            self.status_label.setText("Mapping failed!")
            self.status_label.setStyleSheet("color: #FF3333;")
        self.generate_btn.setEnabled(True)

    def _finish_generate(self):
        self.status_label.setText("Mapping generated successfully!")
        self.status_label.setStyleSheet(f"color: {EDP_COLORS['electric_green']};")
        self._log("[INFO] Mapping generated.", level="info")

    def _log(self, msg, level="info"):
        color = {"info": EDP_COLORS['marine_blue_light'], "warning": EDP_COLORS['ice_blue'], "error": "#FF3333"}.get(level, EDP_COLORS['marine_blue_light'])
        self.log_text.setTextColor(Qt.black if level == "info" else Qt.red)
        self.log_text.append(msg)
        self.log_text.setTextColor(Qt.black)
        self.log_text.ensureCursorVisible()

    def _truncate_path(self, path, maxlen=60):
        if len(path) <= maxlen:
            return path
        return "..." + path[-(maxlen-3):]

    def _build_wsdl2xsd_section(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(8)
        title = QLabel("WSDL to XSD")
        title.setFont(QFont('Mulish', 14, QFont.Bold))
        layout.addWidget(title, alignment=Qt.AlignHCenter)
        subtitle = QLabel("Convert WSDL to XSD for EDP Schema Mapping.")
        subtitle.setFont(QFont('Mulish', 10))
        layout.addWidget(subtitle, alignment=Qt.AlignHCenter)
        # DropZone for WSDL input
        self.wsdl_drop = DropZone("Drop WSDL file here or Browse files", ".wsdl", self._on_wsdl_selected)
        self.wsdl_drop.setMinimumWidth(240)
        self.wsdl_drop.setMaximumWidth(400)
        self.wsdl_drop.setMinimumHeight(60)
        layout.addWidget(self.wsdl_drop, alignment=Qt.AlignHCenter)
        layout.addSpacing(6)
        # Splitter for WSDL input and XSD output
        splitter = QHBoxLayout()
        splitter.setSpacing(8)
        self.wsdl_input = QTextEdit()
        self.wsdl_input.setPlaceholderText("Paste WSDL content here or drop a file above...")
        self.wsdl_input.setMinimumWidth(180)
        self.wsdl_input.setFont(QFont('Mulish', 10))
        self.wsdl_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        splitter.addWidget(self.wsdl_input, stretch=1)
        self.xsd_output = QTextEdit()
        self.xsd_output.setPlaceholderText("Merged XSD will appear here...")
        self.xsd_output.setMinimumWidth(180)
        self.xsd_output.setFont(QFont('Mulish', 10))
        self.xsd_output.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        splitter.addWidget(self.xsd_output, stretch=1)
        layout.addLayout(splitter)
        layout.addSpacing(6)
        # Buttons row
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_row.addStretch(1)
        extract_btn = QPushButton("Extract XSD")
        extract_btn.setObjectName("primary")
        extract_btn.setMinimumWidth(90)
        extract_btn.setMinimumHeight(24)
        extract_btn.clicked.connect(self._on_extract_xsd)
        btn_row.addWidget(extract_btn)
        btn_row.addSpacing(8)
        download_btn = QPushButton("Download XSD")
        download_btn.setMinimumWidth(90)
        download_btn.setMinimumHeight(24)
        download_btn.clicked.connect(self._on_download_xsd)
        btn_row.addWidget(download_btn)
        btn_row.addStretch(1)
        layout.addLayout(btn_row)
        layout.addSpacing(6)
        # Log output area for this tab
        self.wsdl2xsd_log_text = QTextEdit()
        self.wsdl2xsd_log_text.setReadOnly(True)
        self.wsdl2xsd_log_text.setMinimumHeight(40)
        self.wsdl2xsd_log_text.setPlaceholderText("Log output for WSDL to XSD extraction...")
        self.wsdl2xsd_log_text.setProperty('logarea', True)
        layout.addWidget(self.wsdl2xsd_log_text)
        # layout.addStretch()  # Remove excessive stretch
        return w

    def log_to_wsdl_ui(self, message: str):
        if hasattr(self, 'wsdl2xsd_log_text') and self.wsdl2xsd_log_text:
            self.wsdl2xsd_log_text.append(message)

    def _on_wsdl_selected(self, path):
        # Load the WSDL file content into the input area
        try:
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            self.wsdl_input.setPlainText(content)
            self.log_to_wsdl_ui(f"✅ Loaded WSDL file: {os.path.basename(path)}")
        except Exception as e:
            self.log_to_wsdl_ui(f"❌ Error reading WSDL file: {e}")

    def _on_extract_xsd(self):
        wsdl_content = self.wsdl_input.toPlainText()
        if not wsdl_content.strip():
            self.log_to_wsdl_ui("❌ Error: Please paste WSDL content or drop a file.")
            return
        try:
            from microservices.wsdl_to_xsd_extractor import merge_xsd_from_wsdl
            merged_xsd = merge_xsd_from_wsdl(wsdl_content)
            self.xsd_output.setPlainText(merged_xsd)
            self.log_to_wsdl_ui("✅ XSD extraction successful.")
            # --- Compare fields logic ---
            # Extract all element/attribute names from generated XSD
            try:
                import xml.etree.ElementTree as ET
                xsd_root = ET.fromstring(merged_xsd)
                xsd_fields = set()
                for elem in xsd_root.iter():
                    if elem.tag.endswith('element') or elem.tag.endswith('attribute'):
                        name = elem.get('name')
                        if name:
                            xsd_fields.add(name)
                # Try to extract XSD from WSDL (look for <types> or <schema> section)
                import xml.etree.ElementTree as ET
                wsdl_root = ET.fromstring(wsdl_content)
                wsdl_fields = set()
                for schema in wsdl_root.iter():
                    if schema.tag.endswith('schema'):
                        for elem in schema.iter():
                            if elem.tag.endswith('element') or elem.tag.endswith('attribute'):
                                name = elem.get('name')
                                if name:
                                    wsdl_fields.add(name)
                missing_fields = wsdl_fields - xsd_fields
                if missing_fields:
                    for field in sorted(missing_fields):
                        self.log_to_wsdl_ui(f"❌ Missing field in generated XSD: {field}")
                    self.log_to_wsdl_ui(f"⚠️ {len(missing_fields)} field(s) missing in generated XSD compared to WSDL input.")
                else:
                    self.log_to_wsdl_ui("✅ All fields from WSDL are present in the generated XSD.")
            except Exception as e:
                self.log_to_wsdl_ui(f"⚠️ Field comparison failed: {e}")
        except Exception as e:
            import traceback
            error_msg = f"Extraction Error: {e}\n{traceback.format_exc()}"
            self.log_to_wsdl_ui(error_msg)

    def _on_download_xsd(self):
        xsd_content = self.xsd_output.toPlainText()
        if not xsd_content.strip():
            QMessageBox.warning(self, "No XSD", "No XSD to download. Please extract first.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save XSD File", "merged.xsd", "XSD Files (*.xsd)")
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(xsd_content)

    def _wrap_toolbar(self, widget):
        from PySide6.QtWidgets import QToolBar
        tb = QToolBar()
        tb.setMovable(False)
        tb.setFloatable(False)
        tb.addWidget(widget)
        return tb

def excel_sheet_name(name):
    return name[:31]

def main():
    app = QApplication(sys.argv)
    window = ForgeMainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main() 