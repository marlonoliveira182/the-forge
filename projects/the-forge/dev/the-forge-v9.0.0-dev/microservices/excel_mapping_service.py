import openpyxl

class ExcelMappingService:
    def __init__(self, max_structure_depth=8):
        self.max_structure_depth = max_structure_depth
        self.mapping_data = None
        self.headers = None
        self.sheet_names = []
        self.current_sheet = None
        self.header_rows = 2  # Support two header rows as in the template

    def list_sheets(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def load_mapping_file(self, file_path, sheet_name=None):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet_names = wb.sheetnames
        if sheet_name is None:
            sheet = wb.active
            self.current_sheet = sheet.title
        else:
            sheet = wb[sheet_name]
            self.current_sheet = sheet_name
        # Read two header rows
        header_rows = list(sheet.iter_rows(min_row=1, max_row=self.header_rows, values_only=True))
        self.headers = [f'{header_rows[0][i] or ""}\n{header_rows[1][i] or ""}' for i in range(len(header_rows[0]))]
        self.mapping_data = []
        for row in sheet.iter_rows(min_row=self.header_rows+1, values_only=True):
            row_dict = dict(zip(self.headers, row))
            self.mapping_data.append(row_dict)
        return self.mapping_data

    def get_structures(self, structure_type='source'):
        if not self.mapping_data:
            return []
        key = 'Source Structure' if structure_type == 'source' else 'Target Structure'
        return sorted(set(row.get(key) for row in self.mapping_data if row.get(key)))

    def get_mapping_rows(self, source_sheet=None, target_sheet=None):
        if not self.mapping_data:
            return []
        return self.mapping_data

    @staticmethod
    def flatten_schema(schema, parent_key='', sep='.'):
        # Recursively flatten a dict (JSON Schema or XSD structure) into dot notation paths
        items = []
        if isinstance(schema, dict):
            for k, v in schema.items():
                new_key = f'{parent_key}{sep}{k}' if parent_key else k
                if isinstance(v, dict):
                    items.extend(ExcelMappingService.flatten_schema(v, new_key, sep=sep))
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        items.extend(ExcelMappingService.flatten_schema(item, f'{new_key}[{i}]', sep=sep))
                else:
                    items.append(new_key)
        return items

    def generate_mapping_from_schemas(self, source_schema, target_schema, source_struct=None, target_struct=None):
        # source_schema/target_schema: dicts representing the schema (already loaded)
        # source_struct/target_struct: root keys to use (if any)
        # Returns: list of mapping rows (dicts)
        if source_struct:
            source_schema = source_schema.get(source_struct, source_schema)
        if target_struct:
            target_schema = target_schema.get(target_struct, target_schema)
        source_paths = self.flatten_schema(source_schema)
        target_paths = self.flatten_schema(target_schema)
        # For demo: map source to target by index (1:1), fill rest with blanks
        mapping_rows = []
        for i, src in enumerate(source_paths):
            tgt = target_paths[i] if i < len(target_paths) else ''
            mapping_rows.append({'Source Path': src, 'Target Path': tgt})
        return mapping_rows 