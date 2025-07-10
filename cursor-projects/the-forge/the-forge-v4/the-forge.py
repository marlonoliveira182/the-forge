import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, PatternFill
import re

# --- Funções para extração de caminhos de XSD e JSON Schema ---
def extract_paths_from_json_schema(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    paths = []
    def walk(obj, path=''):
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                new_path = f'{path}.{k}' if path else k
                paths.append(new_path)
                walk(v, new_path)
    walk(data)
    return paths

def extract_paths_from_xsd(filepath, keep_case=False):
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    paths = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    def to_camel_case(s):
        if not s:
            return s
        return s[0].lower() + s[1:]
    def get_name(name):
        return name if keep_case else to_camel_case(name)
    def walk_element(element, path='', is_root=False):
        name = element.get('name', '')
        if not name:
            return
        field_name = get_name(name)
        new_path = field_name if is_root else (f'{path}.{field_name}' if path else field_name)
        if not is_root:
        paths.append(new_path)
        typ = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        if typ in complex_types:
            walk_complex_type(complex_types[typ], new_path)
        if complex_type is not None:
            walk_complex_type(complex_type, new_path)
    def walk_complex_type(complex_type, path):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, path)
    for element in root.findall('xs:element', ns):
        walk_element(element, '', is_root=True)
    return paths

# --- Heurística de correspondência ---
try:
    from Levenshtein import ratio as levenshtein_ratio
except ImportError:
    def levenshtein_ratio(a, b):
        # Fallback simples: proporção de prefixo comum
        min_len = min(len(a), len(b))
        common = 0
        for i in range(min_len):
            if a[i] == b[i]:
                common += 1
            else:
                break
        return common / max(len(a), len(b)) if max(len(a), len(b)) > 0 else 0

def map_paths(source_paths, target_paths, threshold=0.7):
    mapping = []
    for s in source_paths:
        # Correspondência exata
        if s in target_paths:
            mapping.append({'source': s, 'target': s, 'similarity': 1.0})
            continue
        # Similaridade
        best = ('', 0.0)
        for t in target_paths:
            sim = levenshtein_ratio(s, t)
            if sim > best[1]:
                best = (t, sim)
        if best[1] >= threshold:
            mapping.append({'source': s, 'target': best[0], 'similarity': round(best[1], 3)})
        else:
            mapping.append({'source': s, 'target': '', 'similarity': 0.0})
    return mapping

# --- Extração detalhada de campos ---
def extract_fields_from_json_schema(filepath):
    import json
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    fields = []
    def walk(obj, levels=None, level=0):
        if levels is None:
            levels = []
        if isinstance(obj, dict):
            for k, v in obj.get('properties', {}).items():
                lvls = levels + [k]  # Use property name as-is
                typ = v.get('type', '')
                desc = v.get('description', '')
                card = '1' if k in obj.get('required', []) else '0..1'
                details = []
                if 'enum' in v:
                    details.append(f"enum: {v['enum']}")
                if 'pattern' in v:
                    details.append(f"pattern: {v['pattern']}")
                if 'minLength' in v:
                    details.append(f"minLength: {v['minLength']}")
                if 'maxLength' in v:
                    details.append(f"maxLength: {v['maxLength']}")
                if 'minimum' in v:
                    details.append(f"minimum: {v['minimum']}")
                if 'maximum' in v:
                    details.append(f"maximum: {v['maximum']}")
                if 'format' in v:
                    details.append(f"format: {v['format']}")
                details_str = '; '.join(details)
                fields.append({
                    'levels': lvls,
                    'Type': typ,
                    'Description': desc,
                    'Cardinality': card,
                    'Details': details_str
                })
                walk(v, lvls, level+1)
    walk(data)
    return fields

def extract_xsd_restrictions(restriction):
    restr_details = []
    for r in restriction:
        tag = r.tag.split('}')[-1]
        val = r.get('value')
        if tag == 'enumeration' and val is not None:
            restr_details.append(f'enum: {val}')
        elif val is not None:
            restr_details.append(f'{tag}: {val}')
    return '; '.join(restr_details)

def extract_fields_from_xsd(filepath, keep_case=False):
    import xml.etree.ElementTree as ET
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    fields = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
    def to_camel_case(s):
        if not s:
            return s
        return s[0].lower() + s[1:]
    def get_name(name):
        return name if keep_case else to_camel_case(name)
    def get_type_and_details(element):
        typ = element.get('type', '')
        details = ''
        real_type = typ
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        elif typ in simple_types:
            restriction = simple_types[typ].find('xs:restriction', ns)
            if restriction is not None:
                real_type = restriction.get('base', typ)
                details = extract_xsd_restrictions(restriction)
        elif element.find('xs:complexType', ns) is not None:
            real_type = 'object'
        return real_type, details
    def walk_element(element, levels=None, omit_root=False):
        if levels is None:
            levels = []
        name = element.get('name', '')
        if not name:
            return
        field_name = get_name(name)
        lvls = levels if omit_root else levels + [field_name]
        desc = ''
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        card = f"{min_occurs}..{max_occurs}" if min_occurs != max_occurs else min_occurs
        typ, details = get_type_and_details(element)
        if not omit_root:
        fields.append({
            'levels': lvls,
            'Type': typ,
            'Description': desc,
            'Cardinality': card,
            'Details': details
        })
        typ_ref = element.get('type', '')
        complex_type = element.find('xs:complexType', ns)
        if typ_ref in complex_types:
            walk_complex_type(complex_types[typ_ref], lvls)
        if complex_type is not None:
            walk_complex_type(complex_type, lvls)
    def walk_complex_type(complex_type, levels):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, levels)
    for element in root.findall('xs:element', ns):
        walk_element(element, [], omit_root=True)
    return fields

def fields_to_row(fields, max_levels):
    # Converte lista de dicts para lista de listas (linhas), preenchendo níveis
    rows = []
    for f in fields:
        lvls = f['levels']
        row = lvls + [''] * (max_levels - len(lvls))
        row += [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return rows

def get_max_levels(fields):
    return max((len(f['levels']) for f in fields), default=1)

def build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path):
    # Descobrir o máximo de níveis
    src_max_levels = get_max_levels(source_fields)
    tgt_max_levels = get_max_levels(target_fields)
    # v3-style headers
    src_headers = [f'Element Level {i+1}' for i in range(src_max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    tgt_headers = [f'Element Level {i+1}' for i in range(tgt_max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(out_headers)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    def get_req_param(fields):
        for f in fields:
            if len(f['levels']) == 1 and f.get('Type') == 'object':
                return 'body (json)' if f['levels'][0].endswith('.json') or f['levels'][0].endswith('json') else 'body (xml)'
        return 'body (json)' if tgt_name.endswith('.json') else 'body (xml)'
    src_req_param = 'body (json)' if src_name.endswith('.json') else 'body (xml)'
    tgt_req_param = 'body (json)' if tgt_name.endswith('.json') else 'body (xml)'
    # Index target fields by full path string (lowercase for comparison)
    tgt_path_map = {'.'.join(f['levels']).lower(): f for f in target_fields}
    prev_src_levels = [''] * src_max_levels
    prev_tgt_levels = [''] * tgt_max_levels
    for i, src in enumerate(source_fields):
        src_path = '.'.join(src['levels'])
        tgt = tgt_path_map.get(src_path.lower(), None)
        # Não repetir parent nos níveis
        src_lvls = src['levels'] + [''] * (src_max_levels - len(src['levels']))
        src_lvls_disp = []
        for idx, val in enumerate(src_lvls):
            src_lvls_disp.append(val if val != prev_src_levels[idx] else '')
        prev_src_levels = [v if v else p for v, p in zip(src_lvls, prev_src_levels)]
        # v3-style row for source
        src_row = (
            src_lvls_disp +
            [src_req_param, '', src.get('Cardinality',''), src.get('Type',''), src.get('Details',''), src.get('Description','')]
        )
        # Preencher o destination field com o caminho do target (mesmo case do target)
        dest_field = '.'.join(tgt['levels']) if tgt else ''
        if tgt:
            tgt_lvls = tgt['levels'] + [''] * (tgt_max_levels - len(tgt['levels']))
            tgt_lvls_disp = []
            for idx, val in enumerate(tgt_lvls):
                tgt_lvls_disp.append(val if val != prev_tgt_levels[idx] else '')
            prev_tgt_levels = [v if v else p for v, p in zip(tgt_lvls, prev_tgt_levels)]
            tgt_row = (
                tgt_lvls_disp +
                [tgt_req_param, '', tgt.get('Cardinality',''), tgt.get('Type',''), tgt.get('Details',''), tgt.get('Description','')]
            )
        else:
            tgt_row = [''] * (tgt_max_levels + 6)
        ws.append(src_row + [dest_field] + tgt_row)
    wb.save(output_path)

# --- Geração do arquivo Excel ---
def write_mapping_excel(mapping, output_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Source Path', 'Target Path', 'Similarity'])
    for row in mapping:
        ws.append([row['source'], row['target'], row['similarity']])
    wb.save(output_path)

def parse_xsd(filepath):
    fields = extract_fields_from_xsd(filepath)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + ['Type', 'Description', 'Cardinality']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels'] + [''] * (max_levels - len(f['levels']))
        lvls_disp = [val if val != prev_levels[idx] else '' for idx, val in enumerate(lvls)]
        prev_levels = [v if v else p for v, p in zip(lvls, prev_levels)]
        row = lvls_disp + [f.get('Type',''), f.get('Description',''), f.get('Cardinality','')]
        rows.append(row)
    return columns, rows

def run_mapping_operation(source_path, target_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    target_ext = os.path.splitext(target_path)[1].lower()
    if source_ext == '.json':
        source_paths = extract_paths_from_json_schema(source_path)
        source_fields = extract_fields_from_json_schema(source_path)
    else:
        source_paths = extract_paths_from_xsd(source_path, keep_case=True)
        source_fields = extract_fields_from_xsd(source_path, keep_case=True)
    if target_ext == '.json':
        target_paths = extract_paths_from_json_schema(target_path)
        target_fields = extract_fields_from_json_schema(target_path)
    else:
        target_paths = extract_paths_from_xsd(target_path, keep_case=True)
        target_fields = extract_fields_from_xsd(target_path, keep_case=True)
    mapping = map_paths(source_paths, target_paths)
    src_name = os.path.splitext(os.path.basename(source_path))[0]
    tgt_name = os.path.splitext(os.path.basename(target_path))[0]
    output_path = os.path.join(dest_dir, f"mapping_{src_name}_to_{tgt_name}.xlsx")
    build_mapping_v2_style(source_fields, target_fields, mapping, src_name, tgt_name, output_path)
    print(f"Mapping file salvo em:\n{output_path}")

def write_excel(rows, output_file, columns):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    # Formatar header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    # Mesclar células de níveis (apenas as células vazias abaixo do pai)
    num_rows = len(rows)
    level_cols = [i+1 for i, col in enumerate(columns) if col.startswith('Element Level')]
    for col in level_cols:
        row = 2
        while row <= num_rows+1:
            if ws.cell(row=row, column=col).value:
                # Encontrar bloco de células vazias abaixo
                start = row + 1
                end = start
                while end <= num_rows+1 and not ws.cell(row=end, column=col).value:
                    end += 1
                if end-1 > start:
                    ws.merge_cells(start_row=start, start_column=col, end_row=end-1, end_column=col)
                row = end
            else:
                row += 1
    wb.save(output_file)

def run_schema_to_excel_operation(source_path, dest_dir):
    source_ext = os.path.splitext(source_path)[1].lower()
    is_json = source_ext == '.json'
    fields = extract_fields_from_json_schema(source_path) if is_json else extract_fields_from_xsd(source_path)
    max_levels = get_max_levels(fields)
    columns = [f'Element Level {i+1}' for i in range(max_levels)] + [
        'Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description']
    rows = []
    prev_levels = [''] * max_levels
    for f in fields:
        lvls = f['levels']
        # Preencher colunas de níveis, só mostrar valor se diferente do anterior
        level_cells = []
        for idx in range(max_levels):
            val = lvls[idx] if idx < len(lvls) else ''
            disp = val if val != prev_levels[idx] else ''
            level_cells.append(disp)
        prev_levels = [v if v else p for v, p in zip([lvls[i] if i < len(lvls) else '' for i in range(max_levels)], prev_levels)]
        req_param = 'body (json)' if is_json else 'body (xml)'
        row = (
            level_cells +
            [req_param, '', f.get('Cardinality',''), f.get('Type',''), f.get('Details',''), f.get('Description','')]
        )
        rows.append(row)
    src_name = os.path.splitext(os.path.basename(source_path))[0]
    output_path = os.path.join(dest_dir, f"schema_{src_name}.xlsx")
    write_excel(rows, output_path, columns)
    # messagebox.showinfo("Concluído", f"Arquivo Excel salvo em:\n{output_path}")
    print(f"Arquivo Excel salvo em:\n{output_path}")

def to_camel_case(s):
    if not s:
        return s
    return s[0].lower() + s[1:]

def run_xsd_to_jsonschema_operation(source_path, dest_dir):
    import xml.etree.ElementTree as ET
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    tree = ET.parse(source_path)
    root = tree.getroot()
    simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}

    def xsd_type_to_json_type(xsd_type):
        if xsd_type.endswith(':string') or xsd_type == 'string':
            return 'string'
        elif xsd_type.endswith(':int') or xsd_type == 'int' or xsd_type.endswith(':integer') or xsd_type == 'integer':
            return 'integer'
        elif xsd_type.endswith(':boolean') or xsd_type == 'boolean':
            return 'boolean'
        elif xsd_type.endswith(':decimal') or xsd_type == 'decimal' or xsd_type.endswith(':float') or xsd_type == 'float':
            return 'number'
        else:
            return 'string'

    def extract_restrictions(element):
        details = {}
        restriction = element.find('xs:restriction', ns)
        if restriction is not None:
            for r in restriction:
                tag = r.tag.split('}')[-1]
                val = r.get('value')
                if tag == 'enumeration' and val is not None:
                    details.setdefault('enum', []).append(val)
                elif val is not None:
                    details[tag] = val
        return details

    def parse_element(element):
        name = to_camel_case(element.get('name', ''))
        typ = element.get('type', '')
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        # Extract description from annotation/documentation
        desc = ''
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        # Se for complexType referenciado
        if typ in complex_types:
            schema = parse_complex_type(complex_types[typ])
            if desc:
                schema['description'] = desc
            return name, schema
        # Se for complexType inline
        complex_type = element.find('xs:complexType', ns)
        if complex_type is not None:
            schema = parse_complex_type(complex_type)
            if desc:
                schema['description'] = desc
            return name, schema
        # Se for simpleType inline
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            base_type = restriction.get('base') if restriction is not None else 'string'
            prop = {'type': xsd_type_to_json_type(base_type)}
            details = extract_restrictions(simple_type)
            if 'enum' in details:
                prop['enum'] = details['enum']
            if 'pattern' in details:
                prop['pattern'] = details['pattern']
            if 'minLength' in details:
                prop['minLength'] = int(details['minLength'])
            if 'maxLength' in details:
                prop['maxLength'] = int(details['maxLength'])
            if desc:
                prop['description'] = desc
            return name, prop
        # Se for simpleType referenciado
        if typ in simple_types:
            restriction = simple_types[typ].find('xs:restriction', ns)
            base_type = restriction.get('base') if restriction is not None else 'string'
            prop = {'type': xsd_type_to_json_type(base_type)}
            details = extract_restrictions(simple_types[typ])
            if 'enum' in details:
                prop['enum'] = details['enum']
            if 'pattern' in details:
                prop['pattern'] = details['pattern']
            if 'minLength' in details:
                prop['minLength'] = int(details['minLength'])
            if 'maxLength' in details:
                prop['maxLength'] = int(details['maxLength'])
            if desc:
                prop['description'] = desc
            return name, prop
        # Se for tipo simples direto
        prop = {'type': xsd_type_to_json_type(typ)} if typ else {'type': 'string'}
        if desc:
            prop['description'] = desc
        return name, prop

    def parse_complex_type(complex_type):
        props = {}
        required = []
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                child_name, child_schema = parse_element(child)
                # Garantir camelCase recursivo
                child_name = to_camel_case(child_name)
                if child_schema.get('type') == 'object' and 'properties' in child_schema:
                    # Recursivamente garantir camelCase nas propriedades aninhadas
                    child_schema['properties'] = {to_camel_case(k): v for k, v in child_schema['properties'].items()}
                props[child_name] = child_schema
                min_occurs = child.get('minOccurs', '1')
                if min_occurs == '1':
                    required.append(child_name)
        return_schema = {'type': 'object', 'properties': props}
        if required:
            return_schema['required'] = required
        return return_schema

    def camelize_json_schema(obj):
        if isinstance(obj, dict):
            new_obj = {}
            for k, v in obj.items():
                # Always keep JSON Schema meta keys as-is
                if k in ('$schema', 'title', 'type', 'required', 'enum', 'pattern', 'minLength', 'maxLength', 'minimum', 'maximum', 'format', 'description', 'allOf', 'anyOf', 'oneOf', 'not', 'default', 'examples'):
                    new_obj[k] = camelize_json_schema(v)
                elif k == 'properties' and isinstance(v, dict):
                    # Recursively camelCase property keys and their values
                    new_obj[k] = {to_camel_case(pk): camelize_json_schema(pv) for pk, pv in v.items()}
                elif k == 'items' and isinstance(v, dict):
                    new_obj[k] = camelize_json_schema(v)
                else:
                    new_obj[to_camel_case(k)] = camelize_json_schema(v)
            return new_obj
        elif isinstance(obj, list):
            return [camelize_json_schema(i) for i in obj]
        else:
            return obj

    # --- Validation: collect all XSD element names (in camelCase) ---
    def collect_xsd_element_names(element, names=None):
        if names is None:
            names = set()
        name = element.get('name')
        if name:
            names.add(to_camel_case(name))
        # Check for children
        ns_seq = element.find('xs:complexType/xs:sequence', ns)
        if ns_seq is not None:
            for child in ns_seq.findall('xs:element', ns):
                collect_xsd_element_names(child, names)
        # Also check for direct xs:sequence (for complexType elements)
        seq = element.find('xs:sequence', ns)
        if seq is not None:
            for child in seq.findall('xs:element', ns):
                collect_xsd_element_names(child, names)
        return names

    # --- Validation: collect all property names from JSON Schema ---
    def collect_json_schema_property_names(schema, names=None):
        if names is None:
            names = set()
        if isinstance(schema, dict):
            for k, v in schema.items():
                if k == 'properties' and isinstance(v, dict):
                    for pk, pv in v.items():
                        names.add(pk)
                        collect_json_schema_property_names(pv, names)
                else:
                    collect_json_schema_property_names(v, names)
        elif isinstance(schema, list):
            for item in schema:
                collect_json_schema_property_names(item, names)
        return names

    # Encontrar o elemento root
    root_element = None
    for el in root.findall('xs:element', ns):
        root_element = el
        break
    if root_element is None:
        # messagebox.showerror("Erro", "Não foi possível encontrar o elemento root no XSD.")
        print("[ERRO] Não foi possível encontrar o elemento root no XSD.")
        return
    root_name = to_camel_case(root_element.get('name', 'Root'))
    root_schema = parse_element(root_element)[1]
    json_schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": root_name,
        **root_schema
    }
    # Recursively camelCase all property names in the schema
    json_schema = camelize_json_schema(json_schema)

    # --- Validation step ---
    xsd_names = collect_xsd_element_names(root_element)
    # Remove the root element name from the set
    root_camel = to_camel_case(root_element.get('name', ''))
    if root_camel in xsd_names:
        xsd_names.remove(root_camel)
    json_names = collect_json_schema_property_names(json_schema)
    missing = sorted(xsd_names - json_names)
    if missing:
        # messagebox.showerror("Erro de Validação", f"Os seguintes campos do XSD não foram encontrados no JSON Schema gerado (camelCase):\n\n" + "\n".join(missing))
        print("[ERRO] Os seguintes campos do XSD não foram encontrados no JSON Schema gerado (camelCase):\n\n" + "\n".join(missing))
        return

    out_name = root_name + '.schema.json'
    output_path = os.path.join(dest_dir, out_name)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_schema, f, indent=2)
    print('[DEBUG] JSON Schema salvo com sucesso')
    # messagebox.showinfo("Concluído", f"JSON Schema salvo em:\n{output_path}")
    print(f"[INFO] JSON Schema salvo em: {output_path}")

def jsonschema_to_xsd(json_schema_path, dest_dir):
    with open(json_schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    def to_pascal_case(s):
        return s[0].upper() + s[1:] if s else s
    def gen_xsd_element(name, obj):
        t = obj.get('type', 'string')
        if t == 'object':
            props = obj.get('properties', {})
            children = [gen_xsd_element(k, v) for k, v in props.items()]
            return f'<xs:element name="{to_pascal_case(name)}"><xs:complexType><xs:sequence>' + ''.join(children) + '</xs:sequence></xs:complexType></xs:element>'
        elif t == 'array':
            items = obj.get('items', {})
            return f'<xs:element name="{to_pascal_case(name)}" minOccurs="0" maxOccurs="unbounded">' + gen_xsd_element(name, items) + '</xs:element>'
        elif t == 'string':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:string"/>'
        elif t == 'integer':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:int"/>'
        elif t == 'number':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:decimal"/>'
        elif t == 'boolean':
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:boolean"/>'
        else:
            return f'<xs:element name="{to_pascal_case(name)}" type="xs:string"/>'
    title = schema.get('title', 'Root')
    xsd = f'<?xml version="1.0" encoding="UTF-8"?>\n<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">' + gen_xsd_element(title, schema) + '</xs:schema>'
    out_name = os.path.splitext(os.path.basename(json_schema_path))[0] + '.fromjson.xsd'
    output_path = os.path.join(dest_dir, out_name)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(xsd)

def main():
    class UnifiedTool(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("The Forge - Unified Tool")
            self.geometry("650x350")
            self.resizable(False, False)
            self.source_path = tk.StringVar()
            self.target_path = tk.StringVar()
            self.dest_dir = tk.StringVar()
            self.op_map = {
                "Gerar Mapping entre Schemas (XSD/JSON)": "mapping",
                "Converter Schema para Excel": "schema_to_excel",
                "Converter XSD para JSON Schema": "xsd_to_jsonschema",
                "Converter JSON Schema para XSD": "jsonschema_to_xsd"
            }
            self.operation = tk.StringVar(value="Selecione a operação...")
            op_combo = ttk.Combobox(self, textvariable=self.operation, state="readonly", width=40)
            op_combo['values'] = ("Selecione a operação...",) + tuple(self.op_map.keys())
            op_combo.pack(pady=(0,10))
            op_combo.bind("<<ComboboxSelected>>", self.update_fields)
            self.frame_fields = tk.Frame(self)
            self.frame_fields.pack(fill='x', padx=10)
            self.build_fields()
            tk.Button(self, text="Executar", command=self.on_continue).pack(pady=15)
            self.result = None
        def build_fields(self):
            for widget in self.frame_fields.winfo_children():
                widget.destroy()
            op = self.op_map.get(self.operation.get(), None)
            if op == "mapping":
                tk.Label(self.frame_fields, text="Source schema (XSD ou JSON Schema):").pack(pady=(5,0))
                frame1 = tk.Frame(self.frame_fields)
                frame1.pack(fill='x')
                tk.Entry(frame1, textvariable=self.source_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame1, text="Procurar", command=self.browse_source).pack(side='left', padx=5)
                tk.Label(self.frame_fields, text="Target schema (XSD ou JSON Schema):").pack(pady=(5,0))
                frame2 = tk.Frame(self.frame_fields)
                frame2.pack(fill='x')
                tk.Entry(frame2, textvariable=self.target_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame2, text="Procurar", command=self.browse_target).pack(side='left', padx=5)
            elif op == "schema_to_excel" or op == "xsd_to_jsonschema" or op == "jsonschema_to_xsd":
                label = {
                    "schema_to_excel": "Schema de entrada (XSD ou JSON Schema):",
                    "xsd_to_jsonschema": "XSD de entrada:",
                    "jsonschema_to_xsd": "JSON Schema de entrada:"
                }[op]
                tk.Label(self.frame_fields, text=label).pack(pady=(5,0))
                frame1 = tk.Frame(self.frame_fields)
                frame1.pack(fill='x')
                tk.Entry(frame1, textvariable=self.source_path, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame1, text="Procurar", command=self.browse_source).pack(side='left', padx=5)
            if op:
                tk.Label(self.frame_fields, text="Diretório de saída:").pack(pady=(5,0))
                frame3 = tk.Frame(self.frame_fields)
                frame3.pack(fill='x')
                tk.Entry(frame3, textvariable=self.dest_dir, width=60, state='readonly').pack(side='left', expand=True, fill='x')
                tk.Button(frame3, text="Procurar", command=self.browse_dest).pack(side='left', padx=5)
        def update_fields(self, event=None):
            self.build_fields()
        def browse_source(self):
            path = filedialog.askopenfilename(title="Selecione o schema SOURCE", filetypes=[("Schemas", "*.xsd *.json"), ("Todos", "*.*")])
            if path:
                self.source_path.set(path)
        def browse_target(self):
            path = filedialog.askopenfilename(title="Selecione o schema TARGET", filetypes=[("Schemas", "*.xsd *.json"), ("Todos", "*.*")])
            if path:
                self.target_path.set(path)
        def browse_dest(self):
            path = filedialog.askdirectory(title="Selecione o diretório de saída")
            if path:
                self.dest_dir.set(path)
        def on_continue(self):
            op = self.op_map.get(self.operation.get(), None)
            if not op:
                messagebox.showerror("Erro", "Selecione a operação desejada.")
                return
            if not self.source_path.get() or not self.dest_dir.get() or (op=="mapping" and not self.target_path.get()):
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios.")
                return
            self.result = (op, self.source_path.get(), self.target_path.get() if op=="mapping" else None, self.dest_dir.get())
            self.destroy()

    selector = UnifiedTool()
    selector.mainloop()
    if not selector.result:
        print("Operação cancelada.")
        return
    op, source_path, target_path, dest_dir = selector.result
    if op == "mapping":
        run_mapping_operation(source_path, target_path, dest_dir)
    elif op == "schema_to_excel":
        run_schema_to_excel_operation(source_path, dest_dir)
    elif op == "xsd_to_jsonschema":
        run_xsd_to_jsonschema_operation(source_path, dest_dir)
    elif op == "jsonschema_to_xsd":
        jsonschema_to_xsd(source_path, dest_dir)

if __name__ == "__main__":
    main() 