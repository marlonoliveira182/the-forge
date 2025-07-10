import argparse
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox

SUPPORTED_TYPES = ['.xsd', '.json', '.xml']

# --- INÍCIO DO AJUSTE ---
# Defina aqui os diretórios de source e target
# SOURCE_DIR = r'C:\Users\marlo\OneDrive - EDP\Documents\the-forge\converter\source'
# TARGET_DIR = r'C:\Users\marlo\OneDrive - EDP\Documents\the-forge\converter\target'
# --- FIM DO AJUSTE ---

import json
from openpyxl import load_workbook, Workbook
import xml.etree.ElementTree as ET

COLUMNS = [
    'Element',
    'Request Parameter',
    'GDPR',
    'Cardinality',
    'Type',
    'Details',
    'Description',
]

def detect_file_type(filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.xsd' or ext == '.xml':
        return ext
    if ext == '.json':
        with open(filename, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
                if isinstance(data, dict) and ('$schema' in data or 'properties' in data):
                    return 'jsonschema'
                else:
                    return 'jsonexample'
            except Exception:
                return None
    return None

def parse_json_schema(data, parent='', level=0):
    rows = []
    required = set(data.get('required', []))
    properties = data.get('properties', {})
    for name, prop in properties.items():
        element = ('    ' * level) + name
        req_param = name
        gdpr = ''
        cardinality = '1' if name in required else '0..1'
        typ = prop.get('type', '')
        details = ''
        if 'enum' in prop:
            details = f"enum: {prop['enum']}"
        elif 'pattern' in prop:
            details = f"pattern: {prop['pattern']}"
        desc = prop.get('description', '')
        rows.append([element, req_param, gdpr, cardinality, typ, details, desc])
        # Nested objects
        if typ == 'object':
            rows += parse_json_schema(prop, parent=name, level=level+1)
        # Arrays
        if typ == 'array' and 'items' in prop:
            item_type = prop['items'].get('type', '')
            rows.append([
                ('    ' * (level+1)) + '[item]', '', '', '', item_type, '', ''
            ])
            if item_type == 'object':
                rows += parse_json_schema(prop['items'], parent=name, level=level+2)
    return rows

def parse_json_example(data, parent='', level=0):
    rows = []
    if isinstance(data, dict):
        for k, v in data.items():
            element = ('    ' * level) + k
            req_param = k
            gdpr = ''
            cardinality = ''
            typ = type(v).__name__
            details = ''
            desc = ''
            rows.append([element, req_param, gdpr, cardinality, typ, details, desc])
            rows += parse_json_example(v, parent=k, level=level+1)
    elif isinstance(data, list):
        element = ('    ' * level) + '[item]'
        req_param = ''
        gdpr = ''
        cardinality = ''
        typ = 'array'
        details = ''
        desc = ''
        rows.append([element, req_param, gdpr, cardinality, typ, details, desc])
        if data:
            rows += parse_json_example(data[0], parent=parent, level=level+1)
    return rows

def parse_xsd(filename):
    tree = ET.parse(filename)
    root = tree.getroot()
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    rows = []
    complex_types = {ct.get('name'): ct for ct in root.findall('xs:complexType', ns)}
    max_depth = [0]
    def walk_element(element, path, level=0):
        name = element.get('name', '')
        typ = element.get('type', '')
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        cardinality = f"{min_occurs}..{max_occurs}" if min_occurs != max_occurs else min_occurs
        details = ''
        desc = ''
        real_type = typ
        annotation = element.find('xs:annotation/xs:documentation', ns)
        if annotation is not None and annotation.text:
            desc = annotation.text.strip()
        # Extrair restrições se houver
        restriction = None
        # Se o tipo for inline simpleType
        simple_type = element.find('xs:simpleType', ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', ns)
            if restriction is not None:
                base = restriction.get('base')
                if base:
                    real_type = base
        # Se o tipo for referenciado, buscar restrição no schema
        if restriction is None and typ:
            simple_types = {st.get('name'): st for st in root.findall('xs:simpleType', ns)}
            if typ in simple_types:
                restriction = simple_types[typ].find('xs:restriction', ns)
                if restriction is not None:
                    base = restriction.get('base')
                    if base:
                        real_type = base
        # Se for complexType inline ou referenciado
        is_object = False
        if typ in complex_types:
            is_object = True
            if not real_type:
                real_type = typ
        complex_type = element.find('xs:complexType', ns)
        if complex_type is not None:
            is_object = True
            if not real_type:
                real_type = 'object'
        if is_object and not real_type:
            real_type = 'object'
        # Extrair detalhes das restrições
        if restriction is not None:
            restr_details = []
            for r in restriction:
                tag = r.tag.split('}')[-1]
                val = r.get('value')
                if val is not None:
                    restr_details.append(f'{tag}="{val}"')
            details = '; '.join(restr_details)
        new_path = path + [name]
        if level > max_depth[0]:
            max_depth[0] = level
        row = [''] * (level) + [name] + [''] * (10 - (level+1))
        row += ['body (xml)', '', cardinality, real_type, details, desc]
        rows.append((list(new_path), row, level))
        if typ in complex_types:
            walk_complex_type(complex_types[typ], new_path, level+1)
        if complex_type is not None:
            walk_complex_type(complex_type, new_path, level+1)
    def walk_complex_type(complex_type, path, level=0):
        sequence = complex_type.find('xs:sequence', ns)
        if sequence is not None:
            for child in sequence.findall('xs:element', ns):
                walk_element(child, path, level)
    for element in root.findall('xs:element', ns):
        walk_element(element, [], 0)
    # Ajustar colunas dinamicamente
    max_level = max((level for _, _, level in rows), default=0)
    element_headers = [f'Element Level {i+1}' for i in range(max_level+1)]
    final_rows = []
    prev_cells = [''] * (max_level+1)
    for path, row, level in rows:
        # Ajustar o tamanho da linha para o número de níveis
        element_cells = [''] * (max_level+1)
        for i, v in enumerate(path):
            element_cells[i] = v
        # Esconder valores repetidos
        for i in range(max_level+1):
            if element_cells[i] == prev_cells[i]:
                element_cells[i] = ''
        prev_cells = [ec if ec else pc for ec, pc in zip(element_cells, prev_cells)]
        final_rows.append(element_cells + row[(10):])
    return element_headers + ['Request Parameter', 'GDPR', 'Cardinality', 'Type', 'Details', 'Description'], final_rows

def parse_xml_example(filename):
    tree = ET.parse(filename)
    root = tree.getroot()
    rows = []
    def walk(element, level=0):
        name = element.tag
        typ = 'object' if list(element) else 'string'
        element_name = ('    ' * level) + name
        rows.append([element_name, name, '', '', typ, '', ''])
        for child in element:
            walk(child, level+1)
    walk(root)
    return rows

def write_excel(rows, output_file, columns):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(output_file)

def double_check(rows, columns):
    issues = []
    name_cols = [i for i, col in enumerate(columns) if col.startswith('Element Level')]
    type_col = columns.index('Type') if 'Type' in columns else -1
    card_col = columns.index('Cardinality') if 'Cardinality' in columns else -1
    for idx, row in enumerate(rows, 2):  # 2 = header + 1-based
        # Nome do campo: último não vazio das colunas de hierarquia
        name = ''
        for i in reversed(name_cols):
            if row[i]:
                name = row[i]
                break
        typ = row[type_col] if type_col >= 0 else ''
        card = row[card_col] if card_col >= 0 else ''
        if not name:
            issues.append(f"Linha {idx}: Campo sem nome.")
        if not typ:
            issues.append(f"Linha {idx}: Campo '{name}' sem tipo.")
        if not card:
            issues.append(f"Linha {idx}: Campo '{name}' sem cardinalidade.")
    return issues

def main():
    class FileDirSelector(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Select Source File and Target Directory")
            self.geometry("500x180")
            self.resizable(False, False)
            self.source_path = tk.StringVar()
            self.target_dir = tk.StringVar()
            tk.Label(self, text="Source file (XSD, JSON, XML):").pack(pady=(10,0))
            frame1 = tk.Frame(self)
            frame1.pack(fill='x', padx=10)
            tk.Entry(frame1, textvariable=self.source_path, width=50, state='readonly').pack(side='left', expand=True, fill='x')
            tk.Button(frame1, text="Browse", command=self.browse_source).pack(side='left', padx=5)
            tk.Label(self, text="Target directory:").pack(pady=(10,0))
            frame2 = tk.Frame(self)
            frame2.pack(fill='x', padx=10)
            tk.Entry(frame2, textvariable=self.target_dir, width=50, state='readonly').pack(side='left', expand=True, fill='x')
            tk.Button(frame2, text="Browse", command=self.browse_target).pack(side='left', padx=5)
            tk.Button(self, text="Continue", command=self.on_continue).pack(pady=15)
            self.result = None
        def browse_source(self):
            path = filedialog.askopenfilename(title="Select source file", filetypes=[("Supported files", "*.xsd *.json *.xml"), ("JSON", "*.json"), ("XSD", "*.xsd"), ("XML", "*.xml"), ("All files", "*.*")])
            if path:
                self.source_path.set(path)
        def browse_target(self):
            path = filedialog.askdirectory(title="Select target directory")
            if path:
                self.target_dir.set(path)
        def on_continue(self):
            if not self.source_path.get() or not self.target_dir.get():
                messagebox.showerror("Error", "Please select both source file and target directory.")
                return
            self.result = (self.source_path.get(), self.target_dir.get())
            self.destroy()

    selector = FileDirSelector()
    selector.mainloop()
    if not selector.result:
        print("Operation cancelled.")
        return
    input_path, target_dir = selector.result
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(target_dir, base_name + ".xlsx")
    file_type = detect_file_type(input_path)
    if not file_type:
        messagebox.showerror("Erro", "Tipo de arquivo não suportado.")
        return
    if file_type == 'jsonschema':
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        rows = parse_json_schema(data)
        columns = COLUMNS
    elif file_type == 'jsonexample':
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        rows = parse_json_example(data)
        columns = COLUMNS
    elif file_type == '.xsd':
        columns, rows = parse_xsd(input_path)
    elif file_type == '.xml':
        rows = parse_xml_example(input_path)
        columns = COLUMNS
    else:
        messagebox.showerror("Erro", "Tipo de arquivo não suportado.")
        return
    issues = double_check(rows, columns)
    write_excel(rows, output_path, columns)
    msg = f"Excel gerado: {output_path}"
    if issues:
        msg += "\n\nDouble check report:\n" + "\n".join(issues)
    else:
        msg += "\n\nDouble check: Nenhum problema encontrado."
    messagebox.showinfo("Concluído", msg)

if __name__ == "__main__":
    main() 