import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# === CONFIGURATION ===
SOURCE_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\source"
TARGET_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\target"
OUTPUT_DIR = r"C:\Users\marlo\OneDrive - EDP\Documents\the-forge\mapping-creator\mapping-file"

# === UTILITY FUNCTIONS ===
def build_path_from_levels(row, level_indices, type_index=None, all_rows=None, row_idx=None):
    # Para cada coluna de nível, busca para cima se estiver vazia
    path_parts = []
    for idx in level_indices:
        val = row[idx]
        if (val is None or str(val).strip() == '') and all_rows is not None and row_idx is not None:
            # Busca para cima
            for up_idx in range(row_idx-1, -1, -1):
                up_val = all_rows[up_idx][idx]
                if up_val is not None and str(up_val).strip() != '':
                    val = up_val
                    break
        if val is not None and str(val).strip() != '':
            path_parts.append(str(val).strip())
    return '.'.join(path_parts)

def normalize_path(path):
    # Lowercase, trim, skip empty, and collapse multiple dots
    parts = [p.strip().lower() for p in path.split('.') if p.strip()]
    return '.'.join(parts)

def get_level_indices(headers):
    return [i for i, h in enumerate(headers) if h.strip().startswith('Element Level')]

def get_destination_field(norm_src_path, tgt_path_display, tgt_row, tgt_rows=None, tgt_level_indices=None, tgt_row_idx=None):
    # Se houver tgt_row, construa o path buscando valores não vazios acima para cada coluna 'Element Level X'
    if tgt_row and tgt_rows is not None and tgt_level_indices is not None and tgt_row_idx is not None:
        path_parts = []
        for col_idx in tgt_level_indices:
            val = tgt_row[col_idx]
            if val is None or str(val).strip() == '':
                # Busca para cima
                for up_idx in range(tgt_row_idx-1, -1, -1):
                    up_val = tgt_rows[up_idx][col_idx]
                    if up_val is not None and str(up_val).strip() != '':
                        val = up_val
                        break
            if val is not None and str(val).strip() != '':
                path_parts.append(str(val).strip())
        return '.'.join(path_parts)
    else:
        return ''

def merge_xlsx(source_path, target_path, output_path):
    src_wb = load_workbook(source_path)
    src_ws = src_wb.active
    src_headers = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]
    src_level_indices = get_level_indices(src_headers)
    src_type_index = src_headers.index('Type') if 'Type' in src_headers else None

    # Build source path map (normalized)
    src_rows = list(src_ws.iter_rows(min_row=2, values_only=True))
    src_path_map = {}
    for idx, row in enumerate(src_rows):
        path = build_path_from_levels(row, src_level_indices, src_type_index, all_rows=src_rows, row_idx=idx)
        norm_path = normalize_path(path)
        src_path_map[norm_path] = row

    # Prepare target
    tgt_rows = []
    tgt_headers = []
    tgt_level_indices = []
    tgt_type_index = None
    tgt_path_to_row = {}
    tgt_path_display = {}
    if target_path and os.path.exists(target_path):
        tgt_wb = load_workbook(target_path)
        tgt_ws = tgt_wb.active
        tgt_headers = [cell.value for cell in next(tgt_ws.iter_rows(min_row=1, max_row=1))]
        tgt_level_indices = get_level_indices(tgt_headers)
        tgt_type_index = tgt_headers.index('Type') if 'Type' in tgt_headers else None
        tgt_rows = list(tgt_ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(tgt_rows):
            tgt_path = build_path_from_levels(row, tgt_level_indices, tgt_type_index, all_rows=tgt_rows, row_idx=idx)
            norm_tgt_path = normalize_path(tgt_path)
            tgt_path_to_row[norm_tgt_path] = row
            tgt_path_display[norm_tgt_path] = tgt_path

    # Prepare output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Merged'
    out_headers = src_headers + ['Destination Field (Target Path)'] + tgt_headers
    out_ws.append(out_headers)

    # Apply header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in out_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # For each source row, find the matching target row by normalized path
    for norm_src_path, src_row in src_path_map.items():
        tgt_row = tgt_path_to_row.get(norm_src_path, None)
        dest_field = get_destination_field(norm_src_path, tgt_path_display, tgt_row, tgt_rows, tgt_level_indices, tgt_rows.index(tgt_row) if tgt_row else None)
        out_row = list(src_row) + [dest_field] + (list(tgt_row) if tgt_row else [''] * len(tgt_headers))
        out_ws.append(out_row)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    out_wb.save(output_path)
    print(f"Merged file saved to: {output_path}")

def get_base_filename(filename):
    # Remove sufixos comuns e extensão, e coloca em minúsculas
    name = os.path.splitext(filename)[0]
    name = re.sub(r'[_-](source|target)$', '', name, flags=re.IGNORECASE)
    return name.lower()

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    source_files = [f for f in os.listdir(SOURCE_DIR) if f.lower().endswith('.xlsx')]
    target_files = [f for f in os.listdir(TARGET_DIR) if f.lower().endswith('.xlsx')]
    # Cria um dicionário base_name -> filename para o target
    target_map = {get_base_filename(f): f for f in target_files}
    for filename in source_files:
        base_name = get_base_filename(filename)
        target_filename = target_map.get(base_name)
        if target_filename:
            target_path = os.path.join(TARGET_DIR, target_filename)
        else:
            target_path = None
        source_path = os.path.join(SOURCE_DIR, filename)
        output_path = os.path.join(OUTPUT_DIR, f"{os.path.splitext(filename)[0]}_merged.xlsx")
        print(f"Merging {filename} with {target_filename if target_filename else 'NO MATCH'} ...")
        merge_xlsx(source_path, target_path, output_path)

if __name__ == "__main__":
    main() 